import os
import re
import json
import math
import time
import base64
import io
import textwrap
import tempfile
import zipfile
import urllib.request
import urllib.parse
import unicodedata
import difflib
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
from html import escape as html_escape
from html.parser import HTMLParser
from xml.sax.saxutils import escape as xml_escape
from zoneinfo import ZoneInfo
import pandas as pd
import requests
import streamlit as st
import folium
from streamlit_folium import st_folium
from sqlalchemy import text

# =====================================================================
# CONFIGURAÇÕES DE TELA E RELÓGIO (VIRADA DE TURNO)
# =====================================================================
st.set_page_config(page_title="Aproar - Torre de Controle", page_icon="🚚", layout="wide")

FUSO_LOCAL = ZoneInfo("America/Fortaleza")
AGORA_REAL = datetime.now(FUSO_LOCAL)

if AGORA_REAL.hour >= 18:
    DATA_REF_ROTA_DATE = AGORA_REAL.date() + timedelta(days=1)
    DATA_REF_ROTA_STR = (AGORA_REAL + timedelta(days=1)).strftime("%d/%m/%Y")
else:
    DATA_REF_ROTA_DATE = AGORA_REAL.date()
    DATA_REF_ROTA_STR = AGORA_REAL.strftime("%d/%m/%Y")

DATA_HOJE_REAL_STR = AGORA_REAL.strftime("%d/%m/%Y")

# =====================================================================
# MOTOR DE BANCO DE DADOS NA NUVEM (POSTGRESQL / SUPABASE)
# =====================================================================
def get_conn():
    # Puxa a conexão automaticamente dos Secrets do Streamlit
    return st.connection("postgresql", type="sql")

def execute_db(query, params=None):
    if params is None: params = {}
    conn_db = get_conn()
    with conn_db.session as s:
        s.execute(text(query), params)
        s.commit()

def fetch_one(query, params=None):
    if params is None: params = {}
    conn_db = get_conn()
    with conn_db.session as s:
        return s.execute(text(query), params).fetchone()

def fetch_all(query, params=None):
    if params is None: params = {}
    conn_db = get_conn()
    with conn_db.session as s:
        return s.execute(text(query), params).fetchall()

def get_df(query, params=None):
    if params is None: params = {}
    conn_db = get_conn()
    with conn_db.session as s:
        res = s.execute(text(query), params)
        keys = res.keys()
        data = res.fetchall()
        return pd.DataFrame(data, columns=keys)

def save_df_to_db(df, table_name):
    conn_db = get_conn()
    with conn_db.session as s:
        s.execute(text(f"TRUNCATE TABLE {table_name} RESTART IDENTITY"))
        s.commit()
    df.to_sql(table_name, conn_db.engine, if_exists="append", index=False)

@st.cache_data(ttl=30, show_spinner=False)
def carregar_abastecimentos_df():
    return get_df("SELECT * FROM abastecimentos")

@st.cache_data(ttl=30, show_spinner=False)
def carregar_registro_km_df():
    return get_df("SELECT * FROM registro_km")

# =====================================================================
# DICIONÁRIO INTELIGENTE DE SINÔNIMOS E ERROS DE DIGITAÇÃO
# =====================================================================
DICIONARIO_SINONIMOS = {
    "DEPOSITO JP": "JP CONSTRUÇÃO",
    "DEPÓSITO JP": "JP CONSTRUÇÃO",
    "JP CONSTRUCOES": "JP CONSTRUÇÃO",
    "ELETRICA FORTALEZA": "ELÉTRICA FORTALEZA",
}

def remover_acentos(txt):
    if not txt: return ""
    return ''.join(c for c in unicodedata.normalize('NFD', str(txt)) if unicodedata.category(c) != 'Mn')

# --- DESIGN SYSTEM APROAR — TORRE + APP DO MOTORISTA -------------------------
def aplicar_estilo_customizado():
    """Aplica o mesmo sistema visual à Torre e à página /davi."""
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Sora:wght@500;600;700;800&display=swap');

        :root {
            --ap-bg: #070913;
            --ap-bg-soft: #0a0e1b;
            --ap-surface: #0f1526;
            --ap-surface-2: #131b30;
            --ap-surface-3: #18223a;
            --ap-line: rgba(148, 163, 184, .16);
            --ap-line-strong: rgba(96, 165, 250, .28);
            --ap-text: #f4f7fb;
            --ap-muted: #94a3b8;
            --ap-blue: #2563eb;
            --ap-blue-2: #3b82f6;
            --ap-green: #22c55e;
            --ap-amber: #f59e0b;
            --ap-red: #ef4444;
            --ap-radius-sm: 10px;
            --ap-radius: 16px;
            --ap-radius-lg: 22px;
            --ap-shadow: 0 18px 45px rgba(0, 0, 0, .28);
        }

        html, body, [class*="css"], .stMarkdown, .stText, p, div,
        h1, h2, h3, h4, h5, h6, label, button, input, textarea {
            font-family: 'Manrope', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
        }
        h1, h2, h3, h4, h5, h6, .aproar-title, .aproar-driver-greeting,
        [data-testid="stMetricValue"], .driver-kpi strong, .driver-stop-copy > strong {
            font-family:'Sora', 'Manrope', sans-serif !important;
        }
        html, body, [data-testid="stAppViewContainer"] { background: var(--ap-bg) !important; }
        [data-testid="stAppViewContainer"] {
            color: var(--ap-text);
            background-image:
                radial-gradient(circle at 84% 2%, rgba(37,99,235,.13), transparent 30rem),
                radial-gradient(circle at 8% 36%, rgba(14,165,233,.06), transparent 28rem),
                linear-gradient(180deg, #070913 0%, #080b15 100%) !important;
        }
        .main .block-container {
            max-width: 1500px;
            padding: 1.15rem 2.35rem 4rem;
        }
        p, li, label, .stCaption { color: #cbd5e1; }
        h1, h2, h3, h4, h5, h6 { color: var(--ap-text) !important; letter-spacing: -.025em; }
        a { color: #93c5fd; }

        span[data-testid="stIconMaterial"], .material-symbols-rounded, .material-symbols-outlined {
            font-family: 'Material Symbols Rounded', 'Material Symbols Outlined' !important;
            font-weight: normal !important; font-style: normal !important;
            letter-spacing: normal !important; text-transform: none !important;
            white-space: nowrap !important; word-wrap: normal !important;
        }

        [data-testid="stHeader"] {
            background: rgba(7, 9, 19, .72) !important;
            border-bottom: 1px solid rgba(148,163,184,.08);
            backdrop-filter: blur(18px);
        }
        [data-testid="stSidebar"] {
            background:
                radial-gradient(circle at 20% 0%, rgba(37,99,235,.13), transparent 18rem),
                #0a0e1b !important;
            border-right: 1px solid var(--ap-line) !important;
        }
        [data-testid="stSidebar"] .block-container { padding-top: 1.25rem; }
        [data-testid="stSidebarNav"] { display: none !important; }
        [data-stale="true"] { opacity: 1 !important; }
        .aproar-sidebar-brand {
            display:flex; align-items:center; gap:11px; margin:0 0 8px; padding:12px;
            border:1px solid var(--ap-line); border-radius:14px; background:rgba(19,27,48,.64);
        }
        .aproar-sidebar-brand span {
            display:grid; place-items:center; width:36px; height:36px; border-radius:11px;
            color:#fff; font-size:15px; font-weight:900; background:linear-gradient(145deg,#3b82f6,#1d4ed8);
            box-shadow:0 8px 18px rgba(37,99,235,.28);
        }
        .aproar-sidebar-brand strong { display:block; color:#f8fafc; font-size:13px; }
        .aproar-sidebar-brand small { display:block; margin-top:2px; color:#94a3b8; font-size:10px; }
        .aproar-logo-sidebar {
            display:block; width:94px; height:35px; flex:0 0 94px;
            object-fit:contain; object-position:left center;
        }

        /* Cabeçalho corporativo */
        .aproar-shell-header {
            display: flex; align-items: center; justify-content: space-between; gap: 22px;
            margin: 0 0 22px; padding: 18px 20px;
            background: linear-gradient(135deg, rgba(19,27,48,.92), rgba(10,14,27,.94));
            border: 1px solid var(--ap-line); border-radius: var(--ap-radius-lg);
            box-shadow: var(--ap-shadow); overflow: hidden; position: relative;
        }
        .aproar-shell-header::after {
            content: ''; position: absolute; inset: auto -70px -105px auto;
            width: 240px; height: 240px; border-radius: 50%;
            background: rgba(37,99,235,.14); filter: blur(15px); pointer-events: none;
        }
        .aproar-brand { display:flex; align-items:center; gap:14px; min-width:0; z-index:1; }
        .aproar-logo-main {
            display:block; width:146px; max-width:22vw; height:48px;
            object-fit:contain; object-position:left center;
            filter:drop-shadow(0 8px 18px rgba(0,0,0,.24));
        }
        .aproar-brand-mark {
            width: 46px; height: 46px; flex: 0 0 46px; display:grid; place-items:center;
            border-radius: 14px; background: linear-gradient(145deg, #3b82f6, #1d4ed8);
            box-shadow: 0 10px 24px rgba(37,99,235,.34); color:#fff;
            font-weight:900; font-size:19px; letter-spacing:-.08em;
        }
        .aproar-logo-fallback {
            display:flex !important; align-items:center; justify-content:flex-start;
            color:#f8fafc; font-family:'Sora', sans-serif !important;
            font-size:14px; font-weight:800; letter-spacing:.09em;
        }
        .aproar-eyebrow { color:#60a5fa; font-size:11px; font-weight:800; letter-spacing:.16em; text-transform:uppercase; }
        .aproar-title { color:var(--ap-text); font-size:21px; line-height:1.14; font-weight:700; margin-top:2px; letter-spacing:-.035em; }
        .aproar-subtitle { color:var(--ap-muted); font-size:12px; margin-top:4px; }
        .aproar-header-meta { display:flex; align-items:center; gap:10px; z-index:1; }
        .aproar-meta-chip {
            display:flex; align-items:center; gap:8px; min-height:38px; padding:0 12px;
            color:#cbd5e1; font-size:12px; font-weight:700; white-space:nowrap;
            background:rgba(255,255,255,.035); border:1px solid var(--ap-line); border-radius:999px;
        }
        .aproar-meta-chip.primary { color:#bfdbfe; border-color:rgba(96,165,250,.28); background:rgba(37,99,235,.11); }
        .aproar-dot { width:7px; height:7px; border-radius:50%; background:var(--ap-green); box-shadow:0 0 0 4px rgba(34,197,94,.12); }

        /* Navegação principal */
        [data-testid="stSegmentedControl"] { margin: 0 0 20px; }
        [data-testid="stSegmentedControl"] > div {
            gap: 5px !important; padding: 6px !important; overflow-x: auto;
            background: rgba(15,21,38,.86) !important; border: 1px solid var(--ap-line) !important;
            border-radius: 15px !important; box-shadow: 0 10px 28px rgba(0,0,0,.16);
        }
        [data-testid="stSegmentedControl"] button {
            min-height: 41px !important; border-radius: 10px !important; color: var(--ap-muted) !important;
            font-size: 12.5px !important; font-weight: 700 !important; letter-spacing:-.01em; white-space: nowrap;
        }
        [data-testid="stSegmentedControl"] button[aria-pressed="true"] {
            color: #fff !important; background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
            box-shadow: 0 7px 18px rgba(37,99,235,.28);
        }

        /* Botões */
        .stButton > button, .stDownloadButton > button, [data-testid="baseButton-secondary"] {
            min-height: 42px; border-radius: 11px !important; font-weight: 700 !important;
            border: 1px solid rgba(148,163,184,.22) !important;
            background: rgba(19,27,48,.82) !important; color: #e2e8f0 !important;
            transition: transform .16s ease, border-color .16s ease, box-shadow .16s ease !important;
        }
        .stButton > button:hover, .stDownloadButton > button:hover {
            transform: translateY(-1px); border-color: rgba(96,165,250,.55) !important;
            box-shadow: 0 8px 20px rgba(0,0,0,.20);
        }
        button[kind="primary"], [data-testid="baseButton-primary"] {
            min-height: 44px; padding: 10px 20px !important; border: 0 !important;
            border-radius: 11px !important; color: #fff !important; font-weight: 800 !important;
            background: linear-gradient(135deg, #2f74f5, #1d4ed8) !important;
            box-shadow: 0 10px 24px rgba(37,99,235,.28) !important;
        }
        button[kind="primary"]:hover { transform: translateY(-1px); box-shadow: 0 14px 28px rgba(37,99,235,.36) !important; }
        button:disabled { opacity:.48 !important; transform:none !important; box-shadow:none !important; }

        /* Cards, métricas e formulários */
        div[data-testid="stMetric"] {
            min-height: 112px; padding: 18px 19px; border-radius: var(--ap-radius) !important;
            background: linear-gradient(145deg, rgba(19,27,48,.92), rgba(12,17,31,.94)) !important;
            border: 1px solid var(--ap-line) !important; box-shadow: 0 12px 30px rgba(0,0,0,.18);
        }
        div[data-testid="stMetric"] label { color:var(--ap-muted) !important; font-weight:700; }
        div[data-testid="stMetric"] [data-testid="stMetricValue"] { color:var(--ap-text) !important; font-weight:800; letter-spacing:-.04em; }
        div[data-testid="stMetric"] [data-testid="stMetricDelta"] { font-weight:700; }
        div[data-testid="stForm"], div[data-testid="stVerticalBlockBorderWrapper"] {
            padding: 20px !important; border-radius: var(--ap-radius) !important;
            background: rgba(15,21,38,.80) !important; border: 1px solid var(--ap-line) !important;
            box-shadow: 0 12px 30px rgba(0,0,0,.14);
        }
        div[data-baseweb="input"] > div, div[data-baseweb="select"] > div,
        div[data-baseweb="textarea"] > div, [data-testid="stNumberInput"] > div > div {
            min-height: 44px; color:var(--ap-text) !important; border-radius:11px !important;
            background:#0d1322 !important; border:1px solid rgba(148,163,184,.22) !important;
            transition:border-color .16s ease, box-shadow .16s ease;
        }
        div[data-baseweb="input"] > div:focus-within, div[data-baseweb="select"] > div:focus-within,
        div[data-baseweb="textarea"] > div:focus-within {
            border-color:#3b82f6 !important; box-shadow:0 0 0 3px rgba(37,99,235,.14) !important;
        }
        [data-testid="stFileUploaderDropzone"] {
            min-height:120px; border-radius:14px !important; border:1.5px dashed rgba(96,165,250,.45) !important;
            background:rgba(37,99,235,.055) !important;
        }

        /* Tabelas, abas, expansores e mensagens */
        [data-testid="stDataFrame"], [data-testid="stTable"] {
            overflow:hidden; border-radius:14px !important; border:1px solid var(--ap-line) !important;
            background:#0d1322 !important; box-shadow:0 12px 28px rgba(0,0,0,.14);
        }
        .stTabs [data-baseweb="tab-list"] {
            gap:5px; padding:6px; border-radius:14px; border:1px solid var(--ap-line);
            background:rgba(15,21,38,.82); overflow-x:auto;
        }
        .stTabs [data-baseweb="tab"] { min-height:40px; padding:8px 15px; color:var(--ap-muted); font-weight:700; border-radius:9px; }
        .stTabs [aria-selected="true"] { color:#fff !important; background:rgba(37,99,235,.9) !important; }
        [data-testid="stExpander"] {
            overflow:hidden; border:1px solid var(--ap-line) !important; border-radius:14px !important;
            background:rgba(15,21,38,.74) !important;
        }
        [data-testid="stAlert"] { border-radius:13px !important; border:1px solid var(--ap-line) !important; }
        hr { border-color:var(--ap-line) !important; }
        iframe[title*="streamlit_folium"], div[data-testid="stIFrame"] iframe {
            width:100% !important; max-width:100% !important; border-radius:16px;
        }

        /* ETA compartilhado */
        .aproar-eta-card {
            display:grid; grid-template-columns:1fr auto 1fr; align-items:center; gap:18px;
            margin: 10px 0 22px; padding:17px 19px; border-radius:16px;
            background:linear-gradient(145deg,rgba(19,27,48,.94),rgba(12,17,31,.96));
            border:1px solid var(--ap-line); box-shadow:0 12px 30px rgba(0,0,0,.18);
        }
        .aproar-eta-label { color:var(--ap-muted); font-size:11px; font-weight:800; letter-spacing:.08em; text-transform:uppercase; }
        .aproar-eta-value { color:var(--ap-text); font-size:22px; font-weight:800; margin-top:3px; letter-spacing:-.04em; }
        .aproar-eta-side.right { text-align:right; }
        .aproar-eta-route { color:#60a5fa; font-size:20px; }

        ::-webkit-scrollbar { width:9px; height:9px; }
        ::-webkit-scrollbar-track { background:#070913; }
        ::-webkit-scrollbar-thumb { background:#283752; border:2px solid #070913; border-radius:999px; }
        ::-webkit-scrollbar-thumb:hover { background:#3b82f6; }

        @media (max-width: 900px) {
            .main .block-container { padding: .9rem 1rem 5rem; }
            .aproar-shell-header { align-items:flex-start; padding:16px; }
            .aproar-header-meta { flex-wrap:wrap; justify-content:flex-end; }
        }
        @media (max-width: 640px) {
            .main .block-container { padding:.7rem .78rem 5.5rem; }
            .aproar-shell-header { border-radius:18px; margin-bottom:14px; }
            .aproar-brand-mark { width:42px; height:42px; flex-basis:42px; }
            .aproar-logo-main { width:112px; max-width:34vw; height:42px; }
            .aproar-title { font-size:18px; }
            .aproar-subtitle { display:none; }
            .aproar-header-meta { display:none; }
            .aproar-eta-card { grid-template-columns:1fr auto 1fr; gap:9px; padding:14px; }
            .aproar-eta-route { font-size:16px; }
            .aproar-eta-value { font-size:19px; }
            div[data-testid="stForm"], div[data-testid="stVerticalBlockBorderWrapper"] { padding:15px !important; }
        }
    </style>
    """, unsafe_allow_html=True)


def _html_logo_aproar(classe_css):
    """Carrega logo.png ao lado do app.py e devolve uma imagem pronta para o HTML."""
    try:
        caminho_logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        with open(caminho_logo, "rb") as arquivo_logo:
            conteudo_logo = base64.b64encode(arquivo_logo.read()).decode("ascii")
        return (
            f'<img class="{html_escape(classe_css, quote=True)}" '
            f'src="data:image/png;base64,{conteudo_logo}" alt="Aproar">'
        )
    except (OSError, ValueError):
        return f'<div class="{html_escape(classe_css, quote=True)} aproar-logo-fallback">APROAR</div>'


def renderizar_cabecalho_torre():
    logo = _html_logo_aproar("aproar-logo-main")
    st.markdown(f"""
        <header class="aproar-shell-header">
            <div class="aproar-brand">
                {logo}
                <div>
                    <div class="aproar-eyebrow">CENTRAL LOGÍSTICA</div>
                    <div class="aproar-title">Torre de Controle</div>
                    <div class="aproar-subtitle">Planejamento, monitoramento e execução em uma única operação</div>
                </div>
            </div>
            <div class="aproar-header-meta">
                <div class="aproar-meta-chip primary">PLANEJAMENTO • {DATA_REF_ROTA_STR} • MOTOR V7</div>
                <div class="aproar-meta-chip"><span class="aproar-dot"></span> OPERAÇÃO ATIVA</div>
            </div>
        </header>
    """, unsafe_allow_html=True)


def renderizar_cabecalho_motorista():
    hora = AGORA_REAL.hour
    saudacao = "Bom dia" if hora < 12 else "Boa tarde" if hora < 18 else "Boa noite"
    logo = _html_logo_aproar("aproar-logo-driver")
    st.markdown(f"""
        <header class="aproar-driver-header">
            <div class="aproar-driver-topline">
                <div class="aproar-driver-brand">{logo}</div>
                <div class="aproar-driver-live"><i></i> ROTA ATIVA</div>
            </div>
            <div class="aproar-driver-greeting">{saudacao}, <strong>Davi</strong></div>
            <div class="aproar-driver-date">Rota oficial • {DATA_REF_ROTA_STR}</div>
        </header>
    """, unsafe_allow_html=True)


def renderizar_resumo_motorista(route_steps, total_km, final_dyn_min, enderecos=None, locais=None):
    paradas = [
        (indice, etapa) for indice, etapa in enumerate(route_steps or [])
        if etapa.get("type") == "stop" and not (indice == 0)
    ]
    concluidas = sum(1 for _, etapa in paradas if etapa.get("is_concluded"))
    proxima = next(((indice, etapa) for indice, etapa in paradas if not etapa.get("is_concluded")), None)
    destino = str(proxima[1].get("destino", "Rota concluída")) if proxima else "Rota concluída"
    chegada = str(proxima[1].get("dyn_chegada", "--:--")) if proxima else format_mins_to_time(final_dyn_min)
    numero = (paradas.index(proxima) + 1) if proxima in paradas else len(paradas)
    progresso = int(round((concluidas / max(1, len(paradas))) * 100))
    acoes_proxima = [str(acao) for acao, _tarefa in (proxima[1].get("actions", []) if proxima else [])]
    tem_coleta = "COLETAR" in acoes_proxima
    tem_entrega = "ENTREGAR" in acoes_proxima
    tipo_proxima = (
        "COLETAR E ENTREGAR" if tem_coleta and tem_entrega
        else "ENTREGAR" if tem_entrega
        else "COLETAR" if tem_coleta
        else "SEGUIR ROTEIRO"
    )
    link_gps = ""
    endereco = str((enderecos or {}).get(destino, "") or "")
    coordenadas = (locais or {}).get(destino, [None, None])
    if endereco.startswith("http"):
        link_gps = endereco
    elif endereco:
        link_gps = f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(endereco)}"
    elif len(coordenadas) >= 2 and coordenadas[0] is not None:
        link_gps = f"https://www.google.com/maps/dir/?api=1&destination={coordenadas[0]},{coordenadas[1]}"
    botao_gps = (
        f'<a class="driver-gps-link" href="{html_escape(link_gps, quote=True)}" target="_blank" rel="noopener">ABRIR GPS&nbsp; ↗</a>'
        if link_gps else ""
    )
    st.markdown(f"""
        <section class="aproar-driver-summary" id="rota">
            <div class="driver-next-stop">
                <div class="driver-stop-index">{numero}</div>
                <div class="driver-stop-copy">
                    <span>PRÓXIMA PARADA</span>
                    <strong>{html_escape(destino)}</strong>
                    <small><b>{tipo_proxima}</b> • chegada {html_escape(chegada)}</small>
                </div>
                {botao_gps}
            </div>
            <div class="driver-simple-progress">
                <span><b>{concluidas}/{len(paradas)}</b> paradas concluídas</span>
                <span>{float(total_km or 0):.1f} km • término {format_mins_to_time(final_dyn_min)}</span>
            </div>
            <div class="driver-progress"><span style="width:{progresso}%"></span></div>
        </section>
    """, unsafe_allow_html=True)


aplicar_estilo_customizado()

def fragmento_independente(func):
    """Mantém compatibilidade e evita recarregar o aplicativo inteiro por um formulário."""
    return st.fragment(func) if hasattr(st, "fragment") else func

# =====================================================================
# FUNÇÕES DE FORMATAÇÃO E ETA DINÂMICO Waze
# =====================================================================
def parse_time_to_mins(time_str):
    if not time_str: return 0
    try:
        h, m = map(int, time_str.split(':'))
        return h * 60 + m
    except: return 0

def format_time(minutes):
    total = int(round(minutes))
    return f"{total // 60:02d}:{total % 60:02d}"

def format_mins_to_time(mins):
    """Formata minutos acumulados sem exibir horários impossíveis como 25:03."""
    try:
        total = max(0, int(round(float(mins))))
    except (TypeError, ValueError):
        total = 0
    dias, resto = divmod(total, 24 * 60)
    hora, minuto = divmod(resto, 60)
    horario = f"{hora:02d}:{minuto:02d}"
    if dias == 1:
        return f"{horario} (+1 dia)"
    if dias > 1:
        return f"{horario} (+{dias} dias)"
    return horario

def plural_pt(qtd, singular, plural=None):
    """Retorna singular/plural sem usar formas como ``demanda(s)`` na interface."""
    try:
        numero = int(qtd)
    except (TypeError, ValueError):
        numero = qtd
    return singular if numero == 1 else (plural or f"{singular}s")


def _limpar_texto_relatorio(valor):
    if valor is None:
        return ""
    if isinstance(valor, (dict, list, tuple, set)):
        valor = json.dumps(valor, ensure_ascii=False, default=str)
    texto = str(valor)
    # Imagens anexadas pelo Trello não agregam ao relatório e deixavam URLs
    # enormes no meio dos materiais.
    texto = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", texto)
    texto = re.sub(r"\s*\|\s*", " • ", texto)
    texto = re.sub(r"[\t\r\n]+", " ", texto)
    texto = re.sub(r"\s{2,}", " ", texto).strip(" •")
    return texto

def _normalizar_tabelas_relatorio(dados):
    """Entrega relatórios limpos, com rótulos amigáveis e sem campos técnicos."""
    itens = [("Dados", dados)] if isinstance(dados, pd.DataFrame) else list((dados or {}).items())
    colunas_tecnicas = {"id", "data_dt", "json_route", "json_locais", "json_geometria", "json_enderecos"}
    nomes_amigaveis = {
        "Urgência": "Prazo",
        "Peso": "Prioridade (1-5)",
        "Tempo_Coleta": "Tempo de coleta (min)",
        "Tempo_Entrega": "Tempo de entrega (min)",
        "obs": "Observação",
        "data": "Data",
        "km": "km",
        "litros": "Litros",
        "valor_litro": "Valor por litro (R$)",
        "manutencao": "Manutenção (R$)",
        "veiculo": "Veículo",
        "apelido": "Local",
        "endereco": "Endereço",
        "lat": "Latitude",
        "lon": "Longitude",
        "🟢 Início da Rota (Hoje)": "Início da rota (hoje)",
        "Última atualização": "Última atualização",
        "Velocidade (km/h)": "Velocidade (km/h)",
        "obra": "Obra",
        "origem": "Origem",
        "destino": "Destino",
        "materiais": "Materiais",
        "data_conclusao": "Data da conclusão",
        "hora_conclusao": "Hora da conclusão",
    }
    tabelas = []
    for nome, tabela in itens:
        if tabela is None:
            continue
        df = tabela.copy() if isinstance(tabela, pd.DataFrame) else pd.DataFrame(tabela)
        remover = [coluna for coluna in df.columns if str(coluna).strip().lower() in colunas_tecnicas]
        df = df.drop(columns=remover, errors="ignore").rename(columns=nomes_amigaveis)
        for coluna in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[coluna]):
                df[coluna] = df[coluna].dt.strftime("%d/%m/%Y %H:%M").fillna("")
            elif df[coluna].dtype == "object":
                df[coluna] = df[coluna].map(_limpar_texto_relatorio)
        df = df.dropna(axis=1, how="all")
        df = df.loc[:, [coluna for coluna in df.columns if not (df[coluna].astype(str).str.strip() == "").all()]]
        tabelas.append((str(nome), df.fillna("")))
    return tabelas or [("Dados", pd.DataFrame({"Informação": ["Nenhum registro disponível."]}))]

def _criar_resumo_analitico_relatorio(titulo, tabelas):
    """Cria indicadores úteis de acordo com o assunto e os campos de cada relatório."""
    def normalizar(valor):
        return re.sub(r"[^a-z0-9]+", " ", remover_acentos(str(valor or "")).lower()).strip()

    def localizar_coluna(df, *candidatos):
        mapa = {coluna: normalizar(coluna) for coluna in df.columns}
        procurados = [normalizar(candidato) for candidato in candidatos]
        for procurado in procurados:
            for coluna, nome in mapa.items():
                if nome == procurado:
                    return coluna
        for procurado in procurados:
            for coluna, nome in mapa.items():
                if procurado and procurado in nome:
                    return coluna
        return None

    def numeros(df, *candidatos):
        coluna = localizar_coluna(df, *candidatos)
        if not coluna:
            return pd.Series(dtype="float64")
        if pd.api.types.is_numeric_dtype(df[coluna]):
            return pd.to_numeric(df[coluna], errors="coerce")

        def converter(valor):
            texto = re.sub(r"[^0-9,.-]", "", str(valor or ""))
            if not texto or texto in {"-", ".", ","}:
                return None
            if "," in texto:
                texto = texto.replace(".", "").replace(",", ".")
            try:
                return float(texto)
            except (TypeError, ValueError):
                return None
        return df[coluna].map(converter).astype("float64")

    def minutos_horario(valor):
        encontrado = re.search(r"(?<!\d)(\d{1,2}):(\d{2})(?!\d)", str(valor or ""))
        if not encontrado:
            return None
        hora, minuto = map(int, encontrado.groups())
        return hora * 60 + minuto if hora < 24 and minuto < 60 else None

    def horario_medio(valores):
        minutos = [minuto for minuto in (minutos_horario(valor) for valor in valores) if minuto is not None]
        if not minutos:
            return None
        media = int(round(sum(minutos) / len(minutos)))
        return f"{media // 60:02d}:{media % 60:02d}"

    def duracao_media(chegadas, saidas):
        duracoes = []
        for chegada, saida in zip(chegadas, saidas):
            inicio, fim = minutos_horario(chegada), minutos_horario(saida)
            if inicio is None or fim is None:
                continue
            if fim < inicio:
                fim += 24 * 60
            if 0 <= fim - inicio <= 12 * 60:
                duracoes.append(fim - inicio)
        return sum(duracoes) / len(duracoes) if duracoes else None

    def numero_br(valor, casas=1):
        return f"{float(valor):,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def moeda_br(valor):
        return f"R$ {numero_br(valor, 2)}"

    indicadores = []
    def adicionar(indicador, resultado, leitura=""):
        if resultado is None or (isinstance(resultado, float) and math.isnan(resultado)):
            return
        indicadores.append({"Indicador": indicador, "Resultado": resultado, "Leitura para análise": leitura})

    por_nome = {normalizar(nome): df for nome, df in tabelas}
    titulo_norm = normalizar(titulo)
    tabelas_preenchidas = [(nome, df) for nome, df in tabelas if not df.empty]
    principal = max(tabelas_preenchidas, key=lambda item: len(item[1]))[1] if tabelas_preenchidas else pd.DataFrame()

    if "rastreador" in titulo_norm:
        velocidades = numeros(principal, "Velocidade (km/h)").dropna()
        em_movimento = velocidades[velocidades > 0]
        total = len(principal)
        adicionar("Veículos monitorados", total, "Quantidade de veículos com leitura atual.")
        adicionar("Veículos em movimento", len(em_movimento), f"{numero_br(100 * len(em_movimento) / total, 1)}% da frota monitorada." if total else "")
        adicionar("Veículos parados", total - len(em_movimento), "Veículos com velocidade igual a zero na última leitura.")
        adicionar("Velocidade média geral", f"{numero_br(velocidades.mean(), 1)} km/h" if len(velocidades) else None, "Inclui veículos parados.")
        adicionar("Velocidade média em movimento", f"{numero_br(em_movimento.mean(), 1)} km/h" if len(em_movimento) else None, "Considera somente velocidades acima de zero.")
        inicio_col = localizar_coluna(principal, "Início da rota (hoje)")
        adicionar("Horário médio de saída", horario_medio(principal[inicio_col]) if inicio_col else None, "Média das saídas registradas no dia.")

    elif "demandas ativas" in titulo_norm:
        total = len(principal)
        status_col = localizar_coluna(principal, "Status da rota", "Status")
        status = principal[status_col].astype(str) if status_col else pd.Series([""] * total)
        concluidas = int(status.str.contains(r"entregue|conclu", case=False, regex=True).sum())
        pendentes = total - concluidas
        adicionar("Demandas analisadas", total, "Volume total presente no relatório.")
        adicionar("Taxa de conclusão", f"{numero_br(100 * concluidas / total, 1)}%" if total else "0,0%", "Boa aderência" if total and concluidas / total >= 0.9 else "Atenção às demandas ainda pendentes.")
        adicionar("Demandas concluídas", concluidas)
        adicionar("Demandas pendentes", pendentes)
        prioridade = numeros(principal, "Prioridade (1-5)").dropna()
        adicionar("Prioridade média", numero_br(prioridade.mean(), 2) if len(prioridade) else None, "Quanto maior, mais urgente é o conjunto de demandas.")
        coleta = numeros(principal, "Tempo de coleta (min)").dropna()
        entrega = numeros(principal, "Tempo de entrega (min)").dropna()
        adicionar("Tempo médio de coleta", f"{numero_br(coleta.mean(), 1)} min" if len(coleta) else None)
        adicionar("Tempo médio de entrega", f"{numero_br(entrega.mean(), 1)} min" if len(entrega) else None)
        if len(coleta) and len(entrega):
            adicionar("Tempo médio total por demanda", f"{numero_br(coleta.mean() + entrega.mean(), 1)} min", "Soma das médias de coleta e entrega.")
        prazo_col = localizar_coluna(principal, "Prazo")
        if prazo_col:
            prazos = principal[prazo_col].astype(str)
            adicionar("Demandas vencidas/atrasadas", int(prazos.str.contains(r"vencid|atrasad", case=False, regex=True).sum()), "Prioridade de regularização.")

    elif "roteiro do davi" in titulo_norm:
        resumo = next((df for nome, df in tabelas if normalizar(nome) == "resumo"), pd.DataFrame())
        paradas = next((df for nome, df in tabelas if "paradas" in normalizar(nome)), principal)
        parada_col = localizar_coluna(paradas, "Parada")
        acao_col = localizar_coluna(paradas, "Ação")
        status_col = localizar_coluna(paradas, "Status")
        paradas_operacionais = set()
        if parada_col:
            paradas_operacionais = {str(valor) for valor in paradas[parada_col] if str(valor).strip().isdigit()}
        entregas = paradas[paradas[acao_col].astype(str).str.contains("ENTREGAR", case=False, na=False)].copy() if acao_col else paradas.copy()
        status = entregas[status_col].astype(str) if status_col else pd.Series([""] * len(entregas))
        concluidas = int(status.str.contains(r"conclu|entregue", case=False, regex=True).sum())
        total_demandas = len(entregas)
        km = numeros(resumo, "Distância planejada (km)").dropna()
        km_total = float(km.iloc[0]) if len(km) else None
        adicionar("Paradas operacionais", len(paradas_operacionais), "Não inclui almoço nem retorno à base.")
        adicionar("Demandas previstas", total_demandas)
        adicionar("Taxa de conclusão da rota", f"{numero_br(100 * concluidas / total_demandas, 1)}%" if total_demandas else "0,0%", "Percentual de entregas previstas que tiveram baixa.")
        adicionar("Demandas concluídas", concluidas)
        adicionar("Demandas pendentes", total_demandas - concluidas)
        adicionar("Distância planejada", f"{numero_br(km_total, 1)} km" if km_total is not None else None)
        adicionar("Distância média por parada", f"{numero_br(km_total / len(paradas_operacionais), 1)} km" if km_total is not None and paradas_operacionais else None, "Indicador de dispersão da rota.")
        if not resumo.empty:
            inicio_col = localizar_coluna(resumo, "Início")
            fim_col = localizar_coluna(resumo, "Término previsto")
            inicio = minutos_horario(resumo.iloc[0][inicio_col]) if inicio_col else None
            fim = minutos_horario(resumo.iloc[0][fim_col]) if fim_col else None
            if inicio is not None and fim is not None:
                adicionar("Duração planejada da rota", f"{numero_br((fim - inicio) / 60, 1)} h", "Tempo estimado entre saída e retorno.")

    elif "fechamento individualizado" in titulo_norm:
        resumo = next((df for nome, df in tabelas if normalizar(nome) == "resumo"), principal)
        km = numeros(resumo, "KM").fillna(0).sum()
        combustivel = numeros(resumo, "Combustível (R$)").fillna(0).sum()
        manutencao = numeros(resumo, "Manutenção (R$)").fillna(0).sum()
        custo_total = numeros(resumo, "Custo total (R$)").fillna(0).sum()
        custo_km = custo_total / km if km > 0 else None
        adicionar("Quilometragem total", f"{numero_br(km, 1)} km")
        adicionar("Custo total da frota", moeda_br(custo_total))
        adicionar("Custo médio por km", moeda_br(custo_km) if custo_km is not None else None, "Dentro da referência de R$ 1,50/km." if custo_km is not None and custo_km <= 1.5 else "Acima da referência de R$ 1,50/km." if custo_km is not None else "")
        adicionar("Total de combustível", moeda_br(combustivel), f"{numero_br(100 * combustivel / custo_total, 1)}% do custo total." if custo_total else "")
        adicionar("Total de manutenção", moeda_br(manutencao), f"{numero_br(100 * manutencao / custo_total, 1)}% do custo total." if custo_total else "")
        gastos = pd.concat([df for nome, df in tabelas if normalizar(nome).startswith("gastos") and not df.empty], ignore_index=True) if any(normalizar(nome).startswith("gastos") and not df.empty for nome, df in tabelas) else pd.DataFrame()
        litros = numeros(gastos, "Litros").dropna()
        adicionar("Litros médios por abastecimento", f"{numero_br(litros.mean(), 1)} L" if len(litros) else None)
        adicionar("Custo médio por lançamento", moeda_br(numeros(gastos, "Total (R$)").dropna().mean()) if len(numeros(gastos, "Total (R$)").dropna()) else None)

    elif "registros e historico da frota" in titulo_norm:
        inicios = next((df for nome, df in tabelas if "inicios de rota" in normalizar(nome)), pd.DataFrame())
        paradas = next((df for nome, df in tabelas if "paradas rastreadas" in normalizar(nome)), pd.DataFrame())
        abastecimentos = next((df for nome, df in tabelas if "abastecimentos" in normalizar(nome)), pd.DataFrame())
        quilometragens = next((df for nome, df in tabelas if "quilometragens" in normalizar(nome)), pd.DataFrame())
        saida_col = localizar_coluna(inicios, "Hora de saída", "Hora Saída") if not inicios.empty else None
        adicionar("Saídas registradas", len(inicios))
        adicionar("Horário médio de saída", horario_medio(inicios[saida_col]) if saida_col else None, "Permite acompanhar a aderência ao início previsto.")
        adicionar("Paradas rastreadas", len(paradas))
        chegada_col = localizar_coluna(paradas, "Chegada") if not paradas.empty else None
        saida_parada_col = localizar_coluna(paradas, "Saída") if not paradas.empty else None
        media_parada = duracao_media(paradas[chegada_col], paradas[saida_parada_col]) if chegada_col and saida_parada_col else None
        adicionar("Tempo médio por parada", f"{numero_br(media_parada, 1)} min" if media_parada is not None else None, "Considera paradas com chegada e saída registradas.")
        locais_col = localizar_coluna(paradas, "Local") if not paradas.empty else None
        adicionar("Locais distintos visitados", int(paradas[locais_col].astype(str).nunique()) if locais_col else None)
        litros = numeros(abastecimentos, "Litros").dropna()
        valor_litro = numeros(abastecimentos, "Valor por litro (R$)", "valor_litro").dropna()
        manutencao = numeros(abastecimentos, "Manutenção (R$)", "manutencao").fillna(0)
        combustivel_total = float((numeros(abastecimentos, "Litros").fillna(0) * numeros(abastecimentos, "Valor por litro (R$)", "valor_litro").fillna(0)).sum()) if not abastecimentos.empty else 0
        adicionar("Litros médios por abastecimento", f"{numero_br(litros.mean(), 1)} L" if len(litros) else None)
        adicionar("Preço médio do litro", moeda_br(valor_litro.mean()) if len(valor_litro) else None)
        adicionar("Custo total registrado", moeda_br(combustivel_total + manutencao.sum()), "Soma do combustível calculado e das manutenções registradas.")
        kms = numeros(quilometragens, "KM").dropna()
        adicionar("Quilometragem total registrada", f"{numero_br(kms.sum(), 1)} km" if len(kms) else None)
        adicionar("Quilometragem média por registro", f"{numero_br(kms.mean(), 1)} km" if len(kms) else None)

    elif "entregas concluidas" in titulo_norm:
        total = len(principal)
        obra_col = localizar_coluna(principal, "Obra")
        destino_col = localizar_coluna(principal, "Destino")
        hora_col = localizar_coluna(principal, "Hora da conclusão", "hora_conclusao")
        obras = int(principal[obra_col].astype(str).nunique()) if obra_col else 0
        destinos = int(principal[destino_col].astype(str).nunique()) if destino_col else 0
        adicionar("Entregas concluídas", total)
        adicionar("Obras atendidas", obras)
        adicionar("Destinos atendidos", destinos)
        adicionar("Média de entregas por obra", numero_br(total / obras, 2) if obras else None, "Ajuda a identificar concentração de atendimento.")
        adicionar("Média de entregas por destino", numero_br(total / destinos, 2) if destinos else None)
        adicionar("Horário médio das conclusões", horario_medio(principal[hora_col]) if hora_col else None)
        if destino_col and total:
            frequencias = principal[destino_col].astype(str).replace("", pd.NA).dropna().value_counts()
            if len(frequencias):
                adicionar("Destino com mais entregas", f"{frequencias.index[0]} ({int(frequencias.iloc[0])})", "Local de maior concentração no período.")

    if not indicadores:
        adicionar("Registros analisados", sum(len(df) for _, df in tabelas), "Total de linhas incluídas no relatório.")
        numericas = principal.select_dtypes(include="number") if not principal.empty else pd.DataFrame()
        for coluna in numericas.columns[:4]:
            serie = pd.to_numeric(numericas[coluna], errors="coerce").dropna()
            adicionar(f"Média de {coluna}", numero_br(serie.mean(), 2) if len(serie) else None)

    return pd.DataFrame(indicadores, columns=["Indicador", "Resultado", "Leitura para análise"])

def _dados_grafico_resumo(df):
    """Seleciona apenas comparações com unidades compatíveis e leitura gerencial clara."""
    if df is None or df.empty or "Indicador" not in df.columns or "Resultado" not in df.columns:
        return None, []

    def normalizar(valor):
        return re.sub(r"[^a-z0-9]+", " ", remover_acentos(str(valor or "")).lower()).strip()

    def valor_numerico(valor):
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            return float(valor) if math.isfinite(float(valor)) else None
        texto = str(valor or "").strip()
        encontrado = re.search(r"-?[0-9][0-9.]*([,][0-9]+)?", texto)
        if not encontrado:
            return None
        numero = encontrado.group(0)
        if "," in numero:
            numero = numero.replace(".", "").replace(",", ".")
        try:
            return float(numero)
        except (TypeError, ValueError):
            return None

    valores = {
        normalizar(linha["Indicador"]): (str(linha["Indicador"]), valor_numerico(linha["Resultado"]))
        for _, linha in df.iterrows()
    }
    comparacoes = [
        ("Demandas concluídas", "Demandas pendentes", "Conclusão das demandas"),
        ("Veículos em movimento", "Veículos parados", "Situação da frota"),
        ("Total de combustível", "Total de manutenção", "Composição dos custos"),
        ("Saídas registradas", "Paradas rastreadas", "Atividade operacional registrada"),
        ("Obras atendidas", "Destinos atendidos", "Cobertura das entregas"),
    ]
    for primeiro, segundo, titulo in comparacoes:
        chave_a, chave_b = normalizar(primeiro), normalizar(segundo)
        if chave_a not in valores or chave_b not in valores:
            continue
        itens = [valores[chave_a], valores[chave_b]]
        if all(valor is not None and valor >= 0 for _, valor in itens) and sum(valor for _, valor in itens) > 0:
            return titulo, itens
    return None, []

def _organizar_secoes_relatorio(titulo, tabelas):
    """Separa o resumo e escolhe a tabela operacional principal de cada relatório."""
    def normalizar(valor):
        return re.sub(r"[^a-z0-9]+", " ", remover_acentos(str(valor or "")).lower()).strip()

    resumo = next((df for nome, df in tabelas if normalizar(nome) == "resumo analitico"), pd.DataFrame())
    dados = [(nome, df) for nome, df in tabelas if normalizar(nome) != "resumo analitico"]
    if not dados:
        return resumo, [], None

    titulo_norm = normalizar(titulo)
    indice_principal = None
    preferencias = []
    if "roteiro" in titulo_norm:
        # No relatório da rota, a leitura principal é a sequência operacional.
        # O detalhamento por demanda vem depois, em uma seção separada.
        preferencias = ["resumo da rota", "ordem da rota", "paradas e demandas", "paradas"]
    elif "fechamento" in titulo_norm:
        preferencias = ["resumo", "gastos"]
    elif "registros e historico" in titulo_norm:
        preferencias = ["paradas rastreadas", "abastecimentos e manutencao"]
    for preferencia in preferencias:
        indice_principal = next((i for i, (nome, _) in enumerate(dados) if preferencia in normalizar(nome)), None)
        if indice_principal is not None:
            break
    if indice_principal is None:
        indice_principal = max(range(len(dados)), key=lambda i: len(dados[i][1]))

    principal = dados[indice_principal]
    ordenadas = [principal] + [item for i, item in enumerate(dados) if i != indice_principal]
    return resumo, ordenadas, principal[0]

def _colunas_relevantes_pdf(df, limite=9):
    """Mantém no PDF as colunas de maior utilidade operacional quando a tabela é muito larga."""
    if df is None or len(df.columns) <= limite:
        return df

    def pontuacao(coluna):
        nome = remover_acentos(str(coluna)).lower()
        regras = [
            ("status", 120), ("situacao", 118), ("data", 115), ("placa", 112),
            ("veiculo", 110), ("obra", 108), ("parada", 106), ("acao", 104),
            ("origem", 102), ("destino", 101), ("local", 100), ("material", 98),
            ("descricao", 96), ("prazo", 94), ("prioridade", 92), ("urgencia", 92),
            ("tempo", 90), ("hora", 89), ("km", 88), ("distancia", 87),
            ("custo", 86), ("valor", 85), ("litro", 84), ("manutencao", 83),
            ("endereco", 80), ("estrategia", 78), ("responsavel", 76),
        ]
        return max((pontos for chave, pontos in regras if chave in nome), default=50)

    posicoes = {coluna: indice for indice, coluna in enumerate(df.columns)}
    escolhidas = sorted(df.columns, key=lambda coluna: (-pontuacao(coluna), posicoes[coluna]))[:limite]
    escolhidas = sorted(escolhidas, key=lambda coluna: posicoes[coluna])
    return df.loc[:, escolhidas]

def _criar_csv_relatorio(tabelas):
    secoes = []
    for nome, df in tabelas:
        tabela = df.copy()
        if "Seção" in tabela.columns:
            tabela = tabela.rename(columns={"Seção": "Seção original"})
        tabela.insert(0, "Seção", nome)
        secoes.append(tabela)
    consolidado = pd.concat(secoes, ignore_index=True, sort=False).fillna("")
    return consolidado.to_csv(index=False, sep=";", decimal=",", lineterminator="\n").encode("utf-8-sig")

def _nome_aba_excel(nome, usados):
    nome_limpo = re.sub(r"[\\/*?:\[\]]", " ", str(nome)).strip() or "Dados"
    nome_limpo = nome_limpo[:31]
    candidato, contador = nome_limpo, 2
    while candidato.lower() in usados:
        sufixo = f" {contador}"
        candidato = f"{nome_limpo[:31-len(sufixo)]}{sufixo}"
        contador += 1
    usados.add(candidato.lower())
    return candidato

def _criar_xlsx_basico(tabelas):
    """Fallback XLSX feito apenas com a biblioteca padrão do Python."""
    def coluna_excel(indice):
        texto = ""
        while indice:
            indice, resto = divmod(indice - 1, 26)
            texto = chr(65 + resto) + texto
        return texto

    def celula_xml(referencia, valor, estilo=0):
        estilo_xml = f' s="{estilo}"' if estilo else ""
        if isinstance(valor, bool):
            return f'<c r="{referencia}" t="b"{estilo_xml}><v>{1 if valor else 0}</v></c>'
        if isinstance(valor, (int, float)) and not isinstance(valor, bool) and math.isfinite(float(valor)):
            return f'<c r="{referencia}"{estilo_xml}><v>{valor}</v></c>'
        texto = xml_escape(str(valor), {'"': '&quot;'})
        return f'<c r="{referencia}" t="inlineStr"{estilo_xml}><is><t xml:space="preserve">{texto}</t></is></c>'

    usados, abas = set(), []
    for nome, df in tabelas:
        nome_aba = _nome_aba_excel(nome, usados)
        linhas = [list(map(str, df.columns))] + df.values.tolist()
        linhas_xml = []
        for numero_linha, valores in enumerate(linhas, start=1):
            celulas = "".join(
                celula_xml(f"{coluna_excel(indice)}{numero_linha}", valor, 1 if numero_linha == 1 else 0)
                for indice, valor in enumerate(valores, start=1)
            )
            linhas_xml.append(f'<row r="{numero_linha}">{celulas}</row>')
        ultima_coluna = coluna_excel(max(1, len(df.columns)))
        ultima_linha = max(1, len(linhas))
        larguras = "".join(
            f'<col min="{i}" max="{i}" width="{min(45, max(12, len(str(col)) + 3))}" customWidth="1"/>'
            for i, col in enumerate(df.columns, start=1)
        )
        filtro = f'<autoFilter ref="A1:{ultima_coluna}{ultima_linha}"/>' if len(df.columns) else ""
        planilha = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<dimension ref="A1:{ultima_coluna}{ultima_linha}"/>'
            '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
            f'<cols>{larguras}</cols><sheetData>{"".join(linhas_xml)}</sheetData>{filtro}</worksheet>'
        )
        abas.append((nome_aba, planilha))

    saida = io.BytesIO()
    with zipfile.ZipFile(saida, "w", zipfile.ZIP_DEFLATED) as arquivo:
        overrides = "".join(
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            for i in range(1, len(abas) + 1)
        )
        arquivo.writestr("[Content_Types].xml", '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            f'{overrides}</Types>')
        arquivo.writestr("_rels/.rels", '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>')
        sheets = "".join(
            f'<sheet name="{xml_escape(nome, {chr(34): "&quot;"})}" sheetId="{i}" r:id="rId{i}"/>'
            for i, (nome, _) in enumerate(abas, start=1)
        )
        arquivo.writestr("xl/workbook.xml", '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<sheets>{sheets}</sheets></workbook>')
        relacionamentos = "".join(
            f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i}.xml"/>'
            for i in range(1, len(abas) + 1)
        )
        arquivo.writestr("xl/_rels/workbook.xml.rels", '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'{relacionamentos}<Relationship Id="rId{len(abas)+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
            '</Relationships>')
        arquivo.writestr("xl/styles.xml", '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font></fonts>'
            '<fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF2563EB"/><bgColor indexed="64"/></patternFill></fill></fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1"/></cellXfs>'
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles></styleSheet>')
        for i, (_, planilha) in enumerate(abas, start=1):
            arquivo.writestr(f"xl/worksheets/sheet{i}.xml", planilha)
    return saida.getvalue()

def _criar_excel_relatorio(tabelas, titulo="Relatório Operacional"):
    def coluna_data_excel(coluna):
        nome = remover_acentos(str(coluna)).lower().strip()
        if "quantidade" in nome or "contagem" in nome:
            return False
        return (
            "data" in nome or nome.startswith("dt.") or nome.startswith("dt ")
            or "limite d+4" in nome or "prazo final" in nome
        )

    def preparar_datas_excel(df):
        preparado = df.copy()
        for coluna in preparado.columns:
            if not coluna_data_excel(coluna):
                continue
            try:
                convertida = pd.to_datetime(preparado[coluna], dayfirst=True, errors="coerce", format="mixed")
            except (TypeError, ValueError):
                convertida = pd.to_datetime(preparado[coluna], dayfirst=True, errors="coerce")
            preenchidos = preparado[coluna].astype(str).str.strip() != ""
            if preenchidos.any() and convertida[preenchidos].notna().mean() >= 0.8:
                preparado[coluna] = convertida
        return preparado

    def coluna_identificador(coluna):
        nome = remover_acentos(str(coluna)).lower()
        return any(chave in nome for chave in ("numero da oc", "numero da cotacao", "solicitacoes"))

    resumo_analitico, secoes_dados, _ = _organizar_secoes_relatorio(titulo, tabelas)
    tabelas_excel = list(secoes_dados)
    if not resumo_analitico.empty:
        tabelas_excel.insert(1 if tabelas_excel else 0, ("Resumo Analítico", resumo_analitico))
    if not tabelas_excel:
        tabelas_excel = list(tabelas)

    try:
        saida = io.BytesIO()
        with pd.ExcelWriter(
            saida, engine="xlsxwriter", date_format="dd/mm/yyyy", datetime_format="dd/mm/yyyy",
            engine_kwargs={"options": {"strings_to_formulas": False, "strings_to_urls": False}},
        ) as escritor:
            workbook = escritor.book
            titulo_fmt = workbook.add_format({"bold": True, "font_size": 15, "font_color": "#FFFFFF", "bg_color": "#2563EB", "align": "center", "valign": "vcenter"})
            meta_fmt = workbook.add_format({"font_size": 9, "font_color": "#64748B", "italic": True})
            texto_fmt = workbook.add_format({"valign": "top", "text_wrap": True, "border": 1, "border_color": "#93C5FD"})
            moeda_fmt = workbook.add_format({"num_format": 'R$ #,##0.00', "valign": "top", "border": 1, "border_color": "#93C5FD"})
            numero_fmt = workbook.add_format({"num_format": '#,##0.00', "valign": "top", "border": 1, "border_color": "#93C5FD"})
            data_fmt = workbook.add_format({"num_format": 'dd/mm/yyyy', "valign": "top", "border": 1, "border_color": "#93C5FD"})
            identificador_fmt = workbook.add_format({"num_format": "@", "valign": "top", "text_wrap": True, "border": 1, "border_color": "#93C5FD"})
            cartao_rotulo_fmt = workbook.add_format({"bold": True, "font_size": 8.5, "font_color": "#FFFFFF", "bg_color": "#0F172A", "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1, "border_color": "#FFFFFF"})
            cartao_valor_fmt = workbook.add_format({"bold": True, "font_size": 13, "font_color": "#0F172A", "bg_color": "#EFF6FF", "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1, "border_color": "#BFDBFE"})
            subtitulo_fmt = workbook.add_format({"bold": True, "font_size": 11, "font_color": "#FFFFFF", "bg_color": "#2563EB", "align": "left", "valign": "vcenter"})
            cabecalho_tabela_fmt = workbook.add_format({"bold": True, "font_size": 9, "font_color": "#FFFFFF", "bg_color": "#2563EB", "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1, "border_color": "#FFFFFF"})
            usados = set()

            for ordem_item, (nome, df_original) in enumerate(tabelas_excel):
                df = preparar_datas_excel(df_original)
                nome_aba = _nome_aba_excel(nome, usados)
                eh_resumo_analitico = nome_aba == "Resumo Analítico"
                eh_principal = ordem_item == 0 and not eh_resumo_analitico
                inicio_tabela = 18 if eh_resumo_analitico else 6 if eh_principal else 3
                df.to_excel(escritor, sheet_name=nome_aba, index=False, startrow=inicio_tabela)
                worksheet = escritor.sheets[nome_aba]
                ultima_coluna = max(0, len(df.columns) - 1)
                if eh_resumo_analitico:
                    worksheet.merge_range(0, 0, 0, 13, f"{titulo.upper()} - ANÁLISES", titulo_fmt)
                    worksheet.merge_range(1, 0, 1, 13, f"Gerado em {datetime.now(FUSO_LOCAL).strftime('%d/%m/%Y às %H:%M')} • {len(df)} {plural_pt(len(df), 'indicador', 'indicadores')}", meta_fmt)
                    worksheet.set_row(0, 29)
                    worksheet.set_row(1, 19)
                    worksheet.hide_gridlines(2)
                    worksheet.set_landscape()
                    worksheet.fit_to_pages(1, 0)
                    worksheet.set_margins(0.35, 0.35, 0.55, 0.55)
                    worksheet.set_footer("&LAPROAR Engenharia&CResumo analítico&R Página &P de &N")
                    worksheet.set_column(0, 5, 14)
                    worksheet.set_column(6, 6, 2)
                    worksheet.set_column(7, 13, 11)

                    for indice, (_, indicador) in enumerate(df.head(6).iterrows()):
                        linha = 3 + (indice // 2) * 4
                        coluna = 0 if indice % 2 == 0 else 3
                        worksheet.merge_range(linha, coluna, linha, coluna + 2, str(indicador.get("Indicador", "Indicador")), cartao_rotulo_fmt)
                        worksheet.merge_range(linha + 1, coluna, linha + 2, coluna + 2, str(indicador.get("Resultado", "-")), cartao_valor_fmt)
                        worksheet.set_row(linha, 22)
                        worksheet.set_row(linha + 1, 23)
                        worksheet.set_row(linha + 2, 23)

                    titulo_grafico, dados_grafico = _dados_grafico_resumo(df)
                    if dados_grafico:
                        linha_auxiliar = 2
                        worksheet.write(linha_auxiliar, 15, "Categoria")
                        worksheet.write(linha_auxiliar, 16, "Valor")
                        for posicao, (rotulo, valor) in enumerate(dados_grafico, start=1):
                            indice_df = next((i for i, texto_indicador in enumerate(df["Indicador"].astype(str)) if texto_indicador == rotulo), None)
                            linha_origem_excel = inicio_tabela + 2 + indice_df if indice_df is not None else None
                            linha_destino = linha_auxiliar + posicao
                            if linha_origem_excel:
                                worksheet.write_formula(linha_destino, 15, f"=A{linha_origem_excel}", None, rotulo)
                                formula_valor = f'=IFERROR(NUMBERVALUE(SUBSTITUTE(SUBSTITUTE(B{linha_origem_excel},"R$ ",""),"%",""),",","."),B{linha_origem_excel})'
                                worksheet.write_formula(linha_destino, 16, formula_valor, None, valor)
                            else:
                                worksheet.write(linha_destino, 15, rotulo)
                                worksheet.write_number(linha_destino, 16, valor)
                        worksheet.set_column(15, 16, None, None, {"hidden": True})
                        grafico = workbook.add_chart({"type": "doughnut"})
                        pontos = [{"fill": {"color": "#2563EB"}}, {"fill": {"color": "#60A5FA"}}]
                        if "custos" in remover_acentos(titulo_grafico).lower(): pontos = [{"fill": {"color": "#2563EB"}}, {"fill": {"color": "#0F172A"}}]
                        grafico.add_series({
                            "name": titulo_grafico,
                            "categories": [nome_aba, linha_auxiliar + 1, 15, linha_auxiliar + len(dados_grafico), 15],
                            "values": [nome_aba, linha_auxiliar + 1, 16, linha_auxiliar + len(dados_grafico), 16],
                            "points": pontos[:len(dados_grafico)],
                            "data_labels": {"percentage": True, "leader_lines": True},
                        })
                        grafico.set_title({"name": titulo_grafico, "name_font": {"size": 12, "bold": True, "color": "#334155"}})
                        grafico.set_hole_size(58)
                        grafico.set_legend({"position": "bottom", "font": {"size": 9}})
                        grafico.set_chartarea({"border": {"none": True}, "fill": {"color": "#FFFFFF"}})
                        grafico.set_plotarea({"border": {"none": True}, "fill": {"color": "#FFFFFF"}})
                        grafico.set_size({"width": 500, "height": 265})
                        worksheet.insert_chart(3, 7, grafico)

                    worksheet.merge_range(16, 0, 16, 13, "Leituras gerenciais e indicadores complementares", subtitulo_fmt)
                    worksheet.set_row(16, 23)
                    if len(df):
                        worksheet.add_table(inicio_tabela, 0, inicio_tabela + len(df), ultima_coluna, {
                            "name": f"Tabela_{len(usados)}_ResumoAnalitico",
                            "style": "Table Style Light 9",
                            "columns": [{"header": str(coluna), "header_format": cabecalho_tabela_fmt} for coluna in df.columns],
                        })
                    worksheet.set_column(0, 0, 34, texto_fmt)
                    worksheet.set_column(1, 1, 22, texto_fmt)
                    worksheet.set_column(2, 2, 52, texto_fmt)
                    worksheet.set_row(inicio_tabela, 26)
                    for linha in range(inicio_tabela + 1, inicio_tabela + 1 + len(df)):
                        worksheet.set_row(linha, 32)
                    worksheet.freeze_panes(inicio_tabela + 1, 0)
                    continue

                largura_modelo = max(ultima_coluna, 5 if eh_principal else ultima_coluna)
                titulo_planilha = titulo.upper() if eh_principal else f"{titulo} - {nome}".upper()
                if largura_modelo:
                    worksheet.merge_range(0, 0, 0, largura_modelo, titulo_planilha, titulo_fmt)
                else:
                    worksheet.write(0, 0, titulo_planilha, titulo_fmt)
                worksheet.write(1, 0, f"Gerado em {datetime.now(FUSO_LOCAL).strftime('%d/%m/%Y às %H:%M')} • {len(df)} {plural_pt(len(df), 'registro', 'registros')} • Seção: {nome}", meta_fmt)
                worksheet.set_row(0, 27)
                worksheet.set_row(inicio_tabela, 28)
                worksheet.freeze_panes(inicio_tabela + 1, 0)
                worksheet.hide_gridlines(2)
                worksheet.set_landscape()
                worksheet.fit_to_pages(1, 0)
                worksheet.set_margins(0.35, 0.35, 0.55, 0.55)
                worksheet.set_footer("&LAPROAR Engenharia&CRelatório operacional&R Página &P de &N")

                if eh_principal:
                    total_colunas_painel = largura_modelo + 1
                    worksheet.set_column(0, largura_modelo, 14)
                    for indice_kpi, (_, indicador) in enumerate(resumo_analitico.head(6).iterrows()):
                        coluna_inicial = (indice_kpi * total_colunas_painel) // 6
                        coluna_final = ((indice_kpi + 1) * total_colunas_painel) // 6 - 1
                        rotulo = str(indicador.get("Indicador", "Indicador"))
                        resultado = str(indicador.get("Resultado", "-"))
                        if coluna_inicial == coluna_final:
                            worksheet.write(3, coluna_inicial, rotulo, cartao_rotulo_fmt)
                            worksheet.write(4, coluna_inicial, resultado, cartao_valor_fmt)
                        else:
                            worksheet.merge_range(3, coluna_inicial, 3, coluna_final, rotulo, cartao_rotulo_fmt)
                            worksheet.merge_range(4, coluna_inicial, 4, coluna_final, resultado, cartao_valor_fmt)
                    worksheet.set_row(3, 31)
                    worksheet.set_row(4, 34)

                if len(df.columns):
                    if len(df):
                        worksheet.add_table(inicio_tabela, 0, inicio_tabela + len(df), ultima_coluna, {
                            "name": f"Tabela_{len(usados)}_{re.sub(r'[^A-Za-z0-9]', '', nome_aba)[:15] or 'Dados'}",
                            "style": "Table Style Light 9",
                            "columns": [{"header": str(coluna), "header_format": cabecalho_tabela_fmt} for coluna in df.columns],
                        })
                    else:
                        for indice, coluna in enumerate(df.columns):
                            worksheet.write(inicio_tabela, indice, str(coluna), cabecalho_tabela_fmt)

                    for indice, coluna in enumerate(df.columns):
                        valores = df[coluna].astype(str).head(250).tolist()
                        maior = max([len(str(coluna))] + [len(valor) for valor in valores])
                        nome_coluna = remover_acentos(str(coluna)).lower()
                        longa = any(chave in nome_coluna for chave in ("material", "endereco", "observacao", "motivo", "leitura"))
                        if nome_aba == "Resumo Analítico" and nome_coluna == "resultado": largura = 22
                        elif nome_aba == "Resumo Analítico" and "indicador" in nome_coluna: largura = 34
                        else: largura = 52 if longa else min(32, max(12, maior + 2))
                        formato = identificador_fmt if coluna_identificador(coluna) else data_fmt if coluna_data_excel(coluna) else moeda_fmt if "r$" in nome_coluna or "custo" in nome_coluna or "valor" in nome_coluna or "manutencao" in nome_coluna else numero_fmt if nome_coluna in {"km", "litros", "distancia (km)"} else texto_fmt
                        worksheet.set_column(indice, indice, largura, formato)
                        if coluna_identificador(coluna):
                            for deslocamento, valor in enumerate(df[coluna].tolist(), start=inicio_tabela + 1):
                                texto_valor = "" if pd.isna(valor) else str(valor)
                                worksheet.write_string(deslocamento, indice, texto_valor, identificador_fmt)

                    for linha, valores in enumerate(df.astype(str).values.tolist(), start=inicio_tabela + 1):
                        maior_texto = max([len(valor) for valor in valores] + [0])
                        altura = min(72, 18 + 12 * max(0, maior_texto // 70))
                        worksheet.set_row(linha, altura)

                    for indice, coluna in enumerate(df.columns):
                        if "status" in remover_acentos(str(coluna)).lower() and len(df):
                            intervalo = (inicio_tabela + 1, indice, inicio_tabela + len(df), indice)
                            worksheet.conditional_format(*intervalo, {"type": "text", "criteria": "containing", "value": "Dentro", "format": workbook.add_format({"bg_color": "#DCFCE7", "font_color": "#166534"})})
                            worksheet.conditional_format(*intervalo, {"type": "text", "criteria": "containing", "value": "Conclu", "format": workbook.add_format({"bg_color": "#DCFCE7", "font_color": "#166534"})})
                            worksheet.conditional_format(*intervalo, {"type": "text", "criteria": "containing", "value": "Fora", "format": workbook.add_format({"bg_color": "#FEE2E2", "font_color": "#991B1B"})})
                            worksheet.conditional_format(*intervalo, {"type": "text", "criteria": "containing", "value": "Sem Recebimento", "format": workbook.add_format({"bg_color": "#E2E8F0", "font_color": "#334155"})})
                            worksheet.conditional_format(*intervalo, {"type": "text", "criteria": "containing", "value": "Pend", "format": workbook.add_format({"bg_color": "#FEF3C7", "font_color": "#92400E"})})
                        if "observacao" in remover_acentos(str(coluna)).lower() and len(df):
                            intervalo = (inicio_tabela + 1, indice, inicio_tabela + len(df), indice)
                            worksheet.conditional_format(*intervalo, {"type": "text", "criteria": "containing", "value": "corre", "format": workbook.add_format({"bg_color": "#FEF3C7", "font_color": "#92400E"})})
        return saida.getvalue()
    except (ImportError, ModuleNotFoundError):
        pass

    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        saida = io.BytesIO()
        with pd.ExcelWriter(saida, engine="openpyxl") as escritor:
            usados = set()
            for ordem_item, (nome, df_original) in enumerate(tabelas_excel):
                df = preparar_datas_excel(df_original)
                nome_aba = _nome_aba_excel(nome, usados)
                eh_resumo_analitico = nome_aba == "Resumo Analítico"
                eh_principal = ordem_item == 0 and not eh_resumo_analitico
                inicio_tabela = 6 if eh_principal else 3
                df.to_excel(escritor, sheet_name=nome_aba, index=False, startrow=inicio_tabela)
                ws = escritor.sheets[nome_aba]
                ultima_coluna = max(6 if eh_principal else 1, len(df.columns))
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_coluna)
                ws.cell(1, 1, titulo.upper() if eh_principal else f"{titulo} - {nome}".upper())
                ws.cell(1, 1).font = Font(bold=True, size=16, color="FFFFFF")
                ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
                for coluna_titulo in range(1, ultima_coluna + 1): ws.cell(1, coluna_titulo).fill = PatternFill("solid", fgColor="08B7B7")
                tipo_linha = plural_pt(len(df), "indicador", "indicadores") if nome_aba == "Resumo Analítico" else plural_pt(len(df), "registro", "registros")
                ws.cell(2, 1, f"Gerado em {datetime.now(FUSO_LOCAL).strftime('%d/%m/%Y às %H:%M')} • {len(df)} {tipo_linha}")
                ws.cell(2, 1).font = Font(size=9, italic=True, color="64748B")
                if eh_principal:
                    for indice_kpi, (_, indicador) in enumerate(resumo_analitico.head(6).iterrows(), start=1):
                        ws.cell(4, indice_kpi, str(indicador.get("Indicador", "Indicador")))
                        ws.cell(5, indice_kpi, str(indicador.get("Resultado", "-")))
                        ws.cell(4, indice_kpi).font = Font(bold=True, color="FFFFFF", size=9)
                        ws.cell(4, indice_kpi).fill = PatternFill("solid", fgColor="4B4B4B")
                        ws.cell(5, indice_kpi).font = Font(bold=True, color="334155", size=12)
                        ws.cell(5, indice_kpi).fill = PatternFill("solid", fgColor="F3FAF3")
                        ws.cell(4, indice_kpi).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws.cell(5, indice_kpi).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                linha_cabecalho = inicio_tabela + 1
                ws.freeze_panes = f"A{linha_cabecalho + 1}"
                ws.sheet_view.showGridLines = False
                ws.auto_filter.ref = f"A{linha_cabecalho}:{get_column_letter(max(1, len(df.columns)))}{max(linha_cabecalho, linha_cabecalho + len(df))}"
                for celula in ws[linha_cabecalho]:
                    celula.font = Font(bold=True, color="FFFFFF")
                    celula.fill = PatternFill("solid", fgColor="08B7B7")
                    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for indice, coluna in enumerate(df.columns, start=1):
                    valores = df[coluna].astype(str).head(250).tolist()
                    maior = max([len(str(coluna))] + [len(valor) for valor in valores])
                    nome_coluna = remover_acentos(str(coluna)).lower()
                    longa = any(chave in nome_coluna for chave in ("material", "endereco", "observacao", "motivo", "leitura"))
                    if nome_aba == "Resumo Analítico" and nome_coluna == "resultado": largura = 22
                    elif nome_aba == "Resumo Analítico" and "indicador" in nome_coluna: largura = 34
                    else: largura = 52 if longa else min(32, max(12, maior + 2))
                    ws.column_dimensions[get_column_letter(indice)].width = largura
                    for linha in range(linha_cabecalho + 1, linha_cabecalho + 1 + len(df)):
                        ws.cell(linha, indice).alignment = Alignment(vertical="top", wrap_text=True)
                        if coluna_data_excel(coluna): ws.cell(linha, indice).number_format = "dd/mm/yyyy"
                        if coluna_identificador(coluna): ws.cell(linha, indice).number_format = "@"
                if nome_aba == "Resumo Analítico":
                    for linha in range(linha_cabecalho + 1, linha_cabecalho + 1 + len(df)): ws.row_dimensions[linha].height = 32
        return saida.getvalue()
    except (ImportError, ModuleNotFoundError):
        pass
    return _criar_xlsx_basico(tabelas_excel)

def _texto_pdf(valor):
    bytes_texto = str(valor).encode("cp1252", errors="replace")
    return "".join(
        chr(byte) if 32 <= byte <= 126 and chr(byte) not in "\\()" else
        f"\\{chr(byte)}" if chr(byte) in "\\()" else f"\\{byte:03o}"
        for byte in bytes_texto
    )

def _criar_pdf_textual(titulo, tabelas):
    """PDF visual independente de bibliotecas externas, usado como contingência."""
    largura_pagina, altura_pagina = 842.0, 595.0
    margem_x, limite_inferior = 34.0, 42.0
    largura_util = largura_pagina - 2 * margem_x
    cores = {
        # Paleta APROAR: azul institucional, azul-marinho e branco.
        "teal": (0.145, 0.388, 0.922),       # #2563EB
        "teal_borda": (0.576, 0.773, 0.992), # #93C5FD
        "cinza": (0.059, 0.090, 0.165),      # #0F172A
        "texto": (0.118, 0.161, 0.231),      # #1E293B
        "borda": (0.749, 0.847, 0.949),      # #BFDBFE
        "branco": (1.0, 1.0, 1.0),
        "fundo": (0.973, 0.980, 0.988),
        "azul_claro": (0.937, 0.965, 1.0),   # #EFF6FF
        "verde": (0.867, 0.957, 0.835),
        "amarelo": (0.996, 0.953, 0.780),
        "vermelho": (0.996, 0.792, 0.792),
    }

    def limpo(valor):
        return str(valor).replace("—", "-").replace("–", "-").replace("‑", "-").replace("•", "-").replace("\n", " ").strip()

    def rgb(cor, operador):
        return f"{cor[0]:.3f} {cor[1]:.3f} {cor[2]:.3f} {operador}"

    def retangulo(comandos, x, y, largura, altura, preenchimento, contorno=None, espessura=0.5):
        comandos.append(rgb(preenchimento, "rg"))
        if contorno:
            comandos.extend([rgb(contorno, "RG"), f"{espessura:.2f} w", f"{x:.2f} {y:.2f} {largura:.2f} {altura:.2f} re B"])
        else:
            comandos.append(f"{x:.2f} {y:.2f} {largura:.2f} {altura:.2f} re f")

    def largura_texto(texto, tamanho):
        return len(str(texto)) * tamanho * 0.50

    def escrever(comandos, texto, x, y, tamanho=8, negrito=False, cor=None, alinhamento="esquerda"):
        texto = limpo(texto)
        cor = cor or cores["texto"]
        if alinhamento == "centro":
            x -= largura_texto(texto, tamanho) / 2
        elif alinhamento == "direita":
            x -= largura_texto(texto, tamanho)
        fonte = "F2" if negrito else "F1"
        comandos.extend([
            rgb(cor, "rg"),
            f"BT /{fonte} {tamanho:.2f} Tf {x:.2f} {y:.2f} Td ({_texto_pdf(texto)}) Tj ET",
        ])

    def quebrar_texto(valor, largura, tamanho=7, limite=240):
        texto = limpo(valor)
        if len(texto) > limite:
            texto = texto[:limite - 3].rstrip() + "..."
        caracteres = max(4, int(largura / max(3.5, tamanho * 0.50)))
        return textwrap.wrap(texto, width=caracteres, break_long_words=True, break_on_hyphens=False) or ["-"]

    def escrever_caixa(comandos, valor, x, y, largura, altura, tamanho=7, negrito=False, cor=None, alinhamento="esquerda", limite=240):
        linhas = quebrar_texto(valor, largura - 8, tamanho, limite)
        entrelinha = tamanho + 1.6
        bloco = len(linhas) * entrelinha
        inicio = y + altura - 5 - tamanho if alinhamento == "esquerda" else y + (altura + bloco) / 2 - tamanho
        for indice, linha in enumerate(linhas):
            eixo_x = x + 4 if alinhamento == "esquerda" else x + largura / 2
            escrever(comandos, linha, eixo_x, inicio - indice * entrelinha, tamanho, negrito, cor, alinhamento)

    paginas, comandos, y = [], [], 0.0

    def nova_pagina():
        nonlocal comandos, y
        if comandos:
            paginas.append(comandos)
        comandos = []
        y = 548.0

    def garantir(altura_necessaria):
        if y - altura_necessaria < limite_inferior:
            nova_pagina()
            return True
        return False

    def desenhar_cabecalho_tabela(colunas, larguras, y_topo):
        altura = 27.0
        x = margem_x
        for coluna, largura in zip(colunas, larguras):
            retangulo(comandos, x, y_topo - altura, largura, altura, cores["teal"], cores["branco"], 0.45)
            escrever_caixa(comandos, coluna, x, y_topo - altura, largura, altura, 7.0, True, cores["branco"], "centro", 100)
            x += largura
        return y_topo - altura

    def desenhar_secao(nome, df_original):
        nonlocal y
        df = _colunas_relevantes_pdf(df_original, limite=9)
        garantir(28)
        retangulo(comandos, margem_x, y - 21, largura_util, 21, cores["teal"])
        escrever(comandos, f"{limpo(nome).upper()} - {len(df_original)} {plural_pt(len(df_original), 'registro', 'registros')}", margem_x + 7, y - 14, 9, True, cores["branco"])
        y -= 21
        if df.empty or not len(df.columns):
            escrever(comandos, "Nenhum registro disponível nesta seção.", margem_x + 6, y - 15, 8, False, cores["texto"])
            y -= 27
            return

        pesos = []
        for coluna in df.columns:
            nome_coluna = remover_acentos(str(coluna)).lower()
            amostra = [len(str(coluna))] + [len(str(valor)) for valor in df[coluna].astype(str).head(80)]
            peso = max(8, min(30, max(amostra)))
            if any(chave in nome_coluna for chave in ("material", "descricao", "endereco", "observacao")):
                peso = max(peso, 25)
            pesos.append(peso)
        soma_pesos = sum(pesos) or 1
        larguras = [largura_util * peso / soma_pesos for peso in pesos]
        y = desenhar_cabecalho_tabela(list(df.columns), larguras, y)

        for indice_registro, (_, registro) in enumerate(df.iterrows(), start=1):
            linhas_celulas = [quebrar_texto(registro[coluna], largura - 8, 6.7, 240) for coluna, largura in zip(df.columns, larguras)]
            altura_linha = max(20.0, min(68.0, 8.3 * max(len(linhas) for linhas in linhas_celulas) + 7))
            if y - altura_linha < limite_inferior:
                nova_pagina()
                retangulo(comandos, margem_x, y - 21, largura_util, 21, cores["teal"])
                escrever(comandos, f"{limpo(nome).upper()} - CONTINUAÇÃO", margem_x + 7, y - 14, 9, True, cores["branco"])
                y -= 21
                y = desenhar_cabecalho_tabela(list(df.columns), larguras, y)
            x = margem_x
            fundo_linha = cores["fundo"] if indice_registro % 2 == 0 else cores["branco"]
            for indice_coluna, (coluna, largura, linhas) in enumerate(zip(df.columns, larguras, linhas_celulas)):
                fundo_celula = fundo_linha
                nome_coluna = remover_acentos(str(coluna)).lower()
                if "status" in nome_coluna or "situacao" in nome_coluna:
                    status = remover_acentos(str(registro[coluna])).lower()
                    if re.search(r"conclu|entregue|dentro", status): fundo_celula = cores["verde"]
                    elif re.search(r"fora|atras|vencid", status): fundo_celula = cores["vermelho"]
                    elif re.search(r"pend|aguard", status): fundo_celula = cores["amarelo"]
                retangulo(comandos, x, y - altura_linha, largura, altura_linha, fundo_celula, cores["teal_borda"], 0.35)
                entrelinha = 8.3
                for indice_linha, linha in enumerate(linhas[:7]):
                    escrever(comandos, linha, x + 4, y - 11 - indice_linha * entrelinha, 6.7, False, cores["texto"])
                x += largura
            y -= altura_linha
        y -= 8

    resumo_analitico, secoes_dados, _ = _organizar_secoes_relatorio(titulo, tabelas)
    total_registros = sum(len(df) for _, df in secoes_dados)
    nova_pagina()

    retangulo(comandos, margem_x, y - 29, largura_util, 29, cores["teal"])
    escrever(comandos, limpo(titulo).upper(), largura_pagina / 2, y - 19, 13, True, cores["branco"], "centro")
    y -= 36
    escrever(comandos, f"Gerado em {datetime.now(FUSO_LOCAL).strftime('%d/%m/%Y às %H:%M')} | {total_registros} {plural_pt(total_registros, 'registro', 'registros')}", margem_x + 6, y - 8, 8, False, cores["texto"])
    y -= 19

    if not resumo_analitico.empty:
        kpis = resumo_analitico.head(6).to_dict("records")
        while len(kpis) < 6:
            kpis.append({"Indicador": "", "Resultado": ""})
        largura_kpi = largura_util / 6
        altura_rotulo, altura_valor = 28.0, 31.0
        x = margem_x
        for kpi in kpis:
            retangulo(comandos, x, y - altura_rotulo, largura_kpi, altura_rotulo, cores["cinza"], cores["branco"], 0.5)
            escrever_caixa(comandos, kpi.get("Indicador", ""), x, y - altura_rotulo, largura_kpi, altura_rotulo, 7.0, True, cores["branco"], "centro", 90)
            retangulo(comandos, x, y - altura_rotulo - altura_valor, largura_kpi, altura_valor, cores["azul_claro"], cores["borda"], 0.5)
            escrever_caixa(comandos, kpi.get("Resultado", ""), x, y - altura_rotulo - altura_valor, largura_kpi, altura_valor, 10.5, True, cores["texto"], "centro", 70)
            x += largura_kpi
        y -= altura_rotulo + altura_valor + 9

    for nome, df in secoes_dados:
        desenhar_secao(nome, df)

    if not resumo_analitico.empty:
        leituras = resumo_analitico[resumo_analitico["Leitura para análise"].astype(str).str.strip() != ""]
        complementares = resumo_analitico.iloc[6:]
        analises = pd.concat([leituras, complementares], ignore_index=True).drop_duplicates(subset=["Indicador"])
        if not analises.empty:
            desenhar_secao("Análises complementares", analises[["Indicador", "Resultado", "Leitura para análise"]].rename(columns={"Leitura para análise": "Interpretação"}))

    if comandos:
        paginas.append(comandos)

    objetos = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>",
    ]
    ids_paginas = []
    total_paginas = len(paginas)
    for numero_pagina, comandos_pagina in enumerate(paginas, start=1):
        moldura = []
        retangulo(moldura, 0, altura_pagina - 28, largura_pagina, 28, cores["cinza"])
        escrever(moldura, "APROAR ENGENHARIA", margem_x, altura_pagina - 18, 8.5, True, cores["branco"])
        escrever(moldura, "Torre de Controle Logístico", largura_pagina - margem_x, altura_pagina - 18, 7, False, cores["branco"], "direita")
        moldura.extend([rgb(cores["borda"], "RG"), "0.5 w", f"{margem_x:.2f} 30 m {largura_pagina - margem_x:.2f} 30 l S"])
        escrever(moldura, "Relatório operacional", margem_x, 18, 7, False, cores["texto"])
        escrever(moldura, f"Página {numero_pagina} de {total_paginas}", largura_pagina - margem_x, 18, 7, False, cores["texto"], "direita")
        conteudo_bytes = ("\n".join(moldura + comandos_pagina)).encode("ascii")
        id_pagina = len(objetos) + 1
        id_conteudo = id_pagina + 1
        ids_paginas.append(id_pagina)
        objetos.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {largura_pagina:.0f} {altura_pagina:.0f}] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {id_conteudo} 0 R >>".encode("ascii"))
        objetos.append(f"<< /Length {len(conteudo_bytes)} >>\nstream\n".encode("ascii") + conteudo_bytes + b"\nendstream")
    objetos[1] = f"<< /Type /Pages /Kids [{' '.join(f'{i} 0 R' for i in ids_paginas)}] /Count {len(ids_paginas)} >>".encode("ascii")

    arquivo = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for indice, objeto in enumerate(objetos, start=1):
        offsets.append(len(arquivo))
        arquivo.extend(f"{indice} 0 obj\n".encode("ascii") + objeto + b"\nendobj\n")
    inicio_xref = len(arquivo)
    arquivo.extend(f"xref\n0 {len(objetos)+1}\n0000000000 65535 f \n".encode("ascii"))
    for offset in offsets[1:]:
        arquivo.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    arquivo.extend(f"trailer\n<< /Size {len(objetos)+1} /Root 1 0 R >>\nstartxref\n{inicio_xref}\n%%EOF".encode("ascii"))
    return bytes(arquivo)

def _criar_pdf_relatorio(titulo, tabelas):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
    except (ImportError, ModuleNotFoundError):
        return _criar_pdf_textual(titulo, tabelas)

    def texto_pdf_limpo(valor):
        return str(valor).replace("—", "-").replace("–", "-").replace("‑", "-").replace("•", "-")

    def paragrafo(valor, estilo, limite=700):
        texto = texto_pdf_limpo(valor)
        if len(texto) > limite:
            texto = texto[:limite - 3].rstrip() + "..."
        texto = html_escape(texto).replace("\n", "<br/>")
        return Paragraph(texto or "-", estilo)

    # Paleta visual dos relatórios APROAR.
    teal = colors.HexColor("#2563EB")
    teal_borda = colors.HexColor("#93C5FD")
    cinza_escuro = colors.HexColor("#0F172A")
    cinza_fundo = colors.HexColor("#F8FAFC")
    cinza_borda = colors.HexColor("#BFDBFE")
    cinza_texto = colors.HexColor("#475569")
    azul_claro = colors.HexColor("#EFF6FF")
    verde_claro = colors.HexColor("#DDF4D5")
    amarelo_claro = colors.HexColor("#FEF3C7")
    vermelho_claro = colors.HexColor("#FECACA")
    pagina = landscape(A4)
    largura_util = pagina[0] - 24 * mm

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("TituloModelo", parent=estilos["Title"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=colors.white, alignment=TA_CENTER)
    estilo_meta = ParagraphStyle("MetaModelo", parent=estilos["Normal"], fontName="Helvetica", fontSize=8, leading=10, textColor=cinza_texto, alignment=TA_LEFT)
    estilo_secao = ParagraphStyle("SecaoModelo", parent=estilos["Heading2"], fontName="Helvetica-Bold", fontSize=9.5, leading=12, textColor=colors.white, backColor=teal, borderPadding=(5, 7, 5, 7), spaceBefore=7, spaceAfter=0, keepWithNext=1)
    estilo_cabecalho = ParagraphStyle("CabecalhoModelo", parent=estilos["Normal"], fontName="Helvetica-Bold", fontSize=7.2, leading=8.5, textColor=colors.white, alignment=TA_CENTER)
    estilo_celula = ParagraphStyle("CelulaModelo", parent=estilos["Normal"], fontName="Helvetica", fontSize=6.8, leading=8.4, textColor=colors.HexColor("#334155"))
    estilo_kpi_rotulo = ParagraphStyle("RotuloKPIModelo", parent=estilos["Normal"], fontName="Helvetica-Bold", fontSize=7.2, leading=8.5, textColor=colors.white, alignment=TA_CENTER)
    estilo_kpi_valor = ParagraphStyle("ValorKPIModelo", parent=estilos["Normal"], fontName="Helvetica-Bold", fontSize=10.5, leading=13, textColor=colors.HexColor("#334155"), alignment=TA_CENTER)
    estilo_vazio = ParagraphStyle("VazioModelo", parent=estilos["Normal"], fontName="Helvetica-Oblique", fontSize=8, textColor=cinza_texto, spaceBefore=5, spaceAfter=7)

    titulo_limpo = texto_pdf_limpo(titulo)
    resumo_analitico, secoes_dados, _ = _organizar_secoes_relatorio(titulo, tabelas)
    total_registros = sum(len(df) for _, df in secoes_dados)
    saida = io.BytesIO()
    documento = SimpleDocTemplate(
        saida, pagesize=pagina, rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=19 * mm, bottomMargin=14 * mm,
        title=titulo_limpo, author="APROAR Engenharia",
    )

    def cabecalho_rodape(canvas, doc):
        largura, altura = pagina
        canvas.saveState()
        canvas.setFillColor(cinza_escuro)
        canvas.rect(0, altura - 10 * mm, largura, 10 * mm, stroke=0, fill=1)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawString(12 * mm, altura - 6.5 * mm, "APROAR ENGENHARIA")
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(largura - 12 * mm, altura - 6.5 * mm, "Torre de Controle Logístico")
        canvas.setStrokeColor(cinza_borda)
        canvas.line(12 * mm, 9 * mm, largura - 12 * mm, 9 * mm)
        canvas.setFillColor(cinza_texto)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(12 * mm, 5 * mm, "Relatório operacional")
        canvas.drawRightString(largura - 12 * mm, 5 * mm, f"Página {doc.page}")
        canvas.restoreState()

    faixa_titulo = Table([[Paragraph(html_escape(titulo_limpo.upper()), estilo_titulo)]], colWidths=[largura_util], rowHeights=[26])
    faixa_titulo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), teal), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos = [
        faixa_titulo,
        Spacer(1, 5),
        Paragraph(f"Gerado em {datetime.now(FUSO_LOCAL).strftime('%d/%m/%Y às %H:%M')} &nbsp;&nbsp;|&nbsp;&nbsp; {total_registros} {plural_pt(total_registros, 'registro', 'registros')}", estilo_meta),
        Spacer(1, 7),
    ]

    if not resumo_analitico.empty:
        kpis = resumo_analitico.head(6)
        dados_kpi = [
            [paragrafo(linha.get("Indicador", "Indicador"), estilo_kpi_rotulo, 90) for _, linha in kpis.iterrows()],
            [paragrafo(linha.get("Resultado", "-"), estilo_kpi_valor, 80) for _, linha in kpis.iterrows()],
        ]
        while len(dados_kpi[0]) < 6:
            dados_kpi[0].append("")
            dados_kpi[1].append("")
        painel_kpi = Table(dados_kpi, colWidths=[largura_util / 6] * 6, rowHeights=[28, 32])
        painel_kpi.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), cinza_escuro),
            ("BACKGROUND", (0, 1), (-1, 1), azul_claro),
            ("GRID", (0, 0), (-1, -1), 0.55, colors.white),
            ("BOX", (0, 0), (-1, -1), 0.55, cinza_borda),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elementos.extend([painel_kpi, Spacer(1, 6)])

    for indice_secao, (nome, df_original) in enumerate(secoes_dados):
        nome_secao_norm = remover_acentos(str(nome)).lower()
        if "roteiro" in remover_acentos(titulo_limpo).lower() and indice_secao > 0 and any(
            chave in nome_secao_norm for chave in ("paradas e demandas", "detalhamento da rota", "detalhamento")
        ):
            elementos.append(PageBreak())
        elementos.append(Paragraph(f"{html_escape(texto_pdf_limpo(nome).upper())} &nbsp; - &nbsp; {len(df_original)} {plural_pt(len(df_original), 'registro', 'registros')}", estilo_secao))
        if df_original.empty or not len(df_original.columns):
            elementos.append(Paragraph("Nenhum registro disponível nesta seção.", estilo_vazio))
            continue

        df = _colunas_relevantes_pdf(df_original, limite=9)
        pesos = []
        for coluna in df.columns:
            nome_coluna = remover_acentos(str(coluna)).lower()
            amostra = [len(str(coluna))] + [len(str(valor)) for valor in df[coluna].astype(str).head(80)]
            peso = max(8, min(30, max(amostra)))
            if any(chave in nome_coluna for chave in ("material", "descricao", "endereco", "observacao")):
                peso = max(peso, 25)
            pesos.append(peso)
        total_pesos = sum(pesos) or 1
        larguras = [largura_util * peso / total_pesos for peso in pesos]
        dados_tabela = [[paragrafo(coluna, estilo_cabecalho, 70) for coluna in df.columns]]
        dados_tabela.extend([[paragrafo(valor, estilo_celula) for valor in linha] for linha in df.astype(str).values.tolist()])
        tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1, hAlign="LEFT")
        comandos = [
            ("BACKGROUND", (0, 0), (-1, 0), teal),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.35, teal_borda),
        ]
        for indice_linha in range(1, len(dados_tabela)):
            comandos.append(("BACKGROUND", (0, indice_linha), (-1, indice_linha), cinza_fundo if indice_linha % 2 == 0 else colors.white))
        for indice_coluna, coluna in enumerate(df.columns):
            nome_coluna = remover_acentos(str(coluna)).lower()
            if "status" not in nome_coluna and "situacao" not in nome_coluna:
                continue
            for indice_linha, valor in enumerate(df[coluna].astype(str), start=1):
                texto_status = remover_acentos(valor).lower()
                cor = verde_claro if re.search(r"conclu|entregue|dentro", texto_status) else vermelho_claro if re.search(r"fora|atras|vencid", texto_status) else amarelo_claro if re.search(r"pend|aguard", texto_status) else None
                if cor:
                    comandos.append(("BACKGROUND", (indice_coluna, indice_linha), (indice_coluna, indice_linha), cor))
        tabela.setStyle(TableStyle(comandos))
        elementos.extend([tabela, Spacer(1, 7)])

    if not resumo_analitico.empty:
        leituras = resumo_analitico[resumo_analitico["Leitura para análise"].astype(str).str.strip() != ""]
        complementares = resumo_analitico.iloc[6:]
        analises = pd.concat([leituras, complementares], ignore_index=True).drop_duplicates(subset=["Indicador"])
        if not analises.empty:
            elementos.append(Paragraph("ANÁLISES COMPLEMENTARES", estilo_secao))
            dados_analise = [[
                paragrafo("Indicador", estilo_cabecalho),
                paragrafo("Resultado", estilo_cabecalho),
                paragrafo("Interpretação", estilo_cabecalho),
            ]]
            dados_analise.extend([
                [paragrafo(linha.get("Indicador", ""), estilo_celula), paragrafo(linha.get("Resultado", ""), estilo_celula), paragrafo(linha.get("Leitura para análise", ""), estilo_celula)]
                for _, linha in analises.iterrows()
            ])
            tabela_analise = Table(dados_analise, colWidths=[largura_util * 0.27, largura_util * 0.18, largura_util * 0.55], repeatRows=1)
            tabela_analise.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), cinza_escuro),
                ("BACKGROUND", (0, 1), (-1, -1), cinza_fundo),
                ("GRID", (0, 0), (-1, -1), 0.35, cinza_borda),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elementos.append(tabela_analise)

    documento.build(elementos, onFirstPage=cabecalho_rodape, onLaterPages=cabecalho_rodape)
    return saida.getvalue()


def _criar_pdf_resumo_rota_tabela(titulo, df_resumo):
    """PDF enxuto da rota: somente a sequência de paradas e o que fazer em cada uma."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, KeepTogether
    except (ImportError, ModuleNotFoundError):
        return _criar_pdf_textual(titulo, [("Resumo da rota", df_resumo)])

    azul = colors.HexColor("#2563EB")
    azul_escuro = colors.HexColor("#0F172A")
    azul_claro = colors.HexColor("#EFF6FF")
    azul_borda = colors.HexColor("#BFDBFE")
    texto = colors.HexColor("#1E293B")
    texto_sec = colors.HexColor("#64748B")
    branco = colors.white
    verde = colors.HexColor("#16A34A")
    laranja = colors.HexColor("#F59E0B")
    cinza = colors.HexColor("#94A3B8")

    def limpar(valor):
        return str(valor or "").replace("—", "-").replace("–", "-").replace("•", "-")

    def esc(valor):
        return html_escape(limpar(valor)).replace("\n", "<br/>")

    def acao_curta(valor):
        texto_acao = limpar(valor).strip()
        substituicoes = [
            (r"^COLETAR\s+materiais\s+para\s+", "COLETAR: "),
            (r"^COLETAR\s+materiais\s*$", "COLETAR"),
            (r"^ENTREGAR\s+materiais\s*-\s*", "ENTREGAR: "),
            (r"^ENTREGAR\s+materiais\s*$", "ENTREGAR"),
            (r"^PAUSA\s+DE\s+1(?:H|\s+HORA)\s+PARA\s+ALMOÇO$", "ALMOÇO - 1 hora"),
            (r"^RETORNAR\s+para\s+", "RETORNAR: "),
        ]
        for padrao, novo in substituicoes:
            texto_acao = re.sub(padrao, novo, texto_acao, flags=re.IGNORECASE)
        return texto_acao

    df = df_resumo.copy() if isinstance(df_resumo, pd.DataFrame) else pd.DataFrame()
    colunas_esperadas = ["Etapa", "Local", "O que fazer", "Horário previsto"]
    for coluna in colunas_esperadas:
        if coluna not in df.columns:
            df[coluna] = ""
    df = df[colunas_esperadas].fillna("")

    estilos = getSampleStyleSheet()
    st_titulo = ParagraphStyle(
        "RotaResumoTitulo", parent=estilos["Title"], fontName="Helvetica-Bold",
        fontSize=15, leading=18, textColor=branco, alignment=TA_CENTER,
    )
    st_sub = ParagraphStyle(
        "RotaResumoSub", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=9.5, leading=12, textColor=azul_escuro, alignment=TA_LEFT,
    )
    st_seq = ParagraphStyle(
        "RotaResumoSeq", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=8.4, leading=11, textColor=azul_escuro, alignment=TA_LEFT,
    )
    st_head = ParagraphStyle(
        "RotaResumoHead", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=7.8, leading=9.2, textColor=branco, alignment=TA_CENTER,
    )
    st_num = ParagraphStyle(
        "RotaResumoNum", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=9.5, leading=11, textColor=branco, alignment=TA_CENTER,
    )
    st_hora = ParagraphStyle(
        "RotaResumoHora", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=8.2, leading=10, textColor=azul_escuro, alignment=TA_CENTER,
    )
    st_local = ParagraphStyle(
        "RotaResumoLocal", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=8.6, leading=10.5, textColor=texto, alignment=TA_LEFT,
    )
    st_acao = ParagraphStyle(
        "RotaResumoAcao", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=8.2, leading=10.5, textColor=texto, alignment=TA_LEFT,
    )
    st_meta = ParagraphStyle(
        "RotaResumoMeta", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=7.2, leading=9, textColor=texto_sec, alignment=TA_LEFT,
    )

    saida = io.BytesIO()
    documento = SimpleDocTemplate(
        saida, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=18 * mm, bottomMargin=13 * mm,
        title=limpar(titulo), author="APROAR Engenharia",
    )
    largura = A4[0] - 24 * mm

    def cabecalho_rodape(canvas, doc):
        w, h = A4
        canvas.saveState()
        canvas.setFillColor(azul_escuro)
        canvas.rect(0, h - 9 * mm, w, 9 * mm, stroke=0, fill=1)
        canvas.setFillColor(branco)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawString(12 * mm, h - 5.8 * mm, "APROAR ENGENHARIA")
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(w - 12 * mm, h - 5.8 * mm, "Torre de Controle Logístico")
        canvas.setStrokeColor(azul_borda)
        canvas.line(12 * mm, 8 * mm, w - 12 * mm, 8 * mm)
        canvas.setFillColor(texto_sec)
        canvas.setFont("Helvetica", 6.8)
        canvas.drawString(12 * mm, 4.5 * mm, "Resumo da rota do Davi")
        canvas.drawRightString(w - 12 * mm, 4.5 * mm, f"Página {doc.page}")
        canvas.restoreState()

    titulo_limpo = limpar(titulo).upper()
    faixa = Table([[Paragraph(esc(titulo_limpo), st_titulo)]], colWidths=[largura], rowHeights=[27])
    faixa.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), azul),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))

    locais = [limpar(v).strip() for v in df["Local"].tolist() if limpar(v).strip()]
    sequencia = "  >  ".join(locais)

    elementos = [
        faixa,
        Spacer(1, 6),
        Paragraph("ORDEM DA ROTA", st_sub),
        Spacer(1, 3),
    ]
    if sequencia:
        caixa_seq = Table([[Paragraph(esc(sequencia), st_seq)]], colWidths=[largura])
        caixa_seq.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), azul_claro),
            ("BOX", (0, 0), (-1, -1), 0.6, azul_borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elementos.extend([caixa_seq, Spacer(1, 7)])

    cabecalho = [[
        Paragraph("#", st_head),
        Paragraph("HORÁRIO", st_head),
        Paragraph("LOCAL", st_head),
        Paragraph("O QUE FAZER", st_head),
    ]]
    dados = list(cabecalho)
    cores_etapa = []
    for _, linha in df.iterrows():
        etapa = limpar(linha.get("Etapa", ""))
        horario = limpar(linha.get("Horário previsto", ""))
        local = limpar(linha.get("Local", ""))
        acao = acao_curta(linha.get("O que fazer", ""))
        acao_norm = remover_acentos(acao).upper()
        if acao_norm.startswith("COLETAR"):
            cor = laranja
        elif acao_norm.startswith("ENTREGAR"):
            cor = verde
        elif "ALMOCO" in acao_norm:
            cor = cinza
        elif acao_norm.startswith("RETORNAR"):
            cor = azul_escuro
        else:
            cor = azul
        cores_etapa.append(cor)
        dados.append([
            Paragraph(esc(etapa), st_num),
            Paragraph(esc(horario), st_hora),
            Paragraph(esc(local), st_local),
            Paragraph(esc(acao), st_acao),
        ])

    larguras = [12 * mm, 31 * mm, 43 * mm, largura - 86 * mm]
    tabela = Table(dados, colWidths=larguras, repeatRows=1, hAlign="LEFT")
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), azul_escuro),
        ("GRID", (0, 0), (-1, -1), 0.45, azul_borda),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]
    for i, cor in enumerate(cores_etapa, start=1):
        fundo = colors.HexColor("#F8FAFC") if i % 2 == 0 else branco
        estilo.extend([
            ("BACKGROUND", (1, i), (-1, i), fundo),
            ("BACKGROUND", (0, i), (0, i), cor),
        ])
    tabela.setStyle(TableStyle(estilo))
    elementos.append(tabela)
    elementos.extend([
        Spacer(1, 6),
        Paragraph(
            f"Gerado em {datetime.now(FUSO_LOCAL).strftime('%d/%m/%Y às %H:%M')}. "
            "Este PDF mostra somente a sequência operacional da rota.",
            st_meta,
        ),
    ])

    documento.build(elementos, onFirstPage=cabecalho_rodape, onLaterPages=cabecalho_rodape)
    return saida.getvalue()

def _conteudo_exportador(titulo, dados, nome_arquivo, chave):
    """Exibe o exportador sem gerar PDF/Excel durante o carregamento da página.

    Antes, cada rerun montava todos os arquivos de todos os módulos, mesmo sem o
    usuário pedir download. Em Streamlit isso pesa porque o conteúdo das abas é
    executado no mesmo ciclo. Agora o arquivo só é construído ao clicar em
    "Preparar arquivo".
    """
    st.markdown("##### 📤 Exportar relatório")
    col_formato, col_acao = st.columns([1, 2])
    formato = col_formato.selectbox("Formato", ["PDF", "CSV", "Excel"], key=f"formato_relatorio_{chave}")

    arquivo_key = f"_arquivo_relatorio_{chave}"
    meta_key = f"_meta_relatorio_{chave}"
    meta_atual = st.session_state.get(meta_key, {}) or {}

    # Mudar o formato invalida somente o arquivo preparado, sem recalcular nada.
    if meta_atual.get("formato") and meta_atual.get("formato") != formato:
        st.session_state.pop(arquivo_key, None)
        st.session_state.pop(meta_key, None)
        meta_atual = {}

    preparar = col_acao.button(
        f"⚙️ Preparar {formato}",
        key=f"preparar_relatorio_{chave}_{formato}",
        use_container_width=True,
        help="O arquivo só é gerado quando você clicar aqui, deixando a plataforma mais rápida.",
    )

    if preparar:
        with st.spinner(f"Gerando {formato}..."):
            tabelas_dados = _normalizar_tabelas_relatorio(dados)
            eh_roteiro = str(chave).lower() == "roteiro"
            resumo_analitico = _criar_resumo_analitico_relatorio(titulo, tabelas_dados)
            if eh_roteiro:
                resumo_rota = next((df for nome, df in tabelas_dados if "resumo da rota" in remover_acentos(str(nome)).lower()), None)
                if resumo_rota is None:
                    resumo_rota = tabelas_dados[0][1] if tabelas_dados else pd.DataFrame(columns=["Etapa", "Local", "O que fazer", "Horário previsto"])
                tabelas = [("Resumo da rota", resumo_rota)]
            else:
                tabelas = [("Resumo Analítico", resumo_analitico)] + tabelas_dados

            if formato == "CSV":
                arquivo, extensao, mime = _criar_csv_relatorio(tabelas), "csv", "text/csv"
            elif formato == "Excel":
                arquivo, extensao, mime = _criar_excel_relatorio(tabelas, titulo), "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                if eh_roteiro:
                    arquivo = _criar_pdf_resumo_rota_tabela(titulo, tabelas[0][1])
                else:
                    arquivo = _criar_pdf_relatorio(titulo, tabelas)
                extensao, mime = "pdf", "application/pdf"

            st.session_state[arquivo_key] = arquivo
            st.session_state[meta_key] = {
                "formato": formato,
                "extensao": extensao,
                "mime": mime,
                "gerado_em": datetime.now(FUSO_LOCAL).strftime("%H:%M"),
            }
            meta_atual = st.session_state[meta_key]

    arquivo_pronto = st.session_state.get(arquivo_key)
    meta_atual = st.session_state.get(meta_key, {}) or {}
    if arquivo_pronto and meta_atual.get("formato") == formato:
        st.caption(f"Arquivo preparado às {meta_atual.get('gerado_em', '')}. Gere novamente se os dados tiverem mudado.")
        st.download_button(
            f"⬇️ Baixar {formato}",
            data=arquivo_pronto,
            file_name=f"{nome_arquivo}_{AGORA_REAL.strftime('%Y-%m-%d')}.{meta_atual.get('extensao', 'dat')}",
            mime=meta_atual.get("mime", "application/octet-stream"),
            use_container_width=True,
            key=f"baixar_relatorio_{chave}_{formato}",
        )
    else:
        st.caption("O arquivo ainda não foi gerado. Clique em **Preparar PDF/CSV/Excel** somente quando precisar exportar.")

@fragmento_independente
def renderizar_exportador(titulo, dados, nome_arquivo, chave):
    _conteudo_exportador(titulo, dados, nome_arquivo, chave)

@fragmento_independente
def renderizar_detalhes_fechamento(veiculo, gastos, quilometragem, chave):
    with st.expander(f"📋 Ver lançamentos detalhados — {veiculo}", expanded=False):
        st.markdown("**⛽ Combustível e manutenções do mês**")
        if gastos.empty:
            st.info("Nenhum gasto lançado no mês.")
        else:
            st.dataframe(gastos, use_container_width=True, hide_index=True)

        st.markdown("**🛣️ Registros de quilometragem do mês**")
        if quilometragem.empty:
            st.info("Nenhuma quilometragem lançada no mês.")
        else:
            st.dataframe(quilometragem, use_container_width=True, hide_index=True)


def _expandir_almoco_explicito_para_relatorio(route_steps):
    """Deixa a pausa de almoço visível nos relatórios, inclusive em rotas antigas.

    Versões anteriores podiam somar 60 minutos ao horário de saída de uma parada
    quando o atendimento atravessava o meio-dia, sem criar um step de almoço.
    Aqui transformamos esse intervalo oculto em uma etapa explícita apenas para
    exibição/exportação, sem alterar demandas, distâncias ou status.
    """
    passos_origem = list(route_steps or [])
    if not passos_origem:
        return []

    # Rotas novas já trazem o almoço como etapa própria; não mexe nelas.
    if any(str(s.get("type", "")) == "lunch" for s in passos_origem):
        return passos_origem

    resultado = []
    almoco_inserido = False
    for step in passos_origem:
        copia = dict(step)
        if str(copia.get("type", "")) != "stop" or almoco_inserido:
            resultado.append(copia)
            continue

        chegada_txt = str(copia.get("dyn_chegada", copia.get("chegada", "")) or "")
        saida_txt = str(copia.get("dyn_saida", copia.get("saida", "")) or "")
        try:
            chegada_min = parse_time_to_mins(chegada_txt)
            saida_min = parse_time_to_mins(saida_txt)
            if saida_min < chegada_min:
                saida_min += 24 * 60
            tempo_local = max(0, int(round(float(copia.get("tempo_local", 0) or 0))))
        except Exception:
            resultado.append(copia)
            continue

        # Sinal típico da lógica antiga: duração exibida = serviço + ~60 min,
        # com chegada antes do meio-dia e saída já depois das 13h.
        extra = (saida_min - chegada_min) - tempo_local
        almoco_oculto = (
            chegada_min < 12 * 60
            and saida_min >= 13 * 60
            and tempo_local > 0
            and extra >= 50
        )
        if not almoco_oculto:
            resultado.append(copia)
            continue

        fim_servico = chegada_min + tempo_local
        inicio_almoco = fim_servico
        fim_almoco = min(saida_min, inicio_almoco + 60)
        if fim_almoco - inicio_almoco < 45:
            resultado.append(copia)
            continue

        # Corrige apenas o horário exibido da parada; a pausa vira uma linha própria.
        fim_servico_txt = format_mins_to_time(fim_servico)
        if copia.get("dyn_saida") not in (None, ""):
            copia["dyn_saida"] = fim_servico_txt
        copia["saida"] = fim_servico_txt
        resultado.append(copia)
        resultado.append({
            "type": "lunch",
            "chegada": format_mins_to_time(inicio_almoco),
            "saida": format_mins_to_time(fim_almoco),
            "dyn_chegada": format_mins_to_time(inicio_almoco),
            "dyn_saida": format_mins_to_time(fim_almoco),
            "virtual_relatorio": True,
        })
        almoco_inserido = True

    return resultado

def montar_relatorio_rota(route_steps, concluidos):
    linhas, numero_parada = [], 0
    for step in _expandir_almoco_explicito_para_relatorio(route_steps):
        if step.get("type") == "lunch":
            linhas.append({"Parada": "Almoço", "Local": "PAUSA PARA ALMOÇO", "Chegada": step.get("dyn_chegada", step.get("chegada", "")), "Saída": step.get("dyn_saida", step.get("saida", "")), "Ação": "PAUSA"})
            continue
        if step.get("type") == "return":
            linhas.append({"Parada": "Retorno", "Local": step.get("destino", ""), "Chegada": step.get("dyn_chegada", step.get("chegada", "")), "Saída": step.get("dyn_saida", step.get("saida", "")), "Ação": "RETORNO", "Distância (km)": step.get("dist", 0)})
            continue
        numero_parada += 1
        acoes = step.get("actions", []) or [("DESLOCAMENTO", {})]
        for acao, tarefa in acoes:
            card_id = str(tarefa.get("id", ""))
            linhas.append({
                "Parada": numero_parada, "Local": step.get("destino", ""),
                "Chegada": step.get("dyn_chegada", step.get("chegada", "")),
                "Saída": step.get("dyn_saida", step.get("saida", "")),
                "Distância (km)": round(float(step.get("dist", 0) or 0), 2),
                "Ação": acao, "Obra": tarefa.get("Obra", ""), "Materiais": tarefa.get("Materiais", ""),
                "Status": f"Concluída às {concluidos[card_id]}" if card_id in concluidos else "Pendente",
            })
    return pd.DataFrame(linhas)


def _lista_natural_rota(itens):
    itens = [str(item).strip() for item in itens if str(item).strip()]
    unicos = []
    vistos = set()
    for item in itens:
        chave = remover_acentos(item).upper()
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(item)
    if not unicos:
        return ""
    if len(unicos) == 1:
        return unicos[0]
    if len(unicos) == 2:
        return f"{unicos[0]} e {unicos[1]}"
    return ", ".join(unicos[:-1]) + f" e {unicos[-1]}"


def _materiais_resumo_rota(tarefas, limite_itens=5, limite_chars=210):
    itens = []
    vistos = set()
    for tarefa in tarefas:
        texto = str(tarefa.get("Materiais", "") or "")
        for parte in re.split(r"\s*\|\s*|[\r\n]+", texto):
            item = re.sub(r"\s+", " ", parte).strip(" -•\t")
            if not item:
                continue
            chave = remover_acentos(item).upper()
            if chave not in vistos:
                vistos.add(chave)
                itens.append(item)
    if not itens:
        return ""
    exibidos = itens[:limite_itens]
    texto = "; ".join(exibidos)
    if len(itens) > limite_itens:
        texto += f"; +{len(itens) - limite_itens} item(ns)"
    if len(texto) > limite_chars:
        texto = texto[:limite_chars - 3].rstrip() + "..."
    return texto


def _descricao_rapida_parada(step):
    """Resume uma parada em uma frase simples para o motorista."""
    acoes = step.get("actions", []) or []
    coletas = [t for acao, t in acoes if acao == "COLETAR"]
    entregas = [t for acao, t in acoes if acao == "ENTREGAR"]
    partes = []

    if coletas:
        destinos = _lista_natural_rota([t.get("Destino", "") for t in coletas])
        partes.append(f"COLETAR materiais para {destinos}" if destinos else "COLETAR materiais")
    if entregas:
        obras = _lista_natural_rota([t.get("Obra", "") for t in entregas])
        partes.append(f"ENTREGAR materiais - {obras}" if obras else "ENTREGAR materiais")
    return " e ".join(partes) or "PARADA OPERACIONAL"


def _passos_resumo_rota(route_steps, ponto_saida="ESCRITÓRIO", retornar_base=True):
    """Converte a rota em passos humanos: primeiro aqui, depois ali, sem repetir cada demanda."""
    passos = []
    ponto_saida = str(ponto_saida or "ESCRITÓRIO").strip() or "ESCRITÓRIO"
    tem_retorno = False

    for step in _expandir_almoco_explicito_para_relatorio(route_steps):
        tipo = step.get("type")
        if tipo == "lunch":
            chegada = step.get("dyn_chegada", step.get("chegada", "12:00"))
            saida = step.get("dyn_saida", step.get("saida", "13:00"))
            passos.append({
                "local": "ALMOÇO",
                "acao": "PAUSA DE 1 HORA PARA ALMOÇO",
                "horario": f"{chegada} - {saida}" if chegada or saida else "",
                "tipo": "pausa",
            })
            continue

        if tipo == "return":
            local = str(step.get("destino", "") or ponto_saida).strip() or ponto_saida
            chegada = step.get("dyn_chegada", step.get("chegada", ""))
            passos.append({
                "local": local,
                "acao": f"RETORNAR para {local}",
                "horario": f"Chegada {chegada}" if chegada else "",
                "tipo": "retorno",
            })
            tem_retorno = True
            continue

        local = str(step.get("destino", "") or "").strip()
        if not local:
            continue
        chegada = step.get("dyn_chegada", step.get("chegada", ""))
        saida = step.get("dyn_saida", step.get("saida", ""))
        if chegada and saida:
            horario = f"Chega {chegada} • sai {saida}"
        else:
            horario = chegada or saida or ""
        passos.append({
            "local": local,
            "acao": _descricao_rapida_parada(step),
            "horario": horario,
            "tipo": "parada",
        })

    # Em algumas rotas o retorno é apenas uma configuração e não aparece como step explícito.
    if retornar_base and not tem_retorno:
        ultimo_local = str(passos[-1].get("local", "")) if passos else ""
        if remover_acentos(ultimo_local).upper() != remover_acentos(ponto_saida).upper():
            passos.append({"local": ponto_saida, "acao": f"RETORNAR para {ponto_saida}", "horario": "", "tipo": "retorno"})

    return passos


def montar_resumo_sequencial_rota(route_steps, ponto_saida="ESCRITÓRIO", retornar_base=True):
    """Tabela curta usada no início do PDF/Excel: uma linha por parada, na ordem real."""
    passos = _passos_resumo_rota(route_steps, ponto_saida, retornar_base)
    linhas = []
    for numero, passo in enumerate(passos, start=1):
        linhas.append({
            "Etapa": numero,
            "Local": passo.get("local", ""),
            "O que fazer": passo.get("acao", ""),
            "Horário previsto": passo.get("horario", ""),
        })
    return pd.DataFrame(linhas, columns=["Etapa", "Local", "O que fazer", "Horário previsto"])


def _criar_pdf_resumo_rota_operacional(route_steps, dados_rota):
    """PDF compacto para o motorista: sequência da rota, ações e materiais essenciais."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, KeepTogether
    except (ImportError, ModuleNotFoundError):
        df = montar_relatorio_rota(route_steps, {})
        return _criar_pdf_relatorio("Resumo Operacional da Rota", [("Rota", df)])

    azul = colors.HexColor("#2563EB")
    azul_escuro = colors.HexColor("#0F172A")
    azul_claro = colors.HexColor("#EFF6FF")
    azul_borda = colors.HexColor("#BFDBFE")
    texto = colors.HexColor("#1E293B")
    texto_secundario = colors.HexColor("#64748B")
    branco = colors.white

    estilos = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        "RotaTitulo", parent=estilos["Title"], fontName="Helvetica-Bold",
        fontSize=15, leading=18, textColor=branco, alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle(
        "RotaMeta", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=8, leading=10, textColor=texto_secundario, alignment=TA_LEFT,
    )
    sequencia_style = ParagraphStyle(
        "RotaSequencia", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=9, leading=12, textColor=azul_escuro, alignment=TA_LEFT,
    )
    parada_titulo_style = ParagraphStyle(
        "ParadaTitulo", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=9.2, leading=11, textColor=azul_escuro,
    )
    parada_acao_style = ParagraphStyle(
        "ParadaAcao", parent=estilos["Normal"], fontName="Helvetica-Bold",
        fontSize=8.3, leading=10.5, textColor=azul,
    )
    parada_texto_style = ParagraphStyle(
        "ParadaTexto", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=7.7, leading=9.5, textColor=texto,
    )
    horario_style = ParagraphStyle(
        "ParadaHorario", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=7.4, leading=9, textColor=texto_secundario,
    )

    def seguro(valor):
        return html_escape(str(valor or "").replace("—", "-").replace("–", "-").replace("•", "-")).replace("\n", "<br/>")

    origem = str(dados_rota.get("ponto_saida", "ESCRITÓRIO") or "ESCRITÓRIO")
    passos = _passos_resumo_rota(route_steps, origem, dados_rota.get("retornar_base", True))
    locais_seq = [str(p.get("local", "")) for p in passos if p.get("local")]
    sequencia = "  >  ".join(locais_seq)

    saida = io.BytesIO()
    documento = SimpleDocTemplate(
        saida, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=18 * mm, bottomMargin=13 * mm,
        title="Resumo Operacional da Rota", author="APROAR Engenharia",
    )
    largura_util = A4[0] - 24 * mm

    def cabecalho_rodape(canvas, doc):
        largura, altura = A4
        canvas.saveState()
        canvas.setFillColor(azul_escuro)
        canvas.rect(0, altura - 9 * mm, largura, 9 * mm, stroke=0, fill=1)
        canvas.setFillColor(branco)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawString(12 * mm, altura - 5.8 * mm, "APROAR ENGENHARIA")
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(largura - 12 * mm, altura - 5.8 * mm, "Torre de Controle Logístico")
        canvas.setStrokeColor(azul_borda)
        canvas.line(12 * mm, 8 * mm, largura - 12 * mm, 8 * mm)
        canvas.setFillColor(texto_secundario)
        canvas.setFont("Helvetica", 6.8)
        canvas.drawString(12 * mm, 4.5 * mm, "Resumo rápido da rota do motorista")
        canvas.drawRightString(largura - 12 * mm, 4.5 * mm, f"Página {doc.page}")
        canvas.restoreState()

    faixa = Table([[Paragraph("RESUMO OPERACIONAL DA ROTA", titulo_style)]], colWidths=[largura_util], rowHeights=[27])
    faixa.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), azul),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))

    data = dados_rota.get("data", "")
    inicio = dados_rota.get("inicio", "")
    termino = dados_rota.get("termino", "")
    km = dados_rota.get("km", 0)
    veiculo = dados_rota.get("veiculo", "")
    meta = f"Data: {data}  |  Veículo: {veiculo}  |  Saída: {inicio}  |  Término previsto: {termino}  |  {float(km or 0):.1f} km"

    elementos = [
        faixa,
        Spacer(1, 5),
        Paragraph(seguro(meta), meta_style),
        Spacer(1, 7),
        Table([[Paragraph("ORDEM DA ROTA - LEIA DE CIMA PARA BAIXO", parada_titulo_style)]], colWidths=[largura_util], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), azul_claro),
            ("BOX", (0, 0), (-1, -1), 0.6, azul_borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])),
        Spacer(1, 3),
        Paragraph(seguro(sequencia), sequencia_style),
        Spacer(1, 8),
    ]

    for numero, passo in enumerate(passos, start=1):
        local = str(passo.get("local", "") or "")
        descricao = str(passo.get("acao", "") or "")
        horario = str(passo.get("horario", "") or "")
        tipo = passo.get("tipo", "parada")

        if tipo == "retorno":
            cor_lateral = azul_escuro
        elif tipo == "pausa":
            cor_lateral = texto_secundario
        elif descricao.upper().startswith("COLETAR"):
            cor_lateral = colors.HexColor("#F59E0B")
        elif descricao.upper().startswith("ENTREGAR"):
            cor_lateral = colors.HexColor("#16A34A")
        else:
            cor_lateral = azul

        linhas = [[
            Paragraph(seguro(f"{numero}. {local}"), parada_titulo_style),
            Paragraph(seguro(horario), horario_style),
        ], [Paragraph(seguro(descricao), parada_acao_style), ""]]

        card = Table(linhas, colWidths=[largura_util * 0.78, largura_util * 0.22])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), branco),
            ("BACKGROUND", (0, 0), (-1, 0), azul_claro),
            ("SPAN", (0, 1), (1, 1)),
            ("BOX", (0, 0), (-1, -1), 0.65, azul_borda),
            ("LINEBEFORE", (0, 0), (0, -1), 3, cor_lateral),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elementos.extend([KeepTogether([card, Spacer(1, 4)])])

    documento.build(elementos, onFirstPage=cabecalho_rodape, onLaterPages=cabecalho_rodape)
    return saida.getvalue()

def calcular_distancia_km(lat1, lon1, lat2, lon2):
    dLat = math.radians(lat2 - lat1)
    dLon = math.radians(lon2 - lon1)
    a = math.sin(dLat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dLon/2)**2
    return 6371.0 * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def resumir_rua_rastreador(endereco):
    """Reduz o endereço da Protege à rua e ao número para caber no marcador."""
    endereco_limpo = re.sub(r"\s+", " ", str(endereco or "")).strip(" ,-|")
    if not endereco_limpo:
        return "Rua atual não informada"

    trecho_rua = re.split(r"\s+(?:-|–|—|\|)\s+", endereco_limpo, maxsplit=1)[0].strip()
    partes = [parte.strip() for parte in trecho_rua.split(",") if parte.strip()]
    if len(partes) >= 2 and re.fullmatch(r"(?:N[º°.]?\s*)?\d+[A-Z]?|S/?N", partes[1], flags=re.IGNORECASE):
        trecho_rua = f"{partes[0]}, {partes[1]}"
    elif len(partes) > 1:
        trecho_rua = partes[0]
    return trecho_rua[:90]


def inferir_destino_provavel_por_distancia(
    lat_atual, lon_atual, route_steps, locais, ponto_saida, concluidos
):
    """Escolhe a parada pendente mais próxima da posição GPS, sem usar a ordem da rota."""
    candidatos = []
    concluidos = concluidos or {}

    for etapa in route_steps or []:
        destino = str(etapa.get("destino", "") or "").strip()
        if etapa.get("type") != "stop" or not destino or destino == ponto_saida:
            continue

        tarefas = etapa.get("actions") or []
        if not tarefas or not any(
            str(tarefa.get("id", "")) not in concluidos for _acao, tarefa in tarefas
        ):
            continue

        coordenadas = (locais or {}).get(destino)
        if not isinstance(coordenadas, (list, tuple)) or len(coordenadas) < 2:
            continue
        try:
            lat_destino, lon_destino = float(coordenadas[0]), float(coordenadas[1])
            distancia = calcular_distancia_km(
                float(lat_atual), float(lon_atual), lat_destino, lon_destino
            )
        except (TypeError, ValueError, IndexError):
            continue
        if math.isfinite(distancia):
            candidatos.append((distancia, destino))

    if not candidatos:
        return "", None
    distancia, destino = min(candidatos, key=lambda item: item[0])
    return destino, distancia

def normalizar_geometria_mapa(geometria, referencias=None):
    """Normaliza uma geometria para o formato [lat, lon] esperado pelo Folium.

    O GeoJSON do OSRM devolve [lon, lat]. Versões anteriores do app armazenaram
    algumas dessas geometrias sem inverter os eixos; nesses casos a linha ficava
    milhares de quilômetros fora do mapa e parecia ter desaparecido. A função
    também recupera automaticamente rotas antigas já salvas no Supabase.
    """
    pontos = []
    for ponto in geometria or []:
        try:
            if ponto is None or len(ponto) < 2:
                continue
            a, b = float(ponto[0]), float(ponto[1])
            if math.isfinite(a) and math.isfinite(b):
                pontos.append([a, b])
        except (TypeError, ValueError, IndexError):
            continue

    if len(pontos) < 2:
        return pontos

    refs = []
    for ponto in referencias or []:
        try:
            if ponto is None or len(ponto) < 2:
                continue
            lat, lon = float(ponto[0]), float(ponto[1])
            if math.isfinite(lat) and math.isfinite(lon):
                refs.append([lat, lon])
        except (TypeError, ValueError, IndexError):
            continue

    if not refs:
        return pontos

    # Compara a geometria como veio e com os eixos invertidos. A primeira e a
    # última coordenada são suficientes para reconhecer o erro do GeoJSON/OSRM.
    def pontuacao(candidatos):
        try:
            return (
                calcular_distancia_km(candidatos[0][0], candidatos[0][1], refs[0][0], refs[0][1])
                + calcular_distancia_km(candidatos[-1][0], candidatos[-1][1], refs[-1][0], refs[-1][1])
            )
        except Exception:
            return float('inf')

    invertidos = [[p[1], p[0]] for p in pontos]
    score_normal = pontuacao(pontos)
    score_invertido = pontuacao(invertidos)
    if score_invertido + 0.5 < score_normal:
        pontos = invertidos
        score_normal = score_invertido

    # Uma geometria sem relação com a rota não deve ser desenhada. O mapa usa
    # então a ligação direta entre as paradas até conseguir consultar a malha viária.
    if score_normal > 80.0:
        return []
    return pontos

PLACA_DAVI = "TIF-2123"
HORA_INICIO_ROTA_DAVI = "08:00"
HORA_PREPARACAO_INICIO = "07:30"
HORA_PREPARACAO_FIM = "08:00"

# Precisa existir antes da renderização do App do Motorista. A rota /davi
# encerra a execução com st.stop() antes de chegar ao bloco da Torre (PC).
# Na V17 esta constante estava declarada somente depois do app mobile, causando
# NameError ao montar o banner de ETA no celular.
LIMITE_EXPEDIENTE_DAVI_MIN = 17 * 60

def ajustar_tempo_deslocamento_operacional(dist_km, duracao_api_min, horario_partida_min=None):
    """Evita ETAs urbanos otimistas demais sem substituir a matriz viária.

    A API continua sendo a fonte principal. Este piso operacional só corrige casos
    incompatíveis com a rotina real (saída da vaga, semáforos, acesso/estacionamento
    e tráfego urbano), principalmente em trechos curtos onde 4–5 km podem aparecer
    como 6–8 minutos na matriz.
    """
    try:
        dist = max(0.0, float(dist_km or 0.0))
    except (TypeError, ValueError):
        dist = 0.0
    try:
        api = max(0.0, float(duracao_api_min or 0.0))
    except (TypeError, ValueError):
        api = 0.0

    if dist <= 0.10:
        return 0.0

    # Em trechos curtos de Fortaleza a velocidade média porta-a-porta é bem menor
    # que a velocidade de circulação pura. Em vias mais longas, o piso fica mais
    # permissivo para não penalizar corredores expressos/BR.
    if dist <= 6.0:
        piso = (dist / 24.0) * 60.0 + 2.5
    elif dist <= 12.0:
        piso = (dist / 30.0) * 60.0 + 3.0
    else:
        piso = (dist / 40.0) * 60.0 + 3.0

    # Pequena margem adicional nos horários de pico. Não é um segundo cálculo de
    # trânsito: apenas impede que a previsão operacional fique excessivamente justa.
    try:
        hm = int(round(float(horario_partida_min))) % (24 * 60)
    except (TypeError, ValueError):
        hm = 8 * 60
    pico = (7 * 60 <= hm <= 9 * 60 + 30) or (16 * 60 + 30 <= hm <= 18 * 60 + 30)
    if pico:
        piso *= 1.06

    # A matriz pode ser mais lenta que o piso; nesse caso respeitamos integralmente
    # o trânsito retornado pela fonte viária. Arredondar para minuto inteiro deixa o
    # roteiro mais legível e evita horários como 08:07 derivados de 6,6 minutos.
    return float(max(1, int(math.ceil(max(api, piso)))))


def atualizar_tempos_deslocamento_operacionais(route_steps, start_time_str="08:00"):
    """Revalida os deslocamentos de rotas novas e já salvas.

    Isso é importante porque uma rota persistida no Supabase pode ter sido calculada
    por uma versão anterior. Guardamos o valor original em ``travel_mins_api`` e
    aplicamos o piso com base no horário planejado de cada perna.
    """
    try:
        atual = parse_time_to_mins(str(start_time_str or "08:00"))
    except Exception:
        atual = 8 * 60

    for indice, step in enumerate(route_steps or []):
        tipo = str(step.get("type", ""))
        if tipo == "lunch":
            ini = str(step.get("chegada", "") or "")
            fim = str(step.get("saida", "") or "")
            try:
                ini_min = parse_time_to_mins(ini) if ini else atual
                fim_min = parse_time_to_mins(fim) if fim else ini_min + 60
                if fim_min <= ini_min:
                    fim_min = ini_min + 60
                atual = max(atual, ini_min) + max(60, fim_min - ini_min)
            except Exception:
                atual += 60
            continue

        dist = float(step.get("dist", 0) or 0)
        original = step.get("travel_mins_api", step.get("travel_mins", 0))
        try:
            original = float(original or 0)
        except (TypeError, ValueError):
            original = 0.0
        step["travel_mins_api"] = original

        eh_preparacao = (
            tipo == "stop" and indice == 0 and dist <= 0.05 and original <= 0.5
        )
        if eh_preparacao:
            step["travel_mins"] = 0.0
            try:
                saida = str(step.get("saida", "") or "")
                if saida:
                    atual = max(atual, parse_time_to_mins(saida))
            except Exception:
                pass
            continue

        ajustado = ajustar_tempo_deslocamento_operacional(dist, original, atual)
        step["travel_mins"] = ajustado
        atual += ajustado

        if tipo == "return":
            continue
        if tipo == "stop":
            try:
                atual += max(0.0, float(step.get("tempo_local", 0) or 0))
            except (TypeError, ValueError):
                pass

    return route_steps

@st.cache_data(ttl=10, show_spinner=False)
def obter_hora_inicio_rota(data_rota):
    """A rota do Davi é planejada para iniciar às 08:00.

    No dia da execução, se o rastreador já tiver registrado a saída real da
    Strada do Davi, esse horário real passa a alimentar as previsões dinâmicas.
    Para rota futura ou sem saída registrada, usa 08:00.
    """
    if data_rota == DATA_HOJE_REAL_STR:
        inicio_davi = fetch_one(
            "SELECT hora_inicio FROM inicio_movimento WHERE placa=:placa AND data=:data",
            {"placa": PLACA_DAVI, "data": data_rota},
        )
        if inicio_davi and inicio_davi[0]:
            return str(inicio_davi[0])

    return HORA_INICIO_ROTA_DAVI

def aplicar_tempos_dinamicos(route_steps, dict_concluidos, start_time_str):
    """Atualiza ETAs sem empurrar a rota de amanhã para o horário de agora.

    Para a rota de hoje, as etapas ainda pendentes partem do horário real atual.
    Para uma rota futura (ex.: planejamento do dia seguinte após 18h), conserva-se
    a linha do tempo planejada. A preparação inicial também permanece no horário
    planejado, evitando mensagens confusas como 20:37–21:07 para uma preparação
    originalmente prevista para 07:30–08:00.
    """
    rota_eh_hoje = DATA_REF_ROTA_DATE == AGORA_REAL.date()
    agora_min = AGORA_REAL.hour * 60 + AGORA_REAL.minute
    agora_min_efetivo = (13 * 60 if 12 * 60 <= agora_min < 13 * 60 else agora_min) if rota_eh_hoje else None
    current_min = parse_time_to_mins(start_time_str) if start_time_str else (8 * 60)

    for indice_step, step in enumerate(route_steps):
        if step['type'] == 'lunch':
            chegada_planejada = str(step.get('chegada', '12:00') or '12:00')
            saida_planejada = str(step.get('saida', '13:00') or '13:00')
            try:
                ini_planejado = parse_time_to_mins(chegada_planejada)
                fim_planejado = parse_time_to_mins(saida_planejada)
                if fim_planejado <= ini_planejado:
                    fim_planejado = ini_planejado + 60
                duracao_almoco = max(60, fim_planejado - ini_planejado)
            except Exception:
                ini_planejado, duracao_almoco = 12 * 60, 60

            inicio_almoco = max(current_min, ini_planejado)
            fim_almoco = inicio_almoco + duracao_almoco
            step['dyn_chegada'] = format_mins_to_time(inicio_almoco)
            step['dyn_saida'] = format_mins_to_time(fim_almoco)
            step['is_concluded'] = False
            current_min = fim_almoco
            continue

        if step['type'] == 'return':
            # Se a rota está sendo acompanhada hoje, o deslocamento de retorno
            # começa do horário REAL atual, nunca de um ETA antigo já ultrapassado.
            if agora_min_efetivo is not None and current_min < agora_min_efetivo:
                current_min = agora_min_efetivo
            arr_min = current_min + step.get('travel_mins', 0)
            if current_min <= 12 * 60 and arr_min > 12 * 60:
                arr_min = max(arr_min + 60, 13 * 60)
            if 12 * 60 <= arr_min < 13 * 60:
                arr_min = 13 * 60

            step['dyn_chegada'] = format_mins_to_time(arr_min)
            step['dyn_saida'] = step['dyn_chegada']
            step['is_concluded'] = False
            current_min = arr_min
            continue

        # A primeira etapa é a preparação no pátio. Ela não deve virar um
        # "ETA atualizado" com o horário atual quando a rota já avançou.
        eh_preparacao_inicial = (
            indice_step == 0
            and float(step.get('dist', 0) or 0) <= 0.05
            and float(step.get('travel_mins', 0) or 0) <= 0.5
        )
        if eh_preparacao_inicial:
            chegada_base = str(step.get('chegada', '') or '')
            saida_base = str(step.get('saida', '') or '')
            step['dyn_chegada'] = chegada_base
            step['dyn_saida'] = saida_base
            step['is_concluded'] = False
            saida_min = parse_time_to_mins(saida_base) if saida_base else 0
            current_min = max(current_min, saida_min)
            continue

        concluded_times = []
        has_pending = False
        for _acao, tarefa in step.get('actions', []):
            card_id = str(tarefa.get('id', ''))
            if card_id in dict_concluidos:
                concluded_times.append(parse_time_to_mins(dict_concluidos[card_id]))
            else:
                has_pending = True

        if not has_pending and concluded_times:
            max_c = max(concluded_times)
            step['dyn_chegada'] = "Concluído"
            step['dyn_saida'] = format_mins_to_time(max_c)
            current_min = max(current_min, max_c)
            step['is_concluded'] = True
            continue

        if 12 * 60 <= current_min < 13 * 60:
            current_min = 13 * 60

        # A primeira etapa ainda pendente deve sair de AGORA. O comportamento
        # anterior apenas elevava a chegada para o horário atual e, com isso,
        # podia até eliminar indevidamente o tempo de deslocamento até a parada.
        if agora_min_efetivo is not None and current_min < agora_min_efetivo:
            current_min = agora_min_efetivo

        travel = step.get('travel_mins', 0)
        arr_min = current_min + travel

        if current_min <= 12 * 60 and arr_min > 12 * 60:
            arr_min = max(arr_min + 60, 13 * 60)
        if 12 * 60 <= arr_min < 13 * 60:
            arr_min = 13 * 60

        service = step.get('tempo_local', 0)
        dep_min = arr_min + service
        proximo_e_almoco = (
            indice_step + 1 < len(route_steps)
            and route_steps[indice_step + 1].get('type') == 'lunch'
        )
        # Compatibilidade com rotas antigas: se não houver uma etapa explícita de
        # almoço, mantém a hora de pausa embutida no ETA. Rotas novas exibem a pausa
        # como etapa própria e, portanto, não somam 60 min aqui.
        if arr_min <= 12 * 60 and dep_min > 12 * 60 and not proximo_e_almoco:
            dep_min = max(dep_min + 60, 13 * 60)

        step['dyn_chegada'] = format_mins_to_time(arr_min)
        step['dyn_saida'] = format_mins_to_time(dep_min)
        current_min = dep_min
        step['is_concluded'] = False

    return route_steps, current_min


@st.cache_data(ttl=5, show_spinner=False)
def carregar_paradas_rastreadas_rota(data_rota, placa=PLACA_DAVI):
    """Carrega as visitas reais do rastreador, mais recentes primeiro."""
    try:
        return get_df(
            """
            SELECT id, local, hora_chegada, hora_saida
            FROM rastreio_paradas
            WHERE data=:data AND placa=:placa
            ORDER BY id DESC
            """,
            {"data": data_rota, "placa": placa},
        )
    except Exception:
        return pd.DataFrame(columns=["id", "local", "hora_chegada", "hora_saida"])


@st.cache_data(ttl=5, show_spinner=False)
def carregar_conclusoes_rota(data_rota):
    """Uma leitura curta alimenta Torre, recálculo e aplicativo do Davi."""
    try:
        return get_df(
            "SELECT id, hora_conclusao FROM historico_concluidos WHERE data_conclusao = :data",
            {"data": data_rota},
        )
    except Exception:
        return pd.DataFrame(columns=["id", "hora_conclusao"])


@st.cache_data(ttl=5, show_spinner=False)
def carregar_rota_publicada_mobile(data_rota):
    """Evita ler novamente todo o JSON da rota a cada toque no celular."""
    try:
        return fetch_one(
            "SELECT json_route, json_locais, json_geometria, json_enderecos, total_km "
            "FROM rota_ativa WHERE id = 1 AND data_rota = :data",
            {"data": data_rota},
        )
    except Exception:
        return None


def _normalizar_local_rastreio(valor):
    return re.sub(r"\s+", " ", remover_acentos(str(valor or "")).upper()).strip()


def _duracao_horarios_minutos(inicio, fim):
    ini = parse_time_to_mins(str(inicio or ""))
    ter = parse_time_to_mins(str(fim or ""))
    if ter < ini:
        ter += 24 * 60
    return max(0, ter - ini)


def _formatar_duracao_parada(minutos):
    minutos = max(0, int(minutos or 0))
    horas, resto = divmod(minutos, 60)
    if horas and resto:
        return f"{horas}h{resto:02d}"
    if horas:
        return f"{horas}h"
    return f"{resto} min"


def obter_status_rastreio_local(df_paradas, local, data_rota):
    """Retorna chegada, saída e permanência real da visita mais recente ao local."""
    if df_paradas is None or df_paradas.empty or not str(local or "").strip():
        return None

    chave = _normalizar_local_rastreio(local)
    candidatos = df_paradas[
        df_paradas["local"].astype(str).map(_normalizar_local_rastreio) == chave
    ]
    if candidatos.empty:
        return None

    linha = candidatos.iloc[0]
    chegada = str(linha.get("hora_chegada", "") or "").strip()
    saida = str(linha.get("hora_saida", "") or "").strip()
    if not chegada:
        return None

    if saida and saida.lower() not in {"none", "nan", "nat"}:
        duracao = _duracao_horarios_minutos(chegada, saida)
        return {
            "aberta": False,
            "chegada": chegada,
            "saida": saida,
            "duracao_min": duracao,
            "duracao": _formatar_duracao_parada(duracao),
        }

    duracao = None
    if str(data_rota) == DATA_HOJE_REAL_STR:
        agora_hm = AGORA_REAL.strftime("%H:%M")
        duracao = _duracao_horarios_minutos(chegada, agora_hm)

    return {
        "aberta": True,
        "chegada": chegada,
        "saida": "",
        "duracao_min": duracao,
        "duracao": _formatar_duracao_parada(duracao) if duracao is not None else "",
    }


def html_status_rastreio_local(status):
    if not status:
        return ""
    chegada = html_escape(str(status.get("chegada", "")))
    if status.get("aberta"):
        duracao = html_escape(str(status.get("duracao", "")))
        detalhe = f" • ⏱️ no local há {duracao}" if duracao else ""
        return f"<div class='rastreio-real'>📍 Chegada real: <b>{chegada}</b>{detalhe}</div>"
    saida = html_escape(str(status.get("saida", "")))
    duracao = html_escape(str(status.get("duracao", "")))
    return f"<div class='rastreio-real'>📍 Chegada: <b>{chegada}</b> • 🚚 Saída: <b>{saida}</b> • ⏱️ Permanência: <b>{duracao}</b></div>"

def renderizar_banner_eta(hora_atual_str, nova_previsao_str, final_dyn_min):
    if not hora_atual_str:
        return
    cor_previsao = "#16a34a" if final_dyn_min <= LIMITE_EXPEDIENTE_DAVI_MIN else "#f59e0b" if final_dyn_min <= (LIMITE_EXPEDIENTE_DAVI_MIN + 30) else "#ef4444"
    rota_futura = DATA_REF_ROTA_DATE > AGORA_REAL.date()
    rotulo_referencia = "📅 Rota planejada:" if rota_futura else "🕒 Horário atual:"
    valor_referencia = DATA_REF_ROTA_STR if rota_futura else hora_atual_str

    st.markdown(f'''
        <div class="aproar-eta-card">
            <div class="aproar-eta-side">
                <div class="aproar-eta-label">{rotulo_referencia.replace('📅 ', '').replace('🕒 ', '')}</div>
                <div class="aproar-eta-value">{valor_referencia}</div>
            </div>
            <div class="aproar-eta-route">→</div>
            <div class="aproar-eta-side right">
                <div class="aproar-eta-label">Término previsto</div>
                <div class="aproar-eta-value" style="color:{cor_previsao};">{nova_previsao_str}</div>
            </div>
        </div>
    ''', unsafe_allow_html=True)

# =====================================================================
# MARCAÇÕES COMPARTILHADAS DO APP DO DAVI
# =====================================================================
SQL_TABELA_CHECKINS_DAVI = """
CREATE TABLE IF NOT EXISTS roteiro_checkins_davi (
    data_rota TEXT NOT NULL,
    etapa_indice INTEGER NOT NULL,
    destino TEXT NOT NULL,
    marcado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    PRIMARY KEY (data_rota, etapa_indice)
)
"""

@st.cache_resource(show_spinner=False)
def garantir_tabela_checkins_davi():
    """Cria uma única vez a estrutura compartilhada entre celular e escritório."""
    execute_db(SQL_TABELA_CHECKINS_DAVI)
    return True

@st.cache_data(ttl=5, show_spinner=False)
def carregar_checkins_davi(data_rota):
    rows = fetch_all(
        """
        SELECT etapa_indice, destino,
               TO_CHAR(marcado_em AT TIME ZONE 'America/Fortaleza', 'HH24:MI') AS hora,
               EXTRACT(EPOCH FROM marcado_em) AS instante
        FROM roteiro_checkins_davi
        WHERE data_rota = :data
        ORDER BY etapa_indice
        """,
        {"data": data_rota},
    )
    return {
        int(row[0]): {"destino": str(row[1]), "hora": str(row[2] or ""), "instante": float(row[3] or 0)}
        for row in rows
    }

def filtrar_checkins_da_rota(route_steps, checkins):
    """Impede que uma marca antiga seja ligada a outro destino após refazer a rota."""
    return {
        indice: dados
        for indice, dados in checkins.items()
        if 0 <= indice < len(route_steps)
        and route_steps[indice].get("type") not in {"lunch", "return"}
        and str(route_steps[indice].get("destino", "")) == dados.get("destino", "")
    }

def salvar_checkin_davi(data_rota, etapa_indice, destino, feita):
    if feita:
        execute_db(
            """
            INSERT INTO roteiro_checkins_davi (data_rota, etapa_indice, destino, marcado_em)
            VALUES (:data, :indice, :destino, NOW())
            ON CONFLICT (data_rota, etapa_indice)
            DO UPDATE SET destino = EXCLUDED.destino, marcado_em = NOW()
            """,
            {"data": data_rota, "indice": etapa_indice, "destino": destino},
        )
    else:
        execute_db(
            """
            DELETE FROM roteiro_checkins_davi
            WHERE data_rota = :data AND etapa_indice = :indice AND destino = :destino
            """,
            {"data": data_rota, "indice": etapa_indice, "destino": destino},
        )
    try:
        carregar_checkins_davi.clear()
    except Exception:
        pass


# =====================================================================
# COMPROVANTE DE ENTREGA — POWER AUTOMATE / ONEDRIVE
# Fluxo simples para o motorista: 1 recebedor por demanda + várias fotos.
# =====================================================================
def _nome_seguro_comprovante(valor, limite=45):
    texto = remover_acentos(str(valor or "")).upper().strip()
    texto = re.sub(r"[^A-Z0-9]+", "-", texto).strip("-")
    return (texto[:limite] or "SEM-NOME")


def _separar_materiais_comprovante(valor):
    """Transforma o campo Materiais em opções curtas para o motorista."""
    texto = str(valor or "").strip()
    if not texto:
        return []

    # O Trello normalmente separa itens por |. Também aceita quebras de linha.
    partes = re.split(r"\s*\|\s*|[\r\n]+", texto)
    itens, vistos = [], set()
    for parte in partes:
        item = re.sub(r"\s+", " ", str(parte or "")).strip(" •-\t")
        if not item:
            continue
        chave = remover_acentos(item).upper()
        if chave not in vistos:
            vistos.add(chave)
            itens.append(item)
    return itens


def agrupar_coletas_preparacao_exibicao(acoes):
    """Agrupa a preparação por unidade de destino + obra, sem misturar obras."""
    grupos = {}
    ordem = []
    for acao, tarefa in (acoes or []):
        if acao != "COLETAR":
            ordem.append(("acao", len(ordem), acao, tarefa))
            continue

        tarefa = dict(tarefa or {})
        demanda_id = str(tarefa.get("id", "") or "").strip()
        destino = canonicalizar_ponto_rota(tarefa.get("Destino", ""))
        obra = re.sub(r"\s+", " ", str(tarefa.get("Obra", "") or "")).strip()
        chave_obra = remover_acentos(obra).upper()
        # Demandas sem obra identificável permanecem separadas para não juntar
        # cartões diferentes apenas porque têm o mesmo destino.
        chave = (destino, chave_obra or f"SEM-OBRA:{demanda_id}")
        if chave not in grupos:
            grupos[chave] = {
                "tarefa": tarefa,
                "ids": [],
                "materiais": [],
                "vistos_materiais": set(),
                "cards": [],
            }
            ordem.append(("grupo", chave))

        grupo = grupos[chave]
        if demanda_id and demanda_id not in grupo["ids"]:
            grupo["ids"].append(demanda_id)
            grupo["cards"].append({
                "id": demanda_id,
                "obra": obra,
                "destino": destino,
                "materiais": _separar_materiais_comprovante(tarefa.get("Materiais", "")),
            })
        for material in _separar_materiais_comprovante(tarefa.get("Materiais", "")):
            chave_material = remover_acentos(material).upper()
            if chave_material not in grupo["vistos_materiais"]:
                grupo["vistos_materiais"].add(chave_material)
                grupo["materiais"].append(material)

    resultado = []
    for item in ordem:
        if item[0] == "acao":
            resultado.append((item[2], item[3]))
            continue
        grupo = grupos[item[1]]
        tarefa_agrupada = dict(grupo["tarefa"])
        tarefa_agrupada["Materiais"] = " | ".join(grupo["materiais"])
        tarefa_agrupada["_ids_agrupados"] = list(grupo["ids"])
        tarefa_agrupada["_cards_agrupados"] = list(grupo["cards"])
        tarefa_agrupada["_qtd_demandas_agrupadas"] = max(1, len(grupo["ids"]))
        resultado.append(("COLETAR", tarefa_agrupada))
    return resultado



def agrupar_acoes_por_obra_exibicao(acoes):
    """Agrupa cards da mesma obra em uma linha visual, sem perder o status individual.

    A chave inclui tipo da ação + unidade física + obra. Assim dois cards do Trello
    para a mesma obra (ex.: luva e pistola) aparecem juntos. A baixa continua sendo
    conferida pelo ID de cada card, então um grupo pode ficar PARCIAL: 1/2 concluído.
    """
    grupos = {}
    ordem = []
    for acao, tarefa_original in (acoes or []):
        tarefa = dict(tarefa_original or {})
        acao = str(acao or '').upper()
        demanda_id = str(tarefa.get('id', '') or '').strip()
        obra = re.sub(r"\s+", " ", str(tarefa.get('Obra', '') or '')).strip()
        unidade = canonicalizar_ponto_rota(
            tarefa.get('Destino', '') if acao == 'ENTREGAR' else tarefa.get('Origem', '')
        )
        chave_obra = remover_acentos(obra).upper().strip()
        # Sem obra confiável, não misturamos cards diferentes por acidente.
        chave = (acao, unidade, chave_obra or f"SEM-OBRA:{demanda_id}")
        if chave not in grupos:
            grupos[chave] = {
                'tarefa': tarefa,
                'ids': [],
                'cards': [],
                'materiais': [],
                'materiais_vistos': set(),
            }
            ordem.append(chave)
        grupo = grupos[chave]
        materiais_card = _separar_materiais_comprovante(tarefa.get('Materiais', ''))
        if demanda_id and demanda_id not in grupo['ids']:
            grupo['ids'].append(demanda_id)
            grupo['cards'].append({
                'id': demanda_id,
                'obra': obra,
                'origem': canonicalizar_ponto_rota(tarefa.get('Origem', '')),
                'destino': canonicalizar_ponto_rota(tarefa.get('Destino', '')),
                'materiais': materiais_card,
            })
        for material in materiais_card:
            chave_material = remover_acentos(material).upper().strip()
            if chave_material and chave_material not in grupo['materiais_vistos']:
                grupo['materiais_vistos'].add(chave_material)
                grupo['materiais'].append(material)

    resultado = []
    for chave in ordem:
        grupo = grupos[chave]
        tarefa = dict(grupo['tarefa'])
        tarefa['Materiais'] = ' | '.join(grupo['materiais'])
        tarefa['_ids_agrupados'] = list(grupo['ids'])
        tarefa['_cards_agrupados'] = list(grupo['cards'])
        tarefa['_qtd_demandas_agrupadas'] = max(1, len(grupo['ids']))
        resultado.append((chave[0], tarefa))
    return resultado


def _dividir_material_quantidade(valor):
    """Separa a quantidade inicial do nome sem confundir medidas como 2,5mm/20kg."""
    item = re.sub(r"\s+", " ", str(valor or "")).strip(" •-\t")
    if not item:
        return "Material não informado", "—"

    numero = r"(\d+(?:[\.,]\d+)?)"
    unidades = r"(un(?:id(?:ades?)?)?|pçs?|pcs?|peças?|sacos?|caixas?|rolos?|metros?|m|kg|litros?|l|kits?|pares?)"

    com_unidade = re.match(
        rf"^{numero}\s+{unidades}\b\s*(?:[-xX:]\s*)?(.+)$",
        item,
        flags=re.IGNORECASE,
    )
    if com_unidade:
        quantidade = f"{com_unidade.group(1)} {com_unidade.group(2)}"
        return com_unidade.group(3).strip(), quantidade

    com_separador = re.match(rf"^{numero}\s*[-xX:]\s*(.+)$", item)
    if com_separador:
        return com_separador.group(2).strip(), com_separador.group(1)

    numero_e_texto = re.match(rf"^{numero}\s+([A-Za-zÀ-ÿ].+)$", item)
    if numero_e_texto:
        return numero_e_texto.group(2).strip(), numero_e_texto.group(1)

    return item, "—"


def enviar_foto_comprovante_power_automate(tarefa, recebedor, foto, material_foto="GERAL", numero_foto=1):
    """Envia uma foto para o Power Automate e retorna o nome salvo no OneDrive."""
    recebedor = str(recebedor or "").strip()
    if not recebedor:
        return False, "Informe quem recebeu o material."
    if foto is None:
        return False, "Tire ou selecione uma foto antes de enviar."

    try:
        flow_url = str(st.secrets["onedrive"]["flow_url"]).strip()
    except Exception:
        return False, "O Secret [onedrive].flow_url não está configurado no Streamlit."

    if not flow_url:
        return False, "O endereço do Power Automate está vazio nos Secrets."

    try:
        foto_bytes = foto.getvalue()
    except Exception:
        try:
            foto_bytes = foto.read()
        except Exception:
            foto_bytes = b""

    if not foto_bytes:
        return False, "A foto está vazia. Tire a foto novamente."

    tipo = str(getattr(foto, "type", "") or "").lower()
    extensao = "png" if "png" in tipo else "webp" if "webp" in tipo else "jpg"

    demanda_id = str(tarefa.get("id", "") or "")
    obra = str(tarefa.get("Obra", "") or "").strip()
    agora = datetime.now(FUSO_LOCAL)

    # Pasta única por obra. A criação/uso da pasta fica a cargo do fluxo já configurado.
    obra_pasta = re.sub(r'[\\/:*?"<>|#%]+', '-', obra).strip(' .-')
    obra_pasta = re.sub(r'\s+', ' ', obra_pasta)[:100] or "OBRA-SEM-NOME"

    material_foto = str(material_foto or "GERAL").strip()
    if material_foto.upper() == "GERAL":
        sufixo_foto = "GERAL"
    else:
        sufixo_foto = f"MAT-{_nome_seguro_comprovante(material_foto, 38)}"

    # Mantém o que foi combinado como identificação principal e acrescenta
    # apenas o número/tipo da foto para facilitar várias imagens na mesma demanda.
    nome_arquivo = (
        f"ID-{_nome_seguro_comprovante(demanda_id or 'SEM-ID', 32)}__"
        f"REC-{_nome_seguro_comprovante(recebedor, 32)}__"
        f"{agora.strftime('%d-%m-%Y_%H-%M-%S')}__"
        f"FOTO-{int(numero_foto):02d}__{sufixo_foto}.{extensao}"
    )

    payload = {
        "nome_arquivo": nome_arquivo,
        "foto_base64": base64.b64encode(foto_bytes).decode("ascii"),
        "recebedor": recebedor,
        "obra": obra,
        "obra_pasta": obra_pasta,
        "demanda_id": demanda_id,
    }

    try:
        resposta = requests.post(flow_url, json=payload, timeout=45)
        if 200 <= resposta.status_code < 300:
            return True, nome_arquivo

        detalhe = (resposta.text or "").replace("\n", " ").strip()
        if len(detalhe) > 220:
            detalhe = detalhe[:220] + "..."
        return False, f"Power Automate respondeu HTTP {resposta.status_code}. {detalhe}".strip()
    except requests.Timeout:
        return False, "O Power Automate demorou demais para responder. Tente novamente."
    except Exception as erro:
        return False, f"Não foi possível enviar a foto: {erro}"


# =====================================================================
# ESTADO PERSISTENTE DOS COMPROVANTES DO DAVI
# Mantém fotos/recebedor/finalização mesmo quando o swipe troca a demanda.
# =====================================================================
@st.cache_resource(show_spinner=False)
def garantir_tabela_comprovantes_davi():
    execute_db(
        """
        CREATE TABLE IF NOT EXISTS comprovantes_entrega_davi (
            id BIGSERIAL PRIMARY KEY,
            data_rota TEXT NOT NULL,
            demanda_id TEXT NOT NULL,
            obra TEXT,
            destino TEXT,
            recebedor TEXT NOT NULL,
            arquivo TEXT NOT NULL UNIQUE,
            tipo_foto TEXT,
            enviado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            finalizado BOOLEAN NOT NULL DEFAULT FALSE
        )
        """
    )
    return True


@st.cache_data(ttl=20, show_spinner=False)
def carregar_resumo_comprovantes_davi(data_rota):
    """Resumo leve para a Torre: não transfere o campo arquivo/foto em cada rerun."""
    estados = {}
    try:
        linhas = fetch_all(
            """
            SELECT demanda_id, MAX(recebedor) AS recebedor, COUNT(*) AS qtd_fotos,
                   BOOL_OR(finalizado) AS finalizado
            FROM comprovantes_entrega_davi
            WHERE data_rota = :data
            GROUP BY demanda_id
            """,
            {"data": data_rota},
        )
    except Exception:
        return estados
    for linha in linhas:
        m = linha._mapping if hasattr(linha, "_mapping") else linha
        demanda_id = str(m["demanda_id"] or "")
        chave = _nome_seguro_comprovante(demanda_id or "SEM-ID", 40)
        qtd = int(m["qtd_fotos"] or 0)
        estados[chave] = {
            "recebedor": str(m["recebedor"] or ""),
            "fotos": [None] * qtd,
            "finalizado": bool(m["finalizado"]),
            "input_version": 0,
        }
    return estados


@st.cache_data(ttl=5, show_spinner=False)
def carregar_comprovantes_davi(data_rota):
    estados = {}
    linhas = fetch_all(
        """
        SELECT demanda_id, obra, destino, recebedor, arquivo, tipo_foto,
               enviado_em, finalizado
        FROM comprovantes_entrega_davi
        WHERE data_rota = :data
        ORDER BY enviado_em ASC, id ASC
        """,
        {"data": data_rota},
    )
    for linha in linhas:
        m = linha._mapping if hasattr(linha, "_mapping") else linha
        demanda_id = str(m["demanda_id"] or "")
        chave = _nome_seguro_comprovante(demanda_id or "SEM-ID", 40)
        estado = estados.setdefault(chave, {
            "recebedor": str(m["recebedor"] or ""),
            "fotos": [],
            "finalizado": False,
            "input_version": 0,
        })
        if not estado.get("recebedor"):
            estado["recebedor"] = str(m["recebedor"] or "")
        enviado_em = m["enviado_em"]
        try:
            hora = enviado_em.astimezone(FUSO_LOCAL).strftime("%H:%M")
        except Exception:
            hora = ""
        estado["fotos"].append({
            "arquivo": str(m["arquivo"] or ""),
            "tipo": str(m["tipo_foto"] or "Foto"),
            "hora": hora,
        })
        estado["finalizado"] = bool(estado.get("finalizado")) or bool(m["finalizado"])
    return estados


def registrar_foto_comprovante_davi(data_rota, tarefa, recebedor, arquivo, tipo_foto):
    execute_db(
        """
        INSERT INTO comprovantes_entrega_davi
            (data_rota, demanda_id, obra, destino, recebedor, arquivo, tipo_foto, enviado_em, finalizado)
        VALUES
            (:data, :demanda_id, :obra, :destino, :recebedor, :arquivo, :tipo_foto, NOW(), FALSE)
        ON CONFLICT (arquivo) DO NOTHING
        """,
        {
            "data": data_rota,
            "demanda_id": str(tarefa.get("id", "") or ""),
            "obra": str(tarefa.get("Obra", "") or ""),
            "destino": str(tarefa.get("Destino", "") or ""),
            "recebedor": str(recebedor or "").strip(),
            "arquivo": str(arquivo or ""),
            "tipo_foto": str(tipo_foto or "Foto"),
        },
    )
    try:
        carregar_resumo_comprovantes_davi.clear()
        carregar_comprovantes_davi.clear()
    except Exception:
        pass


def definir_comprovante_finalizado_davi(data_rota, demanda_id, finalizado=True):
    execute_db(
        """
        UPDATE comprovantes_entrega_davi
        SET finalizado = :finalizado
        WHERE data_rota = :data AND demanda_id = :demanda_id
        """,
        {
            "finalizado": bool(finalizado),
            "data": data_rota,
            "demanda_id": str(demanda_id or ""),
        },
    )
    try:
        carregar_resumo_comprovantes_davi.clear()
        carregar_comprovantes_davi.clear()
    except Exception:
        pass


# =====================================================================
# RENDERIZAÇÃO DO MODO MOBILE (APP DO DAVI)
# =====================================================================
# Compatibilidade: o link antigo ?davi=true continua funcionando, mas o endereço
# oficial do motorista agora é /davi. A página pages/davi.py executa este mesmo
# arquivo com APROAR_DAVI_MODE=True sem depender de parâmetros visíveis na URL.
modo_url = st.query_params.get("davi", "")
try:
    _url_contexto_davi = str(getattr(st.context, "url", "") or "")
    _caminho_contexto_davi = urllib.parse.urlparse(_url_contexto_davi).path.rstrip("/").lower()
except Exception:
    _caminho_contexto_davi = ""
modo_davi = (
    bool(globals().get("APROAR_DAVI_MODE", False))
    or bool(st.session_state.get("_aproar_davi_page", False))
    or modo_url == "true"
    or _caminho_contexto_davi.endswith("/davi")
)

if modo_davi:
    st.markdown("""
        <style>
            [data-testid="stSidebar"] {display: none !important;}
            [data-testid="stHeader"] {display: none !important;}
            [data-testid="stToolbar"] {display:none !important;}
            [data-testid="stAppViewContainer"] {
                background-image:
                    radial-gradient(circle at 100% 0%, rgba(37,99,235,.17), transparent 25rem),
                    linear-gradient(180deg,#070b14 0%,#080b14 100%) !important;
            }
            .main .block-container {
                width:100%; max-width:720px !important;
                padding:14px 14px 92px !important;
            }
            .aproar-driver-header {
                position:relative; overflow:hidden; margin:0 0 8px; padding:14px 15px 15px;
                border:1px solid rgba(96,165,250,.20); border-radius:16px;
                background:linear-gradient(145deg,rgba(15,27,48,.98),rgba(9,14,27,.98));
                box-shadow:0 20px 45px rgba(0,0,0,.26);
            }
            .aproar-driver-header::after {
                content:''; position:absolute; width:190px; height:190px; border-radius:50%;
                right:-82px; bottom:-125px; background:rgba(37,99,235,.24); filter:blur(8px);
            }
            .aproar-driver-topline { display:flex; align-items:center; justify-content:space-between; gap:12px; position:relative; z-index:1; }
            .aproar-driver-brand { display:flex; align-items:center; gap:9px; color:#f8fafc; font-size:14px; font-weight:900; letter-spacing:.12em; }
            .aproar-logo-driver {
                display:block; width:92px; max-width:34vw; height:28px;
                object-fit:contain; object-position:left center;
                filter:drop-shadow(0 8px 16px rgba(0,0,0,.25));
            }
            .aproar-driver-live {
                display:flex; align-items:center; gap:7px; color:#bbf7d0; font-size:10px; font-weight:800;
                letter-spacing:.08em; padding:7px 9px; border:1px solid rgba(34,197,94,.22);
                border-radius:999px; background:rgba(34,197,94,.08);
            }
            .aproar-driver-live i { width:7px; height:7px; border-radius:50%; background:#22c55e; box-shadow:0 0 0 4px rgba(34,197,94,.10); }
            .aproar-driver-greeting { position:relative; z-index:1; margin-top:12px; color:#f8fafc; font-size:22px; font-weight:700; letter-spacing:-.045em; }
            .aproar-driver-greeting strong { color:#60a5fa; font-weight:800; }
            .aproar-driver-date { position:relative; z-index:1; margin-top:5px; color:#94a3b8; font-size:12.5px; font-weight:600; }

            .aproar-driver-summary { margin:6px 0 12px; }
            .aproar-driver-kpis { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:8px; }
            .driver-kpi {
                min-width:0; padding:13px 11px 12px; border-radius:15px;
                background:linear-gradient(145deg,rgba(18,27,47,.96),rgba(12,18,32,.98));
                border:1px solid rgba(148,163,184,.15); box-shadow:0 10px 24px rgba(0,0,0,.15);
            }
            .driver-kpi span { display:block; color:#64748b; font-size:8.5px; font-weight:900; letter-spacing:.10em; }
            .driver-kpi strong { display:block; margin-top:5px; color:#f8fafc; font-size:20px; line-height:1; font-weight:800; letter-spacing:-.04em; }
            .driver-kpi small { display:block; overflow:hidden; margin-top:5px; color:#94a3b8; font-size:9.5px; white-space:nowrap; text-overflow:ellipsis; }
            .driver-progress { height:4px; margin:10px 3px 12px; overflow:hidden; border-radius:999px; background:#172036; }
            .driver-progress span { display:block; height:100%; border-radius:inherit; background:linear-gradient(90deg,#2563eb,#22c55e); }
            .driver-simple-progress { display:flex; justify-content:space-between; gap:10px; margin:9px 3px 0; color:#cbd5e1; font-size:10.5px; }
            .driver-simple-progress b { color:#fff; }
            .driver-next-stop {
                display:grid; grid-template-columns:42px minmax(0,1fr) auto; align-items:center; gap:12px;
                padding:15px; border-radius:17px; border:1px solid rgba(96,165,250,.25);
                background:linear-gradient(135deg,rgba(37,99,235,.12),rgba(15,23,42,.82));
                box-shadow:0 12px 28px rgba(0,0,0,.17);
            }
            .driver-stop-index {
                display:grid; place-items:center; width:42px; height:42px; border-radius:14px;
                background:linear-gradient(145deg,#3b82f6,#1d4ed8); color:#fff; font-size:17px; font-weight:900;
                box-shadow:0 9px 20px rgba(37,99,235,.30);
            }
            .driver-stop-copy { min-width:0; }
            .driver-stop-copy > span { display:block; color:#60a5fa; font-size:9px; font-weight:900; letter-spacing:.11em; }
            .driver-stop-copy > strong { display:block; overflow:hidden; margin-top:3px; color:#f8fafc; font-size:17px; white-space:nowrap; text-overflow:ellipsis; }
            .driver-stop-copy small { display:block; margin-top:4px; color:#94a3b8; font-size:11px; }
            .driver-stop-copy small b { color:#60a5fa; }
            .driver-gps-link {
                display:flex; align-items:center; justify-content:center; min-height:42px; padding:0 13px;
                color:#fff !important; font-size:10.5px; font-weight:900; white-space:nowrap; text-decoration:none !important;
                border-radius:11px; background:linear-gradient(135deg,#2f74f5,#1d4ed8);
                box-shadow:0 9px 20px rgba(37,99,235,.28);
            }

            .aproar-section-anchor { scroll-margin-top:12px; }
            .aproar-section-kicker { color:#60a5fa; font-size:10px; font-weight:900; letter-spacing:.12em; text-transform:uppercase; }
            .aproar-section-title { margin:3px 0 2px; color:#f8fafc; font-size:21px; font-weight:800; letter-spacing:-.035em; }
            .aproar-section-help { margin:0 0 10px; color:#94a3b8; font-size:12px; }
            .aproar-eta-card { margin:10px 0 14px; }
            [data-testid="stExpander"] { margin:6px 0 14px; box-shadow:0 12px 28px rgba(0,0,0,.16); }
            [data-testid="stExpander"] details > summary { min-height:52px; font-weight:800; }
            [data-testid="stCameraInput"] { border-radius:15px; overflow:hidden; }
            [data-testid="stFileUploader"] { margin-top:4px; }
            [data-testid="stFileUploader"] section { min-height:108px !important; }
            [data-testid="stTextInput"] input { min-height:48px; font-size:16px; }
            iframe[title*="streamlit_folium"] { min-height:390px; border:1px solid rgba(148,163,184,.16); }

            .aproar-driver-bottom-nav {
                position:fixed; z-index:999999; left:50%; bottom:10px; transform:translateX(-50%);
                width:min(94%,680px); display:grid; grid-template-columns:repeat(3,1fr); gap:5px;
                padding:7px; border:1px solid rgba(148,163,184,.18); border-radius:18px;
                background:rgba(11,16,29,.94); box-shadow:0 18px 42px rgba(0,0,0,.48);
                backdrop-filter:blur(18px); -webkit-backdrop-filter:blur(18px);
            }
            .aproar-driver-bottom-nav a {
                display:flex; flex-direction:column; align-items:center; justify-content:center; gap:3px;
                min-height:50px; border-radius:12px; color:#94a3b8 !important; font-size:10px; font-weight:800;
                text-decoration:none !important;
            }
            .aproar-driver-bottom-nav a:first-child { color:#bfdbfe !important; background:rgba(37,99,235,.14); }
            .aproar-driver-bottom-nav b { font-size:18px; line-height:1; }

            @media (max-width:380px) {
                .main .block-container { padding-left:10px !important; padding-right:10px !important; }
                .aproar-driver-header { padding:15px; }
                .aproar-driver-greeting { font-size:24px; }
                .driver-kpi { padding:12px 8px; }
                .driver-kpi strong { font-size:18px; }
                .driver-next-stop { grid-template-columns:40px minmax(0,1fr); }
                .driver-gps-link { grid-column:1 / -1; }
            }
        </style>
    """, unsafe_allow_html=True)

    renderizar_cabecalho_motorista()

    if st.button("↻  ATUALIZAR ROTA", use_container_width=True): st.rerun()

    erro_checkin_mobile = ""
    try:
        garantir_tabela_checkins_davi()
        res = carregar_rota_publicada_mobile(DATA_REF_ROTA_STR)
        df_mobile = carregar_conclusoes_rota(DATA_REF_ROTA_STR)
        dict_concluidos_mobile = dict(zip(df_mobile['id'].astype(str), df_mobile['hora_conclusao']))
        hora_inicio_real = obter_hora_inicio_rota(DATA_REF_ROTA_STR)
    except: res, dict_concluidos_mobile, hora_inicio_real = None, {}, HORA_INICIO_ROTA_DAVI

    if not res:
        st.info("Nenhuma rota foi liberada pela Torre de Controle para hoje ainda. Aguarde o cálculo da central e tente atualizar a tela.")
        st.stop()

    route_steps = json.loads(res[0])
    locais_dict = json.loads(res[1])
    geometria_rota = json.loads(res[2])
    enderecos_dict = json.loads(res[3])
    total_km = res[4]
    p_saida = route_steps[0]['destino'] if route_steps else ""
    df_paradas_mobile = carregar_paradas_rastreadas_rota(DATA_REF_ROTA_STR, PLACA_DAVI)

    # O clique no cartão volta ao app com estes parâmetros. A gravação é feita
    # no servidor para que a mesma informação apareça no painel do escritório.
    etapa_param = st.query_params.get("etapa", "")
    feito_param = st.query_params.get("feito", "")
    foco_param_atual = st.query_params.get("foco", "")
    if etapa_param != "" and feito_param in {"0", "1"}:
        try:
            etapa_indice = int(etapa_param)
            if not 0 <= etapa_indice < len(route_steps):
                raise ValueError("Etapa fora da rota atual")
            etapa_escolhida = route_steps[etapa_indice]
            if etapa_escolhida.get("type") in {"lunch", "return"}:
                raise ValueError("Esta etapa não pode ser marcada")
            salvar_checkin_davi(
                DATA_REF_ROTA_STR,
                etapa_indice,
                str(etapa_escolhida.get("destino", "")),
                feito_param == "1",
            )
            st.query_params.clear()
            st.query_params["davi"] = "true"
            if str(foco_param_atual).strip():
                st.query_params["foco"] = str(foco_param_atual).strip()
            st.rerun()
        except Exception:
            erro_checkin_mobile = "Não foi possível salvar a marcação. Tente novamente."
            st.query_params.clear()
            st.query_params["davi"] = "true"
            if str(foco_param_atual).strip():
                st.query_params["foco"] = str(foco_param_atual).strip()

    try:
        dict_checkins_mobile = filtrar_checkins_da_rota(route_steps, carregar_checkins_davi(DATA_REF_ROTA_STR))
    except Exception:
        dict_checkins_mobile = {}
        erro_checkin_mobile = "Não foi possível carregar as marcações compartilhadas agora."

    if erro_checkin_mobile:
        st.error(erro_checkin_mobile)

    # No modo /davi o Streamlit executa este bloco antes de chegar às
    # funções de planejamento declaradas mais abaixo no arquivo. Portanto, não
    # podemos chamar atualizar_tempos_por_parada() diretamente aqui, pois ela
    # ainda não existe neste ponto da execução e causava NameError no app móvel.
    #
    # A rota liberada pela Torre já vem com tempo_local calculado. Aqui apenas
    # saneamos rotas antigas/salvas para a faixa operacional atual (15–25 min),
    # preservando a preparação inicial da base. Quando a função completa estiver
    # disponível (em outros fluxos), ela continua sendo usada normalmente.
    _atualizador_paradas = globals().get("atualizar_tempos_por_parada")
    if callable(_atualizador_paradas):
        route_steps = _atualizador_paradas(route_steps, p_saida)
    else:
        for _idx_step, _step in enumerate(route_steps or []):
            if _step.get("type") != "stop":
                continue
            _destino_step = str(_step.get("destino", "") or "")
            _eh_preparacao_base = (_idx_step == 0 and _destino_step == p_saida)
            if _eh_preparacao_base:
                continue
            try:
                _tempo_salvo = float(_step.get("tempo_local", 20) or 20)
            except (TypeError, ValueError):
                _tempo_salvo = 20.0
            if not math.isfinite(_tempo_salvo):
                _tempo_salvo = 20.0
            _step["tempo_local"] = int(round(min(max(_tempo_salvo, 15.0), 25.0)))
            _step.setdefault("tempo_local_fonte", "estimativa salva da rota")

    route_steps = atualizar_tempos_deslocamento_operacionais(route_steps, hora_inicio_real)
    route_steps, final_dyn_min = aplicar_tempos_dinamicos(route_steps, dict_concluidos_mobile, hora_inicio_real)
    
    hora_atual_str = AGORA_REAL.strftime("%H:%M")
    nova_previsao_str = format_mins_to_time(final_dyn_min)
    renderizar_resumo_motorista(route_steps, total_km, final_dyn_min, enderecos_dict, locais_dict)


    # ---------------------------------------------------------------
    # COMPROVANTE DE ENTREGA — SELEÇÃO AUTOMÁTICA PELO SWIPE
    # O cartão em que o motorista para vira a parada ativa do comprovante.
    # Se houver várias entregas na mesma parada, pega a primeira ainda pendente.
    # ---------------------------------------------------------------
    st.markdown("""
            <div class="aproar-section-anchor" id="comprovante">
                <div class="aproar-section-kicker">REGISTRO DA ENTREGA</div>
            <div class="aproar-section-title">Confirmar entrega</div>
            <div class="aproar-section-help">Digite quem recebeu e envie uma foto</div>
        </div>
    """, unsafe_allow_html=True)

    if "davi_comprovantes_estado" not in st.session_state:
        st.session_state["davi_comprovantes_estado"] = {}
    estados_comprovantes = st.session_state["davi_comprovantes_estado"]

    persistencia_comprovantes_ok = True
    try:
        garantir_tabela_comprovantes_davi()
        estados_banco = carregar_comprovantes_davi(DATA_REF_ROTA_STR)
        for chave_db, estado_db in estados_banco.items():
            estado_local = estados_comprovantes.get(chave_db, {})
            estado_db["input_version"] = int(estado_local.get("input_version", 0))
            if estado_local.get("mensagem"):
                estado_db["mensagem"] = estado_local["mensagem"]
            estados_comprovantes[chave_db] = estado_db
    except Exception:
        persistencia_comprovantes_ok = False

    entregas_por_etapa = {}
    numero_parada_comprovante = 0
    for indice_step, step_comprovante in enumerate(route_steps):
        if step_comprovante.get("type") != "stop":
            continue
        destino_comprovante = str(step_comprovante.get("destino", "") or "")
        is_inicio_comprovante = (indice_step == 0 and destino_comprovante == p_saida)
        if not is_inicio_comprovante:
            numero_parada_comprovante += 1

        for acao_comprovante, tarefa_comprovante in step_comprovante.get("actions", []):
            if acao_comprovante != "ENTREGAR":
                continue

            card_id_comprovante = str(tarefa_comprovante.get("id", "") or "")
            chave_estado = _nome_seguro_comprovante(card_id_comprovante or f"SEM-ID-{indice_step}", 40)
            entregas_por_etapa.setdefault(indice_step, []).append({
                "id": card_id_comprovante,
                "tarefa": tarefa_comprovante,
                "parada": max(1, numero_parada_comprovante),
                "destino": destino_comprovante,
                "chave": chave_estado,
            })

    foco_comprovante = None
    try:
        foco_lido = str(st.query_params.get("foco", "") or "").strip()
        if foco_lido != "":
            foco_comprovante = int(foco_lido)
            if not 0 <= foco_comprovante < len(route_steps):
                foco_comprovante = None
    except Exception:
        foco_comprovante = None

    # Sem seleção manual: abre automaticamente a primeira entrega que ainda
    # não possui comprovante finalizado.
    if foco_comprovante is None:
        for indice_entrega, itens_entrega in sorted(entregas_por_etapa.items()):
            if any(
                not bool(estados_comprovantes.get(item["chave"], {}).get("finalizado"))
                for item in itens_entrega
            ):
                foco_comprovante = indice_entrega
                break

    if entregas_por_etapa:
        with st.expander("📸 REGISTRAR ENTREGA", expanded=True):

            if not persistencia_comprovantes_ok:
                st.warning("O comprovante continua funcionando, mas o histórico interno não pôde ser sincronizado agora. Evite recarregar a página até concluir a entrega.")

            entregas_foco = entregas_por_etapa.get(foco_comprovante, []) if foco_comprovante is not None else []

            if not entregas_foco:
                if foco_comprovante is None:
                    st.success("✅ Todos os comprovantes de entrega foram finalizados.")
                else:
                    destino_foco = str(route_steps[foco_comprovante].get("destino", "") or "")
                    st.info(f"📍 **{destino_foco or 'Esta etapa'}** não possui entrega para registrar.")
            else:
                pendentes_foco = []
                for item_foco in entregas_foco:
                    estado_item = estados_comprovantes.get(item_foco["chave"], {})
                    if not bool(estado_item.get("finalizado")):
                        pendentes_foco.append(item_foco)

                if not pendentes_foco:
                    parada_num = entregas_foco[0]["parada"]
                    destino_foco = entregas_foco[0]["destino"]
                    qtd_entregas_foco = len(entregas_foco)
                    st.success(
                        f"✅ {plural_pt(qtd_entregas_foco, 'A entrega', 'Todas as entregas')} da "
                        f"**Parada {parada_num} — {destino_foco}** já "
                        f"{plural_pt(qtd_entregas_foco, 'tem', 'têm')} comprovante finalizado."
                    )
                    for item_pronto in entregas_foco:
                        estado_pronto = estados_comprovantes.get(item_pronto["chave"], {})
                        obra_pronta = str(item_pronto["tarefa"].get("Obra", "") or "")
                        st.caption(
                            f"✅ {obra_pronta} • Trello {item_pronto['id']} • "
                            f"{len(estado_pronto.get('fotos', []))} {plural_pt(len(estado_pronto.get('fotos', [])), 'foto', 'fotos')} • {estado_pronto.get('recebedor', '')}"
                        )
                else:
                    entrega_sel = pendentes_foco[0]
                    tarefa_sel = entrega_sel["tarefa"]
                    demanda_id_sel = entrega_sel["id"]
                    chave_comprovante = entrega_sel["chave"]

                    estado = estados_comprovantes.setdefault(chave_comprovante, {
                        "recebedor": "",
                        "fotos": [],
                        "finalizado": False,
                        "input_version": 0,
                    })
                    estado.setdefault("recebedor", "")
                    estado.setdefault("fotos", [])
                    estado.setdefault("finalizado", False)
                    estado.setdefault("input_version", 0)

                    obra_sel = str(tarefa_sel.get("Obra", "") or "").strip()
                    materiais_sel = _separar_materiais_comprovante(tarefa_sel.get("Materiais", ""))
                    destino_sel = str(entrega_sel.get("destino", "") or "").strip()
                    indice_na_parada = entregas_foco.index(entrega_sel) + 1

                    materiais_html = "".join(
                        f"<div style='margin:3px 0;'>• {html_escape(item)}</div>"
                        for item in materiais_sel
                    ) or "<div style='color:#94a3b8;'>• Materiais não informados</div>"

                    complemento_demanda = ""
                    if len(entregas_foco) > 1:
                        complemento_demanda = (
                            f"<div style='margin-top:7px;color:#fbbf24;font-size:12px;font-weight:700;'>"
                            f"Entrega {indice_na_parada} de {len(entregas_foco)} nesta parada • ao finalizar, a próxima entra automaticamente"
                            f"</div>"
                        )

                    st.markdown(
                        f"""
                        <div style="background:rgba(37,99,235,.08);border:1px solid rgba(96,165,250,.28);border-radius:12px;padding:12px 14px;margin:6px 0 14px 0;">
                            <div style="font-size:15px;font-weight:700;margin-bottom:5px;">📍 Parada {entrega_sel['parada']} — {html_escape(obra_sel or destino_sel)}</div>
                            <div style="font-size:12px;color:#94a3b8;margin-bottom:8px;">Trello: {html_escape(demanda_id_sel or 'sem ID')} • Destino: {html_escape(destino_sel or '-')}</div>
                            <div style="font-size:13px;font-weight:600;margin-bottom:3px;">Materiais desta demanda:</div>
                            <div style="font-size:13px;line-height:1.35;">{materiais_html}</div>
                            {complemento_demanda}
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                    status_real_comprovante = obter_status_rastreio_local(
                        df_paradas_mobile, destino_sel, DATA_REF_ROTA_STR
                    )
                    if status_real_comprovante:
                        if status_real_comprovante.get("aberta"):
                            texto_chegada = f"📍 Chegou às **{status_real_comprovante['chegada']}**"
                            if status_real_comprovante.get("duracao"):
                                texto_chegada += f" • ⏱️ Está no local há **{status_real_comprovante['duracao']}**"
                            st.info(texto_chegada)
                        else:
                            st.info(
                                f"📍 Chegou às **{status_real_comprovante['chegada']}** • "
                                f"🚚 Saiu às **{status_real_comprovante['saida']}** • "
                                f"⏱️ Ficou **{status_real_comprovante['duracao']}** no local"
                            )

                    mensagem_pendente = estado.pop("mensagem", "") if estado.get("mensagem") else ""
                    if mensagem_pendente:
                        st.success(mensagem_pendente)

                    if estado["fotos"]:
                        st.markdown(
                            f"**1️⃣ Quem recebeu?**  \n👤 **{estado['recebedor']}**",
                        )
                        st.caption("O recebedor fica fixo para todas as fotos desta demanda.")
                        recebedor_comprovante = estado["recebedor"]
                    else:
                        recebedor_comprovante = st.text_input(
                            "1️⃣ Quem recebeu?",
                            placeholder="Ex.: João da Silva",
                            value=estado.get("recebedor", ""),
                            key=f"davi_comprovante_recebedor_{chave_comprovante}",
                            help="Digite uma vez. O mesmo nome será usado em todas as fotos desta demanda.",
                        )

                    material_foto = "GERAL"
                    versao_input = int(estado.get("input_version", 0))
                    foto_comprovante = st.file_uploader(
                        "2️⃣ Tirar ou escolher a foto",
                        type=["jpg", "jpeg", "png", "webp"],
                        accept_multiple_files=False,
                        key=f"davi_comprovante_arquivo_{chave_comprovante}_{versao_input}",
                        help="No celular, escolha Câmera ou Fotos/Galeria.",
                    )

                    numero_proxima_foto = len(estado["fotos"]) + 1
                    if st.button(
                        "✅ REGISTRAR ENTREGA",
                        type="primary",
                        use_container_width=True,
                        key=f"davi_enviar_comprovante_{chave_comprovante}_{versao_input}",
                    ):
                        nome_recebedor = str(recebedor_comprovante or "").strip()
                        if not nome_recebedor:
                            st.error("Informe quem recebeu o material.")
                        elif foto_comprovante is None:
                            st.error("Tire ou selecione uma foto.")
                        else:
                            with st.spinner("Enviando foto para o OneDrive..."):
                                sucesso_comprovante, retorno_comprovante = enviar_foto_comprovante_power_automate(
                                    tarefa_sel,
                                    nome_recebedor,
                                    foto_comprovante,
                                    material_foto=material_foto,
                                    numero_foto=numero_proxima_foto,
                                )
                            if sucesso_comprovante:
                                tipo_registro = "Foto geral" if material_foto == "GERAL" else material_foto
                                finalizacao_automatica_ok = True
                                try:
                                    registrar_foto_comprovante_davi(
                                        DATA_REF_ROTA_STR,
                                        tarefa_sel,
                                        nome_recebedor,
                                        retorno_comprovante,
                                        tipo_registro,
                                    )
                                    definir_comprovante_finalizado_davi(
                                        DATA_REF_ROTA_STR, demanda_id_sel, True
                                    )
                                except Exception:
                                    persistencia_comprovantes_ok = False
                                    finalizacao_automatica_ok = False

                                estado["recebedor"] = nome_recebedor
                                estado["fotos"].append({
                                    "arquivo": retorno_comprovante,
                                    "tipo": tipo_registro,
                                    "hora": datetime.now(FUSO_LOCAL).strftime("%H:%M"),
                                })
                                estado["input_version"] = versao_input + 1
                                estado["finalizado"] = finalizacao_automatica_ok
                                estado["mensagem"] = (
                                    "✅ Entrega registrada com foto e nome do recebedor."
                                    if finalizacao_automatica_ok
                                    else "Foto enviada. Toque em finalizar para concluir o registro."
                                )
                                st.rerun()
                            else:
                                st.error(retorno_comprovante)

                    if estado["fotos"]:
                        st.markdown("**Fotos já enviadas nesta entrega:**")
                        for pos, foto_enviada in enumerate(estado["fotos"], start=1):
                            st.caption(f"✅ Foto {pos} • {foto_enviada['tipo']} • {foto_enviada['hora']}")

                        if st.button(
                            f"✅ FINALIZAR ESTA ENTREGA ({len(estado['fotos'])} {plural_pt(len(estado['fotos']), 'FOTO', 'FOTOS')})",
                            type="primary",
                            use_container_width=True,
                            key=f"davi_finalizar_comprovante_{chave_comprovante}",
                        ):
                            try:
                                definir_comprovante_finalizado_davi(DATA_REF_ROTA_STR, demanda_id_sel, True)
                            except Exception:
                                if persistencia_comprovantes_ok:
                                    st.error("Não consegui registrar a finalização. Tente novamente.")
                                    st.stop()
                            estado["finalizado"] = True
                            estado["mensagem"] = "Comprovante finalizado."
                            st.rerun()
                    else:
                        st.caption("Envie pelo menos uma foto para liberar a finalização desta entrega.")


    st.markdown(f"""
            <div class="aproar-section-anchor" id="roteiro">
            <div class="aproar-section-kicker">ROTA EM EXECUÇÃO</div>
            <div class="aproar-section-title">Roteiro do dia</div>
            <div class="aproar-section-help">{total_km:.1f} km • deslize para os lados para trocar de parada</div>
        </div>
    """, unsafe_allow_html=True)
    MODO_DAVI_SIMPLES = False
    cartoes_mobile = []
    numero_parada_mobile = 1

    for i, step in enumerate(route_steps):
        tipo_step = step.get('type', '')
        destino_step = str(step.get('destino', ''))
        is_start = (i == 0 and destino_step == p_saida)
        tem_entrega_no_cartao = any(acao == "ENTREGAR" for acao, _ in step.get('actions', []))
        classe_card, selo, titulo_card, meta_card = "normal", "ETAPA", destino_step, ""
        botao_gps, botao_feito, botao_comprovante = "", "", ""
        etapa_marcada = False

        if tipo_step == 'lunch':
            classe_card, selo = "almoco", "PAUSA"
            titulo_card = "🍔 Almoço"
            meta_card = f"Horário previsto: {html_escape(str(step.get('dyn_chegada', '12:00')))} às {html_escape(str(step.get('dyn_saida', '13:00')))}"
            corpo_acoes = "<div class='mensagem-etapa'>Pausa programada para descanso e alimentação.</div>"
        elif tipo_step == 'return':
            classe_card, selo = "retorno", "RETORNO"
            titulo_card = f"🏁 Retorno à Base: {html_escape(destino_step)}"
            meta_card = f"Chegada prevista: {html_escape(str(step.get('dyn_chegada', '')))}"
            corpo_acoes = "<div class='mensagem-etapa'>Última etapa do roteiro. Retorne para a base indicada.</div>"
        else:
            endereco_db = enderecos_dict.get(destino_step, "")
            coordenadas = locais_dict.get(destino_step, [None, None])
            if str(endereco_db).startswith("http"):
                link_gps = str(endereco_db)
            elif endereco_db:
                link_gps = f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(str(endereco_db))}"
            elif len(coordenadas) >= 2 and coordenadas[0] is not None:
                link_gps = f"https://www.google.com/maps/dir/?api=1&destination={coordenadas[0]},{coordenadas[1]}"
            else:
                link_gps = ""

            if is_start:
                classe_card, selo = "preparacao", "PREPARAÇÃO"
                titulo_card = f"🏁 {html_escape(destino_step)}"
                meta_card = f"Preparação planejada: {html_escape(str(step.get('chegada', '')))} às {html_escape(str(step.get('saida', '')))}"
            else:
                selo = f"PARADA {numero_parada_mobile}"
                titulo_card = f"📍 {html_escape(destino_step)}"
                meta_card = f"Trecho: {float(step.get('dist', 0) or 0):.1f} km"

            if is_start:
                status_tempo = f"<span class='status pendente'>🕖 Preparação: {html_escape(str(step.get('chegada', '')))} às {html_escape(str(step.get('saida', '')))}</span>"
                status_rastreio_html = ""
            else:
                status_tempo = f"<span class='status concluido'>✅ Concluído às {html_escape(str(step.get('dyn_saida', '')))}</span>" if step.get('is_concluded') else f"<span class='status pendente'>⏳ Previsão atual: {html_escape(str(step.get('dyn_chegada', '')))} às {html_escape(str(step.get('dyn_saida', '')))}</span>"
                status_rastreio_html = html_status_rastreio_local(
                    obter_status_rastreio_local(df_paradas_mobile, destino_step, DATA_REF_ROTA_STR)
                )
            # O tempo é calculado por PARADA FÍSICA, não somando um atendimento
            # completo para cada demanda que acontece no mesmo endereço.
            if not is_start and step.get('actions'):
                tempo_parada_mobile = int(round(float(step.get('tempo_local', 0) or 0)))
                fonte_parada_mobile = str(step.get('tempo_local_fonte', 'média operacional') or 'média operacional')
                status_tempo += (
                    f"<div style='margin:-4px 0 12px;padding:9px 11px;border-radius:10px;"
                    f"color:#cbd5e1;background:rgba(148,163,184,.08);border:1px solid rgba(148,163,184,.18);"
                    f"font-size:12.5px;'>⏱️ Permanência estimada no local: <b>{tempo_parada_mobile} min</b>"
                    f" <span style='color:#94a3b8;'>• {html_escape(fonte_parada_mobile)}</span></div>"
                )

            blocos_acao = []
            for indice_acao, (acao, tarefa) in enumerate(step.get('actions', []), start=1):
                eh_coleta = acao == "COLETAR"
                classe_acao, icone = ("coleta", "📦") if eh_coleta else ("entrega", "📬")
                card_id = str(tarefa.get('id', ''))
                concluido = f"<div class='baixa'>✅ Baixa às {html_escape(str(dict_concluidos_mobile[card_id]))}</div>" if card_id in dict_concluidos_mobile else ""

                materiais_acao = _separar_materiais_comprovante(tarefa.get('Materiais', ''))
                if materiais_acao:
                    materiais_html_acao = "".join(
                        f"<div class='material-item'><span class='material-bullet'>•</span><span>{html_escape(str(material))}</span></div>"
                        for material in materiais_acao
                    )
                else:
                    materiais_html_acao = "<div class='material-item vazio'>Material não informado</div>"

                obra_acao = html_escape(str(tarefa.get('Obra', '') or 'Obra não informada'))
                rotulo_acao = "COLETA" if eh_coleta else "ENTREGA"

                blocos_acao.append(
                    f"<div class='acao {classe_acao}'>"
                    f"<div class='acao-cabecalho'><div class='acao-tipo'>{icone} {rotulo_acao} {indice_acao}</div><div class='acao-obra'>🏗️ {obra_acao}</div></div>"
                    f"<div class='materiais-lista'>{materiais_html_acao}</div>"
                    f"{concluido}</div>"
                )
            corpo_acoes = status_tempo + status_rastreio_html + ("".join(blocos_acao) if blocos_acao else "<div class='mensagem-etapa'>Nenhuma movimentação cadastrada nesta etapa.</div>")
            rotulo_lembrete = "preparação" if is_start else "parada"
            checkin_etapa = dict_checkins_mobile.get(i)
            etapa_marcada = bool(checkin_etapa) or bool(step.get('is_concluded'))
            if step.get('is_concluded'):
                botao_feito = "<button class='marcar-feita ativa' data-feita='1' disabled>✅ Concluída no sistema</button>"
            else:
                novo_estado = "0" if checkin_etapa else "1"
                classe_marcacao = " ativa" if checkin_etapa else ""
                texto_marcacao = (
                    f"✅ Marcada às {html_escape(checkin_etapa['hora'])} • toque para desfazer"
                    if checkin_etapa else f"☐ Marcar {rotulo_lembrete} como feita"
                )
                link_marcacao = html_escape(f"/davi?foco={i}&etapa={i}&feito={novo_estado}", quote=True)
                botao_feito = (
                    f"<a class='marcar-feita{classe_marcacao}' data-feita='{'1' if checkin_etapa else '0'}' "
                    f"href='{link_marcacao}' target='_top' onclick='prepararEnvio(this)'>{texto_marcacao}</a>"
                )
            if not is_start:
                if tem_entrega_no_cartao:
                    rotulo_comprovante = "📸 REGISTRAR ENTREGA" if not step.get('is_concluded') else "📸 VER COMPROVANTE"
                    link_comprovante = html_escape(f"/davi?foco={i}#comprovante", quote=True)
                    botao_comprovante = (
                        f"<a class='comprovante' href='{link_comprovante}' target='_top'>"
                        f"{rotulo_comprovante}</a>"
                    )
                if link_gps:
                    botao_gps = f"<a class='gps' href='{html_escape(link_gps, quote=True)}' target='_blank' rel='noopener'>🧭 ABRIR GPS DA PARADA {numero_parada_mobile}</a>"
                numero_parada_mobile += 1

        botoes_rodape = botao_comprovante + botao_gps + botao_feito
        rodape_card = f"<div class='rodape-card'>{botoes_rodape}</div>" if botoes_rodape else ""
        cartoes_mobile.append(
            f"<article class='cartao {classe_card}{' feita' if etapa_marcada else ''}' data-etapa='{i}' data-entrega='{'1' if tem_entrega_no_cartao else '0'}'><div class='topo-card'><span class='selo'>{html_escape(str(selo))}</span>"
            f"<h2>{titulo_card}</h2><div class='meta'>{meta_card}</div></div>"
            f"<div class='conteudo-card'>{corpo_acoes}</div>{rodape_card}</article>"
        )

    if MODO_DAVI_SIMPLES:
        numero_lista = 0
        for indice_lista, etapa_lista in enumerate(route_steps):
            tipo_lista = str(etapa_lista.get("type", "") or "")
            destino_lista = str(etapa_lista.get("destino", "") or "")
            inicio_lista = indice_lista == 0 and destino_lista == p_saida
            checkin_lista = dict_checkins_mobile.get(indice_lista)
            concluida_lista = bool(etapa_lista.get("is_concluded")) or bool(checkin_lista)

            if tipo_lista == "lunch":
                titulo_lista = "🍔 Pausa para almoço"
            elif tipo_lista == "return":
                titulo_lista = f"🏁 Retorno — {destino_lista}"
            elif inicio_lista:
                titulo_lista = f"🏁 Preparação — {destino_lista}"
            else:
                numero_lista += 1
                titulo_lista = f"{'✅' if concluida_lista else str(numero_lista) + '.'} {destino_lista}"

            expandir_lista = indice_lista == foco_comprovante
            with st.expander(titulo_lista, expanded=expandir_lista):
                if tipo_lista == "lunch":
                    st.write(
                        f"Horário previsto: **{etapa_lista.get('dyn_chegada', '12:00')} às "
                        f"{etapa_lista.get('dyn_saida', '13:00')}**"
                    )
                    continue
                if tipo_lista == "return":
                    st.write(f"Chegada prevista: **{etapa_lista.get('dyn_chegada', '--:--')}**")
                    continue

                if inicio_lista:
                    st.caption(
                        f"Preparação: {etapa_lista.get('chegada', '--:--')} às "
                        f"{etapa_lista.get('saida', '--:--')}"
                    )
                else:
                    st.caption(
                        f"Chegada prevista: {etapa_lista.get('dyn_chegada', '--:--')} • "
                        f"Trecho: {float(etapa_lista.get('dist', 0) or 0):.1f} km"
                    )

                tem_entrega_lista = False
                for acao_lista, tarefa_lista in (etapa_lista.get("actions") or []):
                    eh_entrega_lista = acao_lista == "ENTREGAR"
                    tem_entrega_lista = tem_entrega_lista or eh_entrega_lista
                    icone_lista = "📬" if eh_entrega_lista else "📦"
                    rotulo_lista = "ENTREGAR" if eh_entrega_lista else "COLETAR"
                    st.markdown(
                        f"**{icone_lista} {rotulo_lista} — {tarefa_lista.get('Obra', 'Obra não informada')}**"
                    )
                    materiais_lista = _separar_materiais_comprovante(tarefa_lista.get("Materiais", ""))
                    if materiais_lista:
                        st.markdown("\n".join(f"- {material}" for material in materiais_lista))
                    else:
                        st.caption("Material não informado")

                if not inicio_lista:
                    endereco_lista = str(enderecos_dict.get(destino_lista, "") or "")
                    coordenadas_lista = locais_dict.get(destino_lista, [None, None])
                    if endereco_lista.startswith("http"):
                        link_lista = endereco_lista
                    elif endereco_lista:
                        link_lista = f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(endereco_lista)}"
                    elif len(coordenadas_lista) >= 2 and coordenadas_lista[0] is not None:
                        link_lista = f"https://www.google.com/maps/dir/?api=1&destination={coordenadas_lista[0]},{coordenadas_lista[1]}"
                    else:
                        link_lista = ""

                    if link_lista:
                        st.link_button("🧭 ABRIR GPS", link_lista, use_container_width=True)

                    if tem_entrega_lista and not concluida_lista:
                        if st.button(
                            "📸 REGISTRAR ENTREGA DESTA PARADA",
                            use_container_width=True,
                            key=f"davi_ir_comprovante_simples_{indice_lista}",
                        ):
                            st.query_params.clear()
                            st.query_params["foco"] = str(indice_lista)
                            st.rerun()

                if tipo_lista == "stop" and not etapa_lista.get("is_concluded"):
                    texto_checkin = "↩️ DESFAZER CONCLUSÃO" if checkin_lista else "✅ MARCAR PARADA COMO FEITA"
                    if st.button(
                        texto_checkin,
                        use_container_width=True,
                        key=f"davi_checkin_simples_{indice_lista}_{'1' if checkin_lista else '0'}",
                    ):
                        salvar_checkin_davi(
                            DATA_REF_ROTA_STR, indice_lista, destino_lista, not bool(checkin_lista)
                        )
                        st.rerun()

    elif cartoes_mobile:
        html_carrossel = """
        <!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Sora:wght@600;700;800&display=swap');
            * { box-sizing: border-box; }
            html, body { margin: 0; padding: 0; background: transparent; color: #e4e8f4; font-family: Manrope, Arial, sans-serif; }
            h1, h2, h3, strong, .obra { font-family:Sora, Manrope, Arial, sans-serif; }
            .barra { display:flex; justify-content:space-between; align-items:center; gap:8px; margin:0 2px 10px; color:#94a3b8; font-size:11px; font-weight:700; }
            .resumo-topo { display:flex; align-items:center; gap:6px; }
            .feitas { color:#bbf7d0; font-weight:900; background:rgba(34,197,94,.10); border:1px solid rgba(34,197,94,.23); padding:6px 9px; border-radius:999px; }
            .contador { color:#dbeafe; font-weight:900; background:rgba(37,99,235,.11); border:1px solid rgba(96,165,250,.24); padding:6px 10px; border-radius:999px; }
            .trilho { display:flex; gap:12px; overflow-x:auto; scroll-snap-type:x mandatory; scroll-behavior:smooth; overscroll-behavior-x:contain; -webkit-overflow-scrolling:touch; touch-action:pan-x pan-y; scrollbar-width:none; padding:2px 4px 12px; }
            .trilho::-webkit-scrollbar { display:none; }
            .cartao { flex:0 0 calc(100% - 8px); height:438px; scroll-snap-align:center; scroll-snap-stop:always; display:flex; flex-direction:column; overflow:hidden; background:linear-gradient(145deg,#111a2e,#0a101e); border:1px solid rgba(148,163,184,.18); border-radius:18px; box-shadow:0 16px 34px rgba(0,0,0,.32); }
            .cartao.preparacao { border-color:rgba(59,130,246,.48); }
            .cartao.almoco { border-color:rgba(245,158,11,.48); }
            .cartao.retorno { border-color:rgba(34,197,94,.48); }
            .cartao.feita { border-color:rgba(34,197,94,.62); box-shadow:0 0 0 2px rgba(34,197,94,.12),0 18px 38px rgba(0,0,0,.34); }
            .cartao.feita .topo-card { background:linear-gradient(135deg,rgba(22,163,74,.18),rgba(22,163,74,.03)); }
            .cartao.selecionada { box-shadow:0 0 0 2px rgba(59,130,246,.26),0 20px 42px rgba(0,0,0,.38); }
            .topo-card { padding:18px 18px 14px; border-bottom:1px solid rgba(148,163,184,.13); background:linear-gradient(135deg,rgba(37,99,235,.08),transparent 60%); }
            .selo { display:inline-block; color:#bfdbfe; background:#1d4ed8; font-size:11px; font-weight:900; letter-spacing:.08em; padding:5px 9px; border-radius:999px; }
            .almoco .selo { background:#92400e; color:#fef3c7; }
            .retorno .selo { background:#166534; color:#dcfce7; }
            h2 { margin:11px 0 6px; color:#f8fafc; font-size:21px; line-height:1.18; letter-spacing:-.035em; }
            .meta { color:#8da0b8; font-size:12px; line-height:1.45; }
            .conteudo-card { flex:1; overflow-y:auto; padding:14px 16px 9px; scrollbar-width:thin; scrollbar-color:#334155 transparent; }
            .status { display:block; margin-bottom:12px; padding:9px 11px; border-radius:10px; font-size:13px; font-weight:800; }
            .status.concluido { color:#bbf7d0; background:rgba(22,163,74,.15); }
            .status.pendente { color:#fde68a; background:rgba(245,158,11,.14); }
            .rastreio-real { margin:-4px 0 12px; padding:9px 11px; border-radius:10px; color:#bae6fd; background:rgba(14,165,233,.10); border:1px solid rgba(56,189,248,.22); font-size:12.5px; line-height:1.45; }
            .acao { margin-bottom:10px; padding:0; border-radius:13px; border:1px solid #2b3654; overflow:hidden; background:rgba(255,255,255,.025); }
            .acao.coleta { border-color:rgba(245,158,11,.50); border-left:5px solid #f59e0b; background:rgba(245,158,11,.035); }
            .acao.entrega { border-color:rgba(34,197,94,.52); border-left:5px solid #22c55e; background:rgba(34,197,94,.045); }
            .acao-cabecalho { display:flex; align-items:flex-start; justify-content:space-between; gap:10px; padding:10px 11px 9px; border-bottom:1px solid rgba(141,160,184,.13); }
            .acao.coleta .acao-cabecalho { background:linear-gradient(90deg,rgba(245,158,11,.16),rgba(245,158,11,.035)); }
            .acao.entrega .acao-cabecalho { background:linear-gradient(90deg,rgba(34,197,94,.18),rgba(34,197,94,.035)); }
            .acao-tipo { flex:0 0 auto; font-size:11.5px; font-weight:900; letter-spacing:.04em; white-space:nowrap; padding:4px 7px; border-radius:7px; }
            .acao.coleta .acao-tipo { color:#fde68a; background:rgba(245,158,11,.16); }
            .acao.entrega .acao-tipo { color:#bbf7d0; background:rgba(34,197,94,.16); }
            .acao-obra { min-width:0; color:#cbd5e1; font-size:11.5px; line-height:1.35; text-align:right; font-weight:700; padding-top:3px; }
            .acao-tempo { padding:7px 11px 0; color:#94a3b8; font-size:11.5px; font-weight:700; }
            .materiais-lista { padding:8px 11px 10px; display:grid; gap:6px; }
            .material-item { display:grid; grid-template-columns:10px minmax(0,1fr); gap:6px; color:#e4e8f4; font-size:12.7px; line-height:1.38; }
            .material-bullet { color:#60a5fa; font-weight:900; }
            .acao.entrega .material-bullet { color:#4ade80; }
            .acao.coleta .material-bullet { color:#fbbf24; }
            .material-item.vazio { display:block; color:#8da0b8; font-style:italic; }
            .baixa { color:#86efac; font-size:11.5px; font-weight:800; padding:0 11px 10px; }
            .mensagem-etapa { color:#cbd5e1; font-size:15px; line-height:1.55; padding:18px 6px; }
            .rodape-card { display:grid; grid-template-columns:1fr; gap:7px; padding:10px 13px 13px; border-top:1px solid rgba(148,163,184,.13); background:rgba(5,9,17,.20); }
            .marcar-feita { display:block; width:100%; padding:12px 10px; border-radius:11px; border:1px solid rgba(34,197,94,.48); background:rgba(22,163,74,.08); color:#bbf7d0; font-size:12.5px; font-weight:900; cursor:pointer; text-align:center; text-decoration:none; }
            .marcar-feita.ativa { background:linear-gradient(135deg,#16a34a,#15803d); color:white; }
            .marcar-feita:disabled { cursor:default; opacity:1; background:linear-gradient(135deg,#16a34a,#15803d); color:white; }
            .gps { display:block; margin:0; padding:13px 12px; text-decoration:none; text-align:center; color:white; font-size:13px; font-weight:900; border-radius:11px; background:linear-gradient(135deg,#2f74f5,#1d4ed8); box-shadow:0 9px 20px rgba(37,99,235,.28); }
            .comprovante { display:block; margin:0; padding:13px 12px; text-decoration:none; text-align:center; color:#ecfdf5; font-size:13px; font-weight:900; border-radius:11px; background:linear-gradient(135deg,#16a34a,#15803d); box-shadow:0 8px 18px rgba(22,163,74,.24); }
            .controles { display:grid; grid-template-columns:1fr auto 1fr; align-items:center; gap:10px; padding:0 4px; }
            .controle { border:1px solid rgba(148,163,184,.18); background:#111a2d; color:#e2e8f0; border-radius:11px; padding:10px 8px; font-size:11.5px; font-weight:800; cursor:pointer; }
            .controle:disabled { opacity:.35; }
            .pontos { display:flex; gap:5px; justify-content:center; max-width:130px; overflow:hidden; }
            .ponto { width:7px; height:7px; padding:0; border:0; border-radius:50%; background:#475569; cursor:pointer; }
            .ponto.ativo { width:18px; border-radius:999px; background:#2563eb; }
        </style></head><body>
            <div class="barra"><span>↔️ Deslize para trocar de parada</span><div class="resumo-topo"><span id="feitas" class="feitas">0 feitas</span><span id="contador" class="contador">1 de __TOTAL__</span></div></div>
            <div id="trilho" class="trilho">__CARTOES__</div>
            <div class="controles"><button id="anterior" class="controle" onclick="mover(-1)">← Anterior</button><div id="pontos" class="pontos"></div><button id="proxima" class="controle" onclick="mover(1)">Próxima →</button></div>
        <script>
            const trilho = document.getElementById('trilho');
            const cartoes = Array.from(trilho.querySelectorAll('.cartao'));
            const contador = document.getElementById('contador');
            const anterior = document.getElementById('anterior');
            const proxima = document.getElementById('proxima');
            const pontos = document.getElementById('pontos');
            const feitasEl = document.getElementById('feitas');
            const focoServidor = String('__FOCO__');
            let atual = 0;
            let gestoAtivo = false;

            function indiceMaisProximo() {
                const centro = trilho.scrollLeft + trilho.clientWidth / 2;
                let melhor = 0, dist = Infinity;
                cartoes.forEach((c, i) => {
                    const d = Math.abs(c.offsetLeft + c.offsetWidth / 2 - centro);
                    if (d < dist) { dist = d; melhor = i; }
                });
                return melhor;
            }

            cartoes.forEach((_, i) => {
                const p = document.createElement('button');
                p.className = 'ponto';
                p.type = 'button';
                p.addEventListener('click', () => ir(i));
                pontos.appendChild(p);
            });

            function atualizarFeitas() {
                const botoes = Array.from(document.querySelectorAll('.marcar-feita'));
                const total = botoes.length;
                const feitas = botoes.filter(b => b.dataset.feita === '1').length;
                feitasEl.textContent = `${feitas}/${total} ${feitas === 1 ? 'feita' : 'feitas'}`;
            }
            function prepararEnvio(botao) { botao.textContent='⏳ Salvando...'; botao.style.pointerEvents='none'; }
            function atualizar(i) {
                atual = Math.max(0, Math.min(cartoes.length - 1, i));
                contador.textContent = `${atual + 1} de ${cartoes.length}`;
                anterior.disabled = atual === 0;
                proxima.disabled = atual === cartoes.length - 1;
                Array.from(pontos.children).forEach((p, j) => p.classList.toggle('ativo', j === atual));
                cartoes.forEach((c, j) => c.classList.toggle('selecionada', j === atual));
            }
            function ir(i) {
                const indice = Math.max(0, Math.min(cartoes.length - 1, i));
                const alvo = cartoes[indice];
                trilho.scrollTo({left: alvo.offsetLeft - trilho.offsetLeft, behavior: 'smooth'});
                atualizar(indice);
            }
            function mover(delta) { ir(atual + delta); }

            trilho.addEventListener('pointerdown', () => { gestoAtivo = true; }, {passive:true});
            trilho.addEventListener('pointerup', () => {
                if (!gestoAtivo) return;
                const melhor = indiceMaisProximo();
                atualizar(melhor);
                gestoAtivo = false;
            }, {passive:true});
            trilho.addEventListener('pointercancel', () => { gestoAtivo = false; }, {passive:true});

            let timer;
            trilho.addEventListener('scroll', () => {
                clearTimeout(timer);
                timer = setTimeout(() => atualizar(indiceMaisProximo()), 90);
            }, {passive:true});

            atualizarFeitas();
            const indiceInicial = cartoes.findIndex(c => String(c.dataset.etapa || '') === focoServidor);
            if (indiceInicial >= 0) {
                const alvo = cartoes[indiceInicial];
                trilho.scrollTo({left: alvo.offsetLeft - trilho.offsetLeft, behavior: 'auto'});
                atualizar(indiceInicial);
            } else {
                const primeiraPendente = cartoes.findIndex(c => !c.classList.contains('feita'));
                const indicePadrao = primeiraPendente >= 0 ? primeiraPendente : 0;
                const alvo = cartoes[indicePadrao];
                trilho.scrollTo({left: alvo.offsetLeft - trilho.offsetLeft, behavior: 'auto'});
                atualizar(indicePadrao);
            }
        </script></body></html>
        """.replace("__CARTOES__", "".join(cartoes_mobile)).replace("__TOTAL__", str(len(cartoes_mobile))).replace("__FOCO__", "" if foco_comprovante is None else str(foco_comprovante))
        st.components.v1.html(html_carrossel, height=550, scrolling=False)
    else:
        st.info("A rota ainda não possui etapas para exibir.")

    st.divider()
    mostrar_mapa_davi = st.toggle("🗺️ Mostrar mapa completo", value=False, key="davi_mostrar_mapa_completo")
    if not mostrar_mapa_davi:
        st.caption("O botão **ABRIR GPS** de cada parada leva direto ao destino.")
        st.stop()

    st.markdown("""
        <div class="aproar-section-anchor" id="mapa-rota">
            <div class="aproar-section-kicker">VISÃO GERAL</div>
            <div class="aproar-section-title">Mapa da rota</div>
            <div class="aproar-section-help">Trajeto, sequência e localização das paradas</div>
        </div>
    """, unsafe_allow_html=True)
    m_mobile = folium.Map(location=[-3.7319, -38.5267], zoom_start=12, tiles="OpenStreetMap")
    pontos_reais_mobile = []
    if p_saida in locais_dict:
        pontos_reais_mobile.append([float(locais_dict[p_saida][0]), float(locais_dict[p_saida][1])])
    for i, step in enumerate(route_steps):
        if step.get('destino') in locais_dict and step.get('type') not in ['lunch', 'return'] and not (i == 0 and step.get('destino') == p_saida):
            _lat_r, _lon_r = locais_dict[step['destino']]
            pontos_reais_mobile.append([float(_lat_r), float(_lon_r)])

    def _escala_visual_mobile(pontos):
        if len(pontos) < 2:
            return 0.75
        lat_ref = sum(p[0] for p in pontos) / len(pontos)
        span_lat_km = (max(p[0] for p in pontos) - min(p[0] for p in pontos)) * 111.0
        span_lon_km = (max(p[1] for p in pontos) - min(p[1] for p in pontos)) * 111.0 * max(math.cos(math.radians(lat_ref)), 0.2)
        return max(0.70, min(2.0, max(span_lat_km, span_lon_km, 1.0) * 0.070))

    distancia_visual_mobile = _escala_visual_mobile(pontos_reais_mobile)
    marcadores_posicionados_mobile = []

    def apply_offset_mobile(lat, lon):
        lat, lon = float(lat), float(lon)
        if not marcadores_posicionados_mobile:
            marcadores_posicionados_mobile.append((lat, lon))
            return lat, lon
        if all(calcular_distancia_km(lat, lon, p_lat, p_lon) >= distancia_visual_mobile for p_lat, p_lon in marcadores_posicionados_mobile):
            marcadores_posicionados_mobile.append((lat, lon))
            return lat, lon
        for tentativa in range(1, 49):
            anel = 1 + (tentativa - 1) // 12
            angulo = math.radians(((tentativa - 1) % 12) * 30 + anel * 11)
            raio_km = distancia_visual_mobile * (0.82 + 0.42 * (anel - 1))
            dlat = (raio_km / 111.0) * math.sin(angulo)
            dlon = (raio_km / (111.0 * max(math.cos(math.radians(lat)), 0.2))) * math.cos(angulo)
            candidato = (lat + dlat, lon + dlon)
            if all(calcular_distancia_km(candidato[0], candidato[1], p_lat, p_lon) >= distancia_visual_mobile * 0.92 for p_lat, p_lon in marcadores_posicionados_mobile):
                marcadores_posicionados_mobile.append(candidato)
                return candidato
        candidato = (lat - distancia_visual_mobile / 111.0, lon + distancia_visual_mobile / 111.0)
        marcadores_posicionados_mobile.append(candidato)
        return candidato

    p_num_mapa = 1
    pos_base_mobile = apply_offset_mobile(*locais_dict[p_saida]) if p_saida in locais_dict else None

    # Sempre desenha algum traçado. Também corrige automaticamente geometrias
    # antigas do OSRM que foram salvas no formato [lon, lat].
    geom_mobile = normalizar_geometria_mapa(geometria_rota or [], pontos_reais_mobile)
    geom_mobile_viaria = bool(st.session_state.get('geometria_viaria', False)) and len(geom_mobile) > 2
    if len(geom_mobile) < 2 and len(pontos_reais_mobile) > 1:
        geom_mobile = [list(p) for p in pontos_reais_mobile]
        geom_mobile_viaria = False
    if len(geom_mobile) > 1:
        folium.PolyLine(geom_mobile, color="#FFFFFF", weight=8, opacity=0.80).add_to(m_mobile)
        folium.PolyLine(
            geom_mobile, color="#2563eb", weight=5, opacity=0.98,
            dash_array=None if geom_mobile_viaria else "9,7",
            tooltip="Traçado da rota" if geom_mobile_viaria else "Ligação aproximada entre as paradas",
        ).add_to(m_mobile)

    for i, step in enumerate(route_steps):
        if step.get('destino') and step['destino'] in locais_dict:
            if step.get('type') in ['lunch', 'return']: continue
            if (i == 0 and step['destino'] == p_saida): continue

            lat_orig, lon_orig = map(float, locais_dict[step['destino']])
            lat, lon = apply_offset_mobile(lat_orig, lon_orig)
            if calcular_distancia_km(lat_orig, lon_orig, lat, lon) > 0.01:
                folium.PolyLine(
                    [[lat_orig, lon_orig], [lat, lon]], color="#475569", weight=2,
                    opacity=0.90, dash_array="4,5",
                    tooltip="O círculo foi afastado; a ponta da linha é o local real",
                ).add_to(m_mobile)
                folium.CircleMarker([lat_orig, lon_orig], radius=3, color="#475569", weight=1, fill=True, fill_opacity=0.9).add_to(m_mobile)

            acoes = [a[0] for a in step.get('actions', [])]
            tem_coleta, tem_entrega = "COLETAR" in acoes, "ENTREGAR" in acoes
            fundo_marcador = "linear-gradient(90deg, #f59e0b 0 50%, #16a34a 50% 100%)" if (tem_coleta and tem_entrega) else "#f59e0b" if tem_coleta else "#16a34a"
            popup_html = f"<b>Parada {p_num_mapa}: {html_escape(str(step['destino']))}</b>"
            folium.Marker(
                [lat, lon], popup=folium.Popup(popup_html, max_width=280), tooltip=f"Parada {p_num_mapa}",
                z_index_offset=1200 + p_num_mapa,
                icon=folium.DivIcon(html=f'''<div style="background: {fundo_marcador}; color: white; border: 3px solid white; border-radius: 50%; width: 32px; height: 32px; display: flex; justify-content: center; align-items: center; font-weight: 900; box-shadow: 0 2px 7px rgba(0,0,0,0.65); font-size: 14px;">{p_num_mapa}</div>''')
            ).add_to(m_mobile)
            p_num_mapa += 1

    if len(pontos_reais_mobile) > 1:
        m_mobile.fit_bounds(pontos_reais_mobile, padding=(30, 30), max_zoom=14)
    if p_saida in locais_dict and pos_base_mobile is not None:
        folium.Marker(
            [pos_base_mobile[0], pos_base_mobile[1]], popup=folium.Popup(f"<b>Saída: {html_escape(str(p_saida))}</b>", max_width=280),
            z_index_offset=2500,
            icon=folium.DivIcon(html=f'''<div style="background: linear-gradient(135deg, #2563eb, #1d4ed8); color: white; border: 3px solid white; border-radius: 50%; width: 34px; height: 34px; display: flex; justify-content: center; align-items: center; box-shadow: 0 2px 8px rgba(0,0,0,0.7); font-size: 16px;">🏁</div>''')
        ).add_to(m_mobile)

    st_folium(m_mobile, height=400, use_container_width=True, returned_objects=[])
    st.markdown("<div style='text-align:center;font-size:11px;margin:8px 0 18px;color:#64748b;'><b style='color:#94a3b8;'>LEGENDA</b> &nbsp; 🟡 Coleta &nbsp; 🟢 Entrega &nbsp; 🏁 Início<br>Azul = trajeto • cinza = ajuste visual do marcador</div>", unsafe_allow_html=True)
    st.markdown("""
        <nav class="aproar-driver-bottom-nav" aria-label="Navegação do motorista">
            <a href="#rota"><b>⌂</b><span>Resumo</span></a>
            <a href="#roteiro"><b>↗</b><span>Roteiro</span></a>
            <a href="#mapa-rota"><b>⌖</b><span>Mapa</span></a>
        </nav>
    """, unsafe_allow_html=True)
    st.caption("Central de Logística APROAR • rota sincronizada com a Torre")
    st.stop()

# =====================================================================
# DESIGN INDUSTRIAL MINIMAL — TORRE DE CONTROLE (SOMENTE COMPUTADOR)
# =====================================================================
st.markdown("""
    <style>
        :root {
            --ap-bg:#070913;
            --ap-bg-soft:#0a0e1b;
            --ap-surface:#0f1526;
            --ap-surface-2:#131b30;
            --ap-surface-3:#18223a;
            --ap-line:rgba(148,163,184,.18);
            --ap-line-strong:rgba(96,165,250,.32);
            --ap-text:#f4f7fb;
            --ap-muted:#ffffff;
            --ap-blue:#2563eb;
            --ap-blue-2:#3b82f6;
            --ap-green:#22c55e;
            --ap-amber:#f59e0b;
            --ap-red:#ef4444;
            --ap-radius-sm:4px;
            --ap-radius:6px;
            --ap-radius-lg:8px;
            --ap-shadow:0 14px 34px rgba(0,0,0,.22);
        }

        html, body, [data-testid="stAppViewContainer"] { background:var(--ap-bg) !important; }
        [data-testid="stAppViewContainer"] {
            background-image:linear-gradient(180deg,#080b18 0%,#060812 100%) !important;
        }
        .main .block-container { max-width:1680px; padding:.85rem 1.25rem 4rem; }
        [data-testid="stHeader"] {
            background:rgba(7,9,19,.94) !important;
            border-bottom:1px solid var(--ap-line); backdrop-filter:blur(14px);
        }
        [data-testid="stSidebar"] {
            width:292px !important; min-width:292px !important;
            background:#0a1020 !important; border-right:1px solid var(--ap-line) !important;
        }
        [data-testid="stSidebar"] .block-container { padding:1rem .9rem 2rem; }

        .aproar-shell-header {
            margin:0 0 14px; padding:13px 4px 18px; overflow:visible;
            background:transparent; border:0; border-bottom:1px solid var(--ap-line);
            border-radius:0; box-shadow:none;
        }
        .aproar-shell-header::after { display:none; }
        .aproar-brand { gap:20px; }
        .aproar-logo-main { width:170px; height:52px; filter:none; }
        .aproar-eyebrow { color:#60a5fa; font-size:10px; letter-spacing:.17em; }
        .aproar-title {
            margin-top:1px; color:#f6f7f6; font-size:24px; font-weight:700;
            letter-spacing:.015em; text-transform:uppercase;
        }
        .aproar-subtitle { color:#8ea0ba; font-size:12px; }
        .aproar-meta-chip {
            min-height:42px; border-radius:5px; color:#dbeafe;
            background:#101a31; border:1px solid #263759;
        }
        .aproar-meta-chip.primary { color:#bfdbfe; background:#0f1a31; border-color:#28518d; }
        .aproar-meta-chip:last-child { color:#7be5a9; border-color:rgba(34,197,94,.28); }
        .aproar-dot { background:var(--ap-green); }

        .aproar-sidebar-brand {
            margin:0 0 15px; padding:8px 8px 16px; gap:12px;
            background:transparent; border:0; border-bottom:1px solid var(--ap-line); border-radius:0;
        }
        .aproar-logo-sidebar { width:118px; height:42px; flex-basis:118px; }
        .aproar-sidebar-brand strong { color:#f4f5f4; font-size:12px; text-transform:uppercase; }
        .aproar-sidebar-brand small { color:#71829e; }
        .aproar-sidebar-section {
            margin:4px 8px 7px; color:#62738f; font-size:9px; font-weight:800;
            letter-spacing:.16em; text-transform:uppercase;
        }

        [data-testid="stSidebar"] div[role="radiogroup"] { gap:4px; }
        [data-testid="stSidebar"] div[role="radiogroup"] > label {
            min-height:44px; margin:0; padding:0 11px; border-left:3px solid transparent;
            border-radius:4px; background:transparent; color:#aab2af;
            transition:background .14s ease,color .14s ease,border-color .14s ease;
        }
        [data-testid="stSidebar"] div[role="radiogroup"] > label:hover { background:#121b30; color:#fff; }
        [data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked) {
            color:#fff; background:linear-gradient(90deg,rgba(239,68,68,.20),rgba(239,68,68,.05));
            border-left-color:var(--ap-red);
        }
        [data-testid="stSidebar"] div[role="radiogroup"] > label > div:first-child,
        [data-testid="stSidebar"] [data-baseweb="radio"] > div:first-child { display:none !important; }
        [data-testid="stSidebar"] div[role="radiogroup"] p { font-size:12px; font-weight:700; }

        .stButton > button, .stDownloadButton > button, [data-testid="baseButton-secondary"] {
            min-height:40px; border-radius:4px !important; background:#10172a !important;
            border:1px solid var(--ap-line) !important; box-shadow:none !important;
        }
        .stButton > button:hover, .stDownloadButton > button:hover {
            transform:none; color:#d9e5f3 !important; border-color:rgba(113,149,196,.46) !important;
        }
        button[kind="primary"], [data-testid="baseButton-primary"],
        section[data-testid="stMain"] .stButton > button[kind="primary"] {
            min-height:42px; border-radius:4px !important; color:#f8fbff !important;
            background:#2563eb !important; border:1px solid #3b82f6 !important; box-shadow:none !important;
        }
        button[kind="primary"]:hover { transform:none; color:#fff !important; background:#2f6fec !important; box-shadow:none !important; }

        div[data-testid="stMetric"], div[data-testid="stForm"],
        div[data-testid="stVerticalBlockBorderWrapper"], [data-testid="stExpander"],
        [data-testid="stDataFrame"], [data-testid="stTable"] {
            border-radius:6px !important; background:#0f1526 !important;
            border:1px solid var(--ap-line) !important; box-shadow:none !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"] { padding:14px !important; }
        [data-testid="stAlert"], [data-testid="stAlert"] > div,
        div[data-baseweb="notification"] {
            border-radius:5px !important; color:#d7dcda !important;
            background:#414916 !important; border-color:#59631d !important;
        }
        .aproar-eta-card {
            margin:9px 0 14px; border-radius:6px; background:#10182c;
            border:1px solid var(--ap-line); box-shadow:none;
        }
        .aproar-eta-value { font-size:21px; }
        .aproar-eta-side.right .aproar-eta-value { color:var(--ap-green); }
        iframe[title*="streamlit_folium"], div[data-testid="stIFrame"] iframe {
            border:1px solid var(--ap-line); border-radius:6px;
        }
        [data-testid="stExpander"] details > summary { min-height:46px; }
        [data-testid="stExpander"] details > summary:hover { color:#d9e5f3; }

        .aproar-industrial-heading {
            display:flex; align-items:center; justify-content:space-between; gap:12px;
            margin:0 0 10px; padding:0 0 10px; border-bottom:1px solid var(--ap-line);
        }
        .aproar-industrial-heading h2 {
            margin:0; color:#f4f5f4 !important; font-size:18px; font-weight:700;
            letter-spacing:.035em; text-transform:uppercase;
        }
        .aproar-industrial-heading span {
            padding:4px 8px; color:#aab2af; font-size:10px; font-weight:800;
            letter-spacing:.08em; border:1px solid var(--ap-line); border-radius:4px;
        }
        .aproar-stop-header {
            display:grid; grid-template-columns:auto minmax(0,1fr) auto; align-items:center; gap:10px;
            margin:-2px 0 9px; padding:9px 10px; border:1px solid var(--ap-line);
            border-left-width:4px; border-radius:5px;
        }
        .aproar-stop-header:has(.aproar-stop-action.coleta) {
            background:rgba(245,158,11,.10); border-left-color:#f59e0b;
        }
        .aproar-stop-header:has(.aproar-stop-action.entrega) {
            background:rgba(34,197,94,.10); border-left-color:#22c55e;
        }
        .aproar-stop-action {
            padding:4px 7px; border-radius:4px; font-size:10px; font-weight:900;
            letter-spacing:.08em;
        }
        .aproar-stop-action.coleta { color:#fbbf24; background:rgba(245,158,11,.10); border:1px solid rgba(245,158,11,.35); }
        .aproar-stop-action.entrega { color:#4ade80; background:rgba(34,197,94,.10); border:1px solid rgba(34,197,94,.30); }
        .aproar-stop-copy { min-width:0; }
        .aproar-stop-copy strong { display:block; overflow:hidden; color:#f4f5f4; font-size:13px; text-overflow:ellipsis; white-space:nowrap; }
        .aproar-stop-copy small { display:block; margin-top:3px; color:#87918e; font-size:10px; }
        .aproar-stop-number { color:#67716e; font-size:11px; font-weight:800; }
        .aproar-material-table { margin:5px 0; overflow:hidden; border:1px solid var(--ap-line); border-radius:4px; }
        .aproar-material-row {
            display:grid; grid-template-columns:minmax(0,1fr) auto; gap:10px;
            padding:7px 9px; background:#0b1020; border-bottom:1px solid rgba(148,163,184,.10);
        }
        .aproar-material-row:last-child { border-bottom:0; }
        .aproar-material-row span { color:#d6dbd9; font-size:11px; }
        .aproar-material-row strong { color:#f4f5f4; font-size:11px; white-space:nowrap; }
        .aproar-unit-line {
            display:flex; justify-content:space-between; gap:10px; margin-top:7px;
            color:#7f8986; font-size:10px;
        }
        .aproar-unit-line strong { color:#cfd5d2; text-align:right; }
        .aproar-industrial-summary {
            display:grid; grid-template-columns:1.35fr .75fr; gap:10px; margin-top:10px;
        }
        .aproar-route-panel, .aproar-fleet-panel {
            min-width:0; padding:13px 14px; background:#0f1526;
            border:1px solid var(--ap-line); border-radius:6px;
        }
        .aproar-summary-title {
            display:flex; align-items:center; justify-content:space-between; gap:10px;
            padding-bottom:9px; border-bottom:1px solid var(--ap-line);
        }
        .aproar-summary-title strong { color:#f4f5f4; font-size:13px; text-transform:uppercase; }
        .aproar-summary-title span { color:var(--ap-blue-2); font-size:9px; font-weight:800; letter-spacing:.08em; }
        .aproar-summary-times { display:grid; grid-template-columns:1fr auto 1fr; align-items:center; gap:10px; padding:12px 0; }
        .aproar-summary-times div:last-child { text-align:right; }
        .aproar-summary-times span, .aproar-summary-times strong { display:block; }
        .aproar-summary-times span { color:#7d8784; font-size:9px; text-transform:uppercase; }
        .aproar-summary-times strong { margin-top:3px; color:#f4f5f4; font-size:20px; }
        .aproar-summary-times div:last-child strong { color:var(--ap-green); }
        .aproar-summary-line { width:100px; height:4px; background:#242b29; }
        .aproar-summary-line i { display:block; width:62%; height:100%; background:var(--ap-blue); }
        .aproar-summary-data { display:grid; grid-template-columns:repeat(4,1fr); border-top:1px solid var(--ap-line); }
        .aproar-summary-data div { padding:9px 8px 0; border-right:1px solid var(--ap-line); }
        .aproar-summary-data div:first-child { padding-left:0; }
        .aproar-summary-data div:last-child { border-right:0; }
        .aproar-summary-data span, .aproar-summary-data strong { display:block; }
        .aproar-summary-data span { color:#747e7b; font-size:8px; text-transform:uppercase; }
        .aproar-summary-data strong { margin-top:4px; color:#dce1df; font-size:11px; }
        .aproar-fleet-body { padding-top:12px; }
        .aproar-fleet-body span, .aproar-fleet-body strong, .aproar-fleet-body small { display:block; }
        .aproar-fleet-body > span { color:#737d7a; font-size:9px; text-transform:uppercase; }
        .aproar-fleet-body strong { margin-top:5px; color:#f4f5f4; font-size:14px; }
        .aproar-fleet-body small { margin-top:3px; color:#8b9592; }
        .aproar-fleet-cost { margin-top:12px; padding-top:10px; border-top:1px solid var(--ap-line); }
        .aproar-fleet-cost b { color:#cbd9e9; font-size:15px; }

        /* Legendas e subtextos brancos — somente na plataforma do computador. */
        [data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] p,
        .stCaption, .stCaption p,
        [data-testid="stSidebar"] small, [data-testid="stSidebar"] .stCaption,
        [data-testid="stSidebar"] .stCaption p,
        .aproar-subtitle, .aproar-sidebar-brand small,
        .aproar-stop-copy small, .aproar-stop-number,
        .aproar-unit-line, .aproar-unit-line strong,
        .aproar-eta-label, .aproar-industrial-heading span,
        .aproar-summary-title span, .aproar-summary-times span,
        .aproar-summary-data span, .aproar-fleet-body > span,
        .aproar-fleet-body small {
            color:#ffffff !important; opacity:1 !important;
        }

        ::-webkit-scrollbar-track { background:#080b0b; }
        ::-webkit-scrollbar-thumb { background:#2c3331; border-color:#080b0b; border-radius:2px; }
        ::-webkit-scrollbar-thumb:hover { background:#3b82f6; }

        @media (max-width:1100px) {
            [data-testid="stSidebar"] { width:250px !important; min-width:250px !important; }
            .aproar-title { font-size:20px; }
            .aproar-logo-main { width:142px; }
            .aproar-industrial-summary { grid-template-columns:1fr; }
        }
    </style>
""", unsafe_allow_html=True)

# =====================================================================
# TORRE DE CONTROLE (PC)
# =====================================================================
TRELLO_JSON_URL = "https://trello.com/b/tyR8YgDF.json"
RASTREADOR_LOGIN_URLS = ["https://portal.protegeexpress.com.br/sistema/login.aspx", "http://portal.protegeexpress.com.br/sistema/login.aspx"]
RASTREADOR_VEICULOS_PADRAO = "007046861,807289138"
VELOCIDADE_MEDIA_KMH = 25.0
ROTA_ENGINE_VERSION = 7

COLUNAS_DEMANDAS = ["id", "Obra", "Origem", "Destino", "Materiais", "Urgência", "Peso", "Tempo_Coleta", "Tempo_Entrega", "Supervisor", "_Titulo_Trello"]

# O limite do expediente é declarado junto às constantes do Davi, antes do
# App do Motorista, para estar disponível tanto no mobile quanto na Torre (PC).

# Nomes canônicos das unidades atendidas pela empresa. FIEC/CASA DA INDÚSTRIA
# permanecem por compatibilidade com cadastros antigos já existentes no sistema.
UNIDADES_PROPRIAS = [
    "UNIFOR", "CENTRO", "MUSEU", "BARRA DO CEARÁ", "MARACANAÚ", "COLISEU",
    "HORIZONTE", "SEBRAE", "PARANGABA", "ESCRITÓRIO", "ESCRITÓRIO PROVISÓRIO",
    "SÃO GONÇALO DO AMARANTE", "FIEC", "CASA DA INDÚSTRIA",
]

# O Trello costuma misturar a instituição (FIEC/SENAI/SESI/IEL) com o nome da
# unidade. Todos os aliases abaixo convergem para um único ponto físico/canônico.
ALIASES_UNIDADES_EMPRESA = {
    "UNIFOR": ("UNIFOR",),
    "CENTRO": ("CENTRO", "SENAI CENTRO", "SESI CENTRO", "IEL CENTRO", "FIEC CENTRO", "ESCOLA CENTRO", "NR SAÚDE", "NR SAUDE"),
    "MUSEU": ("MUSEU", "SESI MUSEU", "SENAI MUSEU"),
    "BARRA DO CEARÁ": ("BARRA DO CEARÁ", "BARRA DO CEARA", "SENAI BARRA", "SESI BARRA", "IEL BARRA", "FIEC BARRA", "SENAI BARRA DO CEARÁ", "SESI BARRA DO CEARÁ", "BARRA"),
    "MARACANAÚ": ("MARACANAÚ", "MARACANAU", "SESI ALBANO FRANCO", "SESI CLUBE DA PARCERIA", "SENAI ISTEMM", "SENAI CETAFR"),
    "COLISEU": ("COLISEU", "EDIFÍCIO COLISEU", "EDIFICIO COLISEU", "APARTAMENTO COLISEU", "CONDOMÍNIO COLISEU", "CONDOMINIO COLISEU"),
    "HORIZONTE": ("HORIZONTE", "SESI HORIZONTE", "SENAI HORIZONTE"),
    "SEBRAE": ("SEBRAE",),
    "PARANGABA": ("PARANGABA", "SESI PARANGABA", "SENAI PARANGABA"),
    "ESCRITÓRIO PROVISÓRIO": ("ESCRITÓRIO PROVISÓRIO", "ESCRITORIO PROVISORIO", "ESCRITÓRIO PROVISORIO", "ESCRITORIO PROVISÓRIO"),
    "ESCRITÓRIO": ("ESCRITÓRIO", "ESCRITORIO", "ALMOXARIFADO"),
    "SÃO GONÇALO DO AMARANTE": ("SÃO GONÇALO DO AMARANTE", "SAO GONCALO DO AMARANTE", "SÃO GONÇALO", "SAO GONCALO", "SGA"),
    "FIEC": ("FIEC",),
    "CASA DA INDÚSTRIA": ("CASA DA INDÚSTRIA", "CASA DA INDUSTRIA"),
}


def _chave_busca_unidade(valor):
    texto = remover_acentos(str(valor or "")).upper()
    return re.sub(r"[^A-Z0-9]+", " ", texto).strip()


def identificar_unidade_empresa(valor, permitir_contexto=False):
    """Reconhece a unidade mesmo quando o Trello usa SENAI/SESI/FIEC/IEL ou abreviações.

    Em nomes de ponto físico exigimos correspondência exata para não confundir,
    por exemplo, um fornecedor localizado no bairro Centro com a obra CENTRO.
    No TÍTULO do cartão, ``permitir_contexto=True`` permite encontrar a unidade
    dentro de textos maiores como "OBRA 086 - REFORMA FACHADA - UNIFOR".
    """
    chave = _chave_busca_unidade(valor)
    if not chave:
        return ""

    # Aliases longos primeiro: evita que ESCRITÓRIO seja escolhido antes de
    # ESCRITÓRIO PROVISÓRIO e que BARRA vença BARRA DO CEARÁ.
    candidatos = []
    for oficial, aliases in ALIASES_UNIDADES_EMPRESA.items():
        for alias in aliases:
            alias_chave = _chave_busca_unidade(alias)
            candidatos.append((len(alias_chave), alias_chave, oficial))

    for _tam, alias_chave, oficial in sorted(candidatos, reverse=True):
        if chave == alias_chave:
            return oficial
        if permitir_contexto and re.search(rf"(?<![A-Z0-9]){re.escape(alias_chave)}(?![A-Z0-9])", chave):
            return oficial
    return ""


SUPERVISORES_MAP = {
    "BARRA DO CEARÁ": "Luis Eduardo Rodrigues", "BARRA": "Luis Eduardo Rodrigues", "SESI BARRA DO CEARÁ": "Luis Eduardo Rodrigues", "SENAI BARRA DO CEARÁ": "Luis Eduardo Rodrigues",
    "CENTRO": "Victor Bezerra", "SENAI CENTRO": "Victor Bezerra", "ESCOLA CENTRO": "Victor Bezerra", "NR SAÚDE": "Victor Bezerra", "MUSEU": "Victor Bezerra", "SESI MUSEU": "Victor Bezerra",
    "CASA DA INDÚSTRIA": "Gustavo Souza", "FIEC": "Gustavo Souza",
    "MARACANAÚ": "Neto Porto", "SESI ALBANO FRANCO": "Neto Porto", "SESI CLUBE DA PARCERIA": "Neto Porto", "SENAI ISTEMM": "Neto Porto", "SENAI CETAFR": "Neto Porto",
    "HORIZONTE": "Soares Junior", "SESI HORIZONTE": "Soares Junior", "SENAI HORIZONTE": "Soares Junior", "SEBRAE": "Soares Junior",
    "UNIFOR": "Joel Lima", "ESCRITÓRIO": "Sede / Logística", "PARANGABA": "Sede / Logística"
}

TEAMS_SECRET_KEYS = {"Luis Eduardo Rodrigues": "luis_eduardo", "Victor Bezerra": "victor_bezerra", "Gustavo Souza": "gustavo_souza", "Neto Porto": "neto_porto", "Soares Junior": "soares_junior", "Joel Lima": "joel_lima", "Sede / Logística": "sede_logistica", "Geral / Logística": "geral_logistica"}
LOCAL_BASE_ENDERECO = "Rua Professor Mário Rocha, 84 - Joaquim Távora, Fortaleza - CE, 60120-200"
LOCAL_BASE_COORDS = (-3.752270016704, -38.51537298342)
ALIASES_LOCAL_BASE = {"ALMOXARIFADO", "ESCRITÓRIO"}

# ENDEREÇOS FIXOS DO SISTEMA
ENDERECOS_PADRAO = [
    ("CASA DA INDÚSTRIA", "Av. Barão de Studart, 1980 - Aldeota, Fortaleza - CE"), 
    ("SENAI CENTRO", "R. Padre Ibiapina, 1280 - Jacarecanga, Fortaleza - CE"), 
    ("ESCOLA CENTRO", "R. Padre Ibiapina, 1280 - Jacarecanga, Fortaleza - CE"), 
    ("CENTRO", "R. Padre Ibiapina, 1280 - Jacarecanga, Fortaleza - CE"), 
    ("NR SAÚDE", "R. Padre Ibiapina, 1280 - Jacarecanga, Fortaleza - CE"), 
    ("SESI BARRA DO CEARÁ", "Rua Florencio de Alencar, 900 - Barra do Ceará, Fortaleza - CE"), 
    ("SENAI BARRA DO CEARÁ", "Rua Florencio de Alencar, 900 - Barra do Ceará, Fortaleza - CE"), 
    ("BARRA DO CEARÁ", "Rua Florencio de Alencar, 900 - Barra do Ceará, Fortaleza - CE"),
    ("BARRA", "Rua Florencio de Alencar, 900 - Barra do Ceará, Fortaleza - CE"),
    ("SESI ALBANO FRANCO", "Av. Sen. Virgílio Távora, 1395 - Distrito Industrial I, Maracanaú - CE"), 
    ("SESI CLUBE DA PARCERIA", "Av. Sen. Virgílio Távora, 1395 - Distrito Industrial I, Maracanaú - CE"), 
    ("SENAI ISTEMM", "Av. Sen. Virgílio Távora, 1395 - Distrito Industrial I, Maracanaú - CE"), 
    ("SENAI CETAFR", "Av. Sen. Virgílio Távora, 1395 - Distrito Industrial I, Maracanaú - CE"), 
    ("MARACANAÚ", "Av. Sen. Virgílio Távora, 1395 - Distrito Industrial I, Maracanaú - CE"),
    ("SESI PARANGABA", "Av. João Pessoa, 6754 - Parangaba, Fortaleza - CE"), 
    ("SENAI PARANGABA", "Av. João Pessoa, 6760 - Damas, Fortaleza - CE"), 
    ("PARANGABA", "Av. João Pessoa, 6760 - Damas, Fortaleza - CE"),
    ("SESI MUSEU", "R. Dr. João Moreira, 143 - Centro, Fortaleza - CE, 60030-000"), 
    ("MUSEU", "R. Dr. João Moreira, 143 - Centro, Fortaleza - CE, 60030-000"), 
    ("SESI SOBRAL", "Av. Dr. José Arimathéa Monte e Silva, 1003 - Junco, Sobral - CE"), 
    ("ESCRITÓRIO", LOCAL_BASE_ENDERECO), 
    ("ALMOXARIFADO", LOCAL_BASE_ENDERECO), 
    ("ESPAÇO SMART", "BR-116, 9370 - Barroso, Fortaleza - CE, 60862-735"), 
    ("ALDEOTA", "Rua Dr. José Lourenço, 1990 - Aldeota, Fortaleza - CE"), 
    ("EDSON QUEIROZ", "Av. Dr. Valmir Pontes, 675 - Edson Queiroz, Fortaleza - CE"), 
    ("FIEC", "Rua Dr. José Lourenço, 1990 - Aldeota, Fortaleza - CE"), 
    ("UNIFOR", "Av. Dr. Valmir Pontes, 675 - Edson Queiroz, Fortaleza - CE"), 
    ("HORIZONTE", "R. Raimunda Pontes - Planalto Horizonte, Horizonte - CE"), 
    ("SEBRAE", "Avenida Monsenhor Tabosa, 777 - Meireles, Fortaleza - CE"), 
    ("LECI FERRAGENS", "Rua Gen. Clarindo de Queiroz, 1668 - Centro, Fortaleza - CE"),
    ("ELÉTRICA FORTALEZA", "Centro, Fortaleza - CE"), 
    ("ELETRICA FORTALEZA", "Centro, Fortaleza - CE"), 
    ("DEPÓSITO JP", "Edson Queiroz, Fortaleza - CE"),
    ("DEPOSITO JP", "Edson Queiroz, Fortaleza - CE"),
    ("JP CONSTRUÇÃO", "Edson Queiroz, Fortaleza - CE"),
    ("JP CONSTRUCOES", "Edson Queiroz, Fortaleza - CE")
]

# Fornecedores conhecidos usados apenas como FALLBACK quando o banco ainda não
# possui o local. Diferente das unidades próprias, estes endereços NÃO substituem
# um cadastro feito manualmente pela equipe na aba Endereços.
ENDERECOS_FORNECEDORES_FALLBACK = [
    ("SV ELÉTRICA", "Av. Bezerra de Menezes, 420 - Farias Brito, Fortaleza - CE, 60325-000"),
    ("SV ELETRICA", "Av. Bezerra de Menezes, 420 - Farias Brito, Fortaleza - CE, 60325-000"),
    ("SV ELÉTRICA MATRIZ", "Av. Bezerra de Menezes, 420 - Farias Brito, Fortaleza - CE, 60325-000"),
    ("SV ELÉTRICA CD", "R. Licurgo Montenegro, 585 - Padre Andrade, Fortaleza - CE, 60356-215"),
    ("SV ELÉTRICA WASHINGTON SOARES", "Av. Washington Soares, 6450 - Cambeba, Fortaleza - CE, 60822-142"),
    ("SV ELÉTRICA MARACANAÚ", "Av. Dr. Mendel Steinbruch, 6340 - Aracapé, Fortaleza - CE, 60765-242"),
    ("FORTEX", "Rodovia 4º Anel Viário, 1515 - KM 9,5 - Distrito Industrial III, Maracanaú - CE, 61930-220"),
]

SCHEMA_APP_VERSION = "2026-08-28-v13"

@st.cache_resource(show_spinner=False)
def inicializar_bd():
    """Prepara o banco sem repetir dezenas de DDLs a cada cold start do Streamlit."""
    conn_db = get_conn()
    with conn_db.session as s:
        # Esta tabela pequena permite saber se as migrações desta versão já foram
        # aplicadas no Supabase. Em reinicializações futuras fazemos só duas consultas.
        s.execute(text("CREATE TABLE IF NOT EXISTS app_meta (chave TEXT PRIMARY KEY, valor TEXT)"))
        versao = s.execute(
            text("SELECT valor FROM app_meta WHERE chave='schema_version'")
        ).fetchone()
        if versao and str(versao[0]) == SCHEMA_APP_VERSION:
            s.commit()
            return True
        s.commit()

    queries = [
        "CREATE TABLE IF NOT EXISTS locais (apelido TEXT PRIMARY KEY, endereco TEXT, lat REAL, lon REAL)",
        "CREATE TABLE IF NOT EXISTS locais_removidos (apelido TEXT PRIMARY KEY)",
        "CREATE TABLE IF NOT EXISTS config_frota (id SERIAL PRIMARY KEY, consumo REAL, preco_gasolina REAL)",
        "CREATE TABLE IF NOT EXISTS abastecimentos (id SERIAL PRIMARY KEY, data TEXT, litros REAL, valor_litro REAL, manutencao REAL, obs TEXT, veiculo TEXT DEFAULT 'Strada')",
        "CREATE TABLE IF NOT EXISTS registro_km (id SERIAL PRIMARY KEY, data TEXT, km REAL, obs TEXT, veiculo TEXT DEFAULT 'Strada')",
        "CREATE TABLE IF NOT EXISTS historico_concluidos (id TEXT PRIMARY KEY, obra TEXT, origem TEXT, destino TEXT, materiais TEXT, data_conclusao TEXT, hora_conclusao TEXT)",
        "CREATE TABLE IF NOT EXISTS rastreio_paradas (id SERIAL PRIMARY KEY, data TEXT, placa TEXT, local TEXT, hora_chegada TEXT, hora_saida TEXT)",
        "CREATE TABLE IF NOT EXISTS inicio_movimento (placa TEXT, data TEXT, hora_inicio TEXT, PRIMARY KEY(placa, data))",
        "CREATE TABLE IF NOT EXISTS webhooks_teams (setor TEXT PRIMARY KEY, url TEXT)",
        "CREATE TABLE IF NOT EXISTS config_trello (id SERIAL PRIMARY KEY, api_key TEXT, token TEXT, id_lista_concluida TEXT)",
        "CREATE TABLE IF NOT EXISTS trello_cache (id SMALLINT PRIMARY KEY, dados JSONB NOT NULL DEFAULT '{}'::jsonb, atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS rota_ativa (id SERIAL PRIMARY KEY, data_rota TEXT, json_route TEXT, json_locais TEXT, json_geometria TEXT, json_enderecos TEXT, total_km REAL)",
        SQL_TABELA_CHECKINS_DAVI,
        "ALTER TABLE rota_ativa ADD COLUMN IF NOT EXISTS fonte_matriz TEXT",
        "ALTER TABLE rota_ativa ADD COLUMN IF NOT EXISTS horario_matriz TEXT",
        "ALTER TABLE rota_ativa ADD COLUMN IF NOT EXISTS json_ajustes_manuais TEXT",
        # Controle de entrega ao Teams. `NULL` identifica registros antigos criados
        # antes desta melhoria; novas baixas entram explicitamente como FALSE e são
        # tentadas novamente até o Teams confirmar o recebimento.
        "ALTER TABLE historico_concluidos ADD COLUMN IF NOT EXISTS teams_notificado BOOLEAN",
        "ALTER TABLE historico_concluidos ADD COLUMN IF NOT EXISTS teams_tentativas INTEGER DEFAULT 0",
        "ALTER TABLE historico_concluidos ADD COLUMN IF NOT EXISTS teams_ultimo_erro TEXT",
        "CREATE INDEX IF NOT EXISTS idx_historico_concluidos_data ON historico_concluidos (data_conclusao)",
        "CREATE INDEX IF NOT EXISTS idx_inicio_movimento_data ON inicio_movimento (data)",
        "CREATE INDEX IF NOT EXISTS idx_rastreio_paradas_data_placa ON rastreio_paradas (data, placa)",
        "CREATE INDEX IF NOT EXISTS idx_rastreio_paradas_abertas ON rastreio_paradas (data, placa, hora_saida)",
    ]

    # Uma única sessão/commit evita dezenas de viagens separadas até o Supabase.
    with conn_db.session as s:
        for query in queries:
            s.execute(text(query))

        s.execute(text("INSERT INTO config_frota (id, consumo, preco_gasolina) VALUES (1, 11.5, 5.90) ON CONFLICT (id) DO NOTHING"))
        s.execute(text("INSERT INTO webhooks_teams (setor, url) VALUES ('Geral / Logística', '') ON CONFLICT (setor) DO NOTHING"))
        for sup in set(SUPERVISORES_MAP.values()):
            s.execute(text("INSERT INTO webhooks_teams (setor, url) VALUES (:sup, '') ON CONFLICT (setor) DO NOTHING"), {"sup": sup})

        locais_existentes = {
            row[0]: row[1]
            for row in s.execute(text("SELECT apelido, endereco FROM locais")).fetchall()
        }
        locais_removidos = {
            row[0]
            for row in s.execute(text("SELECT apelido FROM locais_removidos")).fetchall()
        }

        for apelido, end in ENDERECOS_PADRAO:
            if apelido in locais_existentes:
                if locais_existentes[apelido] != end:
                    s.execute(text("UPDATE locais SET endereco = :end, lat = NULL, lon = NULL WHERE apelido = :apelido"), {"end": end, "apelido": apelido})
            elif apelido not in locais_removidos:
                s.execute(text("INSERT INTO locais (apelido, endereco) VALUES (:apelido, :end)"), {"apelido": apelido, "end": end})

        # Fornecedor é diferente de unidade própria: se a equipe já cadastrou uma
        # filial/endereço específico, preservamos o banco. O fallback só preenche
        # um fornecedor ainda inexistente.
        for apelido, end in ENDERECOS_FORNECEDORES_FALLBACK:
            if apelido not in locais_existentes and apelido not in locais_removidos:
                s.execute(text("INSERT INTO locais (apelido, endereco) VALUES (:apelido, :end) ON CONFLICT (apelido) DO NOTHING"), {"apelido": apelido, "end": end})

        s.execute(text("DELETE FROM locais WHERE UPPER(TRIM(apelido)) = 'DESCONHECIDO'"))
        for alias in ALIASES_LOCAL_BASE:
            s.execute(text("INSERT INTO locais (apelido, endereco, lat, lon) VALUES (:alias, :end, :lat, :lon) ON CONFLICT (apelido) DO UPDATE SET endereco=EXCLUDED.endereco, lat=EXCLUDED.lat, lon=EXCLUDED.lon"), {"alias": alias, "end": LOCAL_BASE_ENDERECO, "lat": LOCAL_BASE_COORDS[0], "lon": LOCAL_BASE_COORDS[1]})

        s.execute(
            text("INSERT INTO app_meta (chave, valor) VALUES ('schema_version', :versao) "
                 "ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor"),
            {"versao": SCHEMA_APP_VERSION},
        )
        s.commit()

    return True

# A interface não deve esperar DDL/migrações nem a leitura da rota antes de aparecer.
# O banco já é persistente no Supabase; a rota salva é carregada somente quando o
# módulo Roteiro realmente é aberto. Isso deixa a navegação disponível imediatamente.
def carregar_rota_salva_para_sessao(data_rota):
    if st.session_state.get("rota_gerada") and st.session_state.get("data_rota") == data_rota:
        return True
    try:
        res_rota = fetch_one(
            "SELECT json_route, json_locais, json_geometria, json_enderecos, total_km, fonte_matriz, horario_matriz "
            "FROM rota_ativa WHERE id = 1 AND data_rota = :data",
            {"data": data_rota},
        )
        if not res_rota:
            return False
        st.session_state['route_steps'] = json.loads(res_rota[0] or '[]')
        st.session_state['locais_dict'] = json.loads(res_rota[1] or '{}')
        st.session_state['geometria_rota'] = json.loads(res_rota[2] or '[]')
        st.session_state['enderecos_dict'] = json.loads(res_rota[3] or '{}')
        st.session_state['total_km'] = float(res_rota[4] or 0)
        st.session_state['fonte_matriz_rota'] = res_rota[5] or "OSRM — rota viária"
        st.session_state['horario_matriz_rota'] = res_rota[6] or ""
        if st.session_state['route_steps']:
            st.session_state['p_saida'] = st.session_state['route_steps'][0].get('destino', 'ESCRITÓRIO')
            # Uma rota persistida por um motor antigo pode conter o mesmo endereço
            # em várias paradas (por exemplo, BARRA -> FIEC -> BARRA). Não a tratamos
            # como planejamento definitivo: o módulo Roteiro usará este sinal para
            # carregar as demandas e reconstruí-la com uma visita por local.
            st.session_state['_rota_locais_repetidos_carregada'] = detectar_locais_repetidos_rota(
                st.session_state['route_steps'], st.session_state['p_saida']
            )
            _ultima_saida = str(st.session_state['route_steps'][-1].get('saida', '') or '')
            try:
                h, m = map(int, _ultima_saida.split(':')[:2])
                st.session_state['horario_conclusao_min'] = h * 60 + m
            except Exception:
                pass
        _geom_carregada = st.session_state.get('geometria_rota') or []
        _passos_carregados = st.session_state.get('route_steps') or []
        st.session_state['geometria_viaria'] = len(_geom_carregada) > max(6, len(_passos_carregados) + 3)
        st.session_state['rota_gerada'] = True
        st.session_state['data_rota'] = data_rota
        return True
    except Exception as erro:
        st.session_state['_erro_carregar_rota'] = str(erro)
        return False

# =====================================================================
# LÓGICA DE EXTRAÇÃO E AUTOMAÇÃO DO TRELLO
# =====================================================================
INTERVALO_TRELLO_SEGUNDOS = 2 * 60

@st.cache_data(ttl=20, show_spinner=False)
def ler_cache_trello_supabase():
    """Lê a cópia mantida pelo Cron sem consultar o banco a cada rerun do Streamlit."""
    try:
        registro = fetch_one("SELECT dados FROM trello_cache WHERE id = 1")
        if not registro or registro[0] is None:
            return None
        dados = registro[0]
        if isinstance(dados, str):
            dados = json.loads(dados)
        if isinstance(dados, dict) and isinstance(dados.get("cards"), list) and isinstance(dados.get("lists"), list):
            return dados
    except Exception:
        pass
    return None

def salvar_cache_trello_supabase(dados):
    """Mantém o botão manual compatível com a mesma fonte usada pelo Cron."""
    execute_db(
        """
        INSERT INTO trello_cache (id, dados, atualizado_em)
        VALUES (1, CAST(:dados AS JSONB), NOW())
        ON CONFLICT (id)
        DO UPDATE SET dados=EXCLUDED.dados, atualizado_em=EXCLUDED.atualizado_em
        """,
        {"dados": json.dumps(dados, ensure_ascii=False)},
    )
    try:
        ler_cache_trello_supabase.clear()
    except Exception:
        pass

def obter_dados_trello(forcar=False, somente_cache=False):
    if not forcar:
        dados_cache = ler_cache_trello_supabase()
        if dados_cache is not None:
            return dados_cache
        if somente_cache:
            return None

    # Reserva para a primeira instalação e para o botão manual.
    try:
        resposta = requests.get(TRELLO_JSON_URL, timeout=20)
        resposta.raise_for_status()
        dados = resposta.json()
        if not isinstance(dados, dict) or not isinstance(dados.get("cards"), list) or not isinstance(dados.get("lists"), list):
            return None
        try:
            salvar_cache_trello_supabase(dados)
        except Exception:
            pass
        return dados
    except Exception:
        return None

def identificar_grupo_teams(destino, obra=""):
    texto = normalizar_local(f"{obra} {destino}")
    regras = [(("GERAL / LOGÍSTICA",), "geral_logistica"),(("CASA DA INDÚSTRIA", "FIEC"), "casa_industria"),(("MARACANAÚ",), "maracanau"),(("HORIZONTE",), "horizonte"),(("SEBRAE",), "sebrae"),(("MUSEU",), "museu"),(("BARRA",), "barra"),(("CENTRO", "NR SAÚDE"), "centro"),(("UNIFOR",), "unifor"),(("PARANGABA", "ESCRITÓRIO"), "sede_parangaba")]
    for termos, chave in regras:
        if any(termo in texto for termo in termos): return chave
    return ""

def obter_webhook_teams(setor, supervisor=None, obra=""):
    """Resolve o canal da unidade e nunca perde a baixa por falta de webhook específico.

    Ordem: grupo da unidade -> cadastro do supervisor -> banco -> grupo geral.
    O grupo geral é apenas contingência; não duplica mensagens quando o canal correto existe.
    """
    chave_unidade = identificar_grupo_teams(setor, obra)
    if chave_unidade:
        try:
            url_secret = str(st.secrets["teams_unidades"].get(chave_unidade, "")).strip()
            if url_secret: return url_secret, "Secrets — grupo da unidade"
        except Exception:
            pass

    chave_supervisor = TEAMS_SECRET_KEYS.get(supervisor or setor)
    if chave_supervisor:
        try:
            url_secret = str(st.secrets["teams"].get(chave_supervisor, "")).strip()
            if url_secret: return url_secret, "Secrets — cadastro anterior"
        except Exception:
            pass

    try:
        registro = fetch_one("SELECT url FROM webhooks_teams WHERE setor = :setor", {"setor": supervisor or setor})
        if registro and registro[0]: return registro[0].strip(), "Banco local"
    except Exception:
        pass

    # Se uma unidade ainda não tiver webhook próprio, a baixa vai para o grupo geral
    # em vez de desaparecer silenciosamente.
    try:
        url_geral = str(st.secrets["teams_unidades"].get("geral_logistica", "")).strip()
        if url_geral: return url_geral, "Secrets — grupo geral (contingência)"
    except Exception:
        pass
    try:
        url_geral = str(st.secrets["teams"].get("geral_logistica", "")).strip()
        if url_geral: return url_geral, "Secrets — grupo geral (contingência)"
    except Exception:
        pass
    try:
        registro = fetch_one("SELECT url FROM webhooks_teams WHERE setor = 'Geral / Logística'")
        if registro and registro[0]: return registro[0].strip(), "Banco local — grupo geral (contingência)"
    except Exception:
        pass
    return "", "Não configurado"

def disparar_teams(webhook_url, titulo, mensagem):
    if not webhook_url or not webhook_url.lower().startswith("https://"): return False, "O link precisa ser um webhook HTTPS do Teams Workflows."
    payload = {"type": "message", "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive", "contentUrl": None, "content": {"$schema": "http://adaptivecards.io/schemas/adaptive-card.json", "type": "AdaptiveCard", "version": "1.2", "body": [{"type": "TextBlock", "text": titulo, "size": "Medium", "weight": "Bolder", "wrap": True}, {"type": "TextBlock", "text": mensagem, "wrap": True, "spacing": "Medium"}],},}],}
    ultimo_erro = ""
    for tentativa in range(3):
        try:
            resposta = requests.post(webhook_url, json=payload, timeout=15)
            if 200 <= resposta.status_code < 300: return True, "Mensagem aceita pelo Teams."
            ultimo_erro = f"Teams respondeu com o código {resposta.status_code}."
            if resposta.status_code != 429 and resposta.status_code < 500: break
        except requests.RequestException: ultimo_erro = "Não foi possível alcançar o Teams."
        if tentativa < 2: time.sleep(1 + tentativa)
    return False, ultimo_erro or "Falha desconhecida ao enviar a mensagem."

def _limpar_texto_extra_trello(texto):
    """Remove atividade padrão do Trello e preserva somente conteúdo útil da demanda."""
    if not texto:
        return ""
    linhas = []
    for linha in str(texto).replace("\r", "").split("\n"):
        limpa = re.sub(r"\s+", " ", linha).strip()
        if not limpa:
            continue
        norm = remover_acentos(re.sub(r"[*_`]", "", limpa)).upper()
        # Ex.: **Fulano** adicionou este cartão a EM ROTA [20 de ago...](link)
        if "ADICIONOU ESTE CARTAO A" in norm and "EM ROTA" in norm:
            continue
        linhas.append(limpa)
    return "\n".join(linhas).strip()

def extrair_observacoes_trello(card, acoes=None, limite=1200):
    """Retorna SOMENTE comentários humanos do cartão.

    A descrição do cartão não entra aqui porque ela já é a fonte usada para extrair
    origem, destino e materiais. Colocá-la novamente no alerta do Teams repetia toda
    a lista de materiais. Atividades automáticas (ex.: mover o cartão para EM ROTA,
    CONCLUÍDAS etc.) também não entram: na API do Trello elas não são `commentCard`.
    """
    textos = []
    card_id = str((card or {}).get("id", ""))

    for acao in acoes or []:
        if str(acao.get("type", "")) != "commentCard":
            continue
        dados = acao.get("data", {}) or {}
        card_acao = dados.get("card") or {}
        if card_id and str(card_acao.get("id", "")) != card_id:
            continue

        comentario = _limpar_texto_extra_trello(dados.get("text", ""))
        if comentario:
            textos.append(comentario)

    # Evita repetir o mesmo comentário caso o JSON do Trello traga ações duplicadas.
    unicos, vistos = [], set()
    for texto in textos:
        chave = re.sub(r"\s+", " ", remover_acentos(texto).upper()).strip()
        if chave and chave not in vistos:
            vistos.add(chave)
            unicos.append(texto)

    resultado = "\n\n".join(unicos).strip()
    if len(resultado) > limite:
        resultado = resultado[:limite - 3].rstrip() + "..."
    return resultado


def formatar_materiais_teams(materiais):
    """Formata os materiais em lista vertical para o alerta do Teams."""
    texto = str(materiais or "").replace("\r", "").strip()
    if not texto:
        return "- Ver Trello"

    partes = re.split(r"\s*\|\s*|\n+", texto)
    itens, vistos = [], set()
    for parte in partes:
        item = re.sub(r"^[\-•▪◦]+\s*", "", str(parte)).strip()
        item = re.sub(r"\s+", " ", item)
        if not item:
            continue
        chave = remover_acentos(item).upper()
        if chave in vistos:
            continue
        vistos.add(chave)
        itens.append(item)

    if not itens:
        return "- Ver Trello"
    return "\n".join(f"- {item}" for item in itens)


def informar_entrega_manual_teams(card_id, tarefa):
    """Fallback manual: registra a entrega e dispara o mesmo alerta do Teams.

    A automação do Trello continua sendo o caminho principal. Este método só deve
    ser usado quando a baixa automática não sinalizar. Não move o cartão no Trello;
    apenas registra a entrega no histórico interno e informa o Teams imediatamente.
    """
    card_id = str(card_id or "").strip()
    tarefa = tarefa or {}
    if not card_id:
        return False, "Demanda sem identificador do Trello."

    agora = datetime.now(FUSO_LOCAL)
    data_str = agora.strftime("%d/%m/%Y")
    hora_str = agora.strftime("%H:%M")

    # Usa o cartão real para preservar nome, descrição e comentários sempre que possível.
    card = None
    acoes = []
    try:
        dados_trello = obter_dados_trello(forcar=True) or {}
        card = next((c for c in dados_trello.get("cards", []) if str(c.get("id", "")) == card_id), None)
        acoes = dados_trello.get("actions", []) or []
    except Exception:
        card = None
        acoes = []

    if card:
        short_name, origem, destino, materiais = extrair_dados_completos(
            card.get("desc", ""), card.get("name", "")
        )
        observacao_trello = extrair_observacoes_trello(card, acoes)
    else:
        short_name = str(tarefa.get("Obra", "") or "")
        origem = str(tarefa.get("Origem", "") or "")
        destino = str(tarefa.get("Destino", "") or "")
        materiais = str(tarefa.get("Materiais", "") or "")
        observacao_trello = ""

    # Se o parser do Trello vier incompleto, reaproveita os dados da rota.
    short_name = short_name or str(tarefa.get("Obra", "") or "")
    origem = origem or str(tarefa.get("Origem", "") or "")
    destino = destino or str(tarefa.get("Destino", "") or "")
    materiais = materiais or str(tarefa.get("Materiais", "") or "")
    supervisor = str(tarefa.get("Supervisor", "") or SUPERVISORES_MAP.get(destino, "Sede / Logística"))

    # Registra a baixa interna antes do envio. Assim o roteiro/app também reconhece
    # a entrega informada manualmente, mesmo se o webhook estiver temporariamente fora.
    execute_db(
        "INSERT INTO historico_concluidos "
        "(id, obra, origem, destino, materiais, data_conclusao, hora_conclusao, teams_notificado, teams_tentativas, teams_ultimo_erro) "
        "VALUES (:id, :obra, :origem, :destino, :mat, :data, :hora, FALSE, 0, NULL) "
        "ON CONFLICT (id) DO UPDATE SET obra=EXCLUDED.obra, origem=EXCLUDED.origem, destino=EXCLUDED.destino, "
        "materiais=EXCLUDED.materiais, data_conclusao=EXCLUDED.data_conclusao, hora_conclusao=EXCLUDED.hora_conclusao",
        {
            "id": card_id,
            "obra": short_name,
            "origem": origem,
            "destino": destino,
            "mat": materiais,
            "data": data_str,
            "hora": hora_str,
        },
    )

    url_webhook, fonte_webhook = obter_webhook_teams(destino, supervisor=supervisor, obra=short_name)
    mensagem = (
        "✅ **Os materiais foram informados como entregues pela Torre de Controle.**\n\n"
        f"**Obra:** {short_name}\n\n"
        f"**Local:** {destino}\n\n"
        f"**Materiais:**\n{formatar_materiais_teams(materiais)}\n\n"
        f"**Data e Hora:** {agora.strftime('%d/%m/%Y às %H:%M')}\n\n"
        "**Origem do aviso:** Informado manualmente na aba Demandas Ativas."
    )
    if observacao_trello:
        mensagem += f"\n\n**Comentários do Trello:**\n{observacao_trello}"

    if not url_webhook:
        enviado, detalhe = False, f"Webhook do Teams não configurado ({fonte_webhook})."
    else:
        enviado, detalhe = disparar_teams(url_webhook, f"✅ Entrega concluída — {destino}", mensagem)

    if enviado:
        execute_db(
            "UPDATE historico_concluidos SET teams_notificado=TRUE, "
            "teams_tentativas=COALESCE(teams_tentativas,0)+1, teams_ultimo_erro=NULL WHERE id=:id",
            {"id": card_id},
        )
        return True, "Entrega registrada e alerta enviado ao Teams."

    execute_db(
        "UPDATE historico_concluidos SET teams_notificado=FALSE, "
        "teams_tentativas=COALESCE(teams_tentativas,0)+1, teams_ultimo_erro=:erro WHERE id=:id",
        {"id": card_id, "erro": str(detalhe)[:500]},
    )
    return False, f"Entrega registrada, mas o Teams não confirmou o alerta: {detalhe}"

def is_in_ceara(lat, lon): return -7.5 <= lat <= -2.5 and -42.0 <= lon <= -37.0

@st.cache_data(ttl=24 * 60 * 60, show_spinner=False)
def buscar_coordenadas(endereco):
    if not endereco: return None, None
    endereco_limpo = endereco.strip()
    match_coords = re.search(r'^(-?\d+\.\d+)[\s,;]+(-?\d+\.\d+)$', endereco_limpo)
    if match_coords:
        lat, lon = float(match_coords.group(1)), float(match_coords.group(2))
        if is_in_ceara(lat, lon): return lat, lon
    try:
        url_arcgis = "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates?f=json&singleLine=" + urllib.parse.quote(endereco_limpo + ", Ceará, Brasil") + "&maxLocations=1"
        req = urllib.request.Request(url_arcgis, headers={'User-Agent': 'AproarLogisticsWeb/1.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read())
            if data.get('candidates'):
                lat, lon = float(data['candidates'][0]['location']['y']), float(data['candidates'][0]['location']['x'])
                if is_in_ceara(lat, lon): return lat, lon
    except: pass
    return None, None

def normalizar_local(nome):
    if not nome: return ""
    n = nome.upper().strip()
    if "MARACANAU" in n: n = n.replace("MARACANAU", "MARACANAÚ")
    if "ESCRITORIO" in n: n = n.replace("ESCRITORIO", "ESCRITÓRIO")
    if "DEPÓSITO" in n: n = n.replace("DEPÓSITO", "DEPOSITO")
    if "ESPACO" in n: n = n.replace("ESPACO", "ESPAÇO")
    return n

def canonicalizar_ponto_rota(nome):
    texto = normalizar_local(str(nome or ""))
    texto = re.sub(r"[\\*_`]+", "", texto).strip(" :-\t\r\n")
    texto = re.sub(r'^(?:O|A|OS|AS)\s+', '', texto)

    # Primeiro tenta reconhecer uma unidade própria. Assim, por exemplo,
    # "SENAI BARRA", "SESI BARRA DO CEARÁ" e "BARRA" viram o mesmo ponto.
    unidade_identificada = identificar_unidade_empresa(texto)
    if unidade_identificada:
        return unidade_identificada

    # Corrige artefatos de frases corridas do Trello. Exemplo real:
    # "COLETAR NO ESCRITÓRIO E ENTREGAR O CENTRO" não pode virar o local
    # "ESCRITÓRIO E ENTREGAR O CENTRO". A partir do segundo verbo logístico,
    # ficamos somente com o primeiro ponto físico.
    texto = re.split(
        r'(?i)\s+(?:E\s+)?(?:ENTREGAR|ENTREGA|LEVAR|DEVOLVER|ENCAMINHAR|TRANSPORTAR|DEIXAR)\b',
        texto,
        maxsplit=1,
    )[0].strip(" :-\t\r\n")

    unidade_identificada = identificar_unidade_empresa(texto)
    if unidade_identificada:
        return unidade_identificada
    
    texto_limpo = remover_acentos(texto)
    for sin, oficial in DICIONARIO_SINONIMOS.items():
        if texto_limpo == remover_acentos(sin):
            texto = oficial
            
    if texto in ALIASES_LOCAL_BASE: return "ESCRITÓRIO"
    return texto


def obter_endereco_fornecedor_fallback(nome):
    """Obtém endereço conhecido sem substituir cadastros manuais do banco."""
    chave = canonicalizar_ponto_rota(nome)
    for apelido, endereco in ENDERECOS_FORNECEDORES_FALLBACK:
        if canonicalizar_ponto_rota(apelido) == chave:
            return endereco
    return ""


def detectar_locais_repetidos_rota(route_steps, ponto_saida=""):
    """Lista paradas físicas repetidas, ignorando preparação e retorno à base."""
    vistos = set()
    repetidos = []
    for step in route_steps or []:
        if not isinstance(step, dict) or step.get("type") != "stop":
            continue
        local = canonicalizar_ponto_rota(step.get("destino", ""))
        if not local or local in {"DESCONHECIDO", "NAN", "NONE"}:
            continue
        # Uma eventual parada intermediária na própria base também é erro lógico,
        # salvo a preparação/retorno, que usam outros tipos de etapa.
        if local in vistos and local not in repetidos:
            repetidos.append(local)
        vistos.add(local)
    return repetidos


# =====================================================================
# AJUSTES MANUAIS DA ROTA — ARRASTAR DEMANDAS ENTRE PARADAS
# =====================================================================
@st.cache_data(ttl=20, show_spinner=False)
def carregar_ajustes_manuais_rota(data_rota):
    """Lê os ajustes feitos pelo usuário sem alterar o Trello."""
    try:
        row = fetch_one(
            "SELECT json_ajustes_manuais FROM rota_ativa WHERE id = 1 AND data_rota = :data",
            {"data": data_rota},
        )
        if not row or not row[0]:
            return {}
        dados = json.loads(row[0]) if isinstance(row[0], str) else row[0]
        return dados if isinstance(dados, dict) else {}
    except Exception:
        return {}


def salvar_ajustes_manuais_rota(data_rota, ajustes):
    ajustes = ajustes if isinstance(ajustes, dict) else {}
    execute_db(
        "UPDATE rota_ativa SET json_ajustes_manuais = :ajustes WHERE id = 1 AND data_rota = :data",
        {"ajustes": json.dumps(ajustes, ensure_ascii=False), "data": data_rota},
    )
    try:
        carregar_ajustes_manuais_rota.clear()
    except Exception:
        pass


def limpar_ajustes_manuais_rota(data_rota):
    salvar_ajustes_manuais_rota(data_rota, {})


def _chave_acao_rota(tarefa, acao):
    return f"{str((tarefa or {}).get('id', '') or '')}|{str(acao or '').upper()}"


def aplicar_ajustes_manuais_demandas(df, ajustes, ponto_saida):
    """Aplica endereços e prioridades operacionais antes de montar a rota."""
    if df is None or df.empty or not isinstance(ajustes, dict):
        return df
    mapa = ajustes.get("acoes", {}) or {}
    prioridades_operacionais = {
        str(demanda_id or "").strip()
        for demanda_id in (ajustes.get("prioridades_operacionais", []) or [])
        if str(demanda_id or "").strip()
    }
    if not mapa and not prioridades_operacionais:
        return df
    resultado = df.copy()
    for idx, linha in resultado.iterrows():
        demanda_id = str(linha.get("id", "") or "")
        if not demanda_id:
            continue
        ajuste_coleta = mapa.get(f"{demanda_id}|COLETAR", {}) or {}
        ajuste_entrega = mapa.get(f"{demanda_id}|ENTREGAR", {}) or {}
        alvo_coleta = canonicalizar_ponto_rota(ajuste_coleta.get("destino", ""))
        alvo_entrega = canonicalizar_ponto_rota(ajuste_entrega.get("destino", ""))
        if alvo_coleta:
            resultado.at[idx, "Origem"] = alvo_coleta
        if alvo_entrega:
            resultado.at[idx, "Destino"] = alvo_entrega
        if demanda_id in prioridades_operacionais:
            # A solicitação do engenheiro desempata a seleção do que cabe no dia,
            # mas nunca autoriza ultrapassar o encerramento real das 17h.
            try:
                peso_atual = float(linha.get("Peso", 1) or 1)
            except (TypeError, ValueError):
                peso_atual = 1.0
            resultado.at[idx, "Peso"] = max(peso_atual, 9.0)
            urgencia_atual = str(linha.get("Urgência", "") or "").strip()
            if "PRIORIDADE OPERACIONAL" not in urgencia_atual.upper():
                resultado.at[idx, "Urgência"] = (
                    f"PRIORIDADE OPERACIONAL • {urgencia_atual}"
                    if urgencia_atual else "PRIORIDADE OPERACIONAL"
                )
    return resultado


def _pontuacao_rotulo_obra(rotulo):
    """Prioriza o rótulo que contém o identificador real da obra.

    Ex.: ``086 - UNIFOR`` deve sempre prevalecer sobre ``UNIFOR``. O número
    permanece como texto para conservar zeros à esquerda.
    """
    texto = str(rotulo or "").strip()
    if not texto:
        return -1
    tem_codigo = bool(re.search(r'(?i)(?:^|\s)(?:APR[A-Z0-9._-]*\d[A-Z0-9._-]*|\d+(?:\.\d+)?)(?:\s|$|\s*-)', texto))
    tem_unidade = bool(identificar_unidade_empresa(texto, permitir_contexto=True))
    return (1000 if tem_codigo else 0) + (100 if tem_unidade else 0) + min(len(texto), 80)


def construir_mapa_rotulos_obras_trello(dados_trello=None):
    """Lê o título de TODOS os cartões do quadro, inclusive os concluídos.

    A lista de demandas ativas não contém cartões já movidos para CONCLUÍDAS.
    Por isso o roteiro não pode depender apenas de ``st.session_state.demandas``
    para recuperar o número da obra. O ID do cartão é a chave estável entre o
    Trello, a rota salva e o histórico.
    """
    dados = dados_trello if isinstance(dados_trello, dict) else (obter_dados_trello() or {})
    mapa = {}
    for card in dados.get("cards", []) or []:
        # Não descartamos cartões arquivados/fechados aqui. Se o JSON do quadro
        # ainda os trouxer, o título original continua sendo a melhor fonte para
        # recuperar o número da obra de uma rota/histórico antigo.
        card_id = str(card.get("id", "") or "").strip()
        if not card_id:
            continue
        try:
            rotulo, _origem, _destino, _materiais = extrair_dados_completos(
                card.get("desc", ""), card.get("name", "")
            )
        except Exception:
            rotulo = ""
        rotulo = str(rotulo or "").strip()
        if rotulo:
            anterior = mapa.get(card_id, "")
            if _pontuacao_rotulo_obra(rotulo) >= _pontuacao_rotulo_obra(anterior):
                mapa[card_id] = rotulo
    return mapa


def atualizar_rotulos_obras_route_steps(route_steps, df_demandas=None, mapa_trello=None):
    """Atualiza o rótulo resumido da obra em rotas novas e antigas.

    Usa duas fontes pelo ID do cartão:
    1. demandas ativas da sincronização atual;
    2. títulos brutos do Trello, incluindo cartões em CONCLUÍDAS.

    A fonte mais informativa prevalece. Assim uma tarefa antiga salva como
    ``UNIFOR`` volta a aparecer como ``086 - UNIFOR`` mesmo depois da baixa.
    """
    if not route_steps:
        return route_steps

    mapa = {}
    if df_demandas is not None and not df_demandas.empty and "id" in df_demandas.columns:
        for _, linha in df_demandas.iterrows():
            tarefa_id = str(linha.get("id", "") or "").strip()
            rotulo = str(linha.get("Obra", "") or "").strip()
            if tarefa_id and rotulo:
                mapa[tarefa_id] = rotulo

    for tarefa_id, rotulo in (mapa_trello or {}).items():
        tarefa_id = str(tarefa_id or "").strip()
        rotulo = str(rotulo or "").strip()
        if not tarefa_id or not rotulo:
            continue
        if _pontuacao_rotulo_obra(rotulo) >= _pontuacao_rotulo_obra(mapa.get(tarefa_id, "")):
            mapa[tarefa_id] = rotulo

    if not mapa:
        return route_steps

    resultado = []
    for step in route_steps:
        novo_step = dict(step)
        novas_acoes = []
        for acao, tarefa in step.get("actions", []) or []:
            nova_tarefa = dict(tarefa)
            tarefa_id = str(nova_tarefa.get("id", "") or "").strip()
            rotulo_antigo = str(nova_tarefa.get("Obra", "") or "").strip()

            # Rotas novas carregam o título original do Trello dentro da própria
            # tarefa. Assim o número da obra não se perde depois que o cartão sai
            # das demandas ativas, muda de lista ou é arquivado.
            rotulo_embutido = ""
            titulo_embutido = str(nova_tarefa.get("_Titulo_Trello", "") or "").strip()
            if titulo_embutido:
                try:
                    rotulo_embutido, _o, _d, _m = extrair_dados_completos("", titulo_embutido)
                except Exception:
                    rotulo_embutido = ""

            rotulo_novo = mapa.get(tarefa_id, "")
            candidatos_rotulo = [rotulo_antigo, rotulo_embutido, rotulo_novo]
            melhor_rotulo = max(candidatos_rotulo, key=_pontuacao_rotulo_obra)
            if melhor_rotulo and _pontuacao_rotulo_obra(melhor_rotulo) >= _pontuacao_rotulo_obra(rotulo_antigo):
                nova_tarefa["Obra"] = melhor_rotulo
            novas_acoes.append((acao, nova_tarefa))
        if "actions" in step:
            novo_step["actions"] = novas_acoes
        resultado.append(novo_step)
    return resultado


def aplicar_ordem_manual_route_steps(route_steps, ajustes):
    """Reaplica a ordem dos cartões dentro de cada parada após recalcular."""
    if not route_steps or not isinstance(ajustes, dict):
        return route_steps
    ordem_por_local = ajustes.get("ordem_por_local", {}) or {}
    for step in route_steps:
        if step.get("type") != "stop":
            continue
        local = canonicalizar_ponto_rota(step.get("destino", ""))
        ordem = list(ordem_por_local.get(local, []) or [])
        if not ordem:
            continue
        pos = {chave: i for i, chave in enumerate(ordem)}
        acoes = list(step.get("actions", []) or [])
        acoes.sort(key=lambda item: pos.get(_chave_acao_rota(item[1], item[0]), 10000))
        step["actions"] = acoes
    return route_steps



def consolidar_coletas_base_na_preparacao(route_steps, ajustes, ponto_saida):
    """Move toda COLETA destinada à base para a etapa PREPARAÇÃO.

    A função atua também sobre etapas antigas/concluídas já salvas no Supabase.
    Assim, quando uma coleta foi corrigida manualmente para ESCRITÓRIO, ela não
    reaparece como "PARADA: ESCRITÓRIO": fica dentro da preparação 07:30–08:00.

    A ENTREGA da mesma demanda continua no destino real (ex.: UNIFOR).
    """
    if not route_steps:
        return route_steps

    base = canonicalizar_ponto_rota(ponto_saida)
    if not base:
        return route_steps

    mapa_ajustes = (ajustes or {}).get("acoes", {}) or {}
    passos = []
    for step in route_steps:
        copia = dict(step)
        if step.get("type") == "stop":
            copia["actions"] = list(step.get("actions", []) or [])
        passos.append(copia)

    def _eh_prep(step, indice):
        if step.get("type") != "stop":
            return False
        local = canonicalizar_ponto_rota(step.get("destino", ""))
        if local != base:
            return False
        fonte = remover_acentos(str(step.get("tempo_local_fonte", "") or "")).lower()
        try:
            dist = float(step.get("dist", 0) or 0)
        except Exception:
            dist = 999.0
        try:
            viagem = float(step.get("travel_mins", 0) or 0)
        except Exception:
            viagem = 999.0
        return ("preparacao" in fonte) or (indice == 0 and dist <= 0.10 and viagem <= 0.5)

    idx_prep = next((i for i, s in enumerate(passos) if _eh_prep(s, i)), None)
    prep = passos[idx_prep] if idx_prep is not None else None

    # Tudo que já existe na preparação permanece, sem duplicar.
    acoes_prep = []
    chaves_prep = set()
    if prep is not None:
        for acao, tarefa in prep.get("actions", []) or []:
            chave = _chave_acao_rota(tarefa, acao)
            if chave and chave not in chaves_prep:
                acoes_prep.append((acao, tarefa))
                chaves_prep.add(chave)

    novos_passos = []
    houve_movimento = False

    for indice, step in enumerate(passos):
        if indice == idx_prep:
            continue
        if step.get("type") != "stop":
            novos_passos.append(step)
            continue

        local_atual = canonicalizar_ponto_rota(step.get("destino", ""))
        restantes = []
        for acao, tarefa in step.get("actions", []) or []:
            acao_txt = str(acao or "").upper()
            chave = _chave_acao_rota(tarefa, acao_txt)
            alvo_manual = canonicalizar_ponto_rota((mapa_ajustes.get(chave, {}) or {}).get("destino", ""))
            alvo_efetivo = alvo_manual or local_atual

            # Regra central: COLETA na base nunca é uma parada operacional.
            # Ela pertence à preparação, inclusive quando já foi concluída.
            if acao_txt == "COLETAR" and alvo_efetivo == base:
                if chave and chave not in chaves_prep:
                    acoes_prep.append((acao_txt, tarefa))
                    chaves_prep.add(chave)
                houve_movimento = True
                continue

            # Se a mesma coleta já está na preparação, elimina a cópia duplicada
            # que possa ter sobrado em uma rota antiga.
            if acao_txt == "COLETAR" and chave in chaves_prep:
                houve_movimento = True
                continue

            restantes.append((acao, tarefa))

        if restantes:
            step["actions"] = restantes
            novos_passos.append(step)
        elif step.get("actions"):
            # Parada ficou vazia depois de absorver a coleta na preparação.
            houve_movimento = True
        else:
            novos_passos.append(step)

    if acoes_prep:
        if prep is None:
            prep = {
                "type": "stop",
                "destino": base,
                "dist": 0.0,
                "travel_mins": 0.0,
                "travel_mins_api": 0.0,
                "tempo_local": 30,
                "tempo_local_fonte": "preparação fixa da base",
                "chegada": HORA_PREPARACAO_INICIO,
                "saida": HORA_PREPARACAO_FIM,
                "actions": [],
            }
            houve_movimento = True
        else:
            prep = dict(prep)

        prep.update({
            "type": "stop",
            "destino": base,
            "dist": 0.0,
            "travel_mins": 0.0,
            "travel_mins_api": 0.0,
            "tempo_local": 30,
            "tempo_local_fonte": "preparação fixa da base",
            "chegada": HORA_PREPARACAO_INICIO,
            "saida": HORA_PREPARACAO_FIM,
            "actions": acoes_prep,
        })
        novos_passos.insert(0, prep)

    # Só retorna lista nova; o chamador decide se precisa persistir/recalcular.
    return novos_passos


def registrar_movimento_manual_rota(data_rota, route_steps, demanda_id, acao, destino_alvo, indice_alvo, ponto_saida=""):
    """Persiste um drag-and-drop e devolve True quando o movimento é válido."""
    demanda_id = str(demanda_id or "").strip()
    acao = str(acao or "").upper().strip()
    destino_alvo = canonicalizar_ponto_rota(destino_alvo)
    if not demanda_id or acao not in {"COLETAR", "ENTREGAR"} or not destino_alvo:
        return False

    chave = f"{demanda_id}|{acao}"
    encontrado = False
    locais_validos = set()
    listas_atuais = {}
    for step in route_steps or []:
        if step.get("type") != "stop":
            continue
        local = canonicalizar_ponto_rota(step.get("destino", ""))
        if not local:
            continue
        locais_validos.add(local)
        lista = []
        for acao_step, tarefa_step in step.get("actions", []) or []:
            chave_step = _chave_acao_rota(tarefa_step, acao_step)
            if chave_step:
                lista.append(chave_step)
            if chave_step == chave:
                encontrado = True
        listas_atuais[local] = lista

    base_manual = canonicalizar_ponto_rota(ponto_saida)
    if base_manual:
        locais_validos.add(base_manual)
    if not encontrado or destino_alvo not in locais_validos:
        return False

    ajustes = dict(carregar_ajustes_manuais_rota(data_rota) or {})
    mapa = dict(ajustes.get("acoes", {}) or {})
    ordem_por_local = {k: list(v or []) for k, v in (ajustes.get("ordem_por_local", {}) or {}).items()}
    mapa[chave] = {"destino": destino_alvo}

    for local, lista in listas_atuais.items():
        ordem_por_local[local] = [k for k in lista if k != chave]

    alvo_lista = [k for k in listas_atuais.get(destino_alvo, []) if k != chave]
    try:
        pos = max(0, min(int(indice_alvo), len(alvo_lista)))
    except Exception:
        pos = len(alvo_lista)
    alvo_lista.insert(pos, chave)
    ordem_por_local[destino_alvo] = alvo_lista

    ajustes["acoes"] = mapa
    ajustes["ordem_por_local"] = ordem_por_local
    salvar_ajustes_manuais_rota(data_rota, ajustes)
    return True


def aplicar_movimento_manual_route_steps_imediato(route_steps, demanda_id, acao, destino_alvo, indice_alvo, ponto_saida=""):
    """Reflete o drag imediatamente na rota exibida.

    O ajuste definitivo continua sendo aplicado ao DataFrame e ao otimizador no
    recálculo seguinte. Esta função evita que a tela permaneça mostrando a rota
    antiga entre o drop e esse recálculo: remove a ação da parada anterior e a
    insere na parada escolhida, preservando as demais etapas.
    """
    demanda_id = str(demanda_id or "").strip()
    acao = str(acao or "").upper().strip()
    destino_alvo = canonicalizar_ponto_rota(destino_alvo)
    if not route_steps or not demanda_id or acao not in {"COLETAR", "ENTREGAR"} or not destino_alvo:
        return route_steps

    chave = f"{demanda_id}|{acao}"
    passos = []
    acao_movida = None
    indice_origem = None

    for indice, step in enumerate(route_steps or []):
        copia = dict(step)
        if step.get("type") == "stop":
            novas_acoes = []
            for acao_step, tarefa_step in step.get("actions", []) or []:
                if _chave_acao_rota(tarefa_step, acao_step) == chave and acao_movida is None:
                    acao_movida = (acao_step, tarefa_step)
                    indice_origem = indice
                    continue
                novas_acoes.append((acao_step, tarefa_step))
            copia["actions"] = novas_acoes
        passos.append(copia)

    if acao_movida is None:
        return route_steps

    # Localiza a parada-alvo existente no próprio editor. A preparação é uma
    # parada normal de destino igual à base e também pode receber COLETAS.
    indice_destino = next(
        (
            i for i, step in enumerate(passos)
            if step.get("type") == "stop"
            and canonicalizar_ponto_rota(step.get("destino", "")) == destino_alvo
        ),
        None,
    )
    if indice_destino is None:
        return route_steps

    lista_alvo = list(passos[indice_destino].get("actions", []) or [])
    try:
        pos = max(0, min(int(indice_alvo), len(lista_alvo)))
    except Exception:
        pos = len(lista_alvo)
    lista_alvo.insert(pos, acao_movida)
    passos[indice_destino]["actions"] = lista_alvo

    # Uma parada operacional que ficou sem ações após o movimento não deve
    # continuar aparecendo no roteiro. Nunca remove preparação, almoço ou retorno.
    base = canonicalizar_ponto_rota(ponto_saida)
    resultado = []
    for i, step in enumerate(passos):
        if step.get("type") != "stop":
            resultado.append(step)
            continue
        local = canonicalizar_ponto_rota(step.get("destino", ""))
        eh_preparacao = (
            local == base
            and (
                i == 0
                or "preparacao" in remover_acentos(str(step.get("tempo_local_fonte", "") or "")).lower()
            )
        )
        if not step.get("actions") and not eh_preparacao:
            continue
        resultado.append(step)

    return resultado


def construir_editor_arrastavel_rota(route_steps, ponto_saida, ajustes):
    """Prepara os dados do editor arrastável e envia o drop diretamente ao Python."""
    ponto_saida = canonicalizar_ponto_rota(ponto_saida)
    ajustes_acoes = (ajustes or {}).get("acoes", {}) or {}
    secoes = []
    vistos = set()

    prep_step = next(
        (s for i, s in enumerate(route_steps or [])
         if s.get("type") == "stop" and i == 0
         and canonicalizar_ponto_rota(s.get("destino", "")) == ponto_saida),
        None,
    )

    def preparar_cards(acoes):
        cards = []
        for acao, tarefa in acoes or []:
            did = str(tarefa.get("id", "") or "").strip()
            acao_txt = str(acao or "").upper().strip()
            if not did or acao_txt not in {"COLETAR", "ENTREGAR"}:
                continue
            chave = f"{did}|{acao_txt}"
            cards.append({
                "id": did,
                "acao": acao_txt,
                "obra": str(tarefa.get("Obra", "Demanda") or "Demanda"),
                "materiais": str(tarefa.get("Materiais", "") or ""),
                "manual": chave in ajustes_acoes,
            })
        return cards

    secoes.append({
        "rotulo": "PREPARAÇÃO",
        "local": ponto_saida,
        "cards": preparar_cards(list((prep_step or {}).get("actions", []) or [])),
    })
    vistos.add(ponto_saida)

    numero = 1
    for step in route_steps or []:
        if step.get("type") != "stop":
            continue
        local = canonicalizar_ponto_rota(step.get("destino", ""))
        if not local or local in vistos:
            continue
        secoes.append({
            "rotulo": f"PARADA {numero}",
            "local": local,
            "cards": preparar_cards(list(step.get("actions", []) or [])),
        })
        vistos.add(local)
        numero += 1

    quantidade_cards = sum(len(secao["cards"]) for secao in secoes)
    altura = min(760, max(360, 105 + len(secoes) * 118 + quantidade_cards * 54))
    return {"secoes": secoes}, altura


_COMPONENTE_DRAG_ROTA = None


def _obter_componente_drag_rota():
    """Usa Custom Components para o iframe devolver o drop ao Streamlit."""
    global _COMPONENTE_DRAG_ROTA
    if _COMPONENTE_DRAG_ROTA is not None:
        return _COMPONENTE_DRAG_ROTA

    pasta = os.path.join(tempfile.gettempdir(), "aproar_dragdrop_rota_component")
    os.makedirs(pasta, exist_ok=True)
    index_path = os.path.join(pasta, "index.html")
    frontend = '<!doctype html>\n<html>\n<head>\n<meta charset="utf-8">\n<style>\n*{box-sizing:border-box}\nhtml,body{margin:0;padding:0;height:100vh;overflow:hidden;background:#070913;color:#e5e7eb;font-family:Inter,Arial,sans-serif}\n#app{height:100vh;overflow-y:auto;overflow-x:hidden;padding:0 6px 0 0;scrollbar-gutter:stable}\n#app::-webkit-scrollbar{width:9px}\n#app::-webkit-scrollbar-track{background:#0b1020;border-radius:8px}\n#app::-webkit-scrollbar-thumb{background:#475569;border-radius:8px;border:2px solid #0b1020}\n#app::-webkit-scrollbar-thumb:hover{background:#64748b}\n.aviso{font-size:12px;color:#94a3b8;margin:0 0 10px;line-height:1.45}\n.secao{border:1px solid #263452;background:#0b1020;border-radius:12px;margin:0 0 10px;overflow:hidden}\n.secao-head{display:flex;justify-content:space-between;gap:10px;padding:9px 12px;background:#11182d;border-bottom:1px solid #263452;font-size:13px}\n.secao-head span{color:#9fb1ca;font-weight:700}\n.zona{min-height:62px;padding:8px;transition:.12s ease}\n.zona.over{outline:2px dashed #60a5fa;outline-offset:-4px;background:#0f1c37}\n.demanda{padding:9px 10px;margin:5px 0;border-radius:9px;border:1px solid #334155;background:#10182b;cursor:grab;box-shadow:0 2px 7px rgba(0,0,0,.18);user-select:none}\n.demanda:active{cursor:grabbing}\n.demanda.dragging{opacity:.28}\n.demanda.coleta{border-left:5px solid #f59e0b}\n.demanda.entrega{border-left:5px solid #22c55e}\n.top{display:flex;align-items:center;gap:7px;font-size:12px}\n.handle{font-size:18px;color:#93c5fd;line-height:1}\n.manual{margin-left:auto;background:#1d4ed8;color:#dbeafe;padding:2px 6px;border-radius:999px;font-size:10px}\n.obra{font-size:12.5px;font-weight:800;margin-top:3px;color:#f1f5f9}\n.mat{font-size:11px;color:#94a3b8;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}\n.vazio{font-size:11px;color:#64748b;text-align:center;padding:12px;border:1px dashed #334155;border-radius:8px}\n</style>\n</head>\n<body><div id="app"></div>\n<script>\nconst app=document.getElementById(\'app\');\nlet arrastado=null;\nlet argsAtuais={};\nfunction post(type, extra={}){window.parent.postMessage(Object.assign({isStreamlitMessage:true,type:type},extra),\'*\');}\nfunction ready(){post(\'streamlit:componentReady\',{apiVersion:1});}\nfunction setValue(value){post(\'streamlit:setComponentValue\',{value:value,dataType:\'json\'});}\nfunction setHeight(){\n  const desejada=Number(argsAtuais.requested_height||0);\n  const altura=Math.max(360,desejada>0?desejada:720);\n  post(\'streamlit:setFrameHeight\',{height:altura});\n}\nfunction el(tag, cls, texto){const n=document.createElement(tag);if(cls)n.className=cls;if(texto!==undefined&&texto!==null)n.textContent=String(texto);return n;}\nfunction habilitarCard(card){\n  card.addEventListener(\'dragstart\',e=>{arrastado=card;card.classList.add(\'dragging\');e.dataTransfer.effectAllowed=\'move\';try{e.dataTransfer.setData(\'text/plain\',card.dataset.id||\'demanda\');}catch(_){}});\n  card.addEventListener(\'dragend\',()=>{card.classList.remove(\'dragging\');document.querySelectorAll(\'.zona\').forEach(z=>z.classList.remove(\'over\'));arrastado=null;});\n}\nfunction habilitarZona(zona){\n  zona.addEventListener(\'dragenter\',e=>{e.preventDefault();zona.classList.add(\'over\');});\n  zona.addEventListener(\'dragover\',e=>{e.preventDefault();zona.classList.add(\'over\');e.dataTransfer.dropEffect=\'move\';});\n  zona.addEventListener(\'dragleave\',e=>{if(!zona.contains(e.relatedTarget))zona.classList.remove(\'over\');});\n  zona.addEventListener(\'drop\',e=>{\n    e.preventDefault();e.stopPropagation();zona.classList.remove(\'over\');if(!arrastado)return;\n    const outros=[...zona.querySelectorAll(\'.demanda\')].filter(x=>x!==arrastado);let indice=outros.length;\n    for(let i=0;i<outros.length;i++){const r=outros[i].getBoundingClientRect();if(e.clientY<r.top+r.height/2){indice=i;break;}}\n    zona.querySelectorAll(\'.vazio\').forEach(v=>v.remove());\n    if(indice<outros.length)zona.insertBefore(arrastado,outros[indice]);else zona.appendChild(arrastado);\n    setValue({nonce:String(Date.now())+\'-\'+Math.random().toString(36).slice(2),demanda_id:arrastado.dataset.id||\'\',acao:arrastado.dataset.acao||\'\',destino:zona.dataset.destino||\'\',ordem:indice});\n    setTimeout(setHeight,20);\n  });\n}\nfunction render(args){\n  argsAtuais=args||{};app.replaceChildren();\n  app.appendChild(el(\'div\',\'aviso\',\'Arraste pelo ⠿. Solte dentro de outra parada para mudar o local da ação; solte acima ou abaixo para mudar a ordem. Role dentro deste painel para ver todas as paradas. O Trello não é alterado.\'));\n  const secoes=((argsAtuais.payload||{}).secoes)||[];\n  secoes.forEach(sec=>{\n    const section=el(\'section\',\'secao\');const head=el(\'div\',\'secao-head\');head.appendChild(el(\'b\',\'\',sec.rotulo||\'\'));head.appendChild(el(\'span\',\'\',sec.local||\'\'));section.appendChild(head);\n    const zona=el(\'div\',\'zona\');zona.dataset.destino=sec.local||\'\';const cards=sec.cards||[];\n    if(!cards.length)zona.appendChild(el(\'div\',\'vazio\',\'Solte uma demanda aqui\'));\n    cards.forEach(c=>{const card=el(\'div\',\'demanda \'+(c.acao===\'COLETAR\'?\'coleta\':\'entrega\'));card.draggable=true;card.dataset.id=c.id||\'\';card.dataset.acao=c.acao||\'\';const top=el(\'div\',\'top\');top.appendChild(el(\'span\',\'handle\',\'⠿\'));top.appendChild(el(\'b\',\'\',c.acao===\'COLETAR\'?\'📦 COLETA\':\'📬 ENTREGA\'));if(c.manual)top.appendChild(el(\'span\',\'manual\',\'manual\'));card.appendChild(top);card.appendChild(el(\'div\',\'obra\',c.obra||\'Demanda\'));card.appendChild(el(\'div\',\'mat\',c.materiais||\'\'));habilitarCard(card);zona.appendChild(card);});\n    habilitarZona(zona);section.appendChild(zona);app.appendChild(section);\n  });\n  setTimeout(setHeight,0);\n}\nwindow.addEventListener(\'message\',e=>{const d=e.data||{};if(d.type===\'streamlit:render\')render(d.args||{});});\nready();\n</script></body></html>\n'
    frontend = frontend.replace(
        "<style>\n",
        "<style>\n@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Sora:wght@600;700;800&display=swap');\n",
        1,
    ).replace(
        "font-family:Inter,Arial,sans-serif",
        "font-family:Manrope,Arial,sans-serif",
        1,
    )
    for cor_legenda_editor in ("#94a3b8", "#64748b", "#9fb1ca"):
        frontend = frontend.replace(cor_legenda_editor, "#ffffff")
    with open(index_path, "w", encoding="utf-8") as arquivo:
        arquivo.write(frontend)

    _COMPONENTE_DRAG_ROTA = st.components.v1.declare_component(
        "aproar_dragdrop_rota", path=pasta
    )
    return _COMPONENTE_DRAG_ROTA


def renderizar_editor_arrastavel_rota(payload, altura, key):
    componente = _obter_componente_drag_rota()
    return componente(payload=payload, requested_height=int(altura), key=key, default=None)


@st.cache_data(ttl=600, show_spinner=False)
def carregar_medias_historicas_paradas():
    """Permanência típica por local usando o rastreador com estatística robusta.

    Visitas muito curtas tendem a ser passagem/ruído de GPS. Visitas muito longas
    podem incluir almoço, conversa, café ou falha de leitura. Para a previsão usamos
    mediana + média aparada das visitas recentes, limitando o quanto um dia atípico
    consegue puxar a estimativa.
    """
    try:
        df_hist = get_df(
            "SELECT local, hora_chegada, hora_saida FROM rastreio_paradas "
            "WHERE hora_saida IS NOT NULL ORDER BY id DESC LIMIT 1000"
        )
    except Exception:
        return {}

    if df_hist is None or df_hist.empty:
        return {}

    por_local = {}
    for _, linha in df_hist.iterrows():
        local = canonicalizar_ponto_rota(linha.get('local', ''))
        chegada = str(linha.get('hora_chegada', '') or '')
        saida = str(linha.get('hora_saida', '') or '')
        if not local or not chegada or not saida:
            continue
        try:
            ini = parse_time_to_mins(chegada)
            fim = parse_time_to_mins(saida)
            if fim < ini:
                fim += 24 * 60
            dur = float(fim - ini)
        except Exception:
            continue

        # Menos de 5 min normalmente é passagem. Acima de 90 min é tratado como
        # ocorrência excepcional e não entra no aprendizado operacional.
        if 5 <= dur <= 90:
            por_local.setdefault(local, []).append(dur)

    resultado = {}
    for local, valores in por_local.items():
        valores = valores[:24]  # prioriza as visitas mais recentes
        if not valores:
            continue

        ordenados = sorted(float(v) for v in valores)
        n = len(ordenados)
        meio = n // 2
        mediana = ordenados[meio] if n % 2 else (ordenados[meio - 1] + ordenados[meio]) / 2.0

        # Com histórico suficiente remove aproximadamente 15% em cada ponta.
        corte = max(1, int(n * 0.15)) if n >= 7 else 0
        aparados = ordenados[corte:n-corte] if corte and (n - 2 * corte) >= 3 else ordenados
        media_aparada = sum(aparados) / len(aparados)

        # Mediana recebe maior peso porque resiste melhor a dias com café/conversa.
        tipica = mediana * 0.65 + media_aparada * 0.35
        tipica = min(max(tipica, 10.0), 45.0)
        resultado[local] = {
            'media': tipica,
            'mediana': mediana,
            'amostras': n,
        }
    return resultado

# Tempos típicos informados pela operação. São a base de permanência de uma
# visita normal; volume, peso, coleta e histórico real ajustam esse valor.
TEMPOS_BASE_UNIDADES_DAVI = {
    "FIEC": 20,
    "MARACANAÚ": 25,
    "UNIFOR": 25,
    "MUSEU": 15,
    "CENTRO": 20,
}


def _grupo_tempo_base_local(ponto):
    """Resolve os nomes usados na rota para as unidades de referência do Davi."""
    local = remover_acentos(canonicalizar_ponto_rota(ponto)).upper()
    if "MARACANAU" in local or local in {"SESI ALBANO FRANCO", "SESI CLUBE DA PARCERIA", "SENAI ISTEMM", "SENAI CETAFR"}:
        return "MARACANAÚ"
    if "UNIFOR" in local:
        return "UNIFOR"
    if "MUSEU" in local:
        return "MUSEU"
    if local == "FIEC" or local.startswith("FIEC "):
        return "FIEC"
    if local in {"CENTRO", "SENAI CENTRO", "ESCOLA CENTRO", "NR SAUDE"}:
        return "CENTRO"
    return None


def _quantidade_inicial_material(descricao):
    """Lê apenas a quantidade inicial do item para não confundir 2,5mm/20kg com quantidade."""
    texto = str(descricao or "").strip()
    match = re.match(r"^(\d+(?:[\.,]\d+)?)\b", texto)
    if not match:
        return 1.0
    try:
        return max(1.0, float(match.group(1).replace(',', '.')))
    except (TypeError, ValueError):
        return 1.0


def _analisar_carga_parada(entregas, coletas):
    """Estima volume e esforço de manuseio a partir dos materiais das demandas.

    Não tenta adivinhar peso em kg. O objetivo é classificar operacionalmente a carga
    (leve, média ou pesada/volumosa) usando quantidade + palavras do material.
    """
    tarefas = list(entregas or []) + list(coletas or [])
    itens = []
    for tarefa in tarefas:
        materiais = str(tarefa.get('Materiais', '') or '')
        partes = [p.strip() for p in re.split(r'\s*\|\s*', materiais) if p.strip()]
        itens.extend(partes or ([materiais.strip()] if materiais.strip() else []))

    if not itens:
        return {
            'ajuste': 0, 'volume': 'baixo volume', 'peso': 'carga leve',
            'itens': 0, 'qtd_aprox': 0,
        }

    # Materiais que normalmente exigem mais força, espaço ou tempo de carga/descarga.
    palavras_pesadas = (
        'CIMENTO', 'ARGAMASSA', 'REJUNTE', 'MASSA', 'SACO', 'SACOS', 'SACA', 'SACAS',
        'CERAMICA', 'CERÂMICA', 'PORCELANATO', 'PISO ', 'TELHA', 'BLOCO ',
        'CHAPA', 'PORTA', 'VIDRO', 'MADEIRA', 'PERFIL', 'GALAO', 'GALOES',
        'LATA', 'LATAS', 'TINTA', 'ESMALTE', 'SOLVENTE', 'THINNER', 'BOBINA',
        'TUBO', 'ELETRODUTO', 'CABO', 'MOTOR', 'BOMBA', 'COMPRESSOR',
    )
    palavras_volumosas = (
        'CAIXA', 'CAIXAS', 'ROLO', 'ROLOS', 'BOBINA', 'TUBO', 'ELETRODUTO',
        'PERFIL', 'CHAPA', 'PORTA', 'TELHA', 'CERAMICA', 'CERÂMICA',
        'PORCELANATO', 'PAINEL ', 'LUMINARIA', 'LUMINÁRIA',
    )
    palavras_leves = (
        'LUVA', 'OCULOS', 'PROTETOR AURICULAR', 'LAPIS',
        'PARAFUSO', 'ABRACADEIRA', 'FITA', 'CONECTOR',
        'DISCO DE CORTE', 'TOMADA', 'INTERRUPTOR', 'PLUG', 'BUCHA',
    )

    pontos_volume = 0.0
    pontos_peso = 0.0
    qtd_aprox = 0.0
    pesados = 0
    volumosos = 0
    leves = 0

    for item in itens:
        qtd = _quantidade_inicial_material(item)
        qtd_aprox += qtd
        item_norm = remover_acentos(item).upper()

        # Quantidade aumenta volume de forma saturada: 300 parafusos não equivalem
        # a 300 caixas, mas ainda exigem conferência/manuseio.
        if qtd >= 200:
            pontos_volume += 4.0
        elif qtd >= 100:
            pontos_volume += 3.0
        elif qtd >= 50:
            pontos_volume += 2.2
        elif qtd >= 20:
            pontos_volume += 1.5
        elif qtd >= 10:
            pontos_volume += 1.0
        elif qtd >= 5:
            pontos_volume += 0.5

        eh_pesado = any(remover_acentos(p).upper() in item_norm for p in palavras_pesadas)
        eh_volumoso = any(remover_acentos(p).upper() in item_norm for p in palavras_volumosas)
        eh_leve = any(remover_acentos(p).upper() in item_norm for p in palavras_leves)

        if eh_pesado:
            pesados += 1
            pontos_peso += 2.5
            if qtd >= 20:
                pontos_peso += 2.0
            elif qtd >= 5:
                pontos_peso += 1.0
        elif eh_leve:
            leves += 1
        else:
            # Item sem classificação conhecida conta como carga intermediária.
            pontos_peso += 0.45

        if eh_volumoso:
            volumosos += 1
            pontos_volume += 1.8

    # Muitas linhas diferentes geram conferência, separação e organização, mesmo
    # quando os itens são leves.
    pontos_volume += min(max(len(itens) - 2, 0) * 0.55, 4.0)

    if pontos_volume < 3.0:
        ajuste_volume, rotulo_volume = 0, 'baixo volume'
    elif pontos_volume < 7.0:
        ajuste_volume, rotulo_volume = 2, 'volume moderado'
    elif pontos_volume < 12.0:
        ajuste_volume, rotulo_volume = 4, 'volume alto'
    else:
        ajuste_volume, rotulo_volume = 6, 'volume muito alto'

    if pontos_peso < 2.5 and pesados == 0:
        ajuste_peso, rotulo_peso = 0, 'carga leve'
    elif pontos_peso < 6.0:
        ajuste_peso, rotulo_peso = 2, 'carga de peso médio'
    elif pontos_peso < 11.0:
        ajuste_peso, rotulo_peso = 4, 'carga pesada'
    else:
        ajuste_peso, rotulo_peso = 6, 'carga muito pesada'

    # Se quase tudo for EPI/ferragem pequena, evita penalizar só por uma quantidade
    # numérica grande (ex.: 100 abraçadeiras ou 72 pares de luva).
    if leves >= max(2, len(itens) - 1) and pesados == 0 and volumosos == 0:
        ajuste_volume = min(ajuste_volume, 4)
        rotulo_volume = 'volume moderado' if ajuste_volume else 'baixo volume'
        ajuste_peso = 0
        rotulo_peso = 'carga leve'

    return {
        'ajuste': ajuste_volume + ajuste_peso,
        'volume': rotulo_volume,
        'peso': rotulo_peso,
        'itens': len(itens),
        'qtd_aprox': qtd_aprox,
        'ajuste_volume': ajuste_volume,
        'ajuste_peso': ajuste_peso,
    }


def estimar_tempo_parada(ponto, entregas=None, coletas=None, retornar_fonte=False):
    """Estima UMA permanência por local com base, carga atual e histórico real.

    Os tempos típicos por unidade são a âncora. Demandas no mesmo endereço não
    recebem um tempo cheio individual: são atendidas dentro da mesma permanência.
    A carga atual (volume/peso), a complexidade manual e o fato de haver coleta
    ajustam a estimativa. O histórico do rastreador corrige hábitos reais do local.
    """
    entregas = list(entregas or [])
    coletas = list(coletas or [])

    # Evita contar a mesma demanda duas vezes em casos raros de origem=destino.
    vistos = set()
    entregas_unicas, coletas_unicas = [], []
    for tipo, lista_origem, lista_destino in (('E', entregas, entregas_unicas), ('C', coletas, coletas_unicas)):
        for tarefa in lista_origem:
            chave = (tipo, str(tarefa.get('id', id(tarefa))))
            if chave not in vistos:
                vistos.add(chave)
                lista_destino.append(tarefa)
    entregas, coletas = entregas_unicas, coletas_unicas

    qtd_acoes = len(entregas) + len(coletas)
    if qtd_acoes <= 0:
        return (0, 'sem atendimento') if retornar_fonte else 0

    grupo_local = _grupo_tempo_base_local(ponto)
    if grupo_local:
        tempo_base = float(TEMPOS_BASE_UNIDADES_DAVI[grupo_local])
        fonte_base = f"base {grupo_local.title()} {int(tempo_base)} min"
    elif coletas and not entregas:
        # Fornecedor/loja costuma consumir mais tempo que uma entrega simples,
        # mas a previsão normal deve permanecer na faixa operacional de 15–25 min.
        tempo_base = 22.0
        fonte_base = "base de coleta/fornecedor 22 min"
    elif coletas and entregas:
        tempo_base = 24.0
        fonte_base = "base de coleta + entrega 24 min"
    else:
        tempo_base = 18.0
        fonte_base = "base de entrega 18 min"

    # Histórico corrige a base do local, mas as referências informadas pela operação
    # continuam sendo a principal âncora para FIEC/Maracanaú/Unifor/Museu/Centro.
    historico = carregar_medias_historicas_paradas().get(canonicalizar_ponto_rota(ponto))
    fonte_hist = ""
    if historico:
        media_hist = float(historico.get('media', tempo_base))
        amostras = int(historico.get('amostras', 0))
        if grupo_local:
            # As referências operacionais fornecidas continuam sendo a âncora.
            peso_hist = 0.12 if amostras < 5 else 0.22
        else:
            peso_hist = 0.20 if amostras < 5 else 0.35
        # Histórico real corrige a base, mas não empurra a previsão para fora da
        # faixa operacional normal. Visitas longas continuam registradas como tempo
        # real, porém não viram automaticamente uma previsão longa para o dia seguinte.
        media_hist = min(max(media_hist, 12.0), 30.0)
        tempo_contexto = tempo_base * (1.0 - peso_hist) + media_hist * peso_hist
        fonte_hist = f" + histórico {amostras} visita{'s' if amostras != 1 else ''}"
    else:
        tempo_contexto = tempo_base

    carga = _analisar_carga_parada(entregas, coletas)

    def tempo_num(tarefa, campo, padrao):
        try:
            valor = float(tarefa.get(campo, padrao) or padrao)
            if math.isnan(valor):
                return float(padrao)
            return max(1.0, valor)
        except (TypeError, ValueError):
            return float(padrao)

    # Tempos manuais continuam úteis para uma demanda excepcionalmente complicada,
    # mas não são somados um a um. Só o maior desvio acima do padrão pesa na parada.
    desvios = []
    for tarefa in entregas:
        desvios.append(max(0.0, tempo_num(tarefa, 'Tempo_Entrega', 10) - 10.0))
    for tarefa in coletas:
        desvios.append(max(0.0, tempo_num(tarefa, 'Tempo_Coleta', 20) - 20.0))
    ajuste_manual = min(max(desvios or [0.0]) * 0.25, 5.0)

    # Em unidades conhecidas, uma coleta acrescenta só uma pequena margem. Nos
    # fornecedores a própria base de 25–27 min já inclui espera/separação.
    ajuste_coleta = 0.0
    if grupo_local and coletas:
        ajuste_coleta = 2.0 if not entregas else 3.0

    # Demandas no mesmo endereço compartilham a visita. Só há acréscimo leve de
    # conferência/organização, nunca outro atendimento completo por cartão.
    ajuste_multiplas = min(max(qtd_acoes - 1, 0) * 0.6, 3.0)

    # Volume, peso e complexidade continuam influenciando, mas a estimativa operacional
    # fica deliberadamente entre 15 e 25 minutos. O tempo REAL medido pelo rastreador
    # pode ser maior ou menor; esta faixa é apenas a previsão usada no planejamento.
    ajuste_carga = min(float(carga.get('ajuste', 0) or 0), 6.0)
    estimativa = tempo_contexto + ajuste_carga + ajuste_manual + ajuste_coleta + ajuste_multiplas
    estimativa = int(round(min(max(estimativa, 15.0), 25.0)))

    detalhes = [fonte_base, carga['volume'], carga['peso']]
    if ajuste_coleta > 0:
        detalhes.append('margem de coleta')
    if ajuste_manual >= 2:
        detalhes.append('complexidade manual')
    fonte = " • ".join(detalhes) + fonte_hist

    return (estimativa, fonte) if retornar_fonte else estimativa

def atualizar_tempos_por_parada(route_steps, ponto_saida=''):
    """Recalcula a permanência das paradas já salvas usando o modelo atual."""
    for indice, step in enumerate(route_steps or []):
        if step.get('type') != 'stop':
            continue
        destino = str(step.get('destino', '') or '')
        is_start = indice == 0 and destino == ponto_saida
        if is_start:
            continue
        entregas = [t for acao, t in step.get('actions', []) if acao == 'ENTREGAR']
        coletas = [t for acao, t in step.get('actions', []) if acao == 'COLETAR']
        tempo, fonte = estimar_tempo_parada(destino, entregas, coletas, retornar_fonte=True)
        step['tempo_local'] = tempo
        step['tempo_local_fonte'] = fonte
    return route_steps

def garantir_gps_local_base():
    coordenadas = None
    for alias in ("ESCRITÓRIO", "ALMOXARIFADO"):
        registro = fetch_one("SELECT lat, lon FROM locais WHERE apelido = :alias", {"alias": alias})
        if registro and registro[0] is not None and registro[1] is not None:
            coordenadas = (float(registro[0]), float(registro[1]))
            break
    if coordenadas is None: coordenadas = LOCAL_BASE_COORDS
    if coordenadas is not None:
        for alias in ALIASES_LOCAL_BASE: 
            execute_db("INSERT INTO locais (apelido, endereco, lat, lon) VALUES (:alias, :end, :lat, :lon) ON CONFLICT (apelido) DO UPDATE SET endereco=EXCLUDED.endereco, lat=EXCLUDED.lat, lon=EXCLUDED.lon", {"alias": alias, "end": LOCAL_BASE_ENDERECO, "lat": coordenadas[0], "lon": coordenadas[1]})
    return coordenadas

def carregar_chave_google_routes():
    """Lê a chave sem expô-la no código e aceita nomes usuais nos Secrets."""
    caminhos = [
        ("google_routes", "api_key"),
        ("google_maps", "api_key"),
    ]
    for secao, campo in caminhos:
        try:
            chave = str(st.secrets[secao][campo]).strip()
            if chave:
                return chave
        except Exception:
            pass

    for campo in ("GOOGLE_MAPS_API_KEY", "google_maps_api_key"):
        try:
            chave = str(st.secrets[campo]).strip()
            if chave:
                return chave
        except Exception:
            pass
    return ""

def carregar_chave_tomtom():
    """Lê a chave gratuita da TomTom sem expô-la no código."""
    try:
        chave = str(st.secrets["tomtom"]["api_key"]).strip()
        if chave:
            return chave
    except Exception:
        pass
    for campo in ("TOMTOM_API_KEY", "tomtom_api_key"):
        try:
            chave = str(st.secrets[campo]).strip()
            if chave:
                return chave
        except Exception:
            pass
    return ""

def carregar_token_mapbox():
    """Lê o access token da Mapbox sem expô-lo no código."""
    caminhos = [
        ("mapbox", "access_token"),
        ("mapbox", "api_key"),
        ("mapbox", "token"),
    ]
    for secao, campo in caminhos:
        try:
            token = str(st.secrets[secao][campo]).strip()
            if token:
                return token
        except Exception:
            pass
    for campo in ("MAPBOX_ACCESS_TOKEN", "mapbox_access_token", "MAPBOX_TOKEN", "mapbox_token"):
        try:
            token = str(st.secrets[campo]).strip()
            if token:
                return token
        except Exception:
            pass
    return ""


def _registrar_diagnostico_mapbox(ok, mensagem, detalhes=""):
    """Guarda diagnóstico operacional da Mapbox sem registrar o token."""
    try:
        st.session_state['_mapbox_diag'] = {
            'ok': bool(ok),
            'mensagem': str(mensagem or '').strip(),
            'detalhes': str(detalhes or '').strip()[:700],
            'quando': datetime.now(FUSO_LOCAL).strftime('%H:%M:%S'),
        }
    except Exception:
        pass


def _erro_mapbox_resposta(resposta):
    """Extrai erro útil da Mapbox sem incluir URL nem access token."""
    codigo_http = getattr(resposta, 'status_code', '')
    mensagem = ''
    codigo_api = ''
    try:
        dados = resposta.json()
        if isinstance(dados, dict):
            codigo_api = str(dados.get('code') or '').strip()
            mensagem = str(dados.get('message') or dados.get('error') or '').strip()
    except Exception:
        try:
            mensagem = str(resposta.text or '')[:300].strip()
        except Exception:
            mensagem = ''
    partes = [f"HTTP {codigo_http}" if codigo_http else "Falha HTTP"]
    if codigo_api:
        partes.append(codigo_api)
    if mensagem:
        partes.append(mensagem)
    return " - ".join(partes)


def _partida_mapbox(horario_partida=None):
    """Normaliza a partida para um instante aceito pelas APIs da Mapbox."""
    agora_seguro = datetime.now(FUSO_LOCAL) + timedelta(minutes=1)
    partida = horario_partida if horario_partida else agora_seguro
    try:
        if partida.tzinfo is None:
            partida = partida.replace(tzinfo=FUSO_LOCAL)
    except Exception:
        partida = agora_seguro
    if partida < agora_seguro:
        partida = agora_seguro
    return partida


def _param_depart_at_mapbox(partida):
    return partida.isoformat(timespec="minutes")


def calcular_matriz_mapbox_trafego(coords, horario_partida=None):
    """Matriz Mapbox com trânsito; OSRM entra apenas em células sem rota.

    O perfil driving-traffic aceita até 10 coordenadas por requisição. Para rotas
    maiores, a matriz N x N é montada em blocos 5 x 5 usando sources/destinations.
    Isso preserva nosso próprio otimizador e usa a Mapbox apenas como fonte viária.
    """
    token = carregar_token_mapbox()
    quantidade = len(coords)
    if not token:
        _registrar_diagnostico_mapbox(False, 'Access token da Mapbox não encontrado nos Secrets.')
        return None
    if quantidade < 2:
        return None
    # Evita exceder 30 req/min do driving-traffic em matrizes excepcionalmente grandes.
    if quantidade > 25:
        _registrar_diagnostico_mapbox(False, f'Rota com {quantidade} pontos: matriz de trânsito foi mantida no OSRM para não exceder o limite operacional da Mapbox.')
        return None

    partida = _partida_mapbox(horario_partida)
    depart_at = _param_depart_at_mapbox(partida)
    distancias = [[None for _ in range(quantidade)] for _ in range(quantidade)]
    duracoes = [[None for _ in range(quantidade)] for _ in range(quantidade)]
    chamadas = 0
    usou_depart_at = True

    def requisitar(indices_origem, indices_destino, usar_depart_at=True):
        # Duplicar uma mesma coordenada entre origem/destino é válido e simplifica
        # a indexação dos blocos mantendo o limite máximo de 10 coordenadas.
        coords_req = [coords[i] for i in indices_origem] + [coords[j] for j in indices_destino]
        coords_str = ';'.join(f"{float(lon):.7f},{float(lat):.7f}" for lat, lon in coords_req)
        sources = ';'.join(str(i) for i in range(len(indices_origem)))
        desloc = len(indices_origem)
        destinations = ';'.join(str(desloc + j) for j in range(len(indices_destino)))
        params = {
            'access_token': token,
            'annotations': 'distance,duration',
            'sources': sources,
            'destinations': destinations,
        }
        if usar_depart_at:
            params['depart_at'] = depart_at
        resposta = requests.get(
            f"https://api.mapbox.com/directions-matrix/v1/mapbox/driving-traffic/{coords_str}",
            params=params,
            timeout=25,
        )
        # depart_at da Matrix é beta e pode não estar habilitado na conta.
        if resposta.status_code == 422 and usar_depart_at:
            return requisitar(indices_origem, indices_destino, usar_depart_at=False)
        if not resposta.ok:
            raise RuntimeError(_erro_mapbox_resposta(resposta))
        dados = resposta.json()
        if dados.get('code') != 'Ok':
            raise RuntimeError(f"Mapbox {dados.get('code')}: {dados.get('message', 'matriz sem resposta válida')}")
        return dados, usar_depart_at

    try:
        if quantidade <= 10:
            # Uma única chamada simétrica é mais eficiente e consome N² elementos.
            coords_str = ';'.join(f"{float(lon):.7f},{float(lat):.7f}" for lat, lon in coords)
            params = {'access_token': token, 'annotations': 'distance,duration', 'depart_at': depart_at}
            resposta = requests.get(
                f"https://api.mapbox.com/directions-matrix/v1/mapbox/driving-traffic/{coords_str}",
                params=params,
                timeout=25,
            )
            if resposta.status_code == 422:
                params.pop('depart_at', None)
                usou_depart_at = False
                resposta = requests.get(
                    f"https://api.mapbox.com/directions-matrix/v1/mapbox/driving-traffic/{coords_str}",
                    params=params,
                    timeout=25,
                )
            chamadas = 1
            if not resposta.ok:
                raise RuntimeError(_erro_mapbox_resposta(resposta))
            dados = resposta.json()
            if dados.get('code') != 'Ok':
                raise RuntimeError(f"Mapbox {dados.get('code')}: {dados.get('message', 'matriz sem resposta válida')}")
            ds = dados.get('distances') or []
            ts = dados.get('durations') or []
            for i in range(quantidade):
                for j in range(quantidade):
                    if i < len(ds) and j < len(ds[i]) and ds[i][j] is not None:
                        distancias[i][j] = float(ds[i][j]) / 1000.0
                    if i < len(ts) and j < len(ts[i]) and ts[i][j] is not None:
                        duracoes[i][j] = float(ts[i][j]) / 60.0
        else:
            bloco = 5

            def particionar_indices(total, maximo=5):
                grupos = [list(range(i, min(i + maximo, total))) for i in range(0, total, maximo)]
                # A Matrix não aceita resultado 1x1. Se sobrar apenas um índice no
                # último grupo, traz um índice do grupo anterior (5+1 vira 4+2).
                if len(grupos) > 1 and len(grupos[-1]) == 1:
                    grupos[-1].insert(0, grupos[-2].pop())
                return grupos

            grupos = particionar_indices(quantidade, bloco)
            blocos = [(origens, destinos) for origens in grupos for destinos in grupos]

            # Até 25 pontos => no máximo 25 requisições, dentro do limite nominal
            # de 30 req/min do perfil driving-traffic.
            for origens, destinos in blocos:
                dados, bloco_usou_depart_at = requisitar(origens, destinos, usar_depart_at=usou_depart_at)
                chamadas += 1
                usou_depart_at = usou_depart_at and bloco_usou_depart_at
                ds = dados.get('distances') or []
                ts = dados.get('durations') or []
                for oi_local, oi_global in enumerate(origens):
                    for dj_local, dj_global in enumerate(destinos):
                        if oi_local < len(ds) and dj_local < len(ds[oi_local]) and ds[oi_local][dj_local] is not None:
                            distancias[oi_global][dj_global] = float(ds[oi_local][dj_local]) / 1000.0
                        if oi_local < len(ts) and dj_local < len(ts[oi_local]) and ts[oi_local][dj_local] is not None:
                            duracoes[oi_global][dj_global] = float(ts[oi_local][dj_local]) / 60.0

        for i in range(quantidade):
            distancias[i][i], duracoes[i][i] = 0.0, 0.0

        faltantes = [
            (i, j) for i in range(quantidade) for j in range(quantidade)
            if distancias[i][j] is None or duracoes[i][j] is None
        ]
        if faltantes:
            matriz_osrm = _calcular_matriz_osrm(coords)
            if matriz_osrm:
                dist_osrm, dur_osrm = matriz_osrm
                for i, j in faltantes:
                    if distancias[i][j] is None:
                        distancias[i][j] = dist_osrm[i][j]
                    if duracoes[i][j] is None:
                        duracoes[i][j] = dur_osrm[i][j]

        restantes = sum(
            1 for i in range(quantidade) for j in range(quantidade)
            if distancias[i][j] is None or duracoes[i][j] is None
        )
        if restantes:
            _registrar_diagnostico_mapbox(False, 'Mapbox respondeu, mas a matriz ficou incompleta.', f"{restantes} {plural_pt(restantes, 'trecho sem rota', 'trechos sem rota')}.")
            return None

        try:
            st.session_state['_mapbox_matriz_hibrida'] = len(faltantes)
            st.session_state['_mapbox_matriz_chamadas'] = chamadas
            st.session_state['_mapbox_depart_at'] = bool(usou_depart_at)
        except Exception:
            pass
        detalhe_tempo = 'trânsito no horário planejado' if usou_depart_at else 'trânsito atual (depart_at da Matrix indisponível na conta)'
        if faltantes:
            _registrar_diagnostico_mapbox(True, f"Mapbox ativa • {detalhe_tempo} • {len(faltantes)} {plural_pt(len(faltantes), 'trecho completado', 'trechos completados')} pelo OSRM.")
        else:
            _registrar_diagnostico_mapbox(True, f'Mapbox Matrix ativa • {detalhe_tempo}.')
        return distancias, duracoes
    except Exception as erro:
        _registrar_diagnostico_mapbox(False, 'Mapbox Matrix falhou; OSRM assumiu a rota.', str(erro)[:600])
        return None


def calcular_trecho_mapbox_por_horario(coord_origem, coord_destino, horario_partida_min):
    """Recalcula cada perna pela Mapbox no horário real planejado daquela etapa."""
    token = carregar_token_mapbox()
    if not token or not coord_origem or not coord_destino:
        return None
    try:
        lat1, lon1 = map(float, coord_origem)
        lat2, lon2 = map(float, coord_destino)
        partida = datetime.combine(DATA_REF_ROTA_DATE, datetime.min.time()).replace(tzinfo=FUSO_LOCAL)
        partida += timedelta(minutes=float(horario_partida_min or 0))
        partida = _partida_mapbox(partida)
        coords_str = f"{lon1:.7f},{lat1:.7f};{lon2:.7f},{lat2:.7f}"
        params = {
            'access_token': token,
            'overview': 'false',
            'steps': 'false',
            'depart_at': _param_depart_at_mapbox(partida),
        }
        resposta = requests.get(
            f"https://api.mapbox.com/directions/v5/mapbox/driving-traffic/{coords_str}",
            params=params,
            timeout=20,
        )
        if not resposta.ok:
            return None
        dados = resposta.json()
        rotas = dados.get('routes') or []
        if dados.get('code') != 'Ok' or not rotas:
            return None
        rota = rotas[0]
        dist_m = rota.get('distance')
        dur_s = rota.get('duration')
        if dist_m is None or dur_s is None:
            return None
        return float(dist_m) / 1000.0, float(dur_s) / 60.0
    except Exception:
        return None


def _duracao_google_minutos(valor):
    try:
        return float(str(valor).rstrip("s")) / 60.0
    except (TypeError, ValueError):
        return None

def calcular_matriz_google_trafego(coords, horario_partida):
    """Matriz viária com trânsito ao vivo/preditivo via Google Routes."""
    chave = carregar_chave_google_routes()
    quantidade = len(coords)
    if not chave or quantidade < 2 or quantidade > 100:
        return None

    agora_seguro = datetime.now(FUSO_LOCAL) + timedelta(minutes=1)
    partida = horario_partida if horario_partida and horario_partida > agora_seguro else agora_seguro
    partida_rfc3339 = partida.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")

    destinos = [
        {"waypoint": {"location": {"latLng": {"latitude": float(lat), "longitude": float(lon)}}}}
        for lat, lon in coords
    ]
    distancias = [[None for _ in range(quantidade)] for _ in range(quantidade)]
    duracoes = [[None for _ in range(quantidade)] for _ in range(quantidade)]
    elementos_por_bloco = 100
    origens_por_bloco = max(1, elementos_por_bloco // quantidade)

    url = "https://routes.googleapis.com/distanceMatrix/v2:computeRouteMatrix"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": chave,
        "X-Goog-FieldMask": "originIndex,destinationIndex,duration,staticDuration,distanceMeters,status,condition",
    }

    def consultar_bloco(indices_origem):
        origens = [
            {
                "waypoint": {"location": {"latLng": {"latitude": float(coords[i][0]), "longitude": float(coords[i][1])}}},
                "routeModifiers": {"avoidFerries": True},
            }
            for i in indices_origem
        ]
        payload = {
            "origins": origens,
            "destinations": destinos,
            "travelMode": "DRIVE",
            "routingPreference": "TRAFFIC_AWARE_OPTIMAL",
            "trafficModel": "BEST_GUESS",
            "departureTime": partida_rfc3339,
            "languageCode": "pt-BR",
            "regionCode": "br",
            "units": "METRIC",
        }
        resposta = requests.post(url, headers=headers, json=payload, timeout=20)
        resposta.raise_for_status()
        elementos = resposta.json()
        if not isinstance(elementos, list):
            raise ValueError("Resposta inválida da matriz do Google Routes")
        return indices_origem, elementos

    try:
        blocos = [
            list(range(inicio, min(inicio + origens_por_bloco, quantidade)))
            for inicio in range(0, quantidade, origens_por_bloco)
        ]
        # A API limita a matriz de trânsito ótimo a 100 elementos por chamada.
        # Os lotes independentes rodam juntos para não multiplicar a espera.
        with ThreadPoolExecutor(max_workers=min(4, len(blocos))) as executor:
            futuros = [executor.submit(consultar_bloco, indices) for indices in blocos]
            for futuro in as_completed(futuros):
                indices_origem, elementos = futuro.result()
                for elemento in elementos:
                    if elemento.get("condition") != "ROUTE_EXISTS":
                        continue
                    origem_global = indices_origem[int(elemento.get("originIndex", 0))]
                    destino_global = int(elemento.get("destinationIndex", 0))
                    duracao_min = _duracao_google_minutos(elemento.get("duration"))
                    distancia_m = elemento.get("distanceMeters")
                    if duracao_min is None or distancia_m is None:
                        continue
                    distancias[origem_global][destino_global] = float(distancia_m) / 1000.0
                    duracoes[origem_global][destino_global] = duracao_min

        for i in range(quantidade):
            distancias[i][i], duracoes[i][i] = 0.0, 0.0
        if any(valor is None for linha in distancias for valor in linha):
            return None
        if any(valor is None for linha in duracoes for valor in linha):
            return None
        return distancias, duracoes
    except Exception:
        return None

def _registrar_diagnostico_tomtom(ok, mensagem, detalhes=""):
    """Guarda um diagnóstico curto sem jamais expor a chave da API."""
    try:
        st.session_state['_tomtom_diag'] = {
            'ok': bool(ok),
            'mensagem': str(mensagem or '').strip(),
            'detalhes': str(detalhes or '').strip()[:700],
            'quando': datetime.now(FUSO_LOCAL).strftime('%H:%M:%S'),
        }
    except Exception:
        pass


def _erro_tomtom_resposta(resposta):
    """Extrai uma mensagem útil de erro da TomTom sem incluir URL/chave."""
    codigo = getattr(resposta, 'status_code', '')
    mensagem = ''
    try:
        dados = resposta.json()
        erro = dados.get('detailedError') or dados.get('error') or {}
        if isinstance(erro, dict):
            partes = [erro.get('code'), erro.get('message')]
            mensagem = ' - '.join(str(p) for p in partes if p)
        if not mensagem and isinstance(dados, dict):
            mensagem = str(dados.get('message') or dados.get('errorText') or '')
    except Exception:
        try:
            mensagem = str(resposta.text or '')[:300]
        except Exception:
            mensagem = ''
    base = f"HTTP {codigo}" if codigo else "Falha HTTP"
    return f"{base}: {mensagem}" if mensagem else base


def _calcular_matriz_osrm(coords):
    """Matriz viária principal da rota, usando o servidor público do OSRM."""
    try:
        coords_str = ";".join([f"{lon},{lat}" for lat, lon in coords])
        url = f"https://router.project-osrm.org/table/v1/driving/{coords_str}?annotations=distance,duration"
        req = urllib.request.Request(url, headers={'User-Agent': 'AproarLogisticsWeb/1.0'})
        with urllib.request.urlopen(req, timeout=12) as response:
            res = json.loads(response.read())
        if res.get('code') != 'Ok':
            return None
        distancias = [[(float(v) / 1000.0) if v is not None else None for v in row] for row in res.get('distances', [])]
        duracoes = [[(float(v) / 60.0) if v is not None else None for v in row] for row in res.get('durations', [])]
        if len(distancias) != len(coords) or len(duracoes) != len(coords):
            return None
        return distancias, duracoes
    except Exception:
        return None


def calcular_matriz_tomtom_trafego(coords, horario_partida):
    """Matriz TomTom com trânsito, em blocos de até 100 células.

    A Matrix Routing v2 síncrona aceita no máximo 100 rotas por requisição. A versão
    anterior enviava N x N de uma vez; acima de 10 pontos isso podia fazer a TomTom
    rejeitar a chamada e o app caía silenciosamente para o OSRM. Agora as origens são
    divididas em blocos e falhas isoladas de map matching não descartam a matriz toda.
    """
    chave = carregar_chave_tomtom()
    quantidade = len(coords)
    if not chave:
        _registrar_diagnostico_tomtom(False, 'Chave TomTom não encontrada nos Secrets.')
        return None
    if quantidade < 2:
        return None
    if quantidade > 100:
        _registrar_diagnostico_tomtom(False, f'Matriz com {quantidade} pontos excede o limite síncrono de 100 destinos.')
        return None

    agora_seguro = datetime.now(FUSO_LOCAL) + timedelta(minutes=1)
    partida = horario_partida if horario_partida and horario_partida > agora_seguro else agora_seguro
    partida_rfc3339 = partida.isoformat(timespec='seconds')

    todos_pontos = [
        {'point': {'latitude': float(lat), 'longitude': float(lon)}}
        for lat, lon in coords
    ]
    distancias = [[None for _ in range(quantidade)] for _ in range(quantidade)]
    duracoes = [[None for _ in range(quantidade)] for _ in range(quantidade)]

    # A documentação da Matrix v2 limita a matriz síncrona a 100 células por chamada.
    origens_por_bloco = max(1, 100 // quantidade)
    falhas_api = []

    try:
        for inicio in range(0, quantidade, origens_por_bloco):
            indices_origem = list(range(inicio, min(inicio + origens_por_bloco, quantidade)))
            payload = {
                'origins': [todos_pontos[i] for i in indices_origem],
                'destinations': todos_pontos,
                'options': {
                    'departAt': partida_rfc3339,
                    'routeType': 'fastest',
                    'traffic': 'live',
                    'travelMode': 'car',
                },
            }
            resposta = requests.post(
                'https://api.tomtom.com/routing/matrix/2',
                params={'key': chave},
                headers={'Content-Type': 'application/json', 'Accept-Encoding': 'gzip'},
                json=payload,
                timeout=35,
            )
            if not resposta.ok:
                detalhe = _erro_tomtom_resposta(resposta)
                _registrar_diagnostico_tomtom(False, 'Matrix Routing v2 recusou a requisição.', detalhe)
                return None

            corpo = resposta.json()
            estatisticas = corpo.get('statistics') or {}
            if estatisticas.get('failures'):
                for item in estatisticas.get('failureDetails') or []:
                    falhas_api.append(f"{item.get('code', 'ERRO')} x{item.get('count', 1)}")

            for celula in corpo.get('data', []):
                origem_local = int(celula.get('originIndex', -1))
                destino = int(celula.get('destinationIndex', -1))
                if not (0 <= origem_local < len(indices_origem) and 0 <= destino < quantidade):
                    continue
                origem = indices_origem[origem_local]
                resumo = celula.get('routeSummary') or {}
                distancia_m = resumo.get('lengthInMeters')
                duracao_s = resumo.get('travelTimeInSeconds')
                if distancia_m is not None and duracao_s is not None:
                    distancias[origem][destino] = float(distancia_m) / 1000.0
                    duracoes[origem][destino] = float(duracao_s) / 60.0

        for i in range(quantidade):
            distancias[i][i], duracoes[i][i] = 0.0, 0.0

        faltantes = [
            (i, j) for i in range(quantidade) for j in range(quantidade)
            if distancias[i][j] is None or duracoes[i][j] is None
        ]

        # Uma única célula com MAP_MATCHING_FAILURE não deve jogar fora todo o trânsito
        # da TomTom. Completa somente os trechos problemáticos usando a malha OSRM.
        if faltantes:
            matriz_osrm = _calcular_matriz_osrm(coords)
            if matriz_osrm:
                dist_osrm, dur_osrm = matriz_osrm
                for i, j in faltantes:
                    if distancias[i][j] is None:
                        distancias[i][j] = dist_osrm[i][j]
                    if duracoes[i][j] is None:
                        duracoes[i][j] = dur_osrm[i][j]

        ainda_faltantes = sum(
            1 for i in range(quantidade) for j in range(quantidade)
            if distancias[i][j] is None or duracoes[i][j] is None
        )
        if ainda_faltantes:
            detalhe = ', '.join(sorted(set(falhas_api))) if falhas_api else f"{ainda_faltantes} {plural_pt(ainda_faltantes, 'trecho sem resposta', 'trechos sem resposta')}"
            _registrar_diagnostico_tomtom(False, 'TomTom respondeu, mas a matriz ficou incompleta.', detalhe)
            return None

        qtd_fallback = len(faltantes)
        try:
            st.session_state['_tomtom_matriz_hibrida'] = qtd_fallback
        except Exception:
            pass
        if qtd_fallback:
            detalhe = ', '.join(sorted(set(falhas_api))) if falhas_api else ''
            _registrar_diagnostico_tomtom(True, f"TomTom ativa; {qtd_fallback} {plural_pt(qtd_fallback, 'trecho isolado completado', 'trechos isolados completados')} por contingência do OSRM.", detalhe)
        else:
            _registrar_diagnostico_tomtom(True, 'TomTom Matrix Routing v2 ativa e respondendo normalmente.')
        return distancias, duracoes
    except requests.RequestException as erro:
        _registrar_diagnostico_tomtom(False, 'Falha de conexão com a TomTom.', f'{type(erro).__name__}: {erro}')
        return None
    except Exception as erro:
        _registrar_diagnostico_tomtom(False, 'Erro ao processar a resposta da TomTom.', f'{type(erro).__name__}: {erro}')
        return None

def calcular_matriz_rotas(coords, horario_partida=None):
    """Calcula a matriz viária usando apenas OSRM.

    O OSRM é a fonte principal e gratuita. Os tempos recebidos ainda passam pela
    validação operacional do aplicativo, que evita ETAs urbanos otimistas demais.
    Nenhuma API paga é consultada por esta função.
    """
    matriz_osrm = _calcular_matriz_osrm(coords)
    if matriz_osrm:
        return matriz_osrm[0], matriz_osrm[1], 'OSRM — rota viária'

    # Contingência final caso o servidor público do OSRM esteja indisponível.
    distancias, duracoes = [], []
    for i in range(len(coords)):
        row_d, row_t = [], []
        for j in range(len(coords)):
            dLat = math.radians(coords[j][0] - coords[i][0])
            dLon = math.radians(coords[j][1] - coords[i][1])
            a = math.sin(dLat/2)**2 + math.cos(math.radians(coords[i][0])) * math.cos(math.radians(coords[j][0])) * math.sin(dLon/2)**2
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
            km = (6371 * c) * 1.3
            row_d.append(km)
            row_t.append((km / VELOCIDADE_MEDIA_KMH) * 60)
        distancias.append(row_d)
        duracoes.append(row_t)
    return distancias, duracoes, 'Estimativa geográfica de contingência'


def priorizar_pontos_sem_revisita(candidatos, tarefas_a_coletar):
    """Mantém um destino fechado até todas as coletas que irão para ele estarem no veículo.

    Isso transforma, sempre que a rede origem→destino não possui um ciclo, várias
    idas ao mesmo endereço em uma única visita consolidada. Se existir um ciclo
    real (A entrega em B e B entrega em A), nenhum ponto fica liberado; nesse caso
    devolvemos os candidatos originais para o planejador quebrar o ciclo com a
    menor repetição possível.
    """
    candidatos = set(candidatos or [])
    if len(candidatos) <= 1:
        return candidatos

    consolidados = {
        ponto for ponto in candidatos
        if not any(
            tarefa.get('Destino') == ponto and tarefa.get('Origem') != ponto
            for tarefa in (tarefas_a_coletar or [])
        )
    }
    return consolidados or candidatos



def pontuar_parada_rota(atual, ponto, unpicked, carrying, estrategia, get_dist_dur):
    """Pontua a próxima parada considerando distância, prioridade e retornos evitáveis."""
    distancia, duracao = get_dist_dur(atual, ponto)
    custo_deslocamento = distancia + (duracao * 0.1)

    coletas_aqui = [t for t in unpicked if t['Origem'] == ponto]
    entregas_aqui = [t for t in carrying if t['Destino'] == ponto]
    is_pickup = bool(coletas_aqui)
    is_dropoff = bool(entregas_aqui)
    destinos_no_carro = {t['Destino'] for t in carrying}

    def peso_valido(tarefa):
        try:
            valor = float(tarefa.get('Peso', 1) or 1)
            return 1.0 if math.isnan(valor) else valor
        except (TypeError, ValueError):
            return 1.0

    urgencia_coleta = max([peso_valido(t) for t in coletas_aqui] + [1.0])
    urgencia_entrega = max([peso_valido(t) for t in entregas_aqui] + [1.0])

    # Se ainda há material a coletar para o mesmo destino, entregar agora
    # normalmente cria uma segunda visita e o conhecido efeito de zigue-zague.
    pendentes_mesmo_destino = [
        t for t in unpicked
        if t['Destino'] == ponto and t['Origem'] != ponto
    ]
    penalidade_retorno = 0.0
    if is_dropoff and pendentes_mesmo_destino:
        urgencia_pendente = max([peso_valido(t) for t in pendentes_mesmo_destino] + [1.0])
        # Na rota Equilibrada, uma entrega de hoje/vencida também pode justificar
        # uma visita antes de esperar outra coleta futura para o mesmo destino.
        entrega_urgente_isolada = (
            ("Urgências" in estrategia or "Equilibrada" in estrategia)
            and urgencia_entrega >= 5
            and urgencia_entrega > urgencia_pendente
        )

        if not entrega_urgente_isolada:
            origens_pendentes = {t['Origem'] for t in pendentes_mesmo_destino}
            ciclos_de_retorno = []
            for origem in origens_pendentes:
                ida, _ = get_dist_dur(ponto, origem)
                volta, _ = get_dist_dur(origem, ponto)
                ciclos_de_retorno.append(ida + volta)

            # Repetir uma unidade é uma exceção operacional, não apenas um pequeno
            # custo de distância. A penalidade só atua quando um ciclo impedir a
            # consolidação rígida aplicada pelo planejador.
            penalidade_retorno = max(60.0, min(ciclos_de_retorno or [20.0]) * 3.0)

    coleta_completa_carga = is_pickup and any(
        t['Destino'] in destinos_no_carro for t in coletas_aqui
    )

    if "Menor Distância" in estrategia:
        score = custo_deslocamento
        if coleta_completa_carga:
            score *= 0.85
        score += penalidade_retorno
    else:
        prioridade = 1.0
        if is_dropoff:
            prioridade *= 2.0
        if is_pickup:
            prioridade *= 3.0 if coleta_completa_carga else 1.5
        urgencia_local = max(urgencia_coleta, urgencia_entrega)
        if "Urgências" in estrategia:
            prioridade *= max(urgencia_local, 1.0) ** 2.2
        elif "Equilibrada" in estrategia:
            # Equilibrada = prazo + logística. Até a versão anterior ela quase não
            # considerava urgência nesta decisão local, permitindo que cartões
            # FUTURO passassem na frente de uma demanda HOJE. O multiplicador é
            # forte para prazo crítico, mas ainda divide o custo de deslocamento,
            # preservando a eficiência geográfica entre alternativas equivalentes.
            if urgencia_local >= 9:
                fator_prazo = 24.0
            elif urgencia_local >= 7:
                fator_prazo = 16.0
            elif urgencia_local >= 6:
                fator_prazo = 12.0
            elif urgencia_local >= 5:
                fator_prazo = 8.0
            elif urgencia_local >= 4:
                fator_prazo = 2.4
            elif urgencia_local >= 3:
                fator_prazo = 1.35
            else:
                fator_prazo = 1.0
            prioridade *= fator_prazo
        elif "Descarregar" in estrategia and is_dropoff:
            prioridade *= 5.0

        score = (custo_deslocamento / max(prioridade, 0.001)) + penalidade_retorno

    # Ações no local onde o veículo já está devem ocorrer imediatamente.
    if distancia < 0.1:
        score = -1.0

    return score, distancia, duracao


def identificar_ids_entregues_na_ordem_rota(ordem_pontos, tarefas, tarefas_pre_coletadas=None):
    """Devolve somente demandas cujo ciclo chega à entrega na ordem calculada."""
    tarefas_por_id = {
        str(t.get('id', '') or ''): t
        for t in (tarefas or [])
        if str(t.get('id', '') or '')
    }
    ids_coletados = {
        str(t.get('id', '') or '')
        for t in (tarefas_pre_coletadas or [])
        if str(t.get('id', '') or '')
    }
    ids_entregues = set()

    for ponto in (ordem_pontos or []):
        ponto = canonicalizar_ponto_rota(ponto)

        # Descarrega o que já estava no veículo ao chegar ao ponto.
        for demanda_id in list(ids_coletados - ids_entregues):
            tarefa = tarefas_por_id.get(demanda_id, {})
            if canonicalizar_ponto_rota(tarefa.get('Destino', '')) == ponto:
                ids_entregues.add(demanda_id)

        # Coleta o material do ponto. Origem e destino iguais são resolvidos
        # na mesma visita, de forma idêntica ao motor de otimização.
        for demanda_id, tarefa in tarefas_por_id.items():
            if demanda_id in ids_coletados:
                continue
            if canonicalizar_ponto_rota(tarefa.get('Origem', '')) == ponto:
                ids_coletados.add(demanda_id)
                if canonicalizar_ponto_rota(tarefa.get('Destino', '')) == ponto:
                    ids_entregues.add(demanda_id)

    return ids_entregues


def identificar_coletas_sem_entrega_route_steps(route_steps):
    """Localiza cartões que aparecem como coleta, mas não possuem entrega na rota."""
    ids_coletas = set()
    ids_entregas = set()
    for step in (route_steps or []):
        for acao, tarefa in (step.get('actions', []) or []):
            demanda_id = str(tarefa.get('id', '') or '').strip()
            if not demanda_id:
                continue
            if acao == 'COLETAR':
                ids_coletas.add(demanda_id)
            elif acao == 'ENTREGAR':
                ids_entregas.add(demanda_id)
    return ids_coletas - ids_entregas


def otimizar_sequencia_rota(tarefas, ponto_inicial, estrategia, get_dist_dur, horario_inicio, retornar_base=False, ponto_base=None, tarefas_pre_coletadas=None, locais_bloqueados=None):
    """Busca em feixe para o problema de coleta e entrega com precedência.

    Avalia sequências completas, agrupa ações no mesmo endereço e pondera
    trânsito/tempo, distância, urgência, carga no veículo, almoço e hora extra.
    """
    if not tarefas:
        return []

    def numero_seguro(valor, padrao):
        try:
            numero = float(valor)
            return padrao if math.isnan(numero) else numero
        except (TypeError, ValueError):
            return padrao

    total_tarefas = len(tarefas)
    ids_pre_coletadas = {
        str(t.get('id', '') or '') for t in (tarefas_pre_coletadas or [])
        if str(t.get('id', '') or '')
    }
    locais_bloqueados = {
        canonicalizar_ponto_rota(local)
        for local in (locais_bloqueados or [])
        if canonicalizar_ponto_rota(local)
    }

    def tarefa_ja_coletada(tarefa):
        tarefa_id = str(tarefa.get('id', '') or '')
        return bool(tarefa_id and tarefa_id in ids_pre_coletadas)

    if total_tarefas > 24:
        # Contingência para dias excepcionalmente grandes: mantém as mesmas
        # regras logísticas sem deixar o aplicativo preso em busca combinatória.
        pendentes = [t for t in tarefas if not tarefa_ja_coletada(t)]
        no_carro = [t for t in tarefas if tarefa_ja_coletada(t)]
        atual = ponto_inicial
        ordem = []
        visitados = set(locais_bloqueados)
        for _ in range(total_tarefas * 2 + 5):
            candidatos = ({t['Origem'] for t in pendentes} | {t['Destino'] for t in no_carro}) - visitados
            if not candidatos:
                break
            candidatos = priorizar_pontos_sem_revisita(candidatos, pendentes)
            proximo = min(candidatos, key=lambda p: pontuar_parada_rota(atual, p, pendentes, no_carro, estrategia, get_dist_dur)[0])
            ordem.append(proximo)
            no_carro = [t for t in no_carro if t['Destino'] != proximo]
            coletadas = [t for t in pendentes if t['Origem'] == proximo]
            pendentes = [t for t in pendentes if t['Origem'] != proximo]
            no_carro.extend(coletadas)
            atual = proximo
            visitados.add(proximo)
        return ordem

    origens = [t['Origem'] for t in tarefas]
    destinos = [t['Destino'] for t in tarefas]
    pesos = [numero_seguro(t.get('Peso', 1), 1.0) for t in tarefas]
    tempos_coleta = [numero_seguro(t.get('Tempo_Coleta', 10), 10.0) for t in tarefas]
    tempos_entrega = [numero_seguro(t.get('Tempo_Entrega', 10), 10.0) for t in tarefas]
    mascara_total = (1 << total_tarefas) - 1

    tarefas_por_origem = {}
    tarefas_por_destino = {}
    for indice, (origem, destino) in enumerate(zip(origens, destinos)):
        tarefas_por_origem.setdefault(origem, []).append(indice)
        tarefas_por_destino.setdefault(destino, []).append(indice)

    if total_tarefas <= 8:
        largura_feixe = 700
    elif total_tarefas <= 14:
        largura_feixe = 320
    elif total_tarefas <= 19:
        largura_feixe = 180
    else:
        largura_feixe = 100

    def avancar_relogio(inicio, viagem, servico):
        partida = inicio
        if 12 * 60 <= partida < 13 * 60:
            partida = 13 * 60
        chegada = partida + viagem
        if partida < 12 * 60 < chegada:
            chegada += 60
        if 12 * 60 <= chegada < 13 * 60:
            chegada = 13 * 60
        saida = chegada + servico
        if chegada < 12 * 60 < saida:
            saida += 60
        return chegada, saida

    def pontos_disponiveis(coletadas_mask, entregues_mask):
        pontos = set()
        for i in range(total_tarefas):
            bit = 1 << i
            if not (coletadas_mask & bit):
                pontos.add(origens[i])
            elif not (entregues_mask & bit):
                pontos.add(destinos[i])
        return pontos

    def heuristica_restante(atual, coletadas_mask, entregues_mask):
        pontos = pontos_disponiveis(coletadas_mask, entregues_mask)
        if not pontos:
            return 0.0
        menor_tempo = min(get_dist_dur(atual, p)[1] for p in pontos)
        base = menor_tempo * 0.35 + len(pontos) * 0.8

        # Evita que o beam search descarte cedo justamente os caminhos que já
        # atenderam as demandas mais urgentes. Estados com HOJE/VENCIDA ainda não
        # entregues carregam uma "dívida de prazo" adicional.
        if "Equilibrada" in estrategia or "Urgências" in estrategia:
            divida_prazo = 0.0
            for i, peso in enumerate(pesos):
                bit = 1 << i
                if entregues_mask & bit:
                    continue
                if peso >= 9:
                    divida_prazo += 260.0
                elif peso >= 7:
                    divida_prazo += 150.0
                elif peso >= 6:
                    divida_prazo += 110.0
                elif peso >= 5:
                    divida_prazo += 75.0
                elif peso >= 3:
                    divida_prazo += 8.0
            base += divida_prazo
        return base

    mascara_pre_coletadas = 0
    for indice, tarefa in enumerate(tarefas):
        if tarefa_ja_coletada(tarefa):
            mascara_pre_coletadas |= 1 << indice

    estado_inicial = {
        "atual": ponto_inicial,
        "hora": float(horario_inicio),
        "coletadas": mascara_pre_coletadas,
        "entregues": 0,
        "ordem": tuple(),
        "custo": 0.0,
        "distancia": 0.0,
        "viagem": 0.0,
    }
    estados = [estado_inicial]
    concluidos = []
    inicio_busca = time.perf_counter()
    max_passos = total_tarefas * 2 + 3

    for _ in range(max_passos):
        proximos_por_estado = {}
        for estado in estados:
            if estado["entregues"] == mascara_total:
                concluidos.append(estado)
                continue

            pontos_estado = pontos_disponiveis(estado["coletadas"], estado["entregues"])
            # Uma unidade só pode aparecer uma vez na rota operacional. Se a rede
            # de coleta/entrega criar um ciclo, o beam mantém a melhor parte viável
            # e deixa o restante para outro planejamento, em vez de retornar.
            pontos_estado = {
                ponto for ponto in pontos_estado
                if ponto not in locais_bloqueados and ponto not in estado["ordem"]
            }
            tarefas_a_coletar_estado = [
                tarefas[i] for i in range(total_tarefas)
                if not (estado["coletadas"] & (1 << i))
            ]
            pontos_estado = priorizar_pontos_sem_revisita(
                pontos_estado, tarefas_a_coletar_estado
            )
            for ponto in pontos_estado:
                distancia, duracao = get_dist_dur(estado["atual"], ponto)

                ids_entrega = [
                    i for i in tarefas_por_destino.get(ponto, [])
                    if (estado["coletadas"] & (1 << i)) and not (estado["entregues"] & (1 << i))
                ]
                ids_coleta = [
                    i for i in tarefas_por_origem.get(ponto, [])
                    if not (estado["coletadas"] & (1 << i))
                ]
                if not ids_entrega and not ids_coleta:
                    continue

                novas_coletadas = estado["coletadas"]
                novas_entregues = estado["entregues"]
                for i in ids_coleta:
                    novas_coletadas |= 1 << i
                for i in ids_entrega:
                    novas_entregues |= 1 << i
                # Origem e destino idênticos são resolvidos na mesma parada.
                for i in ids_coleta:
                    if destinos[i] == ponto:
                        novas_entregues |= 1 << i
                        ids_entrega.append(i)

                tarefas_entrega_parada = [tarefas[i] for i in ids_entrega]
                tarefas_coleta_parada = [tarefas[i] for i in ids_coleta]
                tempo_servico = estimar_tempo_parada(ponto, tarefas_entrega_parada, tarefas_coleta_parada)
                _, nova_hora = avancar_relogio(estado["hora"], duracao, tempo_servico)

                # 17h é limite rígido de planejamento. Uma nova parada só entra
                # na busca se o atendimento couber no expediente e, quando houver
                # retorno à base, ainda existir tempo para voltar antes das 17h.
                if nova_hora > LIMITE_EXPEDIENTE_DAVI_MIN:
                    continue
                if retornar_base and ponto_base and ponto != ponto_base:
                    _dist_volta, _dur_volta = get_dist_dur(ponto, ponto_base)
                    _chegada_base, _fim_base = avancar_relogio(nova_hora, _dur_volta, 0)
                    if _fim_base > LIMITE_EXPEDIENTE_DAVI_MIN:
                        continue

                if "Menor Distância" in estrategia:
                    incremento = distancia * 3.2 + duracao * 0.35
                else:
                    incremento = duracao + distancia * 0.18

                # Evita entregar em um destino se ainda falta coletar outro
                # material que também será entregue nele.
                ainda_falta_para_o_ponto = [
                    i for i in tarefas_por_destino.get(ponto, [])
                    if not (novas_coletadas & (1 << i))
                ]
                if ids_entrega and ainda_falta_para_o_ponto:
                    urgente_agora = max([pesos[i] for i in ids_entrega] + [1])
                    urgente_depois = max([pesos[i] for i in ainda_falta_para_o_ponto] + [1])
                    excecao_urgente = "Urgências" in estrategia and urgente_agora >= 4 and urgente_agora > urgente_depois
                    if not excecao_urgente:
                        incremento += 450.0

                tempo_decorrido = max(0.0, nova_hora - horario_inicio)
                if "Menor Distância" not in estrategia:
                    for i in ids_entrega:
                        peso_i = float(pesos[i])
                        if "Urgências" in estrategia:
                            fator_urgencia = max(0.0, peso_i - 2.0) * 1.15
                        elif "Equilibrada" in estrategia:
                            # Custo por adiar uma entrega: quanto mais crítico o
                            # prazo, mais caro é deixá-la para o fim da rota.
                            if peso_i >= 9:
                                fator_urgencia = 3.60
                            elif peso_i >= 7:
                                fator_urgencia = 2.40
                            elif peso_i >= 6:
                                fator_urgencia = 1.85
                            elif peso_i >= 5:
                                fator_urgencia = 1.30
                            elif peso_i >= 3:
                                fator_urgencia = 0.18
                            else:
                                fator_urgencia = 0.0
                        else:
                            fator_urgencia = {7: 1.20, 6: 0.95, 5: 0.70, 4: 0.40, 3: 0.10}.get(int(peso_i), 0.0)
                        incremento += tempo_decorrido * fator_urgencia

                carga_apos = (novas_coletadas & ~novas_entregues).bit_count()
                if "Descarregar" in estrategia:
                    incremento += carga_apos * (duracao + tempo_servico) * 0.55
                elif "Equilibrada" in estrategia:
                    incremento += carga_apos * (duracao + tempo_servico) * 0.05

                novo_estado = {
                    "atual": ponto,
                    "hora": nova_hora,
                    "coletadas": novas_coletadas,
                    "entregues": novas_entregues,
                    "ordem": estado["ordem"] + (ponto,),
                    "custo": estado["custo"] + incremento,
                    "distancia": estado["distancia"] + distancia,
                    "viagem": estado["viagem"] + duracao,
                }
                chave_estado = (ponto, novas_coletadas, novas_entregues)
                anterior = proximos_por_estado.get(chave_estado)
                if anterior is None or (novo_estado["custo"], novo_estado["hora"]) < (anterior["custo"], anterior["hora"]):
                    proximos_por_estado[chave_estado] = novo_estado

        if concluidos:
            break
        if not proximos_por_estado:
            break

        candidatos_ordenados = sorted(
            proximos_por_estado.values(),
            key=lambda e: e["custo"] + heuristica_restante(e["atual"], e["coletadas"], e["entregues"]),
        )
        estados = candidatos_ordenados[:largura_feixe]
        if time.perf_counter() - inicio_busca > 5.0:
            break

    if not concluidos:
        concluidos = [e for e in estados if e["entregues"] == mascara_total]
    if concluidos:
        def custo_final(estado):
            custo = estado["custo"]
            if retornar_base and ponto_base and estado["atual"] != ponto_base:
                distancia, duracao = get_dist_dur(estado["atual"], ponto_base)
                custo += duracao + distancia * 0.18
            return custo
        melhor = min(concluidos, key=custo_final)
        return list(melhor["ordem"])

    # Se nem todas as demandas couberem até 17h, devolve a melhor sequência
    # PARCIAL possível. Na Equilibrada, prazo vem antes de simplesmente maximizar
    # a quantidade de cartões: não faz sentido concluir três FUTURO e deixar uma
    # demanda HOJE de fora apenas porque as três eram geograficamente mais fáceis.
    if estados:
        def _valor_prioridade_entregue(mask):
            valor = 0.0
            prioritarias = 0
            criticas = 0
            vencidas = 0
            for i, peso in enumerate(pesos):
                if not (mask & (1 << i)):
                    continue
                if peso >= 9:
                    valor += 300.0; prioritarias += 1
                elif peso >= 7:
                    valor += 180.0; vencidas += 1; criticas += 1
                elif peso >= 6:
                    valor += 130.0; criticas += 1
                elif peso >= 5:
                    valor += 100.0; criticas += 1
                elif peso >= 3:
                    valor += 16.0
                elif peso >= 2:
                    valor += 4.0
                else:
                    valor += 1.0
            return prioritarias, vencidas, criticas, valor

        def chave_parcial(estado):
            entregues_mask = int(estado["entregues"])
            entregues_qtd = entregues_mask.bit_count()
            carregadas_sem_entrega = int(estado["coletadas"] & ~estado["entregues"]).bit_count()
            prioritarias, vencidas, criticas, valor_prazo = _valor_prioridade_entregue(entregues_mask)
            if "Equilibrada" in estrategia or "Urgências" in estrategia:
                return (-prioritarias, -vencidas, -criticas, -valor_prazo, -entregues_qtd, carregadas_sem_entrega, estado["hora"], estado["custo"])
            return (-entregues_qtd, carregadas_sem_entrega, estado["hora"], estado["custo"])
        melhor_parcial = min(estados, key=chave_parcial)
        if melhor_parcial.get("ordem"):
            return list(melhor_parcial["ordem"])

    # Se a busca atingir o limite de tempo, conclui de forma determinística
    # com o motor guloso seguro, sem travar a geração da rota.
    pendentes = list(tarefas)
    no_carro = []
    atual = ponto_inicial
    ordem = []
    visitados = set(locais_bloqueados)
    for _ in range(total_tarefas * 2 + 5):
        candidatos = ({t['Origem'] for t in pendentes} | {t['Destino'] for t in no_carro}) - visitados
        if not candidatos:
            break
        candidatos = priorizar_pontos_sem_revisita(candidatos, pendentes)
        proximo = min(candidatos, key=lambda p: pontuar_parada_rota(atual, p, pendentes, no_carro, estrategia, get_dist_dur)[0])
        ordem.append(proximo)
        no_carro = [t for t in no_carro if t['Destino'] != proximo]
        coletadas = [t for t in pendentes if t['Origem'] == proximo]
        pendentes = [t for t in pendentes if t['Origem'] != proximo]
        no_carro.extend(coletadas)
        atual = proximo
        visitados.add(proximo)
    return ordem

def _decodificar_polyline_google(polyline):
    pontos = []
    indice = latitude = longitude = 0
    while indice < len(polyline):
        valores = []
        for _ in range(2):
            resultado = deslocamento = 0
            while indice < len(polyline):
                byte = ord(polyline[indice]) - 63
                indice += 1
                resultado |= (byte & 0x1F) << deslocamento
                deslocamento += 5
                if byte < 0x20:
                    break
            valores.append(~(resultado >> 1) if resultado & 1 else resultado >> 1)
        latitude += valores[0]
        longitude += valores[1]
        pontos.append([latitude / 1e5, longitude / 1e5])
    return pontos

def buscar_geometria_google_trafego(coords_limpas, horario_partida=None):
    chave = carregar_chave_google_routes()
    if not chave or len(coords_limpas) < 2 or len(coords_limpas) > 27:
        return None

    def waypoint(coord, parada=False):
        dado = {"location": {"latLng": {"latitude": float(coord[0]), "longitude": float(coord[1])}}}
        if parada:
            dado["vehicleStopover"] = True
        return dado

    agora_seguro = datetime.now(FUSO_LOCAL) + timedelta(minutes=1)
    partida = horario_partida if horario_partida and horario_partida > agora_seguro else agora_seguro
    payload = {
        "origin": waypoint(coords_limpas[0]),
        "destination": waypoint(coords_limpas[-1]),
        "intermediates": [waypoint(c, parada=True) for c in coords_limpas[1:-1]],
        "travelMode": "DRIVE",
        "routingPreference": "TRAFFIC_AWARE_OPTIMAL",
        "trafficModel": "BEST_GUESS",
        "departureTime": partida.astimezone(timezone.utc).isoformat().replace("+00:00", "Z"),
        "polylineQuality": "HIGH_QUALITY",
        "polylineEncoding": "ENCODED_POLYLINE",
        "routeModifiers": {"avoidFerries": True},
        "languageCode": "pt-BR",
        "regionCode": "br",
        "units": "METRIC",
    }
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": chave,
        "X-Goog-FieldMask": "routes.polyline.encodedPolyline",
    }
    try:
        resposta = requests.post("https://routes.googleapis.com/directions/v2:computeRoutes", headers=headers, json=payload, timeout=35)
        resposta.raise_for_status()
        rotas = resposta.json().get("routes", [])
        encoded = rotas[0].get("polyline", {}).get("encodedPolyline", "") if rotas else ""
        pontos = _decodificar_polyline_google(encoded) if encoded else []
        return pontos if len(pontos) > 1 else None
    except Exception:
        return None

def buscar_geometria_mapbox_trafego(coords_limpas, horario_partida=None):
    """Traçado completo pelas ruas usando Mapbox driving-traffic.

    Directions aceita até 25 coordenadas por chamada. Rotas maiores são quebradas
    em blocos sobrepostos, sem perder a continuidade visual do percurso.
    """
    token = carregar_token_mapbox()
    if not token or len(coords_limpas) < 2:
        return None

    partida = _partida_mapbox(horario_partida)
    pontos_finais = []
    inicio = 0
    try:
        while inicio < len(coords_limpas) - 1:
            fim = min(inicio + 25, len(coords_limpas))
            trecho = coords_limpas[inicio:fim]
            coords_str = ';'.join(f"{float(lon):.7f},{float(lat):.7f}" for lat, lon in trecho)
            params = {
                'access_token': token,
                'geometries': 'geojson',
                'overview': 'full',
                'steps': 'false',
                'depart_at': _param_depart_at_mapbox(partida),
            }
            resposta = requests.get(
                f"https://api.mapbox.com/directions/v5/mapbox/driving-traffic/{coords_str}",
                params=params,
                timeout=30,
            )
            if not resposta.ok:
                try:
                    st.session_state['_mapbox_geom_diag'] = _erro_mapbox_resposta(resposta)
                except Exception:
                    pass
                return None
            dados = resposta.json()
            rotas = dados.get('routes') or []
            if dados.get('code') != 'Ok' or not rotas:
                try:
                    st.session_state['_mapbox_geom_diag'] = f"{dados.get('code', 'Erro')}: {dados.get('message', 'sem rota')}"
                except Exception:
                    pass
                return None
            rota = rotas[0]
            geom = ((rota.get('geometry') or {}).get('coordinates') or [])
            # GeoJSON Mapbox = [lon, lat]; Folium = [lat, lon].
            pontos_trecho = [[float(lat), float(lon)] for lon, lat in geom]
            for ponto in pontos_trecho:
                if not pontos_finais or ponto != pontos_finais[-1]:
                    pontos_finais.append(ponto)
            try:
                partida += timedelta(seconds=float(rota.get('duration') or 0))
            except Exception:
                pass
            if fim >= len(coords_limpas):
                break
            inicio = fim - 1  # repete o último waypoint para ligar os blocos

        if len(pontos_finais) > 1:
            try:
                st.session_state['_mapbox_geom_diag'] = 'Mapbox Directions ativa.'
            except Exception:
                pass
            return pontos_finais
        return None
    except Exception as erro:
        try:
            st.session_state['_mapbox_geom_diag'] = f'{type(erro).__name__}: {erro}'[:500]
        except Exception:
            pass
        return None


def buscar_geometria_rota(coords_ordenadas, horario_partida=None):
    """Obtém o traçado pelas ruas via OSRM e entrega [lat, lon] para o Folium."""
    coords_limpas = []
    for coord in coords_ordenadas:
        if not coords_limpas or coord != coords_limpas[-1]:
            coords_limpas.append(coord)
    if len(coords_limpas) < 2:
        return [[float(lat), float(lon)] for lat, lon in coords_limpas], False

    try:
        coords_str = ';'.join(f"{float(lon)},{float(lat)}" for lat, lon in coords_limpas)
        url = f"https://router.project-osrm.org/route/v1/driving/{coords_str}?overview=full&geometries=geojson&steps=false"
        req = urllib.request.Request(url, headers={'User-Agent': 'AproarLogisticsWeb/1.0'})
        with urllib.request.urlopen(req, timeout=18) as response:
            res = json.loads(response.read())
        if res.get('code') == 'Ok' and res.get('routes'):
            coords_geojson = res['routes'][0].get('geometry', {}).get('coordinates', [])
            # GeoJSON/OSRM = [longitude, latitude]; Folium = [latitude, longitude].
            geometria_osrm = [[float(lat), float(lon)] for lon, lat in coords_geojson]
            if len(geometria_osrm) > 1:
                return geometria_osrm, True
    except Exception:
        pass

    # Se o OSRM estiver temporariamente indisponível, o mapa ainda mostra a sequência
    # das paradas em linha reta em vez de ficar totalmente sem traçado.
    return [[float(lat), float(lon)] for lat, lon in coords_limpas], False



def extrair_dados_completos(texto, card_name):
    titulo = str(card_name or "")

    # O identificador da obra deve ser preservado EXATAMENTE como aparece no
    # título: não existe quantidade fixa de dígitos. Exemplos válidos: 086,
    # 2546, 12, 10358, 2112.1 e códigos APR. Priorizamos o valor logo após
    # a palavra OBRA; se ela não existir, aceitamos um identificador no início
    # do título para não capturar números de bloco, endereço, quantidade etc.
    num_match = re.search(
        r'(?i)\bOBRA\s*[-:#]?\s*(APR[A-Z0-9]+|\d+(?:\.\d+)?)\b',
        titulo,
    )
    if not num_match:
        num_match = re.match(
            r'(?i)^\s*[-:#]?\s*(APR[A-Z0-9._-]+|\d+(?:\.\d+)?)\b',
            titulo,
        )
    if not num_match:
        # Há cartões em que a unidade/descrição vem antes do número da obra.
        # Como o identificador sempre está no título, buscamos o primeiro código
        # numérico/APR restante, sem limitar a quantidade de dígitos.
        num_match = re.search(
            r'(?i)(?<![A-Z0-9])(APR[A-Z0-9._-]*\d[A-Z0-9._-]*|\d+(?:\.\d+)?)(?![A-Z0-9])',
            titulo,
        )
    num = num_match.group(1).upper() if num_match else ""

    # Exemplos:
    # "OBRA 086 - REFORMA FACHADA - BLOCO H - UNIFOR" -> "086 - UNIFOR"
    # "OBRA 2546 - ... - BARRA DO CEARÁ" -> "2546 - BARRA DO CEARÁ".
    unidade = identificar_unidade_empresa(titulo, permitir_contexto=True)

    origem, destino = "", ""
    materiais = "Ver Trello"

    if texto:
        texto_limpo = re.sub(r'[*_`]+', '', texto).strip()

        if "TRANSBORDO" in texto_limpo.upper() or "TRANSBORDOS" in texto_limpo.upper():
            if unidade:
                origem = unidade
                destino = "ESCRITÓRIO"
            materiais = texto_limpo
        else:
            # COLETAR EM/NA/NO, PEGAR, RETIRAR e BUSCAR identificam a ORIGEM.
            mo = re.search(
                r'(?i)(?:coletar|pegar|retirar|buscar|coleta)\s+(?:(?:em|no|na|nos|nas|do|da|dos|das|o|a|ao|à|aos|às)\s+)?([^\:\n\.\-]+)',
                texto_limpo,
            )
            if mo:
                origem_bruta = mo.group(1)
                origem_bruta = re.split(
                    r'(?i)\s+(?:e\s+)?(?:levar|entreg(?:ar|a)|devolver|encaminhar|transportar|deixar|entrega)\b',
                    origem_bruta,
                    maxsplit=1,
                )[0].strip()
                origem = canonicalizar_ponto_rota(origem_bruta)

            # LEVAR PARA / ENTREGAR PARA, EM, NO, NA... identificam o DESTINO.
            md = re.search(
                r'(?i)(?:levar|entreg(?:ar|a)|devolver|encaminhar|transportar|deixar|entrega)\s+(?:(?:para|em|no|na|nos|nas|ao|à|aos|às|o|a)\s+)?([^\:\n\.]+)',
                texto_limpo,
            )
            if md:
                destino = canonicalizar_ponto_rota(md.group(1))

            if mo and md:
                start_idx = mo.end()
                end_idx = md.start()
                if start_idx < end_idx:
                    mat_text = texto_limpo[start_idx:end_idx].strip()
                    linhas_limpas = [
                        l.strip().lstrip('-').strip()
                        for l in mat_text.split('\n')
                        if len(l.strip()) >= 2
                        and l.lower() not in ['e', 'e:', 'e -', 'e,', 'para', 'levar para']
                    ]
                    if linhas_limpas:
                        materiais = " | ".join(linhas_limpas)

    # Se o texto não trouxe a unidade mas o destino/origem é uma unidade conhecida,
    # ela também pode completar o nome resumido da obra.
    if not unidade:
        unidade = identificar_unidade_empresa(destino) or identificar_unidade_empresa(origem)

    if not destino and unidade:
        destino = unidade
    if not origem and destino:
        origem = "ESCRITÓRIO"

    origem = canonicalizar_ponto_rota(origem)
    destino = canonicalizar_ponto_rota(destino)

    short_name = (
        f"{num} - {unidade}" if (num and unidade)
        else num if num
        else unidade if unidade
        else titulo[:25] + "..."
    )

    # Regra operacional mantida a pedido da equipe: materiais destinados ao
    # SEBRAE ficam no escritório para o Soares.
    if "SEBRAE" in destino or "SEBRAE" in short_name:
        if destino and destino not in ["ESCRITÓRIO"]:
            destino = "ESCRITÓRIO"
            materiais += " ⚠️[DEIXAR NO ESCRITÓRIO P/ SOARES]"

    return short_name, origem, destino, materiais

def encontrar_endereco_na_descricao(descricao):
    if not descricao: return None
    mo_link = re.search(r'(https?://(?:www\.)?google\.[a-z\.]+/maps[^\s\n]+|https?://goo\.gl/maps/[^\s\n]+|https?://maps\.app\.goo\.gl/[^\s\n]+)', descricao)
    if mo_link: return mo_link.group(1)
    mo_end = re.search(r'(?i)(?:endere[çc]o|local)\s*(?:\:|-)\s*([^\n]+)', descricao)
    if mo_end: return mo_end.group(1).strip()
    return None

def alvo_endereco_trello(descricao, origem, destino):
    """Decide a qual ponta da demanda pertence um endereço escrito no cartão.

    A versão antiga podia salvar o MESMO `Endereço:` tanto na origem quanto no
    destino quando ambos eram externos. Isso contaminava o banco de fornecedores e,
    por consequência, a matriz da rota. Agora só há gravação quando o contexto é
    suficientemente seguro.
    """
    texto = remover_acentos(str(descricao or "")).upper()
    origem_n = remover_acentos(str(origem or "")).upper().strip()
    destino_n = remover_acentos(str(destino or "")).upper().strip()

    # Marcadores explícitos têm prioridade.
    marcadores_origem = ("ENDERECO DA COLETA", "ENDERECO DO FORNECEDOR", "ENDERECO DE RETIRADA", "LOCAL DA COLETA")
    marcadores_destino = ("ENDERECO DA ENTREGA", "ENDERECO DA OBRA", "LOCAL DA ENTREGA", "ENDERECO DO DESTINO")
    if any(m in texto for m in marcadores_origem):
        return "origem"
    if any(m in texto for m in marcadores_destino):
        return "destino"

    # Se apenas uma ponta é externa, um endereço genérico do cartão normalmente
    # pertence justamente a essa ponta; a unidade própria já tem endereço fixo.
    origem_propria = canonicalizar_ponto_rota(origem) in UNIDADES_PROPRIAS
    destino_proprio = canonicalizar_ponto_rota(destino) in UNIDADES_PROPRIAS
    if origem_propria and not destino_proprio:
        return "destino"
    if destino_proprio and not origem_propria:
        return "origem"

    # Nome literal próximo ao conteúdo também permite associação segura.
    if origem_n and origem_n in texto and ("COLET" in texto or "RETIR" in texto or "BUSCAR" in texto):
        return "origem"
    if destino_n and destino_n in texto and ("ENTREG" in texto or "LEVAR" in texto or "DESTINO" in texto):
        return "destino"
    return None

def classificar_prioridade(due_str):
    """Transforma o prazo do Trello em prioridade operacional da rota.

    A estratégia Equilibrada precisa distinguir uma demanda apenas futura de uma
    demanda que vence hoje — especialmente quando o horário-limite está próximo.
    O número retornado é deliberadamente espaçado para que o otimizador consiga
    respeitar prazo sem abandonar distância/tempo de deslocamento.
    """
    if not due_str:
        return 1, "Sem Prazo"
    try:
        due_local = converter_data_trello(due_str)
        diff = (due_local.date() - DATA_REF_ROTA_DATE).days

        if diff < 0:
            return 7, "VENCIDA"

        if diff == 0:
            # Para o planejamento do próprio dia, considera também o relógio do
            # Trello. Ex.: entrega hoje às 16h não pode ter o mesmo peso de uma
            # demanda futura quando já são 14h.
            if DATA_REF_ROTA_DATE == AGORA_REAL.date():
                agora_min = AGORA_REAL.hour * 60 + AGORA_REAL.minute
                prazo_min = due_local.hour * 60 + due_local.minute
                restante = prazo_min - agora_min
                if restante <= 0:
                    return 7, "VENCIDA"
                if restante <= 120:
                    return 6, "HOJE — prazo crítico"
                if restante <= 240:
                    return 5, "HOJE"
            return 5, "HOJE"

        if diff <= 2:
            return 3, f"Em {diff} dias"
        return 2, "Futuro"
    except Exception:
        return 1, "Sem Prazo"

def converter_data_trello(valor):
    if not valor: return None
    data = datetime.fromisoformat(valor.replace("Z", "+00:00"))
    if data.tzinfo is None: data = data.replace(tzinfo=timezone.utc)
    return data.astimezone(FUSO_LOCAL)

def lista_esta_concluida(nome_lista):
    nome = normalizar_local(nome_lista or "")
    return "CONCLU" in nome or "ENTREG" in nome

def encontrar_conclusao_de_hoje(card_id, acoes):
    hoje = AGORA_REAL.date()
    conclusoes = []
    for acao in acoes:
        if acao.get("type") != "updateCard": continue
        dados = acao.get("data", {})
        if dados.get("card", {}).get("id") != card_id: continue
        if not lista_esta_concluida(dados.get("listAfter", {}).get("name", "")): continue
        if lista_esta_concluida(dados.get("listBefore", {}).get("name", "")): continue
        try:
            momento = converter_data_trello(acao.get("date"))
            if momento and momento.date() == hoje: conclusoes.append(momento)
        except: continue
    return max(conclusoes) if conclusoes else None

def sincronizar_demandas(manual=False, forcar=False, geocodificar=True, somente_cache=False):
    data = obter_dados_trello(forcar=forcar, somente_cache=somente_cache)
    if not data:
        if manual: st.error("⚠️ Erro ao acessar o Trello.")
        return False
        
    trello_lists = {l['id']: l['name'] for l in data.get('lists', []) if not l.get('closed')}
    demandas_extraidas = []

    # Mantém um mapa de ID -> "número da obra - unidade" de TODO o quadro.
    # Isso é necessário porque, ao mover um cartão para CONCLUÍDAS, ele deixa
    # de entrar em ``demandas_extraidas``, mas ainda precisa manter o rótulo
    # correto no roteiro (ex.: 086 - UNIFOR).
    try:
        st.session_state["rotulos_obras_trello"] = construir_mapa_rotulos_obras_trello(data)
    except Exception:
        st.session_state.setdefault("rotulos_obras_trello", {})
    
    for c in data.get('cards', []):
        if c.get('closed') or lista_esta_concluida(trello_lists.get(c.get('idList', ''), '').upper()): continue
        short_name, origem, destino, materiais = extrair_dados_completos(c.get('desc', ''), c.get('name', ''))
        peso, status_prazo = classificar_prioridade(c.get('due'))
        endereco_card = encontrar_endereco_na_descricao(c.get('desc', ''))
        alvo_endereco = alvo_endereco_trello(c.get('desc', ''), origem, destino) if endereco_card else None
        if geocodificar and endereco_card and alvo_endereco:
            local_alvo = origem if alvo_endereco == "origem" else destino
            if local_alvo and canonicalizar_ponto_rota(local_alvo) not in UNIDADES_PROPRIAS:
                # Primeiro consulta o banco. Antes, o aplicativo chamava o geocodificador
                # externo em TODA sincronização mesmo quando o GPS já estava salvo, o que
                # deixava a inicialização muito lenta.
                existente = fetch_one(
                    "SELECT lat, lon FROM locais WHERE apelido = :apelido",
                    {"apelido": local_alvo},
                )
                gps_ja_salvo = bool(
                    existente and existente[0] is not None and existente[1] is not None
                )
                if not gps_ja_salvo:
                    lat, lon = buscar_coordenadas(endereco_card)
                    if lat is not None and lon is not None:
                        # Nunca sobrescreve coordenadas já validadas no banco por causa
                        # de um cartão do Trello. A aba Endereços permanece soberana.
                        execute_db(
                            "INSERT INTO locais (apelido, endereco, lat, lon) VALUES (:apelido, :end, :lat, :lon) "
                            "ON CONFLICT (apelido) DO UPDATE SET endereco=EXCLUDED.endereco, lat=EXCLUDED.lat, lon=EXCLUDED.lon",
                            {"apelido": local_alvo, "end": endereco_card, "lat": lat, "lon": lon},
                        )
        
        tc_val = 20 if canonicalizar_ponto_rota(origem) not in UNIDADES_PROPRIAS else 10
        te_val = 10
        if not st.session_state.demandas.empty and c['id'] in st.session_state.demandas['id'].values:
            linha_antiga = st.session_state.demandas[st.session_state.demandas['id'] == c['id']].iloc[0]
            tc_val, te_val = linha_antiga['Tempo_Coleta'], linha_antiga['Tempo_Entrega']
        
        demandas_extraidas.append({
            "id": c['id'], "Obra": short_name, "Origem": origem, "Destino": destino,
            "Materiais": materiais, "Urgência": status_prazo, "Peso": peso,
            "Tempo_Coleta": tc_val, "Tempo_Entrega": te_val,
            "Supervisor": SUPERVISORES_MAP.get(destino, "Sede / Logística"),
            # Fonte imutável para reconstruir o rótulo da obra no roteiro, mesmo
            # depois que o cartão for concluído/arquivado.
            "_Titulo_Trello": str(c.get('name', '') or ''),
        })

    st.session_state.demandas = pd.DataFrame(demandas_extraidas, columns=COLUNAS_DEMANDAS)
    st.session_state.ultima_sincronizacao = time.time()
    return True

class FormularioLoginParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.action = None
        self.inputs = []
    def handle_starttag(self, tag, attrs):
        atributos = {k: (v or "") for k, v in attrs}
        if tag.lower() == "form" and self.action is None: self.action = atributos.get("action", "")
        elif tag.lower() == "input": self.inputs.append(atributos)

def _escolher_campo(campos, palavras):
    if not campos: return None
    for campo in campos:
        identificador = f"{campo.get('name', '')} {campo.get('id', '')}".lower()
        if any(palavra in identificador for palavra in palavras): return campo
    return campos[0]

def _montar_formulario_login(html, usuario, senha):
    parser = FormularioLoginParser()
    parser.feed(html)
    campos_com_nome = [c for c in parser.inputs if c.get("name")]
    campo_usuario = _escolher_campo([c for c in campos_com_nome if c.get("type", "text").lower() in ("text", "email")], ("usu", "user", "login", "email"))
    campo_senha = _escolher_campo([c for c in campos_com_nome if c.get("type", "").lower() == "password"], ("senha", "password", "pass"))
    if not campo_usuario or not campo_senha: raise RuntimeError("Não foi possível identificar os campos de acesso do portal.")
    dados = {c["name"]: c.get("value", "") for c in campos_com_nome if c.get("type", "").lower() == "hidden"}
    dados[campo_usuario["name"]] = usuario
    dados[campo_senha["name"]] = senha
    botao = _escolher_campo([c for c in campos_com_nome if c.get("type", "").lower() in ("submit", "button")], ("entr", "acess", "login", "logar"))
    if botao: dados[botao["name"]] = botao.get("value", "Entrar")
    imagens = [c for c in campos_com_nome if c.get("type", "").lower() == "image"]
    if imagens: dados[f"{imagens[0]['name']}.x"] = "10"; dados[f"{imagens[0]['name']}.y"] = "10"
    return parser.action, dados

def _parsear_resposta_rastreador(texto):
    posicoes = []
    for registro in texto.replace("\r", "").split(";"):
        registro = registro.strip()
        if not registro: continue
        partes = registro.split("|", 8)
        if len(partes) < 9: continue
        try:
            latitude, longitude, velocidade = float(partes[3]), float(partes[4]), float(partes[5].replace(",", "."))
        except: continue
        codigo_status = partes[6].strip().upper()
        
        placa = partes[1].strip().upper()
        if placa.startswith("TIF"): placa = "TIF-2123"
        elif placa.startswith("OSC"): placa = "OSC-3842"
        
        posicoes.append({"ID": partes[0].strip(), "Placa": placa, "Última atualização": partes[2].strip(), "Latitude": latitude, "Longitude": longitude, "Velocidade (km/h)": velocidade, "Situação": {"P": "Parado", "M": "Em movimento", "L": "Ligado", "D": "Desligado"}.get(codigo_status, codigo_status or "Não informado"), "Código": codigo_status, "Ícone": partes[7].strip(), "Endereço": partes[8].strip()})
    return posicoes

def consultar_posicoes_protege(sessao, pagina_atual, veiculos):
    url = urllib.parse.urljoin(pagina_atual, "consultaajax_all.aspx")
    resposta = sessao.post(url, params={"p1": veiculos}, headers={"X-Requested-With": "XMLHttpRequest", "Referer": pagina_atual}, timeout=20)
    resposta.raise_for_status()
    posicoes = _parsear_resposta_rastreador(resposta.text)
    if not posicoes: raise RuntimeError("O portal não devolveu posições. A sessão pode ter expirado.")
    return posicoes

def autenticar_protege(usuario, senha, veiculos):
    ultimo_erro = None
    for login_url in RASTREADOR_LOGIN_URLS:
        try:
            sessao = requests.Session()
            sessao.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/151.0 Safari/537.36"})
            pagina_login = sessao.get(login_url, timeout=20)
            pagina_login.raise_for_status()
            action, dados = _montar_formulario_login(pagina_login.text, usuario, senha)
            url_post = urllib.parse.urljoin(pagina_login.url, action or pagina_login.url)
            resposta_login = sessao.post(url_post, data=dados, timeout=20, allow_redirects=True)
            resposta_login.raise_for_status()
            pagina_atual = resposta_login.url
            return sessao, pagina_atual, consultar_posicoes_protege(sessao, pagina_atual, veiculos)
        except Exception as erro: ultimo_erro = erro
    raise RuntimeError("Não foi possível autenticar ou consultar o rastreador.") from ultimo_erro

def carregar_config_protege():
    try:
        config = st.secrets["protege"]
        usuario, senha, veiculos = str(config.get("usuario", "")).strip(), str(config.get("senha", "")).strip(), config.get("veiculos", RASTREADOR_VEICULOS_PADRAO)
        return usuario, senha, ",".join(str(v).strip() for v in veiculos) if isinstance(veiculos, (list, tuple)) else str(veiculos).strip()
    except: return "", "", RASTREADOR_VEICULOS_PADRAO


@st.cache_resource(show_spinner=False)
def obter_executor_gps_rota():
    """Um único trabalhador consulta o GPS sem bloquear a renderização da Torre."""
    return ThreadPoolExecutor(max_workers=1, thread_name_prefix="aproar-gps-rota")


def consultar_gps_rota_em_background(usuario, senha, veiculos, sessao=None, pagina=None):
    """Consulta/reautentica fora da thread que desenha a página."""
    if sessao is not None and pagina:
        try:
            posicoes = consultar_posicoes_protege(sessao, pagina, veiculos)
            return {"sessao": sessao, "pagina": pagina, "posicoes": posicoes, "erro": ""}
        except Exception:
            pass
    try:
        nova_sessao, nova_pagina, posicoes = autenticar_protege(usuario, senha, veiculos)
        return {"sessao": nova_sessao, "pagina": nova_pagina, "posicoes": posicoes, "erro": ""}
    except Exception as erro:
        return {"sessao": None, "pagina": "", "posicoes": [], "erro": str(erro)[:180]}

def loop_automacoes_background(processar_rastreador=True):
    agora_loop = datetime.now(FUSO_LOCAL)
    try:
        data = obter_dados_trello()
        if data:
            trello_lists = {l['id']: l['name'] for l in data.get('lists', []) if not l.get('closed')}
            cards = data.get('cards', [])
            acoes = data.get('actions', [])
            
            registros_hist = fetch_all(
                "SELECT id, teams_notificado, hora_conclusao FROM historico_concluidos WHERE data_conclusao = :data",
                {"data": DATA_HOJE_REAL_STR},
            )
            historico_hoje = {
                str(r[0]): {"teams_notificado": r[1], "hora": r[2]}
                for r in registros_hist
            }

            novas_entregas = 0
            for c in cards:
                if c.get('closed'):
                    continue
                if not lista_esta_concluida(trello_lists.get(c.get('idList', ''), '').upper()):
                    continue

                momento_conclusao = encontrar_conclusao_de_hoje(c['id'], acoes)
                if not momento_conclusao or momento_conclusao.strftime("%d/%m/%Y") != DATA_HOJE_REAL_STR:
                    continue

                card_id = str(c['id'])
                short_name, origem, destino, materiais = extrair_dados_completos(c.get('desc', ''), c.get('name', ''))
                hora_str = momento_conclusao.strftime("%H:%M")
                registro_existente = historico_hoje.get(card_id)

                # A baixa é registrada imediatamente para atualizar roteiro/app, mesmo se
                # o Teams estiver temporariamente fora. A notificação fica pendente e será
                # tentada novamente nos próximos ciclos até receber HTTP 2xx.
                if registro_existente is None:
                    execute_db(
                        "INSERT INTO historico_concluidos "
                        "(id, obra, origem, destino, materiais, data_conclusao, hora_conclusao, teams_notificado, teams_tentativas, teams_ultimo_erro) "
                        "VALUES (:id, :obra, :origem, :destino, :mat, :data, :hora, FALSE, 0, NULL) "
                        "ON CONFLICT (id) DO UPDATE SET obra=EXCLUDED.obra, origem=EXCLUDED.origem, destino=EXCLUDED.destino, "
                        "materiais=EXCLUDED.materiais, data_conclusao=EXCLUDED.data_conclusao, hora_conclusao=EXCLUDED.hora_conclusao",
                        {"id": c['id'], "obra": short_name, "origem": origem, "destino": destino, "mat": materiais, "data": DATA_HOJE_REAL_STR, "hora": hora_str},
                    )
                    historico_hoje[card_id] = {"teams_notificado": False, "hora": hora_str}
                    registro_existente = historico_hoje[card_id]
                    novas_entregas += 1
                else:
                    # Mantém os dados do histórico coerentes caso o cartão tenha sido
                    # ajustado no Trello antes da baixa.
                    execute_db(
                        "UPDATE historico_concluidos SET obra=:obra, origem=:origem, destino=:destino, materiais=:mat, hora_conclusao=:hora WHERE id=:id",
                        {"id": c['id'], "obra": short_name, "origem": origem, "destino": destino, "mat": materiais, "hora": hora_str},
                    )

                status_teams = registro_existente.get("teams_notificado")
                minutos_desde_baixa = max(0.0, (agora_loop - momento_conclusao).total_seconds() / 60.0)

                # Mantém a automação no mesmo ritmo operacional do Trello: 2 minutos.
                # NULL = registro legado; só tenta recuperar se a baixa ocorreu nos últimos
                # 2 minutos. FALSE = notificação nova pendente; ela é tentada novamente
                # a cada ciclo de 2 minutos até o Teams confirmar o recebimento.
                deve_notificar = (status_teams is False) or (status_teams is None and minutos_desde_baixa <= 2)
                if not deve_notificar:
                    continue

                url_webhook, fonte_webhook = obter_webhook_teams(
                    destino, supervisor=SUPERVISORES_MAP.get(destino, "Sede / Logística"), obra=short_name
                )
                observacao_trello = extrair_observacoes_trello(c, acoes)
                mensagem = (
                    f"✅ **Os materiais foram entregues na obra e a demanda tomou baixa no Trello.**\n\n"
                    f"**Obra:** {short_name}\n\n"
                    f"**Local:** {destino}\n\n"
                    f"**Materiais:**\n{formatar_materiais_teams(materiais)}\n\n"
                    f"**Data e Hora:** {momento_conclusao.strftime('%d/%m/%Y às %H:%M')}"
                )
                if observacao_trello:
                    mensagem += f"\n\n**Comentários do Trello:**\n{observacao_trello}"

                if not url_webhook:
                    enviado, detalhe = False, f"Webhook do Teams não configurado ({fonte_webhook})."
                else:
                    enviado, detalhe = disparar_teams(
                        url_webhook, f"✅ Entrega concluída — {destino}", mensagem
                    )

                if enviado:
                    execute_db(
                        "UPDATE historico_concluidos SET teams_notificado=TRUE, teams_tentativas=COALESCE(teams_tentativas,0)+1, teams_ultimo_erro=NULL WHERE id=:id",
                        {"id": c['id']},
                    )
                    historico_hoje[card_id]["teams_notificado"] = True
                    st.session_state.pop("_teams_ultimo_erro", None)
                else:
                    execute_db(
                        "UPDATE historico_concluidos SET teams_notificado=FALSE, teams_tentativas=COALESCE(teams_tentativas,0)+1, teams_ultimo_erro=:erro WHERE id=:id",
                        {"id": c['id'], "erro": str(detalhe)[:500]},
                    )
                    historico_hoje[card_id]["teams_notificado"] = False
                    st.session_state["_teams_ultimo_erro"] = f"{short_name}: {detalhe}"

            if novas_entregas > 0:
                try:
                    carregar_conclusoes_rota.clear()
                except Exception:
                    pass
                st.toast(
                    f"🔔 {novas_entregas} {plural_pt(novas_entregas, 'nova baixa registrada', 'novas baixas registradas')} no Trello!",
                    icon="✅",
                )
    except: pass

    # O ciclo automático de 2 minutos do Trello/Teams não deve autenticar no
    # rastreador. A Protege é muito mais pesada e continua restrita ao módulo
    # Rastreador/rotinas específicas. Isso mantém a navegação responsiva.
    if not processar_rastreador:
        return

    try:
        sessao, pagina = st.session_state.get("protege_sessao"), st.session_state.get("protege_pagina")
        usuario_protege, senha_protege, ids_veiculos = carregar_config_protege()
        posicoes = []

        # O medidor não depende mais de alguém abrir primeiro a aba do rastreador:
        # se houver credenciais, tenta autenticar automaticamente no ciclo de fundo.
        if ids_veiculos and usuario_protege and senha_protege:
            try:
                if sessao and pagina:
                    posicoes = consultar_posicoes_protege(sessao, pagina, ids_veiculos)
                else:
                    raise RuntimeError("sessão do rastreador ainda não iniciada")
            except Exception:
                sessao, pagina, posicoes = autenticar_protege(usuario_protege, senha_protege, ids_veiculos)
                st.session_state["protege_sessao"] = sessao
                st.session_state["protege_pagina"] = pagina

        if posicoes:
            # 1. Inteligência de Início de Rota (>500m do escritório)
            lat_base, lon_base = LOCAL_BASE_COORDS
            for p in posicoes:
                if p["Velocidade (km/h)"] > 0:
                    dist_base_km = calcular_distancia_km(lat_base, lon_base, p["Latitude"], p["Longitude"])
                    if dist_base_km > 0.5:
                        if not fetch_one("SELECT hora_inicio FROM inicio_movimento WHERE placa=:placa AND data=:data", {"placa": p["Placa"], "data": DATA_HOJE_REAL_STR}):
                            match_time = re.search(r'(\d{1,2}:\d{2})', str(p.get('Última atualização', '')))
                            hora_leitura = match_time.group(1).zfill(5) if match_time else agora_loop.strftime("%H:%M")
                            execute_db("INSERT INTO inicio_movimento (placa, data, hora_inicio) VALUES (:placa, :data, :hora) ON CONFLICT (placa, data) DO NOTHING", {"placa": p["Placa"], "data": DATA_HOJE_REAL_STR, "hora": hora_leitura})

            # 2. GEOFENCE ROBUSTA — chegada/saída com histerese e confirmação.
            # Entrada: <=250m + baixa velocidade por 2 leituras (~1 min).
            # Permanência: não encerra só porque o carro anda dentro da obra.
            # Saída: >350m por 2 leituras; >600m confirma imediatamente.
            RAIO_ENTRADA_KM = 0.25
            RAIO_SAIDA_KM = 0.35
            RAIO_SAIDA_IMEDIATA_KM = 0.60
            VELOCIDADE_MAX_ENTRADA = 8.0
            LEITURAS_CONFIRMAR = 2

            estados_geo = st.session_state.setdefault("_geofence_confirmacao", {})
            res_rota = fetch_one("SELECT json_locais FROM rota_ativa WHERE id=1 AND data_rota=:data", {"data": DATA_HOJE_REAL_STR})
            if res_rota:
                locais_rota = json.loads(res_rota[0])

                for p in posicoes:
                    lat_v, lon_v = p['Latitude'], p['Longitude']
                    placa_v = str(p['Placa'])
                    vel_v = float(p.get('Velocidade (km/h)', 0) or 0)
                    match_time = re.search(r'(\d{1,2}:\d{2})', str(p.get('Última atualização', '')))
                    hora_leitura = match_time.group(1).zfill(5) if match_time else agora_loop.strftime("%H:%M")

                    estado_geo = estados_geo.setdefault(placa_v, {
                        'entrada_local': None,
                        'entrada_contagem': 0,
                        'entrada_hora': '',
                        'saida_contagem': 0,
                        'saida_hora': '',
                    })

                    # Distâncias a todos os locais da rota (menos base).
                    distancias = {}
                    for nome_loc, coords in locais_rota.items():
                        if nome_loc == "ESCRITÓRIO" or not isinstance(coords, (list, tuple)) or len(coords) < 2:
                            continue
                        try:
                            distancias[nome_loc] = calcular_distancia_km(coords[0], coords[1], lat_v, lon_v)
                        except Exception:
                            continue

                    parada_ativa = fetch_one(
                        "SELECT id, local, hora_chegada FROM rastreio_paradas "
                        "WHERE data=:data AND placa=:placa AND hora_saida IS NULL ORDER BY id DESC LIMIT 1",
                        {"data": DATA_HOJE_REAL_STR, "placa": placa_v},
                    )

                    if parada_ativa:
                        id_ativa, local_ativo = parada_ativa[0], str(parada_ativa[1])
                        dist_ativa = distancias.get(local_ativo)
                        if dist_ativa is None:
                            # Tenta casar por nome normalizado caso a rota tenha sido recalculada.
                            chave_ativa = _normalizar_local_rastreio(local_ativo)
                            for nome_loc, dist_loc in distancias.items():
                                if _normalizar_local_rastreio(nome_loc) == chave_ativa:
                                    dist_ativa = dist_loc
                                    break

                        esta_fora = dist_ativa is None or dist_ativa > RAIO_SAIDA_KM
                        if not esta_fora:
                            # Continua no mesmo local mesmo que mova o carro internamente.
                            estado_geo['saida_contagem'] = 0
                            estado_geo['saida_hora'] = ''
                        else:
                            if estado_geo.get('saida_contagem', 0) == 0:
                                estado_geo['saida_hora'] = hora_leitura
                            estado_geo['saida_contagem'] = int(estado_geo.get('saida_contagem', 0)) + 1
                            saida_imediata = dist_ativa is not None and dist_ativa >= RAIO_SAIDA_IMEDIATA_KM
                            if saida_imediata or estado_geo['saida_contagem'] >= LEITURAS_CONFIRMAR:
                                hora_saida = estado_geo.get('saida_hora') or hora_leitura
                                execute_db("UPDATE rastreio_paradas SET hora_saida=:hora WHERE id=:id", {"hora": hora_saida, "id": id_ativa})
                                estado_geo.update({
                                    'entrada_local': None, 'entrada_contagem': 0, 'entrada_hora': '',
                                    'saida_contagem': 0, 'saida_hora': '',
                                })
                    else:
                        # Só abre uma parada após duas leituras coerentes. Isso reduz falsos
                        # positivos causados por semáforo/congestionamento perto da obra.
                        candidatos = [
                            (dist, nome) for nome, dist in distancias.items()
                            if dist <= RAIO_ENTRADA_KM
                        ]
                        candidato = min(candidatos)[1] if candidatos and vel_v <= VELOCIDADE_MAX_ENTRADA else None

                        if candidato:
                            if estado_geo.get('entrada_local') == candidato:
                                estado_geo['entrada_contagem'] = int(estado_geo.get('entrada_contagem', 0)) + 1
                            else:
                                estado_geo['entrada_local'] = candidato
                                estado_geo['entrada_contagem'] = 1
                                estado_geo['entrada_hora'] = hora_leitura

                            if estado_geo['entrada_contagem'] >= LEITURAS_CONFIRMAR:
                                hora_chegada = estado_geo.get('entrada_hora') or hora_leitura
                                execute_db(
                                    "INSERT INTO rastreio_paradas (data, placa, local, hora_chegada) VALUES (:data, :placa, :local, :hora)",
                                    {"data": DATA_HOJE_REAL_STR, "placa": placa_v, "local": candidato, "hora": hora_chegada},
                                )
                                estado_geo.update({
                                    'entrada_local': None, 'entrada_contagem': 0, 'entrada_hora': '',
                                    'saida_contagem': 0, 'saida_hora': '',
                                })
                        else:
                            estado_geo['entrada_local'] = None
                            estado_geo['entrada_contagem'] = 0
                            estado_geo['entrada_hora'] = ''
    except Exception:
        # Rastreador não pode derrubar o restante do sistema; tenta novamente no próximo ciclo.
        pass

# =====================================================================
# INTERFACE STREAMLIT
# =====================================================================
renderizar_cabecalho_torre()

# NAVEGAÇÃO PRIMEIRO: nenhuma consulta de rede/banco deve bloquear a troca de módulo.
MODULOS_PRINCIPAIS = [
    "🗺️ Roteiro do Davi",
    "📡 Rastreador ao vivo",
    "📦 Demandas ativas",
    "📋 Histórico e concluídos",
    "📍 Endereços",
    "🚗 Frota e custos",
]
with st.sidebar:
    logo_sidebar = _html_logo_aproar("aproar-logo-sidebar")
    st.markdown(f"""
        <div class="aproar-sidebar-brand">
            {logo_sidebar}
            <div><strong>Painel de operações</strong><small>Central logística</small></div>
        </div>
        <div class="aproar-sidebar-section">Navegação</div>
    """, unsafe_allow_html=True)
    modulo_principal = st.radio(
        "Módulo", MODULOS_PRINCIPAIS, index=0,
        key="modulo_principal", label_visibility="collapsed",
    )
modulo_principal = modulo_principal or MODULOS_PRINCIPAIS[0]

if "demandas" not in st.session_state: st.session_state.demandas = pd.DataFrame(columns=COLUNAS_DEMANDAS)

with st.sidebar:
    st.markdown('<div class="aproar-sidebar-section">Sincronização e ajustes</div>', unsafe_allow_html=True)
    st.caption(f"📅 Planejamento ativo para: **{DATA_REF_ROTA_STR}**")
    # Sincronização automática leve: Trello + baixas + Teams a cada 2 minutos.
    # O primeiro carregamento NÃO consulta o Trello, para a página abrir rápido.
    # A partir do próximo ciclo, o fragmento roda isoladamente e só força um rerun
    # completo quando a lista/prioridade das demandas realmente mudou.
    if "ultima_sincronizacao" not in st.session_state:
        st.session_state.ultima_sincronizacao = time.time()

    if hasattr(st, "fragment"):
        @st.fragment(run_every="120s")
        def _ciclo_trello_teams_2min():
            agora_ts = time.time()

            # st.fragment executa imediatamente ao ser criado. Na primeira passagem
            # apenas armamos o relógio; assim o cold start continua rápido.
            if not st.session_state.get("_ciclo_trello_2min_armado"):
                st.session_state["_ciclo_trello_2min_armado"] = True
                st.session_state["_ultimo_ciclo_trello_2min_ts"] = agora_ts
                return

            ultimo = float(st.session_state.get("_ultimo_ciclo_trello_2min_ts", 0) or 0)
            if agora_ts - ultimo < 110:
                return
            st.session_state["_ultimo_ciclo_trello_2min_ts"] = agora_ts

            try:
                df_antes = st.session_state.demandas.copy() if isinstance(st.session_state.get("demandas"), pd.DataFrame) else pd.DataFrame()
                cols_sig = [c for c in ["id", "Obra", "Origem", "Destino", "Urgência", "Peso"] if c in df_antes.columns]
                sig_antes = json.dumps(
                    df_antes[cols_sig].fillna("").sort_values("id").to_dict("records") if cols_sig and "id" in cols_sig else [],
                    ensure_ascii=False, sort_keys=True, default=str,
                )

                # Consulta o quadro uma única vez. No ciclo automático não geocodifica
                # fornecedores novos: isso fica para sincronização manual/aba Endereços
                # e evita travamentos periódicos por chamadas externas.
                sincronizou = sincronizar_demandas(forcar=True, geocodificar=False)
                if not sincronizou:
                    st.session_state["_erro_ciclo_trello_2min"] = "Não foi possível consultar o Trello neste ciclo."
                    return

                # Usa o cache recém-gravado pela chamada acima para registrar baixas e
                # enviar/repetir notificações pendentes ao Teams. Não toca na Protege.
                loop_automacoes_background(processar_rastreador=False)

                df_depois = st.session_state.demandas.copy()
                cols_sig2 = [c for c in ["id", "Obra", "Origem", "Destino", "Urgência", "Peso"] if c in df_depois.columns]
                sig_depois = json.dumps(
                    df_depois[cols_sig2].fillna("").sort_values("id").to_dict("records") if cols_sig2 and "id" in cols_sig2 else [],
                    ensure_ascii=False, sort_keys=True, default=str,
                )

                st.session_state["_ultima_rotina_auto"] = datetime.now(FUSO_LOCAL).strftime("%H:%M:%S")
                st.session_state.pop("_erro_ciclo_trello_2min", None)

                if sig_depois != sig_antes:
                    st.session_state["_recalcular_rota_automatico"] = True
                    # Só atualiza a página inteira quando houve mudança real.
                    st.rerun(scope="app")
            except TypeError:
                # Compatibilidade com versões do Streamlit sem scope="app".
                st.rerun()
            except Exception as erro_ciclo:
                st.session_state["_erro_ciclo_trello_2min"] = str(erro_ciclo)[:220]

        _ciclo_trello_teams_2min()

    ultimo_auto = st.session_state.get("_ultima_rotina_auto")
    if ultimo_auto:
        st.caption(f"🔄 Trello + Teams automáticos a cada 2 min • último ciclo: **{ultimo_auto}**")
    else:
        st.caption("🔄 Trello + Teams automáticos a cada 2 min • primeiro ciclo em até 2 min")
    if st.session_state.get("_erro_ciclo_trello_2min"):
        st.caption(f"⚠️ Último ciclo automático: {st.session_state['_erro_ciclo_trello_2min']}")

    st.markdown("---")
    st.markdown("📱 **App do Motorista**")
    st.components.v1.html("""<script>function copyLink() { try { var tempInput = document.createElement("input"); tempInput.value = window.parent.location.origin + "/davi"; document.body.appendChild(tempInput); tempInput.select(); document.execCommand("copy"); document.body.removeChild(tempInput); var btn = document.getElementById("btn"); btn.innerText = "✅ Copiado!"; btn.style.background = "linear-gradient(135deg, #16a34a, #15803d)"; btn.style.color = "white"; btn.style.border = "none"; setTimeout(() => { btn.innerText = "🔗 Copiar link do Davi"; btn.style.background = "transparent"; btn.style.color = "#8da0b8"; btn.style.border = "1px solid rgba(64,116,146,.35)"; }, 2500); } catch (err) { alert("Erro ao copiar."); } }</script><button id="btn" onclick="copyLink()" style="width:100%; padding:10px; background-color:transparent; color:#8da0b8; border:1px solid rgba(64,116,146,.35); border-radius:8px; font-family:sans-serif; font-size:14px; font-weight:bold; cursor:pointer; transition: all 0.2s;">🔗 Copiar link do Davi</button>""", height=50)
    st.markdown("---")

    if st.button("🔄 Sincronizar manualmente (Trello)", use_container_width=True, type="primary"):
        with st.spinner("Puxando demandas ao vivo..."):
            if sincronizar_demandas(manual=True, forcar=True, geocodificar=True):
                # Processa imediatamente qualquer baixa recém-lida e tenta o Teams,
                # sem esperar o próximo ciclo automático de 2 minutos.
                loop_automacoes_background(processar_rastreador=False)
                # A sincronização manual é uma ação explícita do operador; portanto,
                # a rota deve incorporar imediatamente novos prazos/demandas. Isso
                # não volta a pesar na inicialização porque só acontece após o clique.
                st.session_state["_recalcular_rota_automatico"] = True
                st.session_state["_mensagem_ajuste_rota"] = "✅ Trello sincronizado. A rota foi recalculada com os prazos atuais."
                st.rerun()
    
    @fragmento_independente
    def controles_planejamento_rota():
        st.divider()
        # Migração da opção antiga: ela representava carro próprio, então passa a usar
        # automaticamente a nova tarifa correta de R$ 2,50/km.
        if st.session_state.get("cfg_veiculo_rota") == "Carro Próprio/Frete (R$ 1,50/km)":
            st.session_state["cfg_veiculo_rota"] = "Carro Próprio/Frete (R$ 2,50/km)"
        st.radio(
            "🚗 Tipo de custeio da rota",
            [
                "Frota da Empresa (Calcula Gasolina)",
                "Moto Própria/Frete (R$ 1,50/km)",
                "Carro Próprio/Frete (R$ 2,50/km)",
            ],
            key="cfg_veiculo_rota",
        )
        st.divider()
        st.selectbox("🏁 Ponto de saída", ["ESCRITÓRIO", "CASA DA INDÚSTRIA", "SENAI CENTRO", "MARACANAÚ"], key="cfg_ponto_saida")
        estrategia_atual = st.selectbox("🎯 Estratégia da rota", ["⚖️ Equilibrada", "🏢 Foco em Descarregar", "⛽ Menor Distância", "🚨 Priorizar Urgências"], key="cfg_estrategia_rota")
        st.caption(f"ℹ️ *{ {'⚖️ Equilibrada': 'Prioriza primeiro os prazos críticos (vencidas/hoje) e, entre opções de prioridade semelhante, equilibra proximidade, tempo e carga.', '🏢 Foco em Descarregar': 'Prioriza as entregas para reduzir o volume de materiais transportados no veículo.', '⛽ Menor Distância': 'Prioriza a menor distância percorrida, com foco na economia de combustível.', '🚨 Priorizar Urgências': 'Dá peso máximo às demandas vencidas ou programadas para hoje.'}[estrategia_atual] }*")
        st.checkbox("Retornar à base no fim do dia", value=True, key="cfg_retornar_base")

    controles_planejamento_rota()

veiculo_selecionado = st.session_state.get("cfg_veiculo_rota", "Frota da Empresa (Calcula Gasolina)")

# Tarifas de ressarcimento/frete para veículo próprio. Mantidas separadas da análise
# de custo da frota da empresa, que continua usando combustível/manutenção reais.
if "Moto Própria/Frete" in veiculo_selecionado:
    valor_km_veiculo_proprio = 1.50
elif "Carro Próprio/Frete" in veiculo_selecionado:
    valor_km_veiculo_proprio = 2.50
else:
    valor_km_veiculo_proprio = None

ponto_saida = st.session_state.get("cfg_ponto_saida", "ESCRITÓRIO")
estrategia = st.session_state.get("cfg_estrategia_rota", "⚖️ Equilibrada")
retornar_base = st.session_state.get("cfg_retornar_base", True)

if st.session_state.demandas.empty and not st.session_state.get('rota_gerada', False):
    st.info("👋 Bem-vindo(a) à Torre de Controle! Clique no botão **'🔄 Sincronizar manualmente'** no menu lateral para puxar as demandas ao vivo e começar.")
elif st.session_state.demandas.empty and st.session_state.get('rota_gerada', False):
    st.caption("⚡ Rota salva carregada. As demandas ativas serão consultadas automaticamente quando forem necessárias para atualizar o planejamento.")

if modulo_principal == "📡 Rastreador ao vivo":
    st.subheader("📡 Rastreador ao vivo — Protege Express")
    st.caption("Posições consultadas diretamente no portal. Atualização automática a cada 30 segundos.")

    usuario_protege, senha_protege, ids_veiculos = carregar_config_protege()

    if not usuario_protege or not senha_protege:
        st.warning("Configure o usuário e a senha da Protege Express nos Secrets do aplicativo para ativar o login automático.")
    else:
        def exibir_painel_rastreador():
            col_status, col_atualizar = st.columns([4, 1])
            col_status.success("🔒 Login automático configurado")
            if col_atualizar.button("🔄 Reconectar", key="btn_reconectar_protege", use_container_width=True): st.session_state.pop("protege_sessao", None); st.session_state.pop("protege_pagina", None)

            try:
                sessao, pagina = st.session_state.get("protege_sessao"), st.session_state.get("protege_pagina")
                if not sessao or not pagina:
                    sessao, pagina, posicoes = autenticar_protege(usuario_protege, senha_protege, ids_veiculos)
                    st.session_state["protege_sessao"], st.session_state["protege_pagina"] = sessao, pagina
                else:
                    try: posicoes = consultar_posicoes_protege(sessao, pagina, ids_veiculos)
                    except:
                        sessao, pagina, posicoes = autenticar_protege(usuario_protege, senha_protege, ids_veiculos)
                        st.session_state["protege_sessao"], st.session_state["protege_pagina"] = sessao, pagina

                start_times = {row[0]: row[1] for row in fetch_all("SELECT placa, hora_inicio FROM inicio_movimento WHERE data=:data", {"data": DATA_HOJE_REAL_STR})}

                st.markdown("#### ✏️ Corrigir início da rota")
                st.caption(
                    "Use quando o rastreador tiver sido consultado depois da saída. "
                    "O horário salvo aqui passa a valer no roteiro e nas previsões do Davi."
                )
                placas_disponiveis = sorted({str(p.get("Placa", "")).strip() for p in posicoes if p.get("Placa")})
                if PLACA_DAVI not in placas_disponiveis:
                    placas_disponiveis.insert(0, PLACA_DAVI)

                col_placa, col_hora, col_salvar = st.columns([1.2, 1, 1])
                placa_manual = col_placa.selectbox(
                    "Veículo",
                    placas_disponiveis,
                    index=placas_disponiveis.index(PLACA_DAVI) if PLACA_DAVI in placas_disponiveis else 0,
                    key="placa_inicio_manual",
                )
                hora_registrada = str(start_times.get(placa_manual, HORA_INICIO_ROTA_DAVI))
                try:
                    hora_padrao = datetime.strptime(hora_registrada, "%H:%M").time()
                except (TypeError, ValueError):
                    hora_padrao = datetime.strptime(HORA_INICIO_ROTA_DAVI, "%H:%M").time()
                hora_manual = col_hora.time_input(
                    "Horário real de saída",
                    value=hora_padrao,
                    step=timedelta(minutes=1),
                    key=f"hora_inicio_manual_{placa_manual}",
                )
                col_salvar.write("")
                col_salvar.write("")
                if col_salvar.button("💾 Salvar horário", type="primary", use_container_width=True, key="salvar_inicio_manual"):
                    hora_manual_str = hora_manual.strftime("%H:%M")
                    execute_db(
                        """
                        INSERT INTO inicio_movimento (placa, data, hora_inicio)
                        VALUES (:placa, :data, :hora)
                        ON CONFLICT (placa, data)
                        DO UPDATE SET hora_inicio=EXCLUDED.hora_inicio
                        """,
                        {"placa": placa_manual, "data": DATA_HOJE_REAL_STR, "hora": hora_manual_str},
                    )
                    st.session_state["confirmacao_inicio_manual"] = (
                        f"✅ Início da rota de {placa_manual} corrigido para {hora_manual_str}."
                    )
                    st.rerun()

                confirmacao_inicio = st.session_state.pop("confirmacao_inicio_manual", "")
                if confirmacao_inicio:
                    st.success(confirmacao_inicio)

                for p in posicoes: p['🟢 Início da Rota (Hoje)'] = start_times.get(p['Placa'], "Ainda não saiu (raio de 500 m)")

                velocidades = [p["Velocidade (km/h)"] for p in posicoes]
                met1, met2, met3 = st.columns(3)
                met1.metric("Veículos localizados", len(posicoes))
                met2.metric("Em movimento", sum(1 for v in velocidades if v > 0))
                met3.metric("Última leitura", datetime.now(FUSO_LOCAL).strftime("%H:%M:%S"))

                # MAPA CLARO (OPENSTREETMAP) NO RASTREADOR
                mapa = folium.Map(location=[sum(p["Latitude"] for p in posicoes) / len(posicoes), sum(p["Longitude"] for p in posicoes) / len(posicoes)], zoom_start=11, tiles="OpenStreetMap")
                limites = []
                for p in posicoes:
                    cor, icone = ("green", "play") if p["Velocidade (km/h)"] > 0 else ("red", "stop")
                    limites.append([p["Latitude"], p["Longitude"]])
                    folium.Marker([p["Latitude"], p["Longitude"]], popup=folium.Popup(f"<b>{p['Placa']}</b><br>{p['Situação']} — {p['Velocidade (km/h)']:.0f} km/h<br>Atualização: {p['Última atualização']}<br>{p['Endereço']}", max_width=360), tooltip=f"{p['Placa']} — {p['Situação']}", icon=folium.Icon(color=cor, icon=icone, prefix="fa")).add_to(mapa)

                if len(limites) > 1: mapa.fit_bounds(limites, padding=(35, 35))
                st_folium(mapa, height=520, use_container_width=True, returned_objects=[], key="mapa_rastreador_protege")
                df_posicoes_exportacao = pd.DataFrame(posicoes)[["Placa", "🟢 Início da Rota (Hoje)", "Última atualização", "Velocidade (km/h)", "Situação", "Endereço"]]
                st.dataframe(df_posicoes_exportacao, use_container_width=True, hide_index=True)
                _conteudo_exportador(
                    "Rastreador ao Vivo — Protege Express", df_posicoes_exportacao,
                    "rastreador_ao_vivo", "rastreador",
                )

            except:
                st.session_state.pop("protege_sessao", None); st.session_state.pop("protege_pagina", None)
                st.error("Não consegui entrar automaticamente no rastreador. Confira as credenciais.")

        if hasattr(st, "fragment"): st.fragment(run_every="30s")(exibir_painel_rastreador)()
        else: exibir_painel_rastreador()

if modulo_principal == "📦 Demandas ativas":
    st.subheader(f"Gerenciamento de cargas da rota ({DATA_REF_ROTA_STR})")

    @fragmento_independente
    def editor_tempos_demandas():
        st.session_state.demandas = st.data_editor(st.session_state.demandas, column_config={"Tempo_Coleta": st.column_config.NumberColumn("Tempo Coleta (min)", min_value=1, max_value=120), "Tempo_Entrega": st.column_config.NumberColumn("Tempo Entrega (min)", min_value=1, max_value=120), "Peso": None, "id": None, "Supervisor": None, "_Titulo_Trello": None}, disabled=["Obra", "Origem", "Destino", "Materiais", "Urgência"], hide_index=True, use_container_width=True, key="editor_tempos_demandas")

    editor_tempos_demandas()
    st.caption("⏱️ Os tempos de coleta/entrega representam a complexidade de cada demanda. Quando várias demandas acontecem no mesmo endereço, o sistema calcula uma única permanência no local — não soma 10 ou 20 minutos completos para cada cartão.")
    st.divider()
    st.subheader("📣 Monitoramento da rota atual (status do Trello)")
    st.caption("A baixa do Trello continua sendo automática a cada 2 minutos. O botão **Informar entrega** é apenas uma contingência caso o alerta automático não apareça no Teams.")

    df_entregues_hoje = get_df(
        "SELECT id, hora_conclusao, teams_notificado, teams_ultimo_erro FROM historico_concluidos WHERE data_conclusao = :data",
        {"data": DATA_REF_ROTA_STR},
    )
    dict_concluidos_monitor = dict(zip(df_entregues_hoje['id'].astype(str), df_entregues_hoje['hora_conclusao']))
    status_teams_monitor = {
        str(r['id']): {
            "notificado": r.get('teams_notificado'),
            "erro": r.get('teams_ultimo_erro', ''),
        }
        for _, r in df_entregues_hoje.iterrows()
    } if not df_entregues_hoje.empty else {}
    demandas_na_rota = {str(t.get('id', '')): t for step in st.session_state.get('route_steps', []) for acao, t in step.get('actions', [])}

    feedback_manual = st.session_state.pop("_feedback_informar_entrega", None)
    if feedback_manual:
        ok_feedback, texto_feedback = feedback_manual
        (st.success if ok_feedback else st.warning)(texto_feedback)

    if not demandas_na_rota: st.info("Gere uma rota na aba 'Roteiro do Davi' para monitorar o status das entregas aqui.")
    else:
        for card_id, row in demandas_na_rota.items():
            c1, c_status, c_manual = st.columns([3.4, 2.15, 1.55])
            c1.markdown(f"📦 **{row.get('Obra', '')} — {row.get('Destino', '')}** (Responsável: {row.get('Supervisor', 'Sede')}) <br><span style='font-size:12px; color:gray;'>{row.get('Materiais', '')}</span>", unsafe_allow_html=True)

            info_teams = status_teams_monitor.get(card_id, {})
            teams_ok = info_teams.get("notificado") is True
            if card_id in dict_concluidos_monitor:
                c_status.success(f"✅ **Entregue às {dict_concluidos_monitor[card_id]}**")
                if teams_ok:
                    c_status.caption("📣 Teams avisado")
                elif info_teams.get("notificado") is False:
                    c_status.caption("⚠️ Teams pendente")
            else:
                c_status.warning("⏳ Pendente — no veículo")

            # Se o Teams já confirmou, evita duplicidade. Caso contrário, o botão fica
            # disponível tanto para uma entrega ainda sem baixa automática quanto para
            # uma baixa cuja notificação falhou.
            if c_manual.button(
                "📣 Informar entrega",
                key=f"informar_entrega_manual_{card_id}",
                use_container_width=True,
                disabled=teams_ok,
                help=(
                    "O Teams já confirmou este alerta." if teams_ok
                    else "Contingência manual. A automação do Trello continua prioritária e roda a cada 2 minutos."
                ),
            ):
                try:
                    enviado, detalhe = informar_entrega_manual_teams(card_id, row)
                    st.session_state["_feedback_informar_entrega"] = (enviado, detalhe)
                except Exception as erro_manual:
                    st.session_state["_feedback_informar_entrega"] = (False, f"Não foi possível informar a entrega: {erro_manual}")
                st.rerun()

            st.write("---")

    df_relatorio_demandas = st.session_state.demandas.copy()
    if not df_relatorio_demandas.empty:
        df_relatorio_demandas["Status da rota"] = df_relatorio_demandas["id"].astype(str).map(
            lambda card_id: f"Entregue às {dict_concluidos_monitor[card_id]}" if card_id in dict_concluidos_monitor else "Pendente"
        )
    renderizar_exportador(
        f"Demandas Ativas — {DATA_REF_ROTA_STR}", df_relatorio_demandas,
        "demandas_ativas", "demandas",
    )

if modulo_principal == "📋 Histórico e concluídos":
    st.subheader(f"📋 Entregas fisicamente concluídas ({DATA_HOJE_REAL_STR})")
    df_hist = get_df("SELECT * FROM historico_concluidos WHERE data_conclusao = :data ORDER BY id DESC", {"data": DATA_HOJE_REAL_STR})
    if df_hist.empty: st.info("Nenhuma entrega foi registrada como finalizada no Trello no dia de hoje.")
    else: st.dataframe(df_hist, use_container_width=True, hide_index=True)
    renderizar_exportador(
        f"Entregas Concluídas — {DATA_HOJE_REAL_STR}", df_hist,
        "entregas_concluidas", "historico",
    )

if modulo_principal == "📍 Endereços":
    @fragmento_independente
    def painel_enderecos():
        st.subheader("Locais e coordenadas GPS")
        mensagem_local = st.session_state.pop("mensagem_local", "")
        if mensagem_local:
            st.success(mensagem_local)

        col1, col2 = st.columns(2)
        with col1: apelido_input = st.text_input("Nome da loja ou do local (ex.: LECI FERRAGENS)").upper().strip()
        with col2: endereco_input = st.text_input("Endereço completo ou link do Google Maps").strip()
        if st.button("Salvar endereço e extrair GPS"):
            if apelido_input and endereco_input:
                lat, lon = buscar_coordenadas(endereco_input)
                if lat:
                    execute_db("DELETE FROM locais_removidos WHERE apelido = :apelido", {"apelido": apelido_input})
                    execute_db("INSERT INTO locais (apelido, endereco, lat, lon) VALUES (:apelido, :end, :lat, :lon) ON CONFLICT (apelido) DO UPDATE SET endereco=EXCLUDED.endereco, lat=EXCLUDED.lat, lon=EXCLUDED.lon", {"apelido": apelido_input, "end": endereco_input, "lat": lat, "lon": lon})
                    st.success(f"✅ GPS de '{apelido_input}' salvo com sucesso na nuvem!")
                else: st.error("❌ Não consegui localizar as coordenadas com esse texto. Cole o link direto do Google Maps.")
            else: st.warning("Preencha o nome e o endereço.")

        df_locais = get_df("SELECT * FROM locais ORDER BY apelido")
        st.dataframe(df_locais, use_container_width=True, hide_index=True)
        _conteudo_exportador(
            "Locais e coordenadas GPS", df_locais,
            "locais_e_enderecos", "enderecos",
        )
        st.divider()
        st.markdown("#### Remover local")
        locais_removiveis = [apelido for apelido in df_locais["apelido"].tolist() if apelido not in ALIASES_LOCAL_BASE]
        if locais_removiveis:
            local_remover = st.selectbox("Selecione o local que deseja remover", locais_removiveis, index=None, placeholder="Escolha um local...")
            confirmar_remocao = st.checkbox("Confirmo que desejo remover este local e seu GPS", key="confirmar_remocao_local")
            if st.button("🗑️ Remover local selecionado", disabled=not (local_remover and confirmar_remocao)):
                execute_db("INSERT INTO locais_removidos (apelido) VALUES (:apelido) ON CONFLICT (apelido) DO NOTHING", {"apelido": local_remover})
                execute_db("DELETE FROM locais WHERE apelido = :apelido", {"apelido": local_remover})
                st.session_state["mensagem_local"] = f"✅ Local '{local_remover}' removido."
                if hasattr(st, "fragment"):
                    st.rerun(scope="fragment")
                else:
                    st.rerun()

    painel_enderecos()

if modulo_principal == "🚗 Frota e custos":
    st.subheader("🚗 Frota e custos")
    st.caption("Custos, quilometragem, abastecimentos, manutenção e histórico operacional da frota em um só lugar.")

    SUBMODULOS_FROTA = [
        "📊 Resumo e lançamentos",
        "🕒 Operação e paradas",
        "🗂️ Histórico editável",
    ]
    if hasattr(st, "segmented_control"):
        submodulo_frota = st.segmented_control(
            "Seção da frota", SUBMODULOS_FROTA, default=SUBMODULOS_FROTA[0],
            key="submodulo_frota", label_visibility="collapsed",
        )
    else:
        submodulo_frota = st.radio(
            "Seção da frota", SUBMODULOS_FROTA, index=0, horizontal=True,
            key="submodulo_frota", label_visibility="collapsed",
        )
    submodulo_frota = submodulo_frota or SUBMODULOS_FROTA[0]

    if submodulo_frota == "📊 Resumo e lançamentos":
        cfg = get_df("SELECT consumo, preco_gasolina FROM config_frota WHERE id=1").iloc[0]

        @fragmento_independente
        def configuracao_base_frota():
            st.markdown("#### ⚙️ Parâmetros-base do veículo")
            cc1, cc2 = st.columns(2)
            novo_consumo_cfg = cc1.number_input("Consumo médio (km/L)", value=float(cfg['consumo']), step=0.1, key="cfg_consumo_frota")
            novo_preco_cfg = cc2.number_input("Preço-base da gasolina (R$/L)", value=float(cfg['preco_gasolina']), step=0.01, key="cfg_preco_gasolina")
            if st.button("Atualizar parâmetros"):
                execute_db("UPDATE config_frota SET consumo=:c, preco_gasolina=:p WHERE id=1", {"c": novo_consumo_cfg, "p": novo_preco_cfg})
                st.success("✅ Base de cálculo atualizada!")

        configuracao_base_frota()
        novo_preco = float(st.session_state.get("cfg_preco_gasolina", cfg['preco_gasolina']))

        st.divider()
        col_recibo, col_km = st.columns(2)
        with col_recibo:
            st.markdown("#### ⛽ Lançar recibo de gasto")

            @fragmento_independente
            def formulario_recibo():
                with st.form("form_recibo", clear_on_submit=True):
                    f_data = st.date_input("Data do recibo")
                    fc_veic = st.selectbox("Veículo do gasto", ["Strada", "L200"])
                    fc1, fc2 = st.columns(2)
                    f_litros = fc1.number_input("Litros abastecidos", min_value=0.0, step=0.1)
                    f_valor = fc2.number_input("Preço pago (R$/L)", value=novo_preco, step=0.01)
                    f_manut = st.number_input("Gastos com manutenção (R$)", min_value=0.0, step=10.0)
                    f_obs = st.text_input("Observação (ex.: Posto Ipiranga, troca de óleo)")
                    if st.form_submit_button("Lançar no caixa"):
                        execute_db("INSERT INTO abastecimentos (data, litros, valor_litro, manutencao, obs, veiculo) VALUES (:data, :litros, :valor, :manut, :obs, :veic)", {"data": f_data.strftime("%d/%m/%Y"), "litros": f_litros, "valor": f_valor, "manut": f_manut, "obs": f_obs, "veic": fc_veic})
                        carregar_abastecimentos_df.clear()
                        st.success("Recibo salvo com sucesso!")

            formulario_recibo()

        with col_km:
            st.markdown("#### 🛣️ Lançar quilometragem avulsa")

            @fragmento_independente
            def formulario_km_avulso():
                with st.form("form_km", clear_on_submit=True):
                    k_data = st.date_input("Data do deslocamento")
                    k_veic = st.selectbox("Veículo utilizado", ["Strada", "L200"])
                    k_km = st.number_input("Quilometragem total rodada", min_value=0.1, step=1.0)
                    k_obs = st.text_input("Motivo (ex.: ida ao banco, frete extra)")
                    if st.form_submit_button("Lançar quilometragem"):
                        execute_db("INSERT INTO registro_km (data, km, obs, veiculo) VALUES (:data, :km, :obs, :veic)", {"data": k_data.strftime("%d/%m/%Y"), "km": k_km, "obs": k_obs, "veic": k_veic})
                        carregar_registro_km_df.clear()
                        st.success(f"Quilometragem de {k_km} km salva com sucesso!")

            formulario_km_avulso()

        st.divider()
        st.markdown("#### 📅 Lançamento de fechamento de quilometragem (período)")

        @fragmento_independente
        def formulario_fechamento_km():
            with st.form("form_fechamento_km", clear_on_submit=True):
                col_f1, col_f2 = st.columns([1, 2])
                f_veic = col_f1.selectbox("Veículo do fechamento", ["Strada", "L200"])
                f_obs = col_f2.text_input("Observação (ex.: quinzena 1, fechamento mensal)")
                
                col_f3, col_f4, col_f5, col_f6 = st.columns(4)
                f_data_ini = col_f3.date_input("Data inicial")
                f_km_ini = col_f4.number_input("Quilometragem inicial", min_value=0.0, step=1.0)
                f_data_fin = col_f5.date_input("Data final")
                f_km_fin = col_f6.number_input("Quilometragem final", min_value=0.0, step=1.0)
                
                if st.form_submit_button("Calcular e lançar fechamento"):
                    km_rodado = f_km_fin - f_km_ini
                    if km_rodado > 0:
                        obs_final = f"Fechamento ({f_data_ini.strftime('%d/%m')} a {f_data_fin.strftime('%d/%m')}) - {f_obs}"
                        execute_db("INSERT INTO registro_km (data, km, obs, veiculo) VALUES (:data, :km, :obs, :veic)", {"data": f_data_fin.strftime("%d/%m/%Y"), "km": km_rodado, "obs": obs_final, "veic": f_veic})
                        carregar_registro_km_df.clear()
                        st.success(f"✅ Quilometragem calculada: {km_rodado:.1f} km. Lançamento salvo com sucesso para a {f_veic}.")
                    else:
                        st.warning("⚠️ A quilometragem final precisa ser maior que a quilometragem inicial para calcular o trecho.")

        formulario_fechamento_km()

        st.divider()
        # ===============================================================
        # PAINEL GERENCIAL DE FECHAMENTO DA FROTA — MÊS ATUAL
        # ===============================================================
        nomes_meses_pt = {
            1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
            5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
            9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
        }
        mes_atual_str = AGORA_REAL.strftime("%m/%Y")
        mes_atual_nome = f"{nomes_meses_pt.get(AGORA_REAL.month, AGORA_REAL.strftime('%m'))} de {AGORA_REAL.year}"

        st.markdown(f"### 📊 Fechamento da Frota — {mes_atual_nome}")
        st.caption("Visão gerencial do mês atual com quilometragem, combustível, manutenção e custo real por quilômetro de cada veículo.")

        df_km = carregar_registro_km_df().copy()
        if 'veiculo' not in df_km.columns:
            df_km['veiculo'] = 'Strada'
        df_km['data_dt'] = pd.to_datetime(df_km['data'], format="%d/%m/%Y", errors='coerce')
        df_km['km'] = pd.to_numeric(df_km.get('km', 0), errors='coerce').fillna(0)
        df_km_mes = df_km.dropna(subset=['data_dt']).copy()
        df_km_mes = df_km_mes[df_km_mes['data_dt'].dt.strftime('%m/%Y') == mes_atual_str].copy()

        df_abastec = carregar_abastecimentos_df().copy()
        if 'veiculo' not in df_abastec.columns:
            df_abastec['veiculo'] = 'Strada'
        df_abastec['data_dt'] = pd.to_datetime(df_abastec['data'], format="%d/%m/%Y", errors='coerce')
        for coluna_num in ['litros', 'valor_litro', 'manutencao']:
            if coluna_num not in df_abastec.columns:
                df_abastec[coluna_num] = 0.0
            df_abastec[coluna_num] = pd.to_numeric(df_abastec[coluna_num], errors='coerce').fillna(0.0)
        df_abastec_mes = df_abastec.dropna(subset=['data_dt']).copy()
        df_abastec_mes = df_abastec_mes[df_abastec_mes['data_dt'].dt.strftime('%m/%Y') == mes_atual_str].copy()
        df_abastec_mes['custo_combustivel'] = df_abastec_mes['litros'] * df_abastec_mes['valor_litro']
        df_abastec_mes['custo_total'] = df_abastec_mes['custo_combustivel'] + df_abastec_mes['manutencao']

        def resumo_veiculo_mes(veiculo):
            km_df = df_km_mes[df_km_mes['veiculo'].astype(str) == veiculo].copy()
            gasto_df = df_abastec_mes[df_abastec_mes['veiculo'].astype(str) == veiculo].copy()

            km = float(km_df['km'].sum()) if not km_df.empty else 0.0
            litros = float(gasto_df['litros'].sum()) if not gasto_df.empty else 0.0
            combustivel = float(gasto_df['custo_combustivel'].sum()) if not gasto_df.empty else 0.0
            manutencao = float(gasto_df['manutencao'].sum()) if not gasto_df.empty else 0.0
            custo_total = combustivel + manutencao
            custo_km = custo_total / km if km > 0 else None
            preco_medio_litro = combustivel / litros if litros > 0 else None
            abastecimentos = int((gasto_df['litros'] > 0).sum()) if not gasto_df.empty else 0
            manutencoes = int((gasto_df['manutencao'] > 0).sum()) if not gasto_df.empty else 0

            return {
                'veiculo': veiculo,
                'km': km,
                'litros': litros,
                'combustivel': combustivel,
                'manutencao': manutencao,
                'custo_total': custo_total,
                'custo_km': custo_km,
                'preco_medio_litro': preco_medio_litro,
                'abastecimentos': abastecimentos,
                'manutencoes': manutencoes,
                'gastos_df': gasto_df,
                'km_df': km_df,
            }

        resumos_veiculos = {veiculo: resumo_veiculo_mes(veiculo) for veiculo in ['Strada', 'L200']}
        resumo_strada = resumos_veiculos['Strada']
        resumo_l200 = resumos_veiculos['L200']

        km_total_frota = sum(item['km'] for item in resumos_veiculos.values())
        litros_total_frota = sum(item['litros'] for item in resumos_veiculos.values())
        combustivel_total_frota = sum(item['combustivel'] for item in resumos_veiculos.values())
        manutencao_total_frota = sum(item['manutencao'] for item in resumos_veiculos.values())
        custo_total_frota = combustivel_total_frota + manutencao_total_frota
        custo_km_frota = custo_total_frota / km_total_frota if km_total_frota > 0 else None
        preco_medio_frota = combustivel_total_frota / litros_total_frota if litros_total_frota > 0 else None
        lancamentos_total = len(df_abastec_mes) + len(df_km_mes)

        # -------- Resumo consolidado --------
        st.markdown("#### 🧾 Resumo do mês")
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("🛣️ Quilometragem total", f"{km_total_frota:,.1f} km".replace(',', 'X').replace('.', ',').replace('X', '.'))
        k2.metric("💳 Custo total", f"R$ {custo_total_frota:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        k3.metric("📉 Custo médio por km", f"R$ {custo_km_frota:.2f}".replace('.', ',') if custo_km_frota is not None else "—",
                  "Dentro da referência" if custo_km_frota is not None and custo_km_frota <= 1.50 else "Acima de R$ 1,50/km" if custo_km_frota is not None else "Sem quilometragem lançada",
                  delta_color="normal" if custo_km_frota is not None and custo_km_frota <= 1.50 else "inverse")
        k4.metric("⛽ Combustível", f"R$ {combustivel_total_frota:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))

        k5, k6, k7, k8 = st.columns(4)
        k5.metric("🔧 Manutenção", f"R$ {manutencao_total_frota:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        k6.metric("🧪 Litros abastecidos", f"{litros_total_frota:,.1f} L".replace(',', 'X').replace('.', ',').replace('X', '.'))
        k7.metric("🏷️ Preço médio por litro", f"R$ {preco_medio_frota:.2f}".replace('.', ',') if preco_medio_frota is not None else "—")
        k8.metric("🧾 Lançamentos", str(lancamentos_total))

        if custo_total_frota <= 0 and km_total_frota <= 0:
            st.info("Ainda não há lançamentos suficientes neste mês para montar o fechamento da frota.")

        # -------- Comparação por veículo --------
        st.markdown("#### 🚘 Comparativo por veículo")

        veiculos_validos_custo = [
            r for r in resumos_veiculos.values()
            if r['custo_km'] is not None and r['custo_total'] > 0
        ]
        mais_economico = min(veiculos_validos_custo, key=lambda r: r['custo_km'])['veiculo'] if veiculos_validos_custo else None

        def formatar_moeda_br(valor):
            return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        def formatar_numero_br(valor, casas=1):
            return f"{float(valor):,.{casas}f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        def renderizar_cartao_veiculo(resumo, icone, chave):
            veiculo = resumo['veiculo']
            custo_km = resumo['custo_km']
            dentro_referencia = custo_km is not None and custo_km <= 1.50
            participacao = (resumo['custo_total'] / custo_total_frota) if custo_total_frota > 0 else 0.0
            eh_mais_economico = mais_economico == veiculo and len(veiculos_validos_custo) > 1

            with st.container(border=True):
                # Cabeçalho simples e largo para evitar truncamento.
                st.markdown(f"### {icone} {veiculo}")
                if custo_km is None:
                    st.caption("⚪ Sem quilometragem suficiente para calcular o custo por quilômetro")
                elif dentro_referencia:
                    st.caption("🟢 Custo por km dentro da referência de R$ 1,50/km")
                else:
                    st.caption("🔴 Custo por km acima da referência de R$ 1,50/km")
                if eh_mais_economico:
                    st.success("🏆 Menor custo por km da frota neste mês")

                # Evita st.metric aqui: em meia tela o componente corta valores com "...".
                # Os indicadores são exibidos em duas colunas com texto normal, que quebra linha
                # quando necessário e nunca esconde o valor.
                st.markdown("**Indicadores principais**")

                m1, m2 = st.columns(2)
                with m1:
                    with st.container(border=True):
                        st.caption("💳 Custo total")
                        st.markdown(f"### {formatar_moeda_br(resumo['custo_total'])}")
                with m2:
                    with st.container(border=True):
                        st.caption("🛣️ Quilometragem no mês")
                        st.markdown(f"### {formatar_numero_br(resumo['km'])} km")

                m3, m4 = st.columns(2)
                with m3:
                    with st.container(border=True):
                        st.caption("📉 Custo por km")
                        valor_custo_km = formatar_moeda_br(custo_km) if custo_km is not None else "—"
                        st.markdown(f"### {valor_custo_km}")
                with m4:
                    with st.container(border=True):
                        st.caption("📊 Participação no custo da frota")
                        st.markdown(f"### {participacao * 100:.1f}%" if custo_total_frota > 0 else "### —")

                # Demais informações ficam compactas para caber tudo na tela.
                st.markdown("**Resumo do veículo**")
                r1, r2 = st.columns(2)
                with r1:
                    st.markdown(f"⛽ **Combustível:** {formatar_moeda_br(resumo['combustivel'])}")
                    st.markdown(f"🧪 **Litros:** {formatar_numero_br(resumo['litros'])} L")
                with r2:
                    st.markdown(f"🔧 **Manutenção:** {formatar_moeda_br(resumo['manutencao'])}")
                    media_litro = formatar_moeda_br(resumo['preco_medio_litro']) if resumo['preco_medio_litro'] is not None else "—"
                    st.markdown(f"🏷️ **Preço médio por litro:** {media_litro}")

                abastec_txt = f"{resumo['abastecimentos']} abastecimento" + ("s" if resumo['abastecimentos'] != 1 else "")
                manut_txt = (
                    f"{resumo['manutencoes']} manutenção"
                    if resumo['manutencoes'] == 1
                    else f"{resumo['manutencoes']} manutenções"
                )
                st.caption(f"🧾 {abastec_txt}  •  🛠️ {manut_txt}")

                if custo_total_frota > 0:
                    st.progress(
                        min(max(participacao, 0.0), 1.0),
                        text=f"{participacao * 100:.1f}% do custo total da frota",
                    )

                gastos = resumo['gastos_df'].copy()
                quilometragem = resumo['km_df'].copy()
                if not gastos.empty:
                    gastos = gastos[["data", "litros", "valor_litro", "custo_combustivel", "manutencao", "custo_total", "obs"]].rename(columns={
                        "data": "Data", "litros": "Litros", "valor_litro": "Valor/L (R$)",
                        "custo_combustivel": "Combustível (R$)", "manutencao": "Manutenção (R$)",
                        "custo_total": "Total (R$)", "obs": "Observação",
                    })
                    for coluna in ["Valor/L (R$)", "Combustível (R$)", "Manutenção (R$)", "Total (R$)"]:
                        gastos[coluna] = pd.to_numeric(gastos[coluna], errors="coerce").fillna(0).round(2)
                if not quilometragem.empty:
                    quilometragem = quilometragem[["data", "km", "obs"]].rename(columns={"data": "Data", "km": "km", "obs": "Observação"})

                renderizar_detalhes_fechamento(veiculo, gastos, quilometragem, chave)
                return gastos, quilometragem

        col_strada, col_l200 = st.columns(2)
        with col_strada:
            gastos_strada_mes, kms_strada_mes = renderizar_cartao_veiculo(resumo_strada, "🚗", "strada")
        with col_l200:
            gastos_l200_mes, kms_l200_mes = renderizar_cartao_veiculo(resumo_l200, "🚙", "l200")

        # -------- Gráfico rápido --------
        if custo_total_frota > 0:
            st.markdown("#### 📈 Composição dos custos")
            df_composicao_custos = pd.DataFrame({
                "Veículo": ["Strada", "L200"],
                "Combustível": [resumo_strada['combustivel'], resumo_l200['combustivel']],
                "Manutenção": [resumo_strada['manutencao'], resumo_l200['manutencao']],
            }).set_index("Veículo")
            st.bar_chart(df_composicao_custos, use_container_width=True)

        # -------- Leitura gerencial --------
        alertas_frota = []
        for resumo in resumos_veiculos.values():
            if resumo['custo_total'] > 0 and resumo['km'] <= 0:
                alertas_frota.append(f"⚠️ **{resumo['veiculo']}** tem custos lançados, mas ainda não possui quilometragem no mês; o custo por km não pode ser calculado.")
            if resumo['custo_km'] is not None and resumo['custo_km'] > 1.50:
                alertas_frota.append(f"🔴 **{resumo['veiculo']}** está em **{formatar_moeda_br(resumo['custo_km'])}/km**, acima da referência de R$ 1,50/km.")
            if resumo['custo_total'] > 0 and resumo['manutencao'] / resumo['custo_total'] >= 0.35:
                perc_manut = resumo['manutencao'] / resumo['custo_total'] * 100
                alertas_frota.append(f"🔧 Manutenção representa **{perc_manut:.0f}%** do custo da **{resumo['veiculo']}** neste mês.")

        if mais_economico and len(veiculos_validos_custo) > 1:
            melhor = resumos_veiculos[mais_economico]
            outro = next(r for r in veiculos_validos_custo if r['veiculo'] != mais_economico)
            diferenca = outro['custo_km'] - melhor['custo_km']
            if diferenca > 0:
                st.success(
                    f"🏆 **{mais_economico}** está com o menor custo por km do mês: "
                    f"**{formatar_moeda_br(melhor['custo_km'])}/km**, economia de aproximadamente "
                    f"**{formatar_moeda_br(diferenca)}/km** em relação ao outro veículo."
                )

        for alerta in alertas_frota:
            st.warning(alerta)

        df_resumo_fechamento = pd.DataFrame([
            {
                "Veículo": resumo['veiculo'],
                "Quilometragem (km)": round(resumo['km'], 2),
                "Litros": round(resumo['litros'], 2),
                "Preço médio por litro (R$)": round(resumo['preco_medio_litro'], 2) if resumo['preco_medio_litro'] is not None else 0.0,
                "Combustível (R$)": round(resumo['combustivel'], 2),
                "Manutenção (R$)": round(resumo['manutencao'], 2),
                "Custo total (R$)": round(resumo['custo_total'], 2),
                "Custo por km (R$)": round(resumo['custo_km'], 2) if resumo['custo_km'] is not None else 0.0,
                "Abastecimentos": resumo['abastecimentos'],
                "Manutenções": resumo['manutencoes'],
            }
            for resumo in resumos_veiculos.values()
        ])

        renderizar_exportador(
            f"Fechamento Individualizado — {mes_atual_str}",
            {
                "Resumo": df_resumo_fechamento,
                "Gastos Strada": gastos_strada_mes,
                "Quilometragem Strada": kms_strada_mes,
                "Gastos L200": gastos_l200_mes,
                "Quilometragem L200": kms_l200_mes,
            },
            "fechamento_mensal_frota", "custos",
        )

    if submodulo_frota == "🕒 Operação e paradas":
        st.markdown("### 🕒 Operação do rastreador")
        st.caption("Saídas do pátio e permanência nas obras registradas automaticamente pelo rastreador.")
        st.markdown("#### 🕒 Horários da operação (rastreador)")
        c_inicio, c_paradas = st.columns([1, 1.8])

        with c_inicio:
            st.markdown("**🏁 Início da rota (saídas do pátio)**")
            st.caption("Marcado quando o veículo se afasta a mais de 500 m do escritório.")
            df_inicio = get_df("SELECT data as Data, placa as Placa, hora_inicio as \"Hora de saída\" FROM inicio_movimento ORDER BY data DESC, hora_inicio DESC")
            if not df_inicio.empty:
                st.dataframe(df_inicio, use_container_width=True, hide_index=True)
            else:
                st.info("Nenhum registro de início encontrado.")
                
        with c_paradas:
            st.markdown("**📍 Paradas realizadas nas obras (geocerca)**")
            st.caption("Registra o tempo de permanência dentro de um raio de 250 m do destino.")
            df_paradas_tbl = get_df("SELECT data as Data, placa as Placa, local as Local, hora_chegada as Chegada, hora_saida as Saída FROM rastreio_paradas ORDER BY id DESC LIMIT 150")
            if not df_paradas_tbl.empty:
                st.dataframe(df_paradas_tbl, use_container_width=True, hide_index=True)
            else:
                st.info("Nenhum registro de parada do rastreador encontrado.")

    if submodulo_frota == "🗂️ Histórico editável":
        st.markdown("### 🗂️ Histórico e correções")
        st.caption("Consulte, ajuste e exporte os registros consolidados da frota.")
        st.markdown("#### 💰 Histórico de custos e abastecimentos (editável)")
        st.caption("Você pode alterar os valores nas tabelas abaixo ou apagar linhas inteiras. Para salvar, clique no botão azul correspondente.")

        cx_abast, cx_km = st.columns(2)
        with cx_abast:
            st.markdown("**⛽ Combustível e manutenções**")

            @fragmento_independente
            def editor_abastecimentos():
                df_abastec_all = carregar_abastecimentos_df().sort_values("id", ascending=False).reset_index(drop=True)
                if not df_abastec_all.empty:
                    edited_abastec = st.data_editor(df_abastec_all, num_rows="dynamic", use_container_width=True, hide_index=True, key="edit_abastec")
                    if st.button("💾 Salvar alterações (abastecimentos)", type="primary"):
                        edited_abastec_clean = edited_abastec.drop(columns=['id'], errors='ignore')
                        save_df_to_db(edited_abastec_clean, "abastecimentos")
                        carregar_abastecimentos_df.clear()
                        st.success("Abastecimentos atualizados na nuvem com sucesso!")
                else:
                    st.info("Não há abastecimentos nem manutenções registradas.")

            editor_abastecimentos()
                
        with cx_km:
            st.markdown("**🛣️ Quilometragem rodada**")

            @fragmento_independente
            def editor_quilometragem():
                df_km_all = carregar_registro_km_df().sort_values("id", ascending=False).reset_index(drop=True)
                if not df_km_all.empty:
                    edited_km = st.data_editor(df_km_all, num_rows="dynamic", use_container_width=True, hide_index=True, key="edit_km")
                    if st.button("💾 Salvar alterações (quilometragem)", type="primary"):
                        edited_km_clean = edited_km.drop(columns=['id'], errors='ignore')
                        save_df_to_db(edited_km_clean, "registro_km")
                        carregar_registro_km_df.clear()
                        st.success("Registros de quilometragem atualizados na nuvem com sucesso!")
                else:
                    st.info("Nenhuma quilometragem registrada.")

            editor_quilometragem()

        df_abastecimentos_relatorio = carregar_abastecimentos_df().sort_values("id", ascending=False).reset_index(drop=True)
        df_quilometragens_relatorio = carregar_registro_km_df().sort_values("id", ascending=False).reset_index(drop=True)
        renderizar_exportador(
            "Registros e histórico da frota",
            {
                "Inícios de rota": df_inicio,
                "Paradas rastreadas": df_paradas_tbl,
                "Abastecimentos e manutenção": df_abastecimentos_relatorio,
                "Registros de quilometragem": df_quilometragens_relatorio,
            },
            "registros_da_frota", "registros",
        )

if modulo_principal == "🗺️ Roteiro do Davi":
    # O GPS é buscado em outra thread. A Torre abre primeiro com a última posição
    # conhecida e recebe a nova leitura depois, sem spinner nem espera de rede.
    if hasattr(st, "fragment"):
        @st.fragment(run_every="5s")
        def _ciclo_gps_mapa_rota():
            agora_gps = time.time()
            job = st.session_state.get("_gps_rota_future")

            if job is not None and job.done():
                try:
                    resultado = job.result()
                except Exception as erro_job:
                    resultado = {"sessao": None, "pagina": "", "posicoes": [], "erro": str(erro_job)}
                st.session_state.pop("_gps_rota_future", None)
                st.session_state["_gps_rota_ultimo_erro"] = resultado.get("erro", "")
                if resultado.get("sessao") is not None:
                    st.session_state["protege_sessao"] = resultado["sessao"]
                    st.session_state["protege_pagina"] = resultado.get("pagina", "")
                posicoes_novas = resultado.get("posicoes") or []
                if posicoes_novas:
                    assinatura_antiga = json.dumps(
                        st.session_state.get("_gps_rota_posicoes") or [],
                        ensure_ascii=False, sort_keys=True, default=str,
                    )
                    assinatura_nova = json.dumps(
                        posicoes_novas, ensure_ascii=False, sort_keys=True, default=str,
                    )
                    st.session_state["_gps_rota_posicoes"] = posicoes_novas
                    st.session_state["_gps_rota_atualizado_em"] = agora_gps
                    if assinatura_nova != assinatura_antiga:
                        try:
                            st.rerun(scope="app")
                        except TypeError:
                            st.rerun()

            sem_job = st.session_state.get("_gps_rota_future") is None
            ultimo_envio = float(st.session_state.get("_gps_rota_consulta_em", 0) or 0)
            if sem_job and agora_gps - ultimo_envio >= 30:
                usuario, senha, veiculos = carregar_config_protege()
                if usuario and senha and veiculos:
                    st.session_state["_gps_rota_consulta_em"] = agora_gps
                    st.session_state["_gps_rota_future"] = obter_executor_gps_rota().submit(
                        consultar_gps_rota_em_background,
                        usuario,
                        senha,
                        veiculos,
                        st.session_state.get("protege_sessao"),
                        st.session_state.get("protege_pagina"),
                    )

        _ciclo_gps_mapa_rota()

    if (st.session_state.get('rota_gerada', False) and st.session_state.get('data_rota') != DATA_REF_ROTA_STR):
        st.session_state['rota_gerada'] = False
    if not st.session_state.get('rota_gerada', False):
        carregar_rota_salva_para_sessao(DATA_REF_ROTA_STR)

    df_ativos = st.session_state.demandas.copy()
    # Uma rota salva não pode esconder cartões que chegaram depois. Ao abrir o
    # roteiro de hoje, fazemos uma única leitura atual do quadro mesmo quando a
    # sessão já contém outras demandas. A geocodificação continua desativada aqui;
    # o objetivo é garantir que todos os prazos do dia disputem a matriz atual.
    if (
        st.session_state.get('rota_gerada', False)
        and st.session_state.get('data_rota') == DATA_REF_ROTA_STR
        and DATA_REF_ROTA_DATE == AGORA_REAL.date()
        and (AGORA_REAL.hour * 60 + AGORA_REAL.minute) < LIMITE_EXPEDIENTE_DAVI_MIN
        and not st.session_state.get('_tentou_hidratar_demandas_direto_v2')
    ):
        st.session_state['_tentou_hidratar_demandas_direto_v2'] = True
        if sincronizar_demandas(forcar=True, geocodificar=False):
            df_ativos = st.session_state.demandas.copy()

    _locais_repetidos_rota = detectar_locais_repetidos_rota(
        st.session_state.get('route_steps') or [], ponto_saida
    )
    st.session_state['_rota_locais_repetidos_carregada'] = _locais_repetidos_rota
    _coletas_sem_entrega_rota = identificar_coletas_sem_entrega_route_steps(
        st.session_state.get('route_steps') or []
    )

    # No primeiro acesso, a rota salva chega antes do quadro do Trello. Se ela foi
    # criada pelo motor antigo e repete locais, hidratamos as demandas pelo cache do
    # Supabase agora mesmo para que o recálculo V7 não dependa de um clique manual.
    _chave_hidratacao_repetidos = f"_hidratou_repetidos_{DATA_REF_ROTA_STR}"
    if (
        _locais_repetidos_rota
        and isinstance(df_ativos, pd.DataFrame)
        and df_ativos.empty
        and not st.session_state.get(_chave_hidratacao_repetidos)
    ):
        st.session_state[_chave_hidratacao_repetidos] = True
        if sincronizar_demandas(forcar=False, geocodificar=False, somente_cache=True):
            df_ativos = st.session_state.demandas.copy()

    ajustes_manuais = carregar_ajustes_manuais_rota(DATA_REF_ROTA_STR)
    if not df_ativos.empty:
        df_ativos["Origem"] = df_ativos["Origem"].apply(canonicalizar_ponto_rota)
        df_ativos["Destino"] = df_ativos["Destino"].apply(canonicalizar_ponto_rota)
        # O que foi arrastado manualmente prevalece sobre a interpretação automática do Trello.
        df_ativos = aplicar_ajustes_manuais_demandas(df_ativos, ajustes_manuais, ponto_saida)

        origem_invalida, destino_invalido = df_ativos["Origem"].fillna("").isin(["", "DESCONHECIDO"]), df_ativos["Destino"].fillna("").isin(["", "DESCONHECIDO"])
        if not df_ativos[origem_invalida | destino_invalido].empty:
            st.warning(f"⚠️ Estas demandas estão sem origem ou destino legível no Trello e ficaram fora da rota: **{', '.join(df_ativos[origem_invalida | destino_invalido]['Obra'].astype(str).tolist())}**.")
            df_ativos = df_ativos[~(origem_invalida | destino_invalido)].copy()

    # Atualiza imediatamente nomes antigos da rota com a leitura mais recente do
    # título do Trello. Isso corrige, por exemplo, uma rota salva como "HORIZONTE"
    # quando a sincronização atual já reconhece "2506 - HORIZONTE".
    if st.session_state.get('route_steps'):
        _rota_rotulos_antes = json.dumps(st.session_state.get('route_steps') or [], ensure_ascii=False, sort_keys=True, default=str)
        # O mapa de rótulos é reconstruído/mesclado em TODA renderização do roteiro,
        # não apenas quando a sessão está vazia. Isso evita manter um mapa antigo
        # incompleto após deploy/reinício parcial do Streamlit.
        _mapa_rotulos_trello = dict(st.session_state.get("rotulos_obras_trello", {}) or {})

        def _mesclar_rotulos(_destino, _fonte):
            for _id_card, _rotulo in (_fonte or {}).items():
                _id_card = str(_id_card or "").strip()
                _rotulo = str(_rotulo or "").strip()
                if not _id_card or not _rotulo:
                    continue
                if _pontuacao_rotulo_obra(_rotulo) >= _pontuacao_rotulo_obra(_destino.get(_id_card, "")):
                    _destino[_id_card] = _rotulo

        _dados_rotulos_cache = {}
        # Sincronizar_demandas já constrói este mapa. Só lê o cache completo do Trello
        # quando a sessão ainda não possui mapa de rótulos, evitando varrer >1.500 cartões
        # a cada clique/rerun do roteiro.
        if not _mapa_rotulos_trello:
            try:
                _dados_rotulos_cache = obter_dados_trello() or {}
                _mesclar_rotulos(_mapa_rotulos_trello, construir_mapa_rotulos_obras_trello(_dados_rotulos_cache))
            except Exception:
                pass

        # Se qualquer tarefa salva ainda aparece apenas como unidade (ex.:
        # HORIZONTE/BARRA DO CEARÁ), buscamos o quadro diretamente uma vez.
        # O objetivo é recuperar o título atual pelo ID do cartão e eliminar
        # rótulos sem número em TODAS as unidades, não só UNIFOR.
        _ids_sem_numero_obra = set()
        for _step_rotulo in st.session_state.get('route_steps') or []:
            for _acao_rotulo, _tarefa_rotulo in _step_rotulo.get('actions', []) or []:
                _id_rotulo = str((_tarefa_rotulo or {}).get('id', '') or '').strip()
                _obra_rotulo = str((_tarefa_rotulo or {}).get('Obra', '') or '').strip()
                if _id_rotulo and _pontuacao_rotulo_obra(_obra_rotulo) < 1000:
                    _ids_sem_numero_obra.add(_id_rotulo)

        if _ids_sem_numero_obra and _dados_rotulos_cache:
            try:
                # Não baixa o quadro inteiro de novo durante a renderização.
                # A sincronização automática/manual já renova este cache.
                _dados_rotulos_frescos = _dados_rotulos_cache
                _mapa_fresco = construir_mapa_rotulos_obras_trello(_dados_rotulos_frescos)
                _mesclar_rotulos(_mapa_rotulos_trello, _mapa_fresco)

                # Além do rótulo, injeta o título original nas tarefas encontradas.
                # Depois desta gravação, a rota fica autossuficiente para manter
                # ``número - unidade`` mesmo se o cartão sair do quadro visível.
                _titulos_por_id = {
                    str(_c.get('id', '') or ''): str(_c.get('name', '') or '')
                    for _c in _dados_rotulos_frescos.get('cards', []) or []
                    if str(_c.get('id', '') or '')
                }
                _rota_com_titulos = []
                for _step_titulo in st.session_state.get('route_steps') or []:
                    _novo_step_titulo = dict(_step_titulo)
                    _novas_acoes_titulo = []
                    for _acao_titulo, _tarefa_titulo in _step_titulo.get('actions', []) or []:
                        _nova_tarefa_titulo = dict(_tarefa_titulo)
                        _id_titulo = str(_nova_tarefa_titulo.get('id', '') or '')
                        if _id_titulo in _titulos_por_id and _titulos_por_id[_id_titulo]:
                            _nova_tarefa_titulo['_Titulo_Trello'] = _titulos_por_id[_id_titulo]
                        _novas_acoes_titulo.append((_acao_titulo, _nova_tarefa_titulo))
                    if 'actions' in _step_titulo:
                        _novo_step_titulo['actions'] = _novas_acoes_titulo
                    _rota_com_titulos.append(_novo_step_titulo)
                st.session_state['route_steps'] = _rota_com_titulos
            except Exception:
                pass

        st.session_state["rotulos_obras_trello"] = _mapa_rotulos_trello

        _rota_rotulos_atualizada = atualizar_rotulos_obras_route_steps(
            st.session_state.get('route_steps') or [], df_ativos, _mapa_rotulos_trello
        )
        _rota_rotulos_depois = json.dumps(_rota_rotulos_atualizada, ensure_ascii=False, sort_keys=True, default=str)
        st.session_state['route_steps'] = _rota_rotulos_atualizada
        if _rota_rotulos_depois != _rota_rotulos_antes and st.session_state.get('data_rota') == DATA_REF_ROTA_STR:
            try:
                execute_db(
                    "UPDATE rota_ativa SET json_route=:route WHERE id=1 AND data_rota=:data",
                    {"route": json.dumps(_rota_rotulos_atualizada, ensure_ascii=False), "data": DATA_REF_ROTA_STR},
                )
            except Exception:
                pass

    rota_ativa_hoje = st.session_state.get('rota_gerada', False) and st.session_state.get('data_rota') == DATA_REF_ROTA_STR

    # Quando o volume de HOJE/VENCIDA excede o expediente, a equipe pode registrar
    # a prioridade combinada com o engenheiro. O campo fica recolhido para não
    # poluir a torre e altera somente o planejamento desta data, nunca o Trello.
    if isinstance(df_ativos, pd.DataFrame) and not df_ativos.empty and 'Urgência' in df_ativos.columns:
        df_priorizaveis = df_ativos[
            df_ativos['Urgência'].astype(str).str.contains(r'HOJE|VENCIDA', case=False, na=False)
        ].copy()
    else:
        df_priorizaveis = pd.DataFrame()

    if not df_priorizaveis.empty:
        ids_priorizaveis = df_priorizaveis['id'].astype(str).drop_duplicates().tolist()
        rotulos_prioridade = {}
        for _, linha_prioridade in df_priorizaveis.drop_duplicates(subset=['id']).iterrows():
            demanda_id_prioridade = str(linha_prioridade.get('id', '') or '')
            obra_prioridade = str(linha_prioridade.get('Obra', '') or 'Demanda sem nome')
            origem_prioridade = str(linha_prioridade.get('Origem', '') or '?')
            destino_prioridade = str(linha_prioridade.get('Destino', '') or '?')
            rotulos_prioridade[demanda_id_prioridade] = (
                f"{obra_prioridade} · {origem_prioridade} → {destino_prioridade}"
            )

        ids_prioritarios_atuais = {
            str(demanda_id or '').strip()
            for demanda_id in (ajustes_manuais.get('prioridades_operacionais', []) or [])
            if str(demanda_id or '').strip()
        }
        with st.expander("📌 Prioridades combinadas com os engenheiros", expanded=False):
            st.caption(
                "Use somente quando nem todas as demandas do dia couberem até 17h. "
                "As marcadas entram primeiro na melhor rota possível; o limite do expediente permanece obrigatório."
            )
            with st.form(f"prioridades_operacionais_{DATA_REF_ROTA_STR}"):
                prioridades_escolhidas = st.multiselect(
                    "Demandas que devem ser priorizadas hoje",
                    options=ids_priorizaveis,
                    default=[demanda_id for demanda_id in ids_priorizaveis if demanda_id in ids_prioritarios_atuais],
                    format_func=lambda demanda_id: rotulos_prioridade.get(demanda_id, demanda_id),
                )
                aplicar_prioridades = st.form_submit_button("Aplicar e recalcular a rota")

            if aplicar_prioridades:
                novos_ajustes = dict(ajustes_manuais)
                if prioridades_escolhidas:
                    novos_ajustes['prioridades_operacionais'] = list(prioridades_escolhidas)
                else:
                    novos_ajustes.pop('prioridades_operacionais', None)
                salvar_ajustes_manuais_rota(DATA_REF_ROTA_STR, novos_ajustes)
                st.session_state['_recalcular_rota_automatico'] = True
                st.session_state['_mensagem_ajuste_rota'] = (
                    "Prioridades operacionais atualizadas. A rota foi refeita respeitando o encerramento às 17h."
                )
                st.rerun()

    # Rotas gravadas por uma versão anterior — ou que ainda tragam o mesmo local
    # em várias paradas — são recalculadas uma única vez por data quando as
    # demandas ativas estiverem disponíveis.
    versao_rota_salva = max(
        [int(step.get('_motor_rota_versao', 0) or 0) for step in (st.session_state.get('route_steps') or [])]
        + [0]
    )
    if (
        rota_ativa_hoje
        and (
            versao_rota_salva < ROTA_ENGINE_VERSION
            or bool(_locais_repetidos_rota)
            or bool(_coletas_sem_entrega_rota)
        )
        and isinstance(df_ativos, pd.DataFrame)
        and not df_ativos.empty
        and st.session_state.get('_motor_rota_v7_solicitado_em') != DATA_REF_ROTA_STR
    ):
        st.session_state['_motor_rota_v7_solicitado_em'] = DATA_REF_ROTA_STR
        st.session_state['_recalcular_rota_automatico'] = True
        _nomes_repetidos = ', '.join(_locais_repetidos_rota)
        st.session_state['_mensagem_ajuste_rota'] = (
            '✅ Motor V7 aplicado. A rota foi reorganizada com ciclos completos de coleta e entrega '
            f'em uma única visita sempre que possível{": " + _nomes_repetidos if _nomes_repetidos else "."}'
        )

    # Recebe um movimento do editor arrastável e o transforma em regra persistente.
    _mr_id = str(st.query_params.get("mr_demanda", "") or "").strip()
    _mr_acao = str(st.query_params.get("mr_acao", "") or "").strip().upper()
    _mr_destino = str(st.query_params.get("mr_destino", "") or "").strip()
    _mr_ordem = str(st.query_params.get("mr_ordem", "") or "").strip()
    if _mr_id and st.session_state.get('route_steps'):
        try:
            movimento_ok = registrar_movimento_manual_rota(
                DATA_REF_ROTA_STR, st.session_state.get('route_steps') or [],
                _mr_id, _mr_acao, _mr_destino, int(_mr_ordem or 0), ponto_saida,
            )
            st.query_params.clear()
            if movimento_ok:
                _rota_pos_movimento = aplicar_movimento_manual_route_steps_imediato(
                    st.session_state.get('route_steps') or [],
                    _mr_id, _mr_acao, _mr_destino, int(_mr_ordem or 0), ponto_saida,
                )
                _ajustes_pos_movimento = carregar_ajustes_manuais_rota(DATA_REF_ROTA_STR)
                _rota_pos_movimento = consolidar_coletas_base_na_preparacao(
                    _rota_pos_movimento, _ajustes_pos_movimento, ponto_saida
                )
                _rota_pos_movimento = aplicar_ordem_manual_route_steps(
                    _rota_pos_movimento, _ajustes_pos_movimento
                )
                st.session_state['route_steps'] = _rota_pos_movimento
                try:
                    execute_db(
                        "UPDATE rota_ativa SET json_route=:route WHERE id=1 AND data_rota=:data",
                        {"route": json.dumps(_rota_pos_movimento, ensure_ascii=False), "data": DATA_REF_ROTA_STR},
                    )
                except Exception:
                    pass
                st.session_state["_recalcular_rota_automatico"] = True
                st.session_state["_mensagem_ajuste_rota"] = "✅ Demanda movida. Roteiro, horários e mapa foram atualizados com o ajuste manual."
                st.rerun()
        except Exception as _erro_movimento_rota:
            st.query_params.clear()
            st.warning(f"Não foi possível mover essa demanda agora: {_erro_movimento_rota}")

    # Se um cartão de hoje/vencido chegou após a criação da rota salva, ele precisa
    # disputar imediatamente o tempo restante. O otimizador pode retirar demandas
    # futuras para preservar o expediente, mas a urgente deixa de ficar invisível.
    ids_rota_atuais = {
        str(tarefa.get('id', '') or '')
        for step in (st.session_state.get('route_steps') or [])
        for _acao, tarefa in (step.get('actions', []) or [])
        if str(tarefa.get('id', '') or '')
    }
    if isinstance(df_ativos, pd.DataFrame) and not df_ativos.empty and 'Urgência' in df_ativos.columns:
        df_criticas_fora_rota = df_ativos[
            df_ativos['Urgência'].astype(str).str.contains(r'HOJE|VENCIDA', case=False, na=False)
            & ~df_ativos['id'].astype(str).isin(ids_rota_atuais)
        ].copy()
    else:
        df_criticas_fora_rota = pd.DataFrame()

    ids_criticos_fora_rota = sorted(df_criticas_fora_rota['id'].astype(str).tolist()) if not df_criticas_fora_rota.empty else []
    assinatura_criticos_fora_rota = '|'.join(ids_criticos_fora_rota)
    if (
        rota_ativa_hoje
        and DATA_REF_ROTA_DATE == AGORA_REAL.date()
        and (AGORA_REAL.hour * 60 + AGORA_REAL.minute) < LIMITE_EXPEDIENTE_DAVI_MIN
        and ids_criticos_fora_rota
        and st.session_state.get('_ultimo_lote_urgente_incorporado_v6') != assinatura_criticos_fora_rota
    ):
        # A chave versionada também libera uma nova tentativa depois de mudanças
        # no critério de viabilidade, mesmo que o conjunto de cartões seja igual.
        st.session_state['_ultimo_lote_urgente_incorporado_v6'] = assinatura_criticos_fora_rota
        st.session_state['_recalcular_rota_automatico'] = True
        nomes_criticos = ', '.join(df_criticas_fora_rota['Obra'].astype(str).tolist())
        st.session_state['_mensagem_ajuste_rota'] = (
            f"📌 Demanda de hoje/vencida incorporada ao recálculo: {nomes_criticos}. "
            "O restante do roteiro foi reorganizado por prazo e viabilidade."
        )

    # A rota pode ser recalculada sem apagar as baixas já registradas.
    # As etapas concluídas são reaproveitadas logo abaixo a partir de historico_concluidos.
    txt_botao = "🔄 Recalcular / Atualizar Rota" if rota_ativa_hoje else "🚀 Calcular Rota Otimizada"

    recalculo_automatico = bool(st.session_state.pop("_recalcular_rota_automatico", False))
    recalculo_manual = st.button(txt_botao, type="primary", disabled=df_ativos.empty)

    if (recalculo_manual or recalculo_automatico) and not df_ativos.empty:
        if recalculo_automatico:
            st.toast("🔄 Recalculando o restante do expediente...", icon="🗺️")
        with st.spinner("Analisando histórico e inteligência de nomes para traçar rota..."):
            st.session_state['demandas_adiadas'] = []
            garantir_gps_local_base()
            
            df_torre = carregar_conclusoes_rota(DATA_REF_ROTA_STR)
            dict_concluidos_torre = dict(zip(df_torre['id'].astype(str), df_torre['hora_conclusao']))
            
            past_route_steps = []

            current_time_tsp = parse_time_to_mins(obter_hora_inicio_rota(DATA_REF_ROTA_STR))
            current_point = ponto_saida

            # Tudo que sai da base deve ser carregado UMA VEZ na preparação das
            # 07:30–08:00. Ao recalcular a rota durante o dia, essas coletas não
            # podem voltar como uma nova "PARADA: ESCRITÓRIO".
            registros_ativos_rota = df_ativos.to_dict('records')
            base_canonica = canonicalizar_ponto_rota(ponto_saida)
            tarefas_base_ativas = [
                t for t in registros_ativos_rota
                if str(t.get('id', '')) not in dict_concluidos_torre
                and canonicalizar_ponto_rota(t.get('Origem', '')) == base_canonica
            ]
            ids_base_ativos = {str(t.get('id', '')) for t in tarefas_base_ativas}

            rota_salva = fetch_one("SELECT json_route FROM rota_ativa WHERE id = 1 AND data_rota = :data", {"data": DATA_REF_ROTA_STR})
            preparacao_salva = None
            if rota_salva:
                old_steps = json.loads(rota_salva[0])
                for indice_old, step in enumerate(old_steps):
                    if step.get('type') != 'stop':
                        continue

                    destino_step = canonicalizar_ponto_rota(step.get('destino', ''))
                    fonte_step = remover_acentos(str(step.get('tempo_local_fonte', '') or '')).lower()
                    eh_preparacao = (
                        destino_step == base_canonica
                        and (indice_old == 0 or 'preparacao' in fonte_step)
                        and float(step.get('dist', 0) or 0) <= 0.10
                    )

                    if eh_preparacao and preparacao_salva is None:
                        preparacao_salva = step.copy()
                        continue

                    # Fora da preparação, só preservamos fisicamente o que já foi
                    # concluído. O restante será reotimizado a partir do ponto atual.
                    c_acts = [
                        (a, t) for a, t in step.get('actions', [])
                        if str(t.get('id', '')) in dict_concluidos_torre
                    ]
                    if c_acts:
                        new_s = step.copy()
                        new_s['actions'] = c_acts
                        past_route_steps.append(new_s)

            # REGRA OPERACIONAL DEFINITIVA DA PREPARAÇÃO:
            # se uma demanda ATIVA já constava como COLETAR na preparação salva do
            # escritório, ela já está fisicamente dentro do carro. Isso prevalece
            # sobre a Origem textual do Trello durante o restante da rota. Assim uma
            # demanda preparada no escritório nunca reaparece depois como COLETA na
            # UNIFOR, FIEC, CENTRO ou qualquer outro ponto.
            ativos_por_id = {
                str(t.get('id', '')): t
                for t in registros_ativos_rota
                if str(t.get('id', '')) and str(t.get('id', '')) not in dict_concluidos_torre
            }
            ids_preparados_salvos_ativos = set()
            if preparacao_salva:
                for acao_salva, tarefa_salva in preparacao_salva.get('actions', []) or []:
                    tarefa_id_salva = str(tarefa_salva.get('id', ''))
                    if (
                        acao_salva == 'COLETAR'
                        and tarefa_id_salva
                        and tarefa_id_salva in ativos_por_id
                        # Se a COLETA foi arrastada manualmente para outro local,
                        # o ajuste atual prevalece sobre a preparação antiga salva.
                        and canonicalizar_ponto_rota(
                            ativos_por_id[tarefa_id_salva].get('Origem', '')
                        ) == base_canonica
                    ):
                        ids_preparados_salvos_ativos.add(tarefa_id_salva)

            # União entre o que naturalmente tem origem na base e o que já foi
            # efetivamente colocado na preparação da base. Mantemos os dados atuais
            # do Trello para obra/destino/materiais, mudando apenas o estado logístico
            # de "a coletar" para "já no carro".
            ids_base_ativos |= ids_preparados_salvos_ativos
            tarefas_base_ativas = [
                t for t in registros_ativos_rota
                if str(t.get('id', '')) in ids_base_ativos
                and str(t.get('id', '')) not in dict_concluidos_torre
            ]

            # Reconstrói uma única preparação no início da rota. Mantém as coletas
            # já concluídas que existiam na preparação salva e acrescenta TODAS as
            # demandas ainda ativas cuja origem é a base. Assim elas aparecem na
            # preparação, mas seguem pendentes até a entrega no destino.
            acoes_preparacao = []
            ids_prep_vistos = set()
            if preparacao_salva:
                for acao, tarefa in preparacao_salva.get('actions', []):
                    tarefa_id = str(tarefa.get('id', ''))
                    if acao == 'COLETAR' and tarefa_id in dict_concluidos_torre and tarefa_id not in ids_prep_vistos:
                        acoes_preparacao.append((acao, tarefa))
                        ids_prep_vistos.add(tarefa_id)

            for tarefa in tarefas_base_ativas:
                tarefa_id = str(tarefa.get('id', ''))
                if tarefa_id and tarefa_id not in ids_prep_vistos:
                    acoes_preparacao.append(('COLETAR', tarefa))
                    ids_prep_vistos.add(tarefa_id)

            if acoes_preparacao:
                prep = (preparacao_salva or {}).copy()
                prep.update({
                    'type': 'stop',
                    'destino': ponto_saida,
                    'dist': 0.0,
                    'travel_mins': 0.0,
                    'travel_mins_api': 0.0,
                    'tempo_local': 30,
                    'tempo_local_fonte': 'preparação fixa da base',
                    'chegada': HORA_PREPARACAO_INICIO,
                    'saida': HORA_PREPARACAO_FIM,
                    'actions': acoes_preparacao,
                })
                past_route_steps.insert(0, prep)

            # Corrige também o histórico já salvo: qualquer COLETA que hoje esteja
            # em uma parada ESCRITÓRIO (ou tenha sido arrastada para a base) é
            # absorvida pela PREPARAÇÃO, inclusive se a demanda já tomou baixa.
            past_route_steps = consolidar_coletas_base_na_preparacao(
                past_route_steps, ajustes_manuais, ponto_saida
            )

            # A posição/horário corrente deve ser a última parada realmente concluída.
            # Se só existe a preparação, continuamos na base às 08:00.
            if past_route_steps:
                # Se há alguma parada operacional concluída, partimos dela.
                # Se só existe a preparação, preservamos o horário real de saída
                # detectado/manual (a preparação visual continua fixa 07:30–08:00).
                operacionais_passadas = [
                    s for i_s, s in enumerate(past_route_steps)
                    if not (
                        s.get('type') == 'stop'
                        and canonicalizar_ponto_rota(s.get('destino', '')) == base_canonica
                        and (
                            'preparacao' in remover_acentos(str(s.get('tempo_local_fonte', '') or '')).lower()
                            or (i_s == 0 and float(s.get('dist', 0) or 0) <= 0.10)
                        )
                    )
                ]
                if operacionais_passadas:
                    ultima_operacional = operacionais_passadas[-1]
                    current_point = ultima_operacional['destino']

                    # Para uma parada já concluída, o relógio do recálculo parte da
                    # conclusão REAL do Trello/histórico, não do horário planejado.
                    # Isso é essencial perto das 17h: se Davi concluiu às 17:08,
                    # preservamos essa conclusão e nenhuma nova parada é criada.
                    conclusoes_reais_ultima = []
                    for _acao_real, _tarefa_real in ultima_operacional.get('actions', []) or []:
                        _id_real = str(_tarefa_real.get('id', ''))
                        _hora_real = dict_concluidos_torre.get(_id_real)
                        if _hora_real:
                            try:
                                conclusoes_reais_ultima.append(parse_time_to_mins(str(_hora_real)))
                            except Exception:
                                pass
                    if conclusoes_reais_ultima:
                        current_time_tsp = max(conclusoes_reais_ultima)
                    else:
                        try:
                            h, m = map(int, str(ultima_operacional['saida']).split(':'))
                            current_time_tsp = h * 60 + m
                        except Exception:
                            pass
                else:
                    current_point = ponto_saida
                    current_time_tsp = max(current_time_tsp, parse_time_to_mins(HORA_PREPARACAO_FIM))

            # Em uma rota de HOJE, o recálculo das etapas pendentes precisa partir
            # do relógio real. Antes, o sistema podia reconstruir o restante da rota
            # a partir da última baixa (por exemplo, 10:29), mesmo já sendo 14:00.
            # A sequência cabia até 17h no relógio antigo, mas o ETA dinâmico depois
            # a empurrava para 18h/19h. Agora o próprio planejador usa o horário atual.
            if DATA_REF_ROTA_DATE == AGORA_REAL.date():
                agora_planejamento_min = AGORA_REAL.hour * 60 + AGORA_REAL.minute
                if 12 * 60 <= agora_planejamento_min < 13 * 60:
                    agora_planejamento_min = 13 * 60
                current_time_tsp = max(current_time_tsp, agora_planejamento_min)

            # As coletas da base já estão fisicamente no veículo desde a preparação:
            # elas saem de "a coletar" e entram direto em "carrying".
            unpicked = [
                t for t in registros_ativos_rota
                if str(t.get('id', '')) not in dict_concluidos_torre
                and str(t.get('id', '')) not in ids_base_ativos
            ]

            pontos_brutos = (
                [ponto_saida]
                + [s['destino'] for s in past_route_steps]
                + [t['Origem'] for t in unpicked]
                + [t['Destino'] for t in unpicked]
                + [t['Destino'] for t in tarefas_base_ativas]
            )
            pontos_necessarios = {canonicalizar_ponto_rota(p) for p in pontos_brutos if canonicalizar_ponto_rota(p) not in {"", "DESCONHECIDO", "NAN", "NONE"}}
            
            locais_dict, enderecos_dict = {}, {}
            
            locais_db_raw = fetch_all("SELECT apelido, endereco, lat, lon FROM locais")
            locais_db = {row[0]: (row[1], row[2], row[3]) for row in locais_db_raw}
            
            for p in pontos_necessarios:
                alvo = p
                if alvo not in locais_db:
                    # Primeiro compara pela mesma inteligência de aliases usada no
                    # Trello. Assim um endereço salvo como "CONDOMÍNIO COLISEU" pode
                    # atender o ponto canônico "COLISEU", por exemplo.
                    encontrado = next(
                        (loc for loc in locais_db.keys() if canonicalizar_ponto_rota(loc) == p),
                        None,
                    )
                    p_sem_acento = remover_acentos(p)
                    if not encontrado:
                        encontrado = next((loc for loc in locais_db.keys() if remover_acentos(loc) == p_sem_acento), None)
                    if not encontrado:
                        matches = difflib.get_close_matches(p, locais_db.keys(), n=1, cutoff=0.8)
                        if matches: encontrado = matches[0]
                    if encontrado: alvo = encontrado

                # Fornecedores recorrentes podem vir no cartão sem endereço. O
                # fallback oficial é usado sob demanda e nunca substitui um local
                # que a equipe já cadastrou/ajustou na aba Endereços.
                if alvo in locais_db:
                    endereco_existente, lat_existente, lon_existente = locais_db[alvo]
                    if not str(endereco_existente or '').strip() and lat_existente is None and lon_existente is None:
                        endereco_fallback = obter_endereco_fornecedor_fallback(p)
                        if endereco_fallback:
                            execute_db(
                                "UPDATE locais SET endereco=:end WHERE apelido=:apelido "
                                "AND (endereco IS NULL OR TRIM(endereco)='') AND lat IS NULL AND lon IS NULL",
                                {"apelido": alvo, "end": endereco_fallback},
                            )
                            locais_db[alvo] = (endereco_fallback, None, None)

                if alvo not in locais_db:
                    endereco_fallback = obter_endereco_fornecedor_fallback(p)
                    if endereco_fallback:
                        execute_db(
                            "INSERT INTO locais (apelido, endereco) VALUES (:apelido, :end) "
                            "ON CONFLICT (apelido) DO NOTHING",
                            {"apelido": p, "end": endereco_fallback},
                        )
                        locais_db[p] = (endereco_fallback, None, None)
                        alvo = p

                if alvo in locais_db:
                    end_str, lat_db, lon_db = locais_db[alvo]
                    if lat_db is not None and lon_db is not None:
                        locais_dict[p] = (lat_db, lon_db)
                        enderecos_dict[p] = end_str
                    elif end_str:
                        lat, lon = buscar_coordenadas(end_str)
                        if lat is not None and lon is not None:
                            execute_db("UPDATE locais SET lat=:lat, lon=:lon WHERE apelido=:apelido", {"lat": lat, "lon": lon, "apelido": alvo})
                            locais_dict[p] = (lat, lon)
                            enderecos_dict[p] = end_str
                            
            st.session_state['enderecos_dict'] = enderecos_dict
            
            faltando = sorted(p for p in pontos_necessarios if p not in locais_dict and p not in {"", "DESCONHECIDO", "NAN", "NONE"})
            if faltando: st.warning(f"⚠️ Faltam endereços cadastrados na aba Endereços para: **{', '.join(faltando)}**"); st.stop()

            pontos_unicos = list(locais_dict.keys())
            coords = [locais_dict[p] for p in pontos_unicos]
            horario_partida_matriz = datetime.combine(DATA_REF_ROTA_DATE, datetime.min.time()).replace(tzinfo=FUSO_LOCAL) + timedelta(minutes=current_time_tsp)
            if DATA_REF_ROTA_DATE == AGORA_REAL.date() and horario_partida_matriz < AGORA_REAL:
                horario_partida_matriz = AGORA_REAL + timedelta(minutes=1)

            dist_matrix, dur_matrix, fonte_matriz = calcular_matriz_rotas(coords, horario_partida_matriz)

            def get_dist_dur_bruto(p1, p2):
                if p1 == p2:
                    return 0.0, 0.0
                i, j = pontos_unicos.index(p1), pontos_unicos.index(p2)
                return float(dist_matrix[i][j]), float(dur_matrix[i][j])

            # O otimizador continua usando a matriz viária, mas com um piso operacional
            # realista para não preferir uma sequência baseada em tempos urbanos irreais.
            def get_dist_dur(p1, p2):
                dist, dur = get_dist_dur_bruto(p1, p2)
                return dist, ajustar_tempo_deslocamento_operacional(dist, dur, current_time_tsp)

            tarefas_planejamento = list(unpicked) + list(tarefas_base_ativas)
            ordem_otimizada = otimizar_sequencia_rota(
                tarefas_planejamento,
                current_point,
                estrategia,
                get_dist_dur,
                current_time_tsp,
                retornar_base=retornar_base,
                ponto_base=ponto_saida,
                tarefas_pre_coletadas=tarefas_base_ativas,
            )
            st.session_state['fonte_matriz_rota'] = fonte_matriz
            st.session_state['horario_matriz_rota'] = horario_partida_matriz.strftime("%d/%m/%Y %H:%M")

            # Materiais coletados no escritório durante a preparação já começam
            # no veículo e não geram uma parada futura na base.
            carrying = list(tarefas_base_ativas)
            current = current_point
            route_steps_new = []
            total_km = sum(p_step.get('dist', 0.0) for p_step in past_route_steps)
            
            current_time = current_time_tsp
            lunch_taken = any(s.get('type') == 'lunch' for s in past_route_steps)

            def _registrar_adiadas(tarefas):
                existentes = {str(t.get('id', '')) for t in st.session_state['demandas_adiadas']}
                for tarefa in tarefas:
                    tarefa_id = str(tarefa.get('id', ''))
                    if tarefa_id and tarefa_id not in existentes:
                        st.session_state['demandas_adiadas'].append(tarefa)
                        existentes.add(tarefa_id)

            # O otimizador pode devolver uma sequência parcial quando o expediente
            # não comporta todas as demandas. Identificamos quais cartões chegam de
            # fato à ENTREGA nessa sequência e retiramos os demais antes de montar a
            # preparação. Isso impede a exibição de COLETAS soltas para demandas que
            # já ficaram para o próximo planejamento.
            tarefas_planejamento_por_id = {
                str(t.get('id', '') or ''): t
                for t in tarefas_planejamento
                if str(t.get('id', '') or '')
            }
            ids_planejados_completos = identificar_ids_entregues_na_ordem_rota(
                ordem_otimizada, tarefas_planejamento, tarefas_base_ativas
            )
            tarefas_sem_ciclo_completo = [
                tarefa for demanda_id, tarefa in tarefas_planejamento_por_id.items()
                if demanda_id not in ids_planejados_completos
            ]
            _registrar_adiadas(tarefas_sem_ciclo_completo)
            unpicked = [
                tarefa for tarefa in unpicked
                if str(tarefa.get('id', '') or '') in ids_planejados_completos
            ]
            carrying = [
                tarefa for tarefa in carrying
                if str(tarefa.get('id', '') or '') in ids_planejados_completos
            ]

            # Regra física: depois que Davi atende uma unidade, ela não volta a
            # aparecer mais tarde no mesmo roteiro. Paradas já realizadas também
            # entram neste bloqueio quando o restante do dia é recalculado.
            base_canonica_rota = canonicalizar_ponto_rota(ponto_saida)
            locais_visitados_operacionais = {
                canonicalizar_ponto_rota(passo.get('destino', ''))
                for passo in past_route_steps
                if passo.get('type') == 'stop'
                and canonicalizar_ponto_rota(passo.get('destino', ''))
                and canonicalizar_ponto_rota(passo.get('destino', '')) != base_canonica_rota
            }

            # A preparação já havia sido reconstruída antes da otimização. Agora ela
            # fica limitada às demandas cujo destino também está planejado; coletas
            # históricas concluídas continuam preservadas no resumo do dia.
            passos_passados_filtrados = []
            for passo_passado in past_route_steps:
                if passo_passado.get('type') != 'stop':
                    passos_passados_filtrados.append(passo_passado)
                    continue
                acoes_passadas_validas = []
                for acao_passada, tarefa_passada in (passo_passado.get('actions', []) or []):
                    demanda_id_passada = str(tarefa_passada.get('id', '') or '')
                    coleta_pendente_sem_entrega = (
                        acao_passada == 'COLETAR'
                        and demanda_id_passada not in dict_concluidos_torre
                        and demanda_id_passada not in ids_planejados_completos
                    )
                    if not coleta_pendente_sem_entrega:
                        acoes_passadas_validas.append((acao_passada, tarefa_passada))
                if acoes_passadas_validas:
                    novo_passo_passado = passo_passado.copy()
                    novo_passo_passado['actions'] = acoes_passadas_validas
                    passos_passados_filtrados.append(novo_passo_passado)
            past_route_steps = passos_passados_filtrados

            def _avaliar_candidato_expediente(ponto):
                """Simula a próxima parada e reserva o retorno à base até 17h."""
                dist, dur_api = get_dist_dur_bruto(current, ponto)
                dur = ajustar_tempo_deslocamento_operacional(dist, dur_api, current_time)
                arr = current_time + dur
                pausa_consumida = lunch_taken

                if current_time <= 12 * 60 and arr > 12 * 60 and not pausa_consumida:
                    arr = max(arr + 60, 13 * 60)
                    pausa_consumida = True
                if 12 * 60 <= arr < 13 * 60 and not pausa_consumida:
                    arr = 13 * 60
                    pausa_consumida = True

                entregas = [t for t in carrying if t['Destino'] == ponto]
                coletas = [t for t in unpicked if t['Origem'] == ponto]
                if not entregas and not coletas:
                    return None

                servico = estimar_tempo_parada(ponto, entregas, coletas)
                fim = arr + servico
                if arr < 12 * 60 <= fim and not pausa_consumida:
                    fim += 60
                    pausa_consumida = True

                if fim > LIMITE_EXPEDIENTE_DAVI_MIN:
                    return None

                if retornar_base and ponto != ponto_saida:
                    d_volta, dur_volta_api = get_dist_dur_bruto(ponto, ponto_saida)
                    dur_volta = ajustar_tempo_deslocamento_operacional(d_volta, dur_volta_api, fim)
                    if fim + dur_volta > LIMITE_EXPEDIENTE_DAVI_MIN:
                        return None

                return dist, dur_api, dur

            while unpicked or carrying:
                # Não existe mais corte artificial às 15h30. Enquanto uma demanda
                # ainda couber operacionalmente no tempo restante — deslocamento,
                # atendimento e, quando configurado, retorno à base até 17h — ela
                # continua elegível. A prioridade decide QUAL entra primeiro, mas
                # demandas futuras podem preencher o restante do expediente para
                # evitar ociosidade do motorista.

                # Se o relógio entrou no almoço, a pausa acontece antes de escolher
                # a próxima parada. Isso deixa a simulação de viabilidade consistente.
                if 12 * 60 <= current_time < 13 * 60 and not lunch_taken:
                    route_steps_new.append({"type": "lunch", "chegada": "12:00", "saida": "13:00"})
                    current_time = 13 * 60
                    lunch_taken = True

                candidates = set([t['Origem'] for t in unpicked] + [t['Destino'] for t in carrying])
                if not candidates:
                    break

                # Nunca retorna a uma unidade já atendida. Se uma dependência tardia
                # exigiria BARRA → outros locais → BARRA, o card dependente fica para
                # o próximo planejamento em vez de criar uma segunda visita.
                ids_bloqueados_revisita = {
                    str(t.get('id', '') or '')
                    for t in unpicked
                    if canonicalizar_ponto_rota(t.get('Origem', '')) in locais_visitados_operacionais
                } | {
                    str(t.get('id', '') or '')
                    for t in carrying
                    if canonicalizar_ponto_rota(t.get('Destino', '')) in locais_visitados_operacionais
                }
                ids_bloqueados_revisita.discard('')
                if ids_bloqueados_revisita:
                    bloqueadas_revisita = [
                        t for t in (list(unpicked) + list(carrying))
                        if str(t.get('id', '') or '') in ids_bloqueados_revisita
                    ]
                    _registrar_adiadas(bloqueadas_revisita)
                    unpicked = [t for t in unpicked if str(t.get('id', '') or '') not in ids_bloqueados_revisita]
                    carrying = [t for t in carrying if str(t.get('id', '') or '') not in ids_bloqueados_revisita]
                    candidates = set([t['Origem'] for t in unpicked] + [t['Destino'] for t in carrying])
                    if not candidates:
                        break

                avaliacoes = {p: _avaliar_candidato_expediente(p) for p in candidates}
                candidates_viaveis = {
                    p for p, avaliacao in avaliacoes.items()
                    if avaliacao is not None
                    and (
                        canonicalizar_ponto_rota(p) == base_canonica_rota
                        or canonicalizar_ponto_rota(p) not in locais_visitados_operacionais
                    )
                }

                if not candidates_viaveis:
                    # Nada restante cabe no expediente considerando deslocamento,
                    # atendimento e retorno à base até 17h. Só nesse caso o restante
                    # fica para o próximo planejamento.
                    _registrar_adiadas(list(unpicked) + list(carrying))
                    unpicked.clear()
                    carrying.clear()
                    break

                # Não fecha um destino enquanto ainda existe material a buscar
                # em outra unidade para esse mesmo destino. Assim FIEC, Barra do
                # Ceará etc. recebem tudo em uma única visita sempre que a cadeia
                # de coletas permitir uma ordem sem ciclos.
                candidates_viaveis = priorizar_pontos_sem_revisita(
                    candidates_viaveis, unpicked
                )

                best_point = None
                while ordem_otimizada:
                    ponto_planejado = ordem_otimizada.pop(0)
                    if ponto_planejado in candidates_viaveis:
                        best_point = ponto_planejado
                        break

                if best_point is None:
                    best_point = min(
                        candidates_viaveis,
                        key=lambda p: pontuar_parada_rota(current, p, unpicked, carrying, estrategia, get_dist_dur)[0],
                    )

                # O trecho vem da matriz OSRM. A validação operacional aplica apenas
                # um piso realista para deslocamentos urbanos, sem depender de API paga.
                best_dist, best_dur_api = get_dist_dur_bruto(current, best_point)
                best_dur = ajustar_tempo_deslocamento_operacional(best_dist, best_dur_api, current_time)

                arr_time = current_time + best_dur
                
                if current_time <= 12*60 and arr_time > 12*60 and not lunch_taken:
                    route_steps_new.append({"type": "lunch", "chegada": "12:00", "saida": "13:00"})
                    arr_time = max(arr_time + 60, 13 * 60)
                    lunch_taken = True
                    
                current_time = arr_time
                total_km += best_dist
                actions_here = []

                entregas_here = [t for t in carrying if t['Destino'] == best_point]
                coletas_here = [t for t in unpicked if t['Origem'] == best_point]
                for t in entregas_here:
                    actions_here.append(("ENTREGAR", t)); carrying.remove(t)
                for t in coletas_here:
                    actions_here.append(("COLETAR", t)); unpicked.remove(t); carrying.append(t)

                service_mins, fonte_tempo_local = estimar_tempo_parada(
                    best_point, entregas_here, coletas_here, retornar_fonte=True
                )

                is_start_load = (best_point == ponto_saida and current_time == current_time_tsp and not any(a[0] == "ENTREGAR" for a in actions_here) and len(past_route_steps) == 0)
                
                pausa_almoco_depois = False
                if is_start_load:
                    chegada_str, saida_str, tempo_local_exibicao = format_time(current_time_tsp - 30), format_time(current_time_tsp), 30
                    service_mins = 0
                    dep_time = current_time_tsp
                else:
                    dep_time = current_time + service_mins
                    # Se o atendimento atravessar o meio-dia, termina a atividade
                    # atual e registra a pausa de 1h como uma etapa separada. Isso
                    # mantém o mesmo término da rota que a lógica antiga, mas deixa
                    # o almoço visível no roteiro, no app e nos PDFs.
                    if current_time < 12 * 60 <= dep_time and not lunch_taken:
                        pausa_almoco_depois = True
                    chegada_str, saida_str, tempo_local_exibicao = format_time(current_time), format_time(dep_time), service_mins

                route_steps_new.append({"type": "stop", "destino": best_point, "dist": best_dist, "travel_mins": best_dur, "travel_mins_api": best_dur_api, "tempo_local": tempo_local_exibicao, "tempo_local_fonte": ("preparação fixa da base" if is_start_load else fonte_tempo_local), "chegada": chegada_str, "saida": saida_str, "actions": actions_here})
                current_time = dep_time
                current = best_point
                local_visitado = canonicalizar_ponto_rota(best_point)
                if local_visitado and local_visitado != base_canonica_rota:
                    locais_visitados_operacionais.add(local_visitado)

                if pausa_almoco_depois:
                    inicio_almoco = current_time
                    fim_almoco = inicio_almoco + 60
                    route_steps_new.append({
                        "type": "lunch",
                        "chegada": format_time(inicio_almoco),
                        "saida": format_time(fim_almoco),
                    })
                    current_time = fim_almoco
                    lunch_taken = True

            st.session_state['retorno_omitido_expediente'] = False
            if retornar_base and current != ponto_saida:
                d, dur_api = get_dist_dur_bruto(current, ponto_saida)
                dur = ajustar_tempo_deslocamento_operacional(d, dur_api, current_time)
                chegada_base = current_time + dur
                if current_time < LIMITE_EXPEDIENTE_DAVI_MIN and chegada_base <= LIMITE_EXPEDIENTE_DAVI_MIN:
                    total_km += d
                    route_steps_new.append({"type": "return", "destino": ponto_saida, "dist": d, "travel_mins": dur, "travel_mins_api": dur_api, "chegada": format_time(chegada_base), "saida": format_time(chegada_base), "actions": []})
                    current_time = chegada_base
                else:
                    # Pode acontecer quando uma conclusão REAL já foi registrada após
                    # as 17h. Preservamos a conclusão, mas não inventamos nova rota.
                    st.session_state['retorno_omitido_expediente'] = True

            route_steps = past_route_steps + route_steps_new

            # Última barreira de integridade: uma coleta PENDENTE só pode ser
            # publicada quando a entrega correspondente também existe na rota.
            # Se uma diferença de trânsito/ETA inviabilizar o destino durante a
            # montagem final, o cartão inteiro volta para as adiadas — nunca fica
            # uma coleta isolada na tela do Davi ou na Torre.
            ids_entregas_publicadas = {
                str(tarefa.get('id', '') or '')
                for step_publicado in route_steps
                for acao_publicada, tarefa in (step_publicado.get('actions', []) or [])
                if acao_publicada == 'ENTREGAR' and str(tarefa.get('id', '') or '')
            }
            coletas_orfas_removidas = []
            route_steps_com_pares = []
            for step_publicado in route_steps:
                if step_publicado.get('type') != 'stop':
                    route_steps_com_pares.append(step_publicado)
                    continue
                acoes_publicadas_validas = []
                for acao_publicada, tarefa_publicada in (step_publicado.get('actions', []) or []):
                    demanda_id_publicada = str(tarefa_publicada.get('id', '') or '')
                    coleta_orfa_pendente = (
                        acao_publicada == 'COLETAR'
                        and demanda_id_publicada not in dict_concluidos_torre
                        and demanda_id_publicada not in ids_entregas_publicadas
                    )
                    if coleta_orfa_pendente:
                        coletas_orfas_removidas.append(tarefa_publicada)
                    else:
                        acoes_publicadas_validas.append((acao_publicada, tarefa_publicada))
                if acoes_publicadas_validas:
                    step_publicado_valido = step_publicado.copy()
                    step_publicado_valido['actions'] = acoes_publicadas_validas
                    route_steps_com_pares.append(step_publicado_valido)
            if coletas_orfas_removidas:
                _registrar_adiadas(coletas_orfas_removidas)
            route_steps = route_steps_com_pares

            route_steps = consolidar_coletas_base_na_preparacao(
                route_steps, ajustes_manuais, ponto_saida
            )
            route_steps = aplicar_ordem_manual_route_steps(route_steps, ajustes_manuais)
            if route_steps:
                route_steps[0]['_motor_rota_versao'] = ROTA_ENGINE_VERSION
            st.session_state['_rota_locais_repetidos_carregada'] = detectar_locais_repetidos_rota(
                route_steps, ponto_saida
            )

            # Recalcula a quilometragem a partir das etapas que realmente ficaram
            # na rota; uma antiga "PARADA: ESCRITÓRIO" absorvida pela preparação
            # não pode continuar somando quilômetros.
            total_km = sum(
                float(s.get('dist', 0) or 0)
                for s in route_steps
                if s.get('type') in {'stop', 'return'}
            )

            coords_ordenadas_rota = [locais_dict[ponto_saida]]
            for step in route_steps:
                if step.get("destino") in locais_dict: coords_ordenadas_rota.append(locais_dict[step.get("destino")])
            geometria_rota, geometria_viaria = buscar_geometria_rota(coords_ordenadas_rota, horario_partida_matriz)
            geometria_rota = normalizar_geometria_mapa(geometria_rota, coords_ordenadas_rota)
            if len(geometria_rota) < 2 and len(coords_ordenadas_rota) > 1:
                geometria_rota = [list(map(float, p)) for p in coords_ordenadas_rota]
                geometria_viaria = False
            
            st.session_state['rota_gerada'] = True
            st.session_state['route_steps'] = route_steps
            st.session_state['total_km'] = total_km
            st.session_state['locais_dict'] = locais_dict
            st.session_state['p_saida'] = ponto_saida
            st.session_state['horario_conclusao_min'] = current_time
            st.session_state['geometria_rota'] = geometria_rota
            st.session_state['geometria_viaria'] = geometria_viaria
            st.session_state['data_rota'] = DATA_REF_ROTA_STR

            execute_db(
                "INSERT INTO rota_ativa (id, data_rota, json_route, json_locais, json_geometria, json_enderecos, total_km, fonte_matriz, horario_matriz, json_ajustes_manuais) "
                "VALUES (1, :data, :route, :locs, :geom, :end, :km, :fonte, :horario, :ajustes) "
                "ON CONFLICT (id) DO UPDATE SET data_rota=EXCLUDED.data_rota, json_route=EXCLUDED.json_route, json_locais=EXCLUDED.json_locais, "
                "json_geometria=EXCLUDED.json_geometria, json_enderecos=EXCLUDED.json_enderecos, total_km=EXCLUDED.total_km, "
                "fonte_matriz=EXCLUDED.fonte_matriz, horario_matriz=EXCLUDED.horario_matriz, json_ajustes_manuais=EXCLUDED.json_ajustes_manuais",
                {
                    "data": DATA_REF_ROTA_STR, "route": json.dumps(route_steps), "locs": json.dumps(locais_dict),
                    "geom": json.dumps(geometria_rota), "end": json.dumps(enderecos_dict), "km": total_km,
                    "fonte": fonte_matriz, "horario": horario_partida_matriz.strftime("%d/%m/%Y %H:%M"),
                    "ajustes": json.dumps(ajustes_manuais, ensure_ascii=False),
                },
            )
            try:
                carregar_ajustes_manuais_rota.clear()
            except Exception:
                pass

    if st.session_state.get('rota_gerada', False):
        route_steps, total_km, locais_dict = st.session_state['route_steps'], st.session_state['total_km'], st.session_state['locais_dict']
        enderecos_dict, p_saida = st.session_state.get('enderecos_dict', {}), st.session_state['p_saida']
        ajustes_manuais_atual = carregar_ajustes_manuais_rota(DATA_REF_ROTA_STR)

        # Saneia rotas antigas já persistidas antes de renderizar. É justamente o
        # caso do print em que havia "PARADA: ESCRITÓRIO" com a coleta da UNIFOR:
        # a coleta é fundida na preparação e a entrega continua na UNIFOR.
        _route_antes_normalizacao = json.dumps(route_steps, ensure_ascii=False, sort_keys=True, default=str)
        route_steps = consolidar_coletas_base_na_preparacao(
            route_steps, ajustes_manuais_atual, p_saida
        )
        route_steps = aplicar_ordem_manual_route_steps(route_steps, ajustes_manuais_atual)
        _route_depois_normalizacao = json.dumps(route_steps, ensure_ascii=False, sort_keys=True, default=str)
        st.session_state['route_steps'] = route_steps

        if _route_depois_normalizacao != _route_antes_normalizacao:
            # Atualiza somente o JSON da rota; não regrava cache/Trello nem cria
            # uma nova rota. Na próxima recálculo, a mesma regra já atua na origem.
            execute_db(
                "UPDATE rota_ativa SET json_route=:route WHERE id=1 AND data_rota=:data",
                {"route": json.dumps(route_steps, ensure_ascii=False), "data": DATA_REF_ROTA_STR},
            )

        # A regra acima é genérica: não depende de data, obra ou nome UNIFOR.
        # Qualquer COLETA arrastada para a base vira PREPARAÇÃO e nunca uma
        # parada operacional separada.

        # Reconstitui a lista em toda abertura, inclusive depois de reiniciar a
        # sessão: uma demanda de hoje nunca some só porque não coube na rota salva.
        demandas_fora_rota = {
            str(t.get('id', '') or ''): t
            for t in (st.session_state.get('demandas_adiadas') or [])
            if str(t.get('id', '') or '')
        }
        ids_na_rota_final = {
            str(tarefa.get('id', '') or '')
            for etapa in route_steps
            for _acao, tarefa in (etapa.get('actions', []) or [])
            if str(tarefa.get('id', '') or '')
        }
        if isinstance(df_ativos, pd.DataFrame) and not df_ativos.empty and 'Urgência' in df_ativos.columns:
            df_hoje_fora = df_ativos[
                df_ativos['Urgência'].astype(str).str.contains(r'HOJE|VENCIDA', case=False, na=False)
                & ~df_ativos['id'].astype(str).isin(ids_na_rota_final)
            ]
            for tarefa_fora in df_hoje_fora.to_dict('records'):
                demanda_id_fora = str(tarefa_fora.get('id', '') or '')
                if demanda_id_fora:
                    demandas_fora_rota[demanda_id_fora] = tarefa_fora

        if demandas_fora_rota:
            demandas_adiadas = list(demandas_fora_rota.values())
            qtd_adiadas = len(demandas_adiadas)
            avisos_compactos = [f"{qtd_adiadas} fora da rota até 17h"]
            if st.session_state.get('retorno_omitido_expediente'):
                avisos_compactos.append("conclusão após 17h preservada")
            st.caption("⏰ " + " • ".join(avisos_compactos))
            with st.expander(
                f"⏭️ FORA DA ROTA ATÉ 17H · {qtd_adiadas} "
                f"{plural_pt(qtd_adiadas, 'demanda', 'demandas')} — ver detalhes",
                expanded=False,
            ):
                df_adiadas_compacto = pd.DataFrame([
                    {
                        "Demanda": str(t.get('Obra', 'Demanda sem nome') or 'Demanda sem nome'),
                        "Prazo": str(t.get('Urgência', 'sem prazo') or 'sem prazo'),
                        "Percurso": (
                            f"{canonicalizar_ponto_rota(t.get('Origem', ''))} → "
                            f"{canonicalizar_ponto_rota(t.get('Destino', ''))}"
                        ),
                    }
                    for t in demandas_adiadas
                ])
                st.dataframe(
                    df_adiadas_compacto,
                    use_container_width=True,
                    hide_index=True,
                    height=min(320, 38 + 35 * qtd_adiadas),
                )

        if not demandas_fora_rota and st.session_state.get('retorno_omitido_expediente'):
            st.caption("⏰ Conclusão após 17h preservada • nenhuma nova parada foi criada")
        
        df_torre = carregar_conclusoes_rota(DATA_REF_ROTA_STR)
        dict_concluidos_torre = dict(zip(df_torre['id'].astype(str), df_torre['hora_conclusao']))
        try:
            dict_checkins_torre = filtrar_checkins_da_rota(route_steps, carregar_checkins_davi(DATA_REF_ROTA_STR))
        except Exception:
            dict_checkins_torre = {}

        # Comprovantes da rota: usados pela Torre para mostrar status e permitir reabertura.
        try:
            garantir_tabela_comprovantes_davi()
            comprovantes_torre = carregar_resumo_comprovantes_davi(DATA_REF_ROTA_STR)
        except Exception:
            comprovantes_torre = {}
        
        hora_inicio_real = obter_hora_inicio_rota(DATA_REF_ROTA_STR)
        
        df_paradas = carregar_paradas_rastreadas_rota(DATA_REF_ROTA_STR, PLACA_DAVI)

        route_steps = atualizar_tempos_por_parada(route_steps, ponto_saida)
        # Revalida também rotas antigas carregadas do Supabase: se um trecho de 4–5 km
        # veio como 6–8 min, o ETA passa a usar um tempo operacional plausível.
        route_steps = atualizar_tempos_deslocamento_operacionais(route_steps, hora_inicio_real)
        route_steps, final_dyn_min = aplicar_tempos_dinamicos(route_steps, dict_concluidos_torre, hora_inicio_real)

        # Uma rota que cabia às 08h pode deixar de caber depois de atrasos reais.
        # Se o ETA atual ultrapassar 17h e ainda houver demanda pendente, refazemos
        # automaticamente o restante a partir de AGORA. O motor de planejamento
        # então remove do trajeto o que não couber no expediente e deixa para o
        # próximo planejamento. Conclusões reais já registradas são preservadas.
        pendencias_na_rota = any(
            step.get('type') == 'stop'
            and any(
                str(tarefa.get('id', '')) not in dict_concluidos_torre
                for _acao, tarefa in (step.get('actions', []) or [])
            )
            for step in route_steps
        )

        # Em uma sessão recém-aberta, a rota salva aparece antes de carregarmos o
        # quadro inteiro do Trello para manter o site rápido. Porém, duas situações
        # exigem conhecer imediatamente as demandas ativas: (1) o ETA ficou acima
        # das 17h; (2) o Davi terminou o roteiro antes das 17h e pode receber mais
        # trabalho. Nesses casos usamos SOMENTE o cache do Supabase (sem chamada
        # externa e sem geocodificação) e recalculamos a rota.
        _agora_min_hidratacao = AGORA_REAL.hour * 60 + AGORA_REAL.minute
        _precisa_demandas_ativas = (
            DATA_REF_ROTA_DATE == AGORA_REAL.date()
            and _agora_min_hidratacao < LIMITE_EXPEDIENTE_DAVI_MIN
            and isinstance(df_ativos, pd.DataFrame)
            and df_ativos.empty
            and (final_dyn_min > LIMITE_EXPEDIENTE_DAVI_MIN or not pendencias_na_rota)
        )
        if _precisa_demandas_ativas and not st.session_state.get('_tentou_hidratar_demandas_cache_turno'):
            st.session_state['_tentou_hidratar_demandas_cache_turno'] = True
            if sincronizar_demandas(forcar=False, geocodificar=False, somente_cache=True):
                st.session_state['_recalcular_rota_automatico'] = True
                st.session_state['_mensagem_ajuste_rota'] = (
                    '🕒 O planejamento foi atualizado com as demandas ativas para respeitar '
                    'o expediente até as 17h e aproveitar o tempo restante.'
                )
                st.rerun()
        elif not df_ativos.empty:
            st.session_state.pop('_tentou_hidratar_demandas_cache_turno', None)

        # Se o Davi terminou tudo o que estava no roteiro antes das 17h, não o
        # mandamos automaticamente encerrar o dia enquanto ainda houver demandas
        # ativas. O sistema tenta montar uma extensão do roteiro a partir de AGORA.
        # A própria simulação de expediente abaixo só aceita uma nova parada quando
        # deslocamento + atendimento + retorno (se habilitado) ainda couberem até 17h.
        ids_ativos_restantes = sorted({
            str(valor)
            for valor in (df_ativos['id'].tolist() if isinstance(df_ativos, pd.DataFrame) and 'id' in df_ativos.columns else [])
            if str(valor) and str(valor) not in dict_concluidos_torre
        })
        assinatura_ativos_restantes = '|'.join(ids_ativos_restantes)
        agora_min_turno = AGORA_REAL.hour * 60 + AGORA_REAL.minute
        rota_operacional_concluida = not pendencias_na_rota

        if (
            DATA_REF_ROTA_DATE == AGORA_REAL.date()
            and agora_min_turno < LIMITE_EXPEDIENTE_DAVI_MIN
            and rota_operacional_concluida
            and ids_ativos_restantes
            and st.session_state.get('_ultima_expansao_turno_solicitada') != assinatura_ativos_restantes
        ):
            # Marca antes do rerun para impedir loop caso nenhuma das demandas
            # restantes seja viável dentro do tempo disponível. Uma mudança no
            # conjunto de cartões ativos gera outra assinatura e libera nova tentativa.
            st.session_state['_ultima_expansao_turno_solicitada'] = assinatura_ativos_restantes
            st.session_state['_recalcular_rota_automatico'] = True
            st.session_state['_mensagem_ajuste_rota'] = (
                '🕒 O roteiro atual foi concluído antes das 17h. O sistema está ' 
                'aproveitando o tempo restante e buscando novas demandas que ainda ' 
                'caibam no expediente.'
            )
            st.rerun()

        # Se existe novamente uma etapa pendente, a expansão funcionou. Removemos o
        # bloqueio para que, quando esse novo lote for concluído e a lista de ativos
        # mudar, o expediente possa ser preenchido outra vez.
        if pendencias_na_rota:
            st.session_state.pop('_ultima_expansao_turno_solicitada', None)

        if (
            DATA_REF_ROTA_DATE == AGORA_REAL.date()
            and final_dyn_min > LIMITE_EXPEDIENTE_DAVI_MIN
            and pendencias_na_rota
            and not df_ativos.empty
        ):
            _ids_pendentes_limite = sorted({
                str(tarefa.get('id', ''))
                for step in route_steps
                if step.get('type') == 'stop'
                for _acao, tarefa in (step.get('actions', []) or [])
                if str(tarefa.get('id', '')) not in dict_concluidos_torre
            })
            chave_recalculo_17h = (
                f"{DATA_REF_ROTA_STR}-{AGORA_REAL.strftime('%H:%M')}-"
                f"{int(round(final_dyn_min))}-{'|'.join(_ids_pendentes_limite)}"
            )
            if st.session_state.get('_ultimo_recalculo_limite_17h') != chave_recalculo_17h:
                st.session_state['_ultimo_recalculo_limite_17h'] = chave_recalculo_17h
                st.session_state['_recalcular_rota_automatico'] = True
                st.session_state['_mensagem_ajuste_rota'] = (
                    "⏰ A previsão ultrapassou 17h. O restante da rota foi recalculado "
                    "a partir do horário atual e o que não couber ficará para o próximo planejamento."
                )
                st.rerun()

        # O mapa/resumo ocupa o topo e o roteiro vem abaixo em largura total.
        # Dois containers preservam o código existente sem criar a coluna vazia
        # que aparecia quando as paradas eram muito mais altas que o mapa.
        col_mapa = st.container()
        col_paradas = st.container()
        with col_paradas:
            st.markdown(
                f'<div class="aproar-industrial-heading"><h2>Paradas</h2>'
                f'<span>{sum(1 for etapa in route_steps if etapa.get("type") == "stop")} NA ROTA</span></div>',
                unsafe_allow_html=True,
            )
            st.caption(f"🕖 Expediente: das 07:00 às 17:00  •  🚚 Início da rota do Davi: {hora_inicio_real}")
            st.caption("✅ As demandas concluídas ficam reunidas em uma tabela compacta; abaixo aparecem somente as etapas ainda pendentes.")

            # A marca invisível colocada dentro de uma etapa concluída pelo Davi
            # acende a borda do próprio cartão, sem criar um painel separado.
            st.markdown("""
                <style>
                    div[data-testid="stVerticalBlockBorderWrapper"]:has(.davi-etapa-feita) {
                        border: 2px solid #39ff88 !important;
                        background: linear-gradient(145deg, rgba(34,197,94,.10), rgba(13,16,37,.78)) !important;
                        outline: 1px solid rgba(134,239,172,.85);
                        outline-offset: 1px;
                        box-shadow: 0 0 8px rgba(57,255,136,.95),
                                    0 0 22px rgba(34,197,94,.55),
                                    inset 0 0 18px rgba(34,197,94,.08);
                        animation: pulso-davi 2.2s ease-in-out infinite;
                    }
                    .davi-etapa-feita { display:none; }
                    .selo-davi-feita {
                        display:inline-block; margin:7px 0 11px; padding:5px 10px;
                        color:#d1fae5; background:rgba(22,163,74,.24);
                        border:1px solid rgba(74,222,128,.65); border-radius:999px;
                        font-size:12px; font-weight:800;
                    }
                    @keyframes pulso-davi {
                        0%, 100% { box-shadow:0 0 7px rgba(57,255,136,.75), 0 0 18px rgba(34,197,94,.38), inset 0 0 18px rgba(34,197,94,.06); }
                        50% { box-shadow:0 0 12px rgba(57,255,136,1), 0 0 30px rgba(34,197,94,.68), inset 0 0 22px rgba(34,197,94,.12); }
                    }
                </style>
            """, unsafe_allow_html=True)

            fonte_matriz_exibicao = st.session_state.get('fonte_matriz_rota', 'OSRM — rota viária')
            horario_matriz_exibicao = st.session_state.get('horario_matriz_rota', '')
            referencia_txt = f" • referência {horario_matriz_exibicao}" if horario_matriz_exibicao else ""
            st.caption(f"🛣️ Otimização viária: **{fonte_matriz_exibicao}**{referencia_txt} • ETAs com validação operacional de trecho")

            hora_atual_str = AGORA_REAL.strftime("%H:%M")
            nova_previsao_str = format_mins_to_time(final_dyn_min)
            renderizar_banner_eta(hora_atual_str, nova_previsao_str, final_dyn_min)

            st.caption(
                "🧭 Roteiro completo em ordem operacional: cada demanda aparece na coleta e novamente na entrega."
            )

            _msg_ajuste = st.session_state.pop("_mensagem_ajuste_rota", "")
            if _msg_ajuste:
                st.caption(f"ℹ️ {_msg_ajuste}")

            with st.expander("✋ Ajustar rota manualmente — arraste as demandas", expanded=False):
                st.caption(
                    "Arraste uma COLETA ou ENTREGA para cima ou para baixo, ou solte-a dentro de outra parada. "
                    "O ajuste fica salvo no Supabase e continua valendo quando o Trello atualizar a rota."
                )
                _payload_editor, _altura_editor = construir_editor_arrastavel_rota(
                    route_steps, p_saida, carregar_ajustes_manuais_rota(DATA_REF_ROTA_STR)
                )
                _evento_drag = renderizar_editor_arrastavel_rota(
                    _payload_editor, _altura_editor, key=f"editor_drag_rota_{DATA_REF_ROTA_STR}"
                )
                if isinstance(_evento_drag, dict) and _evento_drag.get("nonce"):
                    _nonce_drag = str(_evento_drag.get("nonce"))
                    if _nonce_drag != st.session_state.get("_ultimo_evento_drag_rota"):
                        st.session_state["_ultimo_evento_drag_rota"] = _nonce_drag
                        try:
                            _movimento_drag_ok = registrar_movimento_manual_rota(
                                DATA_REF_ROTA_STR, route_steps,
                                _evento_drag.get("demanda_id", ""),
                                _evento_drag.get("acao", ""),
                                _evento_drag.get("destino", ""),
                                int(_evento_drag.get("ordem", 0) or 0), p_saida,
                            )
                            if _movimento_drag_ok:
                                # Reflete a mudança imediatamente no roteiro atual, antes mesmo
                                # do recálculo viário completo. Assim editor, cartões e rota não
                                # ficam divergentes após o drop.
                                _rota_pos_drag = aplicar_movimento_manual_route_steps_imediato(
                                    route_steps,
                                    _evento_drag.get("demanda_id", ""),
                                    _evento_drag.get("acao", ""),
                                    _evento_drag.get("destino", ""),
                                    int(_evento_drag.get("ordem", 0) or 0),
                                    p_saida,
                                )
                                _ajustes_pos_drag = carregar_ajustes_manuais_rota(DATA_REF_ROTA_STR)
                                _rota_pos_drag = consolidar_coletas_base_na_preparacao(
                                    _rota_pos_drag, _ajustes_pos_drag, p_saida
                                )
                                _rota_pos_drag = aplicar_ordem_manual_route_steps(
                                    _rota_pos_drag, _ajustes_pos_drag
                                )
                                st.session_state["route_steps"] = _rota_pos_drag
                                try:
                                    execute_db(
                                        "UPDATE rota_ativa SET json_route=:route WHERE id=1 AND data_rota=:data",
                                        {
                                            "route": json.dumps(_rota_pos_drag, ensure_ascii=False),
                                            "data": DATA_REF_ROTA_STR,
                                        },
                                    )
                                except Exception:
                                    pass
                                st.session_state["_recalcular_rota_automatico"] = True
                                st.session_state["_mensagem_ajuste_rota"] = "✅ Demanda movida. Roteiro, horários e mapa foram atualizados com o ajuste manual."
                                st.rerun()
                            else:
                                st.warning("Não foi possível aplicar esse movimento. Atualize a rota e tente novamente.")
                        except Exception as _erro_drag_rota:
                            st.warning(f"Não foi possível mover essa demanda agora: {_erro_drag_rota}")
                if st.button("♻️ Limpar ajustes manuais e voltar ao automático", key="limpar_ajustes_rota_manual"):
                    limpar_ajustes_manuais_rota(DATA_REF_ROTA_STR)
                    st.session_state["_recalcular_rota_automatico"] = True
                    st.session_state["_mensagem_ajuste_rota"] = "♻️ Ajustes manuais removidos. A rota voltou para o planejamento automático."
                    st.rerun()
            
            texto_whatsapp = f"🚚 *ROTEIRO DE LOGÍSTICA - DAVI*\n📅 Data: {DATA_REF_ROTA_STR}\n🕖 Expediente: das 07:00 às 17:00\n🚚 Início da rota do Davi: {hora_inicio_real}\n🚗 Veículo: {veiculo_selecionado.split('(')[0].strip()}\n\n"
            
            num_parada = 1
            for i, step in enumerate(route_steps):
                if step['type'] == 'lunch':
                    st.warning(f"🍔 **Pausa para almoço** (previsão: {step['dyn_chegada']} às {step['dyn_saida']})")
                    texto_whatsapp += f"🍔 Almoço: {step['dyn_chegada']} às {step['dyn_saida']}\n\n"
                    continue
                if step['type'] == 'return':
                    st.info(f"🏁 **Retorno à base:** {step['destino']} (Chegada prevista: {step['dyn_chegada']})")
                    texto_whatsapp += f"🏁 Retorno: {step['destino']} ({step['dyn_chegada']})\n"
                    continue

                is_start = (i == 0 and step['destino'] == p_saida)
                acoes_etapa_torre = list(step.get('actions', []) or [])
                if not acoes_etapa_torre:
                    if not is_start:
                        num_parada += 1
                    continue

                # A sequência física é preservada inteira, inclusive depois da
                # baixa: COLETA permanece na origem e ENTREGA permanece no destino.
                step = dict(step)
                step['actions'] = acoes_etapa_torre
                ids_acoes_etapa = {
                    str(tarefa_etapa.get('id', '') or '')
                    for _acao_etapa, tarefa_etapa in acoes_etapa_torre
                    if str(tarefa_etapa.get('id', '') or '')
                }
                etapa_totalmente_concluida = bool(ids_acoes_etapa) and all(
                    demanda_id_etapa in dict_concluidos_torre
                    for demanda_id_etapa in ids_acoes_etapa
                )
                step['is_concluded'] = etapa_totalmente_concluida
                endereco_db = enderecos_dict.get(step['destino'], "")
                link_parada = endereco_db if endereco_db.startswith("http") else f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(endereco_db)}" if endereco_db else f"https://www.google.com/maps/dir/?api=1&destination={locais_dict[step['destino']][0]},{locais_dict[step['destino']][1]}"

                tipos_acoes_etapa = []
                if any(acao_etapa == 'COLETAR' for acao_etapa, _ in acoes_etapa_torre):
                    tipos_acoes_etapa.append('COLETA')
                if any(acao_etapa == 'ENTREGAR' for acao_etapa, _ in acoes_etapa_torre):
                    tipos_acoes_etapa.append('ENTREGA')
                resumo_acoes_etapa = ' + '.join(tipos_acoes_etapa) or 'ETAPA'
                if etapa_totalmente_concluida and not is_start:
                    titulo_etapa_recolhida = (
                        f"✅ PARADA {num_parada}: {step['destino']} · {resumo_acoes_etapa} — concluída às {step['dyn_saida']}"
                    )
                    contexto_etapa_torre = st.expander(titulo_etapa_recolhida, expanded=True)
                else:
                    contexto_etapa_torre = st.container(border=True)

                with contexto_etapa_torre:
                    if is_start:
                        st.markdown(f"<h3 style='margin:0; color:#e4e8f4;'>🏁 PREPARAÇÃO: {step['destino']}</h3>", unsafe_allow_html=True)
                        st.caption(f"🕖 Preparação planejada: **{step['chegada']} às {step['saida']}**")
                        texto_whatsapp += f"🏁 *PREPARAÇÃO: {step['destino']}* ({step['chegada']} às {step['saida']})\n"
                    else:
                        status_tempo = f"<span style='color: #16a34a; font-weight: 600;'>✅ Concluído às {step['dyn_saida']}</span>" if step.get('is_concluded') else f"<span style='color: #aeb7b4; font-weight: 600;'>⏳ Previsão atual: {step['dyn_chegada']} às {step['dyn_saida']}</span>"
                        st.markdown(f"<h3 style='margin:0; color:#e4e8f4;'>📍 PARADA {num_parada}: {step['destino']}</h3>", unsafe_allow_html=True)
                        st.caption(f"{status_tempo} | Trecho: {step['dist']:.1f} km", unsafe_allow_html=True)

                        status_real = obter_status_rastreio_local(df_paradas, step['destino'], DATA_REF_ROTA_STR)
                        if status_real:
                            if status_real.get('aberta'):
                                texto_real = f"📍 Chegada real: **{status_real['chegada']}**"
                                if status_real.get('duracao'):
                                    texto_real += f"  •  ⏱️ No local há **{status_real['duracao']}**"
                                st.markdown(texto_real)
                            else:
                                st.markdown(
                                    f"📍 Chegada: **{status_real['chegada']}**  •  "
                                    f"🚚 Saída: **{status_real['saida']}**  •  "
                                    f"⏱️ Permanência: **{status_real['duracao']}**"
                                )

                        texto_whatsapp += f"📍 *PARADA {num_parada}: {step['destino']}* ({step['dyn_chegada']} às {step['dyn_saida']})\n🧭 *GPS:* {link_parada}\n"

                    checkin_davi = dict_checkins_torre.get(i)
                    if checkin_davi:
                        st.markdown(
                            f"<span class='davi-etapa-feita'></span>"
                            f"<span class='selo-davi-feita'>✓ Feita pelo Davi às {html_escape(checkin_davi['hora'])}</span>",
                            unsafe_allow_html=True,
                        )
                    
                    # Uma visita ao endereço tem uma única permanência estimada.
                    # Várias demandas no mesmo local compartilham esse tempo.
                    if not is_start and step.get('actions'):
                        tempo_parada_torre = int(round(float(step.get('tempo_local', 0) or 0)))
                        fonte_parada_torre = str(step.get('tempo_local_fonte', 'média operacional') or 'média operacional')
                        st.caption(
                            f"⏱️ **Permanência estimada no local: {tempo_parada_torre} min** • {fonte_parada_torre}. "
                            f"{len(step.get('actions', []))} {plural_pt(len(step.get('actions', [])), 'demanda desta parada compartilha', 'demandas desta parada compartilham')} esse período."
                        )

                    # Uma obra pode ter vários cards no Trello. Exibimos a obra uma
                    # única vez e mantemos, dentro dela, o status individual de cada card.
                    acoes_exibicao_torre = agrupar_acoes_por_obra_exibicao(step['actions'])
                    for indice_demanda, (acao, t) in enumerate(acoes_exibicao_torre, start=1):
                        eh_coleta_torre = acao == "COLETAR"
                        icone_torre = "📦" if eh_coleta_torre else "📬"
                        rotulo_torre = "COLETA" if eh_coleta_torre else "ENTREGA"
                        card_id_torre = str(t.get('id', ''))
                        ids_card_torre = [
                            str(demanda_id or '')
                            for demanda_id in (t.get('_ids_agrupados', []) or [card_id_torre])
                            if str(demanda_id or '')
                        ]
                        concluida = bool(ids_card_torre) and all(
                            demanda_id in dict_concluidos_torre for demanda_id in ids_card_torre
                        )
                        horas_baixa_card = [
                            str(dict_concluidos_torre[demanda_id])
                            for demanda_id in ids_card_torre
                            if demanda_id in dict_concluidos_torre
                        ]
                        hora_baixa_card = max(horas_baixa_card) if horas_baixa_card else ""
                        cards_agrupados_torre = list(t.get('_cards_agrupados', []) or [])
                        qtd_cards_baixados = sum(
                            1 for card_grupo in cards_agrupados_torre
                            if str(card_grupo.get('id', '') or '') in dict_concluidos_torre
                        )
                        qtd_cards_pendentes = max(0, len(cards_agrupados_torre) - qtd_cards_baixados)
                        materiais_torre = _separar_materiais_comprovante(t.get('Materiais', ''))
                        obra_torre_texto = str(t.get('Obra', 'Obra não informada') or 'Obra não informada')
                        obra_torre_html = html_escape(obra_torre_texto)
                        # Na preparação, o local físico já é o escritório; mostrar o
                        # DESTINO deixa claro para qual unidade/obra o material irá.
                        campo_unidade_torre = 'Destino' if (is_start or not eh_coleta_torre) else 'Origem'
                        rotulo_unidade_torre = 'DESTINO' if campo_unidade_torre == 'Destino' else 'ORIGEM'
                        unidade_torre = canonicalizar_ponto_rota(
                            t.get(campo_unidade_torre, '') or step.get('destino', '')
                        )
                        unidade_torre_html = html_escape(str(unidade_torre or 'Unidade não informada'))
                        classe_acao_torre = "coleta" if eh_coleta_torre else "entrega"
                        qtd_demandas_card = int(t.get('_qtd_demandas_agrupadas', 1) or 1)
                        complemento_grupo_torre = (
                            f" • {qtd_demandas_card} demandas agrupadas"
                            if qtd_demandas_card > 1 else ""
                        )

                        # Status por OBRA, e não apenas por card. Isso deixa explícito
                        # quando a visita ficou parcial: ex. luva baixada e pistola pendente.
                        total_cards_grupo = len(cards_agrupados_torre)
                        grupo_parcial = total_cards_grupo > 1 and 0 < qtd_cards_baixados < total_cards_grupo
                        if total_cards_grupo > 1:
                            qtd_itens_grupo = len(materiais_torre)
                            if grupo_parcial:
                                icone_estado_grupo = "⚠️"
                                resumo_estado_grupo = (
                                    f"{qtd_cards_baixados}/{total_cards_grupo} baixados · "
                                    f"FALTA {qtd_cards_pendentes}"
                                )
                            elif qtd_cards_pendentes:
                                icone_estado_grupo = "⏳"
                                resumo_estado_grupo = f"{qtd_cards_pendentes} pendentes"
                            else:
                                icone_estado_grupo = "✅"
                                resumo_estado_grupo = f"{total_cards_grupo}/{total_cards_grupo} baixados"
                            contexto_demanda_torre = st.expander(
                                f"{icone_estado_grupo} {rotulo_torre} · {obra_torre_texto} · "
                                f"{resumo_estado_grupo} · {qtd_itens_grupo} {plural_pt(qtd_itens_grupo, 'item', 'itens')}",
                                expanded=grupo_parcial,
                            )
                        elif concluida and not etapa_totalmente_concluida:
                            contexto_demanda_torre = st.expander(
                                f"✅ {rotulo_torre} · {obra_torre_texto} · baixa às {hora_baixa_card}",
                                expanded=False,
                            )
                        else:
                            contexto_demanda_torre = st.container(border=True)

                        with contexto_demanda_torre:
                            st.markdown(
                                f"""
                                <div class="aproar-stop-header">
                                    <span class="aproar-stop-action {classe_acao_torre}">{icone_torre} {rotulo_torre}</span>
                                    <div class="aproar-stop-copy">
                                        <strong>{obra_torre_html}</strong>
                                        <small>{rotulo_unidade_torre} · {unidade_torre_html}{html_escape(complemento_grupo_torre)}</small>
                                    </div>
                                    <span class="aproar-stop-number">#{indice_demanda:02d}</span>
                                </div>
                                """,
                                unsafe_allow_html=True,
                            )

                            if len(cards_agrupados_torre) > 1:
                                blocos_cards_trello = []
                                for indice_card_grupo, card_grupo in enumerate(cards_agrupados_torre, start=1):
                                    card_grupo_id = str(card_grupo.get('id', '') or '')
                                    card_grupo_concluido = card_grupo_id in dict_concluidos_torre
                                    hora_card_grupo = str(dict_concluidos_torre.get(card_grupo_id, '') or '')
                                    cor_card_grupo = "#22c55e" if card_grupo_concluido else "#f59e0b"
                                    fundo_card_grupo = "rgba(34,197,94,.08)" if card_grupo_concluido else "rgba(245,158,11,.09)"
                                    status_card_grupo = (
                                        f"✅ BAIXA ÀS {html_escape(hora_card_grupo)}"
                                        if card_grupo_concluido
                                        else "⚠️ PENDENTE — FALTOU DAR BAIXA"
                                    )
                                    materiais_card_grupo = card_grupo.get('materiais', []) or []
                                    materiais_card_html = ''.join(
                                        f"<li>{html_escape(str(material_card))}</li>"
                                        for material_card in materiais_card_grupo
                                    ) or "<li>Material não informado</li>"
                                    resumo_material_card = " • ".join(str(m) for m in materiais_card_grupo) or "Material não informado"
                                    status_curto = (
                                        f"✅ {resumo_material_card} · baixa {hora_card_grupo}"
                                        if card_grupo_concluido
                                        else f"⚠️ FALTA: {resumo_material_card}"
                                    )
                                    blocos_cards_trello.append(
                                        f'<div style="margin:5px 0;padding:7px 9px;border-left:3px solid {cor_card_grupo};'
                                        f'border-radius:5px;background:{fundo_card_grupo};color:{cor_card_grupo};'
                                        f'font-size:11px;font-weight:800;">{html_escape(status_curto)}</div>'
                                    )
                                st.markdown(''.join(blocos_cards_trello), unsafe_allow_html=True)
                            elif materiais_torre:
                                linhas_materiais_torre = []
                                for material_torre in materiais_torre:
                                    nome_material_torre, quantidade_material_torre = _dividir_material_quantidade(material_torre)
                                    linhas_materiais_torre.append(
                                        f'<div class="aproar-material-row">'
                                        f'<span>{html_escape(nome_material_torre)}</span>'
                                        f'<strong>{html_escape(quantidade_material_torre)}</strong></div>'
                                    )
                                st.markdown(
                                    '<div class="aproar-material-table">' + ''.join(linhas_materiais_torre) + '</div>',
                                    unsafe_allow_html=True,
                                )
                            else:
                                st.markdown(
                                    '<div class="aproar-material-table"><div class="aproar-material-row">'
                                    '<span>Material não informado</span><strong>—</strong></div></div>',
                                    unsafe_allow_html=True,
                                )

                            if concluida:
                                st.success(f"✅ Baixa registrada às {hora_baixa_card}")

                            # Para entregas, exibe o comprovante e permite reabri-lo sem apagar as fotos.
                            if not eh_coleta_torre and card_id_torre:
                                chave_comp_torre = _nome_seguro_comprovante(card_id_torre, 40)
                                estado_comp_torre = comprovantes_torre.get(chave_comp_torre, {})
                                fotos_comp_torre = estado_comp_torre.get("fotos", []) or []
                                if estado_comp_torre.get("finalizado"):
                                    recebedor_comp_torre = str(estado_comp_torre.get("recebedor", "") or "").strip()
                                    info_comp = f"📸 Comprovante finalizado • {len(fotos_comp_torre)} {plural_pt(len(fotos_comp_torre), 'foto', 'fotos')}"
                                    if recebedor_comp_torre:
                                        info_comp += f" • Recebedor: {recebedor_comp_torre}"
                                    st.caption(info_comp)
                                    if st.button(
                                        "↩️ Reabrir comprovante",
                                        key=f"reabrir_comprovante_torre_{DATA_REF_ROTA_STR}_{card_id_torre}_{i}_{indice_demanda}",
                                        help="Reabre a entrega no App do Davi sem apagar as fotos já enviadas.",
                                    ):
                                        try:
                                            definir_comprovante_finalizado_davi(DATA_REF_ROTA_STR, card_id_torre, False)
                                            st.success("✅ Comprovante reaberto. As fotos existentes foram preservadas.")
                                            st.rerun()
                                        except Exception as erro_reabrir:
                                            st.error(f"Não foi possível reabrir o comprovante: {erro_reabrir}")
                                elif fotos_comp_torre:
                                    qtd_fotos_comp = len(fotos_comp_torre)
                                    st.caption(
                                        f"📸 Comprovante em aberto • {qtd_fotos_comp} "
                                        f"{plural_pt(qtd_fotos_comp, 'foto já enviada', 'fotos já enviadas')}"
                                    )

                        texto_whatsapp += f" - {'✅ ' if concluida else ''}{acao.capitalize()}: {t['Materiais']} (Obra: {t['Obra']})\n"
                    
                    texto_whatsapp += "\n"
                    if not is_start: num_parada += 1

            horario_base_fim = format_time(st.session_state.get('horario_conclusao_min', 17*60))
            horario_dyn_fim = format_mins_to_time(final_dyn_min)
            
            situacao_eta = "🟢 dentro do expediente" if final_dyn_min <= LIMITE_EXPEDIENTE_DAVI_MIN else "⏰ após 17h — sem novas paradas"
            st.info(
                f"🕒 **Rota:** planejado **{horario_base_fim}** • previsão atual **{horario_dyn_fim}** "
                f"({situacao_eta}) • **{total_km:.1f} km**"
            )
            if valor_km_veiculo_proprio is not None:
                custo_estimado_veiculo_proprio = float(total_km) * valor_km_veiculo_proprio
                tipo_veiculo_proprio = "Moto" if "Moto Própria/Frete" in veiculo_selecionado else "Carro"
                custo_estimado_txt = f"R$ {custo_estimado_veiculo_proprio:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                valor_km_txt = f"R$ {valor_km_veiculo_proprio:.2f}/km".replace('.', ',')
                st.info(
                    f"💰 **Custeio estimado — {tipo_veiculo_proprio}:** "
                    f"{custo_estimado_txt} (**{valor_km_txt}**)."
                )

            if len(route_steps) > 1:
                waypts_addr = []
                for s in route_steps:
                    if s['type'] != 'lunch':
                        addr = enderecos_dict.get(s['destino'], "")
                        waypts_addr.append(urllib.parse.quote(addr) if addr and not addr.startswith("http") else f"{locais_dict[s['destino']][0]},{locais_dict[s['destino']][1]}")
                link_maps = f"https://www.google.com/maps/dir/?api=1&origin={waypts_addr[0]}&destination={waypts_addr[-1]}&travelmode=driving"
                if len(waypts_addr) > 2: link_maps += f"&waypoints={'|'.join(waypts_addr[1:-1][:9])}"
                texto_whatsapp += f"\n🗺️ *LINK DO ROTEIRO COMPLETO:*\n{link_maps}\n"

            @fragmento_independente
            def formulario_fechamento_rota():
                with st.form("fechamento_km_rota"):
                    st.caption("Informe o KM real e o veículo usado.")
                    total_acoes = sum(len(step.get('actions', [])) for step in route_steps if step['type'] != 'lunch')
                    acoes_concluidas = sum(1 for step in route_steps for acao, t in step.get('actions', []) if str(t.get('id', '')) in dict_concluidos_torre)
                    
                    status_fechamento = "✅" if acoes_concluidas >= total_acoes else "⚠️"
                    st.caption(f"{status_fechamento} Baixas da rota: {acoes_concluidas}/{total_acoes}")
                        
                    km_real = st.number_input("Quilometragem efetivamente rodada na rota", value=float(total_km), step=1.0)
                    veiculo_fechamento = st.selectbox("Qual carro rodou esta rota?", ["Strada", "L200"])
                    if st.form_submit_button("Registrar quilometragem no painel de custos"):
                        execute_db("INSERT INTO registro_km (data, km, obs, veiculo) VALUES (:data, :km, :obs, :veic)", {"data": DATA_REF_ROTA_STR, "km": km_real, "obs": f"Fechamento Automático ({acoes_concluidas}/{total_acoes})", "veic": veiculo_fechamento})
                        carregar_registro_km_df.clear()
                        st.success(f"✅ Quilometragem de {km_real:.1f} km registrada para o veículo {veiculo_fechamento} na nuvem!")

            with st.expander("💾 Fechar quilometragem da rota", expanded=False):
                formulario_fechamento_rota()

            url_geral, _ = obter_webhook_teams("Geral / Logística")

            @fragmento_independente
            def compartilhamento_rota():
                if url_geral:
                    if st.button("📢 Enviar roteiro ao grupo geral (Teams)", use_container_width=True):
                        resumo = f"O roteiro do Davi já está pronto.\n\n**Data da rota:** {DATA_REF_ROTA_STR}\n\n**Saída real do pátio (TIF-2123 - Strada):** {hora_inicio_real}\n\n**Previsão dinâmica de conclusão:** {nova_previsao_str}\n\n**Total de paradas:** {num_parada-1}\n\n**Quilometragem:** {total_km:.1f} km\n\n[Abrir GPS da rota completa]({link_maps})"
                        enviado, detalhe = disparar_teams(url_geral, "🚚 Roteiro diário atualizado!", resumo)
                        if enviado: st.success("✅ Roteiro enviado!")
                        else: st.error(f"Erro ao enviar: {detalhe}")

                st.text_area("📋 Texto pronto para WhatsApp", value=texto_whatsapp, height=150)

            compartilhamento_rota()

        with col_mapa:
            st.markdown(
                '<div class="aproar-industrial-heading"><h2>Mapa</h2><span>TRAJETO EM TEMPO REAL</span></div>',
                unsafe_allow_html=True,
            )
            # MAPA DA ROTA — OpenStreetMap não exige API key e permanece com as
            # cores cartográficas originais (vias, parques, água e bairros).
            m = folium.Map(location=[-3.7319, -38.5267], zoom_start=12, tiles=None)
            folium.TileLayer(
                tiles="https://tile.openstreetmap.org/{z}/{x}/{y}.png",
                attr="© OpenStreetMap contributors",
                name="OpenStreetMap — sem chave",
                overlay=False,
                control=False,
                max_zoom=19,
            ).add_to(m)

            # O enquadramento usa SEMPRE as posições reais. Os deslocamentos abaixo
            # existem somente para impedir que um número fique escondido por outro.
            pontos_reais_mapa = []
            if p_saida in locais_dict:
                pontos_reais_mapa.append([float(locais_dict[p_saida][0]), float(locais_dict[p_saida][1])])
            for i, step in enumerate(route_steps):
                if step.get('destino') in locais_dict and step.get('type') not in ['lunch', 'return'] and not (i == 0 and step.get('destino') == p_saida):
                    _lat_real, _lon_real = locais_dict[step['destino']]
                    pontos_reais_mapa.append([float(_lat_real), float(_lon_real)])

            def _escala_visual_mapa(pontos):
                if len(pontos) < 2:
                    return 0.85
                lat_ref = sum(p[0] for p in pontos) / len(pontos)
                span_lat_km = (max(p[0] for p in pontos) - min(p[0] for p in pontos)) * 111.0
                span_lon_km = (max(p[1] for p in pontos) - min(p[1] for p in pontos)) * 111.0 * max(math.cos(math.radians(lat_ref)), 0.2)
                span_km = max(span_lat_km, span_lon_km, 1.0)
                # Em um mapa de ~450 px, 30–34 px de marcador equivalem a cerca
                # de 6–8% da largura útil. Esse limite cresce junto com o zoom geral.
                return max(0.75, min(2.20, span_km * 0.065))

            distancia_visual_min = _escala_visual_mapa(pontos_reais_mapa)
            marcadores_posicionados = []

            def apply_offset(lat, lon):
                """Mantém cada número visível, deslocando somente o ícone quando necessário."""
                lat, lon = float(lat), float(lon)
                if not marcadores_posicionados:
                    marcadores_posicionados.append((lat, lon))
                    return lat, lon

                conflito = any(
                    calcular_distancia_km(lat, lon, p_lat, p_lon) < distancia_visual_min
                    for p_lat, p_lon in marcadores_posicionados
                )
                if not conflito:
                    marcadores_posicionados.append((lat, lon))
                    return lat, lon

                # Espalha os pontos em anéis ao redor da posição real. Como o
                # fit_bounds usa as coordenadas verdadeiras, esse afastamento não
                # faz o mapa dar zoom para fora novamente.
                for tentativa in range(1, 49):
                    anel = 1 + (tentativa - 1) // 12
                    angulo = math.radians(((tentativa - 1) % 12) * 30 + anel * 11)
                    raio_km = distancia_visual_min * (0.82 + 0.42 * (anel - 1))
                    dlat = (raio_km / 111.0) * math.sin(angulo)
                    dlon = (raio_km / (111.0 * max(math.cos(math.radians(lat)), 0.2))) * math.cos(angulo)
                    candidato = (lat + dlat, lon + dlon)
                    if all(
                        calcular_distancia_km(candidato[0], candidato[1], p_lat, p_lon) >= distancia_visual_min * 0.92
                        for p_lat, p_lon in marcadores_posicionados
                    ):
                        marcadores_posicionados.append(candidato)
                        return candidato

                # Contingência para aglomerações muito grandes.
                raio_km = distancia_visual_min * 1.8
                dlat = raio_km / 111.0
                candidato = (lat - dlat, lon + dlat)
                marcadores_posicionados.append(candidato)
                return candidato

            p_num = 1
            pos_base_visual = None
            if p_saida in locais_dict:
                pos_base_visual = apply_offset(*locais_dict[p_saida])

            # Traçado primeiro: fica por baixo dos marcadores e permanece visível.
            geometria_rota = st.session_state.get('geometria_rota') or []
            geometria_viaria = bool(st.session_state.get('geometria_viaria', False))
            coords_ordem_real = []
            if p_saida in locais_dict:
                coords_ordem_real.append(locais_dict[p_saida])
            coords_ordem_real.extend([
                locais_dict[s['destino']] for s in route_steps
                if s.get('destino') in locais_dict and s.get('type') != 'lunch'
            ])
            # Repara geometrias antigas gravadas com eixos invertidos pelo OSRM.
            geometria_rota = normalizar_geometria_mapa(geometria_rota, coords_ordem_real)
            if len(geometria_rota) < 2:
                geometria_rota, geometria_viaria = buscar_geometria_rota(coords_ordem_real)
                geometria_rota = normalizar_geometria_mapa(geometria_rota, coords_ordem_real)
            if len(geometria_rota) < 2 and len(coords_ordem_real) > 1:
                geometria_rota = [list(map(float, p)) for p in coords_ordem_real]
                geometria_viaria = False

            # Mantém a versão reparada em memória; ao recalcular a rota ela também
            # será gravada corretamente no banco.
            st.session_state['geometria_rota'] = geometria_rota
            st.session_state['geometria_viaria'] = geometria_viaria

            if len(geometria_rota) > 1:
                # Contorno claro + azul da referência sobre as cores reais do mapa.
                folium.PolyLine(geometria_rota, color="#FFFFFF", weight=9, opacity=0.88).add_to(m)
                folium.PolyLine(
                    geometria_rota,
                    color="#2563EB", weight=5.5, opacity=0.98,
                    dash_array=None if geometria_viaria else "9,7",
                    tooltip="Traçado viário da rota" if geometria_viaria else "Ligação aproximada entre as paradas",
                ).add_to(m)

            for i, step in enumerate(route_steps):
                if step.get('destino') in locais_dict and step.get('type') not in ['lunch', 'return'] and not (i == 0 and step.get('destino') == p_saida):
                    lat_orig, lon_orig = map(float, locais_dict[step['destino']])
                    lat, lon = apply_offset(lat_orig, lon_orig)
                    deslocado = calcular_distancia_km(lat_orig, lon_orig, lat, lon) > 0.01
                    if deslocado:
                        folium.PolyLine(
                            [[lat_orig, lon_orig], [lat, lon]],
                            color="#475569", weight=2.0, opacity=0.90, dash_array="4,5",
                            tooltip="O círculo foi afastado; a ponta da linha é o local real",
                        ).add_to(m)
                        folium.CircleMarker(
                            [lat_orig, lon_orig], radius=3, color="#475569", weight=1,
                            fill=True, fill_opacity=0.9, tooltip=f"Posição real — {step['destino']}"
                        ).add_to(m)

                    acoes = [a[0] for a in step.get('actions', [])]
                    tem_coleta, tem_entrega = "COLETAR" in acoes, "ENTREGAR" in acoes
                    fundo_marcador = "linear-gradient(90deg, #f59e0b 0 50%, #22c55e 50% 100%)" if (tem_coleta and tem_entrega) else "#f59e0b" if tem_coleta else "#22c55e"
                    popup_html = f"<b>Parada {p_num}: {html_escape(str(step['destino']))}</b><br>Previsão: {step.get('dyn_chegada', step.get('chegada', ''))}<br>Ação: {html_escape(' e '.join(sorted(set(acoes))).title())}"
                    folium.Marker(
                        [lat, lon], popup=folium.Popup(popup_html, max_width=280), tooltip=f"Parada {p_num}",
                        z_index_offset=1200 + p_num,
                        icon=folium.DivIcon(html=f'''<div style="background: {fundo_marcador}; color: white; border: 3px solid white; border-radius: 50%; width: 32px; height: 32px; display: flex; justify-content: center; align-items: center; font-weight: 900; box-shadow: 0 2px 7px rgba(0,0,0,0.65); font-size: 14px;">{p_num}</div>''')
                    ).add_to(m)
                    p_num += 1

            # Última posição real do Davi. Ela é atualizada pela consulta em
            # background acima; desenhar o caminhão nunca bloqueia o mapa.
            posicoes_gps_rota = st.session_state.get("_gps_rota_posicoes") or []
            placa_davi_normalizada = re.sub(r"[^A-Z0-9]", "", PLACA_DAVI.upper())
            posicao_davi = next(
                (
                    pos for pos in posicoes_gps_rota
                    if re.sub(r"[^A-Z0-9]", "", str(pos.get("Placa", "")).upper())
                    == placa_davi_normalizada
                ),
                None,
            )

            if posicao_davi:
                try:
                    lat_caminhao = float(posicao_davi.get("Latitude"))
                    lon_caminhao = float(posicao_davi.get("Longitude"))
                    coordenada_gps_valida = (
                        -90 <= lat_caminhao <= 90 and -180 <= lon_caminhao <= 180
                        and abs(lat_caminhao) + abs(lon_caminhao) > 0.01
                    )
                except (TypeError, ValueError):
                    coordenada_gps_valida = False

                if coordenada_gps_valida:
                    endereco_atual_davi = re.sub(
                        r"\s+", " ", str(posicao_davi.get("Endereço", "") or "")
                    ).strip(" ,-|—–") or "Endereço atual não informado"
                    endereco_atual_davi_html = html_escape(endereco_atual_davi)
                    popup_caminhao = (
                        "<div style='max-width:310px;white-space:normal;line-height:1.35'>"
                        f"📍 {endereco_atual_davi_html}</div>"
                    )
                    tooltip_caminhao = (
                        "<div style='min-width:245px;max-width:310px;white-space:normal;line-height:1.3;"
                        "font-size:12px;font-weight:700;color:#0f172a'>"
                        f"📍 {endereco_atual_davi_html}</div>"
                    )
                    folium.Marker(
                        [lat_caminhao, lon_caminhao],
                        popup=folium.Popup(popup_caminhao, max_width=310),
                        tooltip=tooltip_caminhao,
                        z_index_offset=5000,
                        icon=folium.DivIcon(
                            icon_size=(52, 52),
                            icon_anchor=(26, 26),
                            html='''
                                <div style="width:52px;height:52px;display:flex;align-items:center;justify-content:center;
                                            border-radius:50%;background:radial-gradient(circle at 38% 28%,#ffffff 0,#eff6ff 72%);
                                            border:3px solid #ffffff;box-shadow:0 0 0 2px #2563eb,
                                            0 4px 13px rgba(15,23,42,.48);box-sizing:border-box;">
                                    <svg viewBox="0 0 70 38" width="44" height="29" aria-label="Veículo do Davi"
                                         style="display:block;overflow:visible">
                                        <defs>
                                            <linearGradient id="pickupBody" x1="0" y1="0" x2="1" y2="1">
                                                <stop offset="0" stop-color="#60a5fa"/>
                                                <stop offset="1" stop-color="#1d4ed8"/>
                                            </linearGradient>
                                        </defs>
                                        <path d="M4 24V13h29l8-9h14l7 14h4c2 0 3 1 3 3v7H4z"
                                              fill="url(#pickupBody)" stroke="#1e3a8a" stroke-width="1.8"
                                              stroke-linejoin="round"/>
                                        <path d="M8 15h23v7H8z" fill="#bfdbfe" opacity=".92"/>
                                        <path d="M42 7h11l5 11H36z" fill="#dbeafe" stroke="#1e40af" stroke-width="1.4"/>
                                        <path d="M48 7v11M35 18h24" stroke="#1e40af" stroke-width="1.3"/>
                                        <path d="M5 25h63" stroke="#172554" stroke-width="2" stroke-linecap="round"/>
                                        <path d="M64 20h4" stroke="#fde68a" stroke-width="2.6" stroke-linecap="round"/>
                                        <circle cx="17" cy="28" r="6" fill="#0f172a" stroke="#ffffff" stroke-width="1.5"/>
                                        <circle cx="17" cy="28" r="2.3" fill="#94a3b8"/>
                                        <circle cx="55" cy="28" r="6" fill="#0f172a" stroke="#ffffff" stroke-width="1.5"/>
                                        <circle cx="55" cy="28" r="2.3" fill="#94a3b8"/>
                                    </svg>
                                </div>
                            ''',
                        ),
                    ).add_to(m)
                    pontos_reais_mapa.append([lat_caminhao, lon_caminhao])

            if len(pontos_reais_mapa) > 1:
                m.fit_bounds(pontos_reais_mapa, padding=(45, 45), max_zoom=14)
            if p_saida in locais_dict and pos_base_visual is not None:
                folium.Marker(
                    [pos_base_visual[0], pos_base_visual[1]],
                    popup=folium.Popup(f"<b>Saída/retorno: {html_escape(str(p_saida))}</b>", max_width=280),
                    z_index_offset=2500,
                    icon=folium.DivIcon(html=f'''<div style="background: linear-gradient(135deg, #2563eb, #1d4ed8); color: white; border: 3px solid #dbeafe; border-radius: 50%; width: 34px; height: 34px; display: flex; justify-content: center; align-items: center; box-shadow: 0 2px 8px rgba(0,0,0,0.55); font-size: 16px;">🏁</div>''')
                ).add_to(m)

            st_folium(
                m, height=540, use_container_width=True, returned_objects=[],
                key=f"mapa_rota_{DATA_REF_ROTA_STR}",
            )
            total_paradas_industrial = sum(
                1 for etapa in route_steps
                if etapa.get('type') == 'stop' and etapa.get('destino') != p_saida
            )
            distancia_industrial = f"{float(total_km):.1f}".replace('.', ',')
            veiculo_industrial = html_escape(veiculo_selecionado.split('(')[0].strip())
            if valor_km_veiculo_proprio is not None:
                custo_industrial = float(total_km) * float(valor_km_veiculo_proprio)
                custo_industrial_txt = f"R$ {custo_industrial:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            else:
                custo_industrial_txt = "Frota da empresa"

            st.markdown(
                f"""
                <div class="aproar-industrial-summary">
                    <section class="aproar-route-panel">
                        <div class="aproar-summary-title"><strong>Rota 01 — Davi</strong><span>PLANEJADA</span></div>
                        <div class="aproar-summary-times">
                            <div><span>Horário atual</span><strong>{hora_atual_str}</strong></div>
                            <div class="aproar-summary-line"><i></i></div>
                            <div><span>Término previsto</span><strong>{nova_previsao_str}</strong></div>
                        </div>
                        <div class="aproar-summary-data">
                            <div><span>Início</span><strong>{html_escape(hora_inicio_real)}</strong></div>
                            <div><span>Expediente</span><strong>07:00–17:00</strong></div>
                            <div><span>Paradas</span><strong>{total_paradas_industrial}</strong></div>
                            <div><span>Distância</span><strong>{distancia_industrial} km</strong></div>
                        </div>
                    </section>
                    <section class="aproar-fleet-panel">
                        <div class="aproar-summary-title"><strong>Frota</strong><span>EM ROTA</span></div>
                        <div class="aproar-fleet-body">
                            <span>Motorista</span><strong>Davi · {veiculo_industrial}</strong>
                            <small>Rota sincronizada com a central</small>
                            <div class="aproar-fleet-cost"><span>Custeio estimado</span><b>{custo_industrial_txt}</b></div>
                        </div>
                    </section>
                </div>
                """,
                unsafe_allow_html=True,
            )
            legenda_tracado = "trajeto viário" if geometria_viaria else "ligação de contingência entre as paradas"
            st.markdown(f"<div style='text-align: center; font-size: 14px; margin-top: 10px; color: #8da0b8;'><b>Legenda:</b> 🟡 Coleta | 🟢 Entrega | 🏁 Início/Retorno | 🟡🟢 Ambos<br><span style='font-size:12px;'>Azul = {legenda_tracado}. Linha cinza pontilhada = marcador afastado da posição real para não esconder outro número.</span></div>", unsafe_allow_html=True)

        df_relatorio_rota = montar_relatorio_rota(route_steps, dict_concluidos_torre)
        df_resumo_rota = pd.DataFrame([{
            "Data": DATA_REF_ROTA_STR,
            "Ponto de saída": p_saida,
            "Veículo": veiculo_selecionado.split('(')[0].strip(),
            "Valor por km (R$)": round(valor_km_veiculo_proprio, 2) if valor_km_veiculo_proprio is not None else None,
            "Custeio estimado da rota (R$)": round(float(total_km) * valor_km_veiculo_proprio, 2) if valor_km_veiculo_proprio is not None else None,
            "Estratégia": estrategia,
            "Distância planejada (km)": round(float(total_km), 2),
            "Início": hora_inicio_real,
            "Término previsto": format_mins_to_time(final_dyn_min),
            "Fonte viária": st.session_state.get('fonte_matriz_rota', 'OSRM — rota viária'),
        }])
        df_resumo_sequencial = montar_resumo_sequencial_rota(route_steps, p_saida, retornar_base=True)
        st.caption("🛣️ Os horários do resumo usam distância viária e piso operacional para deslocamentos urbanos, evitando estimativas excessivamente otimistas. 🍽️ A pausa de 1 hora para almoço aparece como etapa própria.")
        renderizar_exportador(
            f"Roteiro do Davi - {DATA_REF_ROTA_STR}",
            {"Resumo da rota": df_resumo_sequencial},
            "roteiro_do_davi", "roteiro",
        )
